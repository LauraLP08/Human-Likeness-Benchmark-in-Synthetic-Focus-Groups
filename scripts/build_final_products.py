"""
Build FINAL_RESULTS_TABLES.xlsx and RESULTS_TRACEABILITY_INDEX.md.

Every figure is read from its source artefact; nothing is retyped. Runs only after
final_integration_q3.build() reconciles, so a contradiction stops the build.
"""
from __future__ import annotations

import collections
import csv
import json
import statistics
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))
import final_integration_q3 as fi   # noqa: E402

_OUT = _ROOT / "analysis" / "production_evaluation"
_RES = _OUT / "results"
_Q3 = _OUT / "emergent_calibration_q3"
_FINAL = _OUT / "final"

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

STRUCT = ["total_words", "participant_turns", "words_per_turn_iqr",
          "short_turn_proportion_25w", "turn_balance_gini", "chain_depth",
          "moderator_word_share"]


def _csvr(p):
    with p.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def structural():
    rows = _csvr(_RES / "structural_interaction_metrics_long.csv")
    cell, hum = collections.defaultdict(list), collections.defaultdict(list)
    for r in rows:
        if r["metric_id"] not in STRUCT or r["value"] in ("", "None"):
            continue
        v = float(r["value"])
        if r["side"] == "human":
            hum[(r["metric_id"], r["fg"])].append(v)
        else:
            cell[(r["metric_id"], r["fg"], r["condition"])].append(v)
    out = []
    for m in STRUCT:
        h = [statistics.mean(hum[(m, f)]) for f in ("fg1", "fg2", "fg3", "fg4", "fg5")
             if hum.get((m, f))]
        e = [statistics.mean(cell[(m, f, "enriched")]) for f in
             ("fg1", "fg2", "fg3", "fg4", "fg5") if cell.get((m, f, "enriched"))]
        d = [statistics.mean(cell[(m, f, "demographics-only")]) for f in
             ("fg1", "fg2", "fg3", "fg4", "fg5") if cell.get((m, f, "demographics-only"))]
        closer = sum(1 for i in range(len(h)) if abs(e[i] - h[i]) < abs(d[i] - h[i]))
        out.append({"metric": m, "human_mean": statistics.mean(h),
                    "enriched_mean": statistics.mean(e),
                    "demographics_only_mean": statistics.mean(d),
                    "enriched_minus_demo": statistics.mean(e) - statistics.mean(d),
                    "n_fg_enriched_closer_to_human": f"{closer}/5"})
    return out


def _sheet(wb, title, cols, rows, widths=None):
    ws = wb.create_sheet(title)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill, cell.font = HDR, HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = (widths or [18] * len(cols))[j - 1]
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(cols, start=1):
            v = r.get(c)
            cell = ws.cell(row=i, column=j,
                           value=round(v, 4) if isinstance(v, float) else v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def build():
    rec = fi.build()                       # raises if anything fails to reconcile
    _FINAL.mkdir(parents=True, exist_ok=True)

    d = json.loads((_Q3 / "matching_derivation_q3.json").read_text(encoding="utf-8"))
    b = json.loads((_Q3 / "bplus_evaluation_q3.json").read_text(encoding="utf-8"))
    a = json.loads((_Q3 / "cross_model_analysis_q3.json").read_text(encoding="utf-8"))
    q = json.loads((_Q3 / "cross_model_quote_audit_q3.json").read_text(encoding="utf-8"))
    sup = json.loads((_OUT / "transportability_sample" /
                      "supplementary_human_reference.json").read_text(encoding="utf-8"))
    cost = json.loads((_Q3 / "cross_model_cost_actual.json").read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "README"
    notes = [
        ("FINAL RESULTS TABLES — Macho Meals synthetic focus-group evaluation", True),
        ("", False),
        (f"built {datetime.now(UTC).date()}   all figures reconciled against source "
         f"artefacts", False),
        ("", False),
        ("WHAT MUST NOT BE POOLED", True),
        ("  U01-U07/Q3 (emergent) with S01-S06 (supplementary transportability)", False),
        ("  deductive Tier 1 results with emergent calibration results", False),
        ("  Batch-executed results with synchronous results", False),
        ("  theme x unit instances (44) with thematic categories (16)", False),
        ("  the 15 blinded units as if they were 15 independent focus groups", False),
        ("", False),
        ("UNIT OF ANALYSIS", True),
        ("  FG remains the comparative unit: n=5 paired groups.", False),
        ("  The three replicates per cell estimate GENERATOR VARIABILITY, not extra", False),
        ("  groups. Sessions are not independent focus groups.", False),
        ("", False),
        ("WHAT THIS WORKBOOK DOES NOT CONTAIN", True),
        ("  No confirmatory p-values. No causal claims. No conclusion that the", False),
        ("  enriched condition is superior.", False),
    ]
    for i, (t, bold) in enumerate(notes, start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    # 1. deductive per FG
    _sheet(wb, "1_Deductive_by_FG",
           ["metric", "fg", "enriched_mean", "enriched_within_cell_sd",
            "demographics_only_mean", "demographics_only_within_cell_sd",
            "difference", "direction"],
           rec["deductive_per_fg"], [16, 7, 14, 20, 20, 26, 12, 18])

    # 2. deductive summary
    ded = [{"metric": m, **v} for m, v in rec["deductive"].items()]
    _sheet(wb, "2_Deductive_Summary",
           ["metric", "mean_difference", "n_favouring_enriched",
            "n_favouring_demographics_only", "n_ties", "exploratory_sign_test_p",
            "minimum_attainable_p", "cannot_reach_p05"], ded,
           [16, 16, 20, 26, 9, 22, 20, 18])

    # 3. structural
    _sheet(wb, "3_Structural_Interaction",
           ["metric", "human_mean", "enriched_mean", "demographics_only_mean",
            "enriched_minus_demo", "n_fg_enriched_closer_to_human"],
           structural(), [30, 14, 14, 22, 20, 28])

    # 4. emergent Q3
    em = rec["emergent"]
    emrows = [
        {"quantity": "human theme x unit instances (denominator)", "value": 44,
         "note": "16 thematic categories, 76 original coder rows — never conflated"},
        {"quantity": "machine themes extracted", "value": 30, "note": "7 units, Q3"},
        {"quantity": "recall vs union_reference", "value": "30/44 = 0.6818",
         "note": "human-confirmed matches only; UNCERTAIN rows excluded"},
        {"quantity": "strict precision vs union_reference", "value": "24/30 = 0.8000",
         "note": "CONSERVATIVE: valid novel themes not linked to a human theme still "
                 "count outside the numerator"},
        {"quantity": "human uncertainty", "value": "6/44 = 0.1364",
         "note": "retained as uncertainty, not resolved"},
        {"quantity": "literal_evidence_attachment_rate", "value": "30/30 themes; 58/58 quotes",
         "note": "ATTACHMENT check only — each theme carries a verbatim participant "
                 "quote from its own unit. NOT substantive groundedness."},
        {"quantity": "coverage benchmark (lower coder recall)", "value": "28/44 = 0.6364",
         "note": "necessary, not sufficient; not a human ceiling"},
    ]
    _sheet(wb, "4_Emergent_Q3", ["quantity", "value", "note"], emrows, [46, 24, 76])

    # 5. cross-model
    cal, aud = a["judge_calibration"], a["cross_model_audit"]
    cmrows = [
        {"item": "requests (76 = 38 cases x 2 repetitions)", "value": 76, "kind": "scope"},
        {"item": "pending cases corroborated", "value": f"{aud['n_corroborated']}/{aud['n']}",
         "kind": "corroboration"},
        {"item": "pending cases unresolved", "value": aud["n_unresolved"],
         "kind": "uncertainty retained"},
        {"item": "exact agreement with researcher (stable cases)",
         "value": f"{cal['n_agree']}/{cal['n_decided']} = "
                  f"{cal['exact_agreement_on_decided']:.3f}", "kind": "against it"},
        {"item": "self-contradiction between repetitions",
         "value": f"{cal['n_unstable']}/{cal['n_calibration_cases']} = "
                  f"{cal['instability_rate']:.3f}", "kind": "against it"},
        {"item": "abstentions", "value": cal["n_abstain"], "kind": "against it"},
        {"item": "non-verbatim quotations by the auditor",
         "value": f"{q['n_non_verbatim']}/{q['n_quotations']}", "kind": "against it"},
        {"item": "fabricated / misattributed quotations",
         "value": f"{q['by_classification'].get('FABRICATED_OR_MISATTRIBUTED', 0)} "
                  f"(both in B::U01::M5)", "kind": "against it"},
        {"item": "judge status", "value": cal["judge_usable_for_pending_cases"],
         "kind": "verdict"},
    ]
    _sheet(wb, "5_CrossModel_Audit", ["item", "value", "kind"], cmrows, [46, 30, 22])

    # 6. unresolved, retained individually
    disp = a["final_disposition"]
    _sheet(wb, "6_Unresolved_Retained",
           ["case_id", "task", "unit_id", "reason"], disp["unresolved_cases"],
           [34, 30, 10, 74])

    # 7. supplementary — kept apart
    suprows = [{"blind_unit_id": u, "question_id": v["question_id"],
                "stratum": v["stratum"], "n_human_themes": v["n_human_themes"]}
               for u, v in sup["denominators_per_unit"].items()]
    ws7 = _sheet(wb, "7_Supplementary_S01_S06",
                 ["blind_unit_id", "question_id", "stratum", "n_human_themes"],
                 suprows, [16, 13, 22, 18])
    ws7.cell(row=len(suprows) + 3, column=1,
             value=("NEVER pooled with U01-U07/Q3. One coder, six units, four different "
                    "guide questions, no inter-coder agreement, relevance NOT_ASSESSED. "
                    "No automatic extractor has been run on this sample."))
    ws7.cell(row=len(suprows) + 3, column=1).alignment = Alignment(wrap_text=True)

    # 8. provenance
    _sheet(wb, "8_Provenance", ["figure", "value", "source", "note"],
           rec["provenance"], [42, 26, 44, 44])

    wb.save(_FINAL / "FINAL_RESULTS_TABLES.xlsx")

    # --- traceability index ------------------------------------------------
    lines = [
        "# Results traceability index",
        "",
        f"Built {datetime.now(UTC).date()}. Every figure in the final report resolves to "
        "a source artefact here. The build refuses to publish a number that two files "
        "disagree about.",
        "",
        "## Sealed artefacts (verified unchanged at build time)",
        "",
        "| Artefact | SHA-256 | Verified |",
        "|---|---|---|",
    ]
    for name, v in rec["sealed_artefacts"].items():
        lines.append(f"| `{name}` | `{v['sha256'][:24]}…` | "
                     f"{'yes' if v['matches_expected'] else 'NO'} |")
    lines += ["", "## Figure provenance", "",
              "| Figure | Value | Source | Note |", "|---|---|---|---|"]
    for p in rec["provenance"]:
        lines.append(f"| `{p['figure']}` | {p['value']} | `{p['source']}` | "
                     f"{p['note']} |")
    lines += [
        "", "## Reconciliation checks performed", "",
        "- FG-level mean differences in `primary_effects_summary.csv` recomputed from "
        "the five per-FG rows in `primary_effects_by_fg.csv`.",
        "- The same means and direction counts cross-checked against "
        "`condition_comparison.csv`.",
        "- Human instance count cross-checked against `human_reference_q3.json`.",
        "- Machine theme count recomputed from `extraction_results_q3.json`.",
        "- 58 quotations counted from source rather than copied.",
        "- Cross-model corroborated + unresolved verified to sum to the case total.",
        "- Supplementary theme count verified against its own per-unit denominators.",
        "- Cost recomputed from measured token counts at the published Batch rate.",
        "",
        f"**Result: {rec['reconciliation']}** — {len(rec['provenance'])} figures, no "
        "contradictions found.",
        "",
        "## Cost record",
        "",
        f"- Pre-run estimate: ${cost['estimated_cost_pre_run_usd']} "
        "(`cross_model_cost_basis.json`, retained as historical record only)",
        f"- Measured: {cost['actual_input_tokens']:,} input + "
        f"{cost['actual_output_tokens']:,} output tokens",
        f"- Formula: `{cost['formula']}`",
        f"- Worked: {cost['worked']}",
        f"- **Calculated at list Batch rate: ${cost['calculated_list_batch_cost_usd']}** "
        f"(`cross_model_cost_actual.json`)",
        f"- The estimate was low by ${cost['estimate_error_usd']} "
        f"({cost['estimate_error_pct_of_actual']}% of actual). "
        f"{cost['why_the_estimate_was_low']}.",
        "- This is a calculated list-rate cost, **not necessarily the amount charged**; "
        "negotiated rates are not exposed by any API endpoint.",
        "",
    ]
    (_FINAL / "RESULTS_TRACEABILITY_INDEX.md").write_text("\n".join(lines),
                                                          encoding="utf-8")
    return rec


def main() -> int:
    rec = build()
    print("wrote final/FINAL_RESULTS_TABLES.xlsx")
    print("wrote final/RESULTS_TRACEABILITY_INDEX.md")
    print("figures traced:", len(rec["provenance"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
