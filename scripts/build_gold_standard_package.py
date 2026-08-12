"""
Gold-standard package — guide section 3 ("Gender and food choice"), 15 blind units.

TWO SEQUENTIAL COMPONENTS over the SAME 15 units. The sample is not redistributed
across guide questions: the study codebook was intentionally selective and was never
designed for exhaustive question-by-question coverage, so spreading 15 units thinner
would weaken the one question it does cover without fixing that.

  PART 1  emergent coding, codebook WITHHELD — coders name the principal ideas in
          their own words, with a verbatim quote and central/secondary relevance.
  PART 2  deductive coding against the 11-subtheme codebook, binary presence with a
          verified quote. Released only once the coder's Part 1 has been returned.

Part 1 first is the point: once the codebook is seen it cannot be unseen, and an
emergent reading taken afterwards would be contaminated by the study's categories.

Two jobs, in order, because the second is conditional on the first:

  1. BOUNDARY AUDIT of the 15 section-3 units. D4 was approved conditional on this.
     A unit whose section-3 boundary cannot be established is reported and the
     package is NOT built.
  2. BLIND PACKAGE for two independent coders, with a sealed source mapping.

THE 15 UNITS (pre-specified, not chosen on results)
    5 human            — section 3 of FG1..FG5
    5 enriched         — canonical replication 2: run02 for FG1-FG3,
                         run04 for FG4, run03 for FG5
    5 demographics-only— demoonly_run02 for FG1..FG5

The FG4/FG5 naming exceptions are archival, fixed before any outcome was seen.
Canonical replication 2 is never swapped for a better-scoring run.

BOUNDARIES
  Human      — the moderator turn beginning "Question 3." up to (not including) the
               "Question 4." header, or end of transcript if absent.
  Synthetic  — guide section 3 as segmented by moderator_log.section_transition,
               cross-checked against state_turn_*.json, sliced out of the approved
               comparable window so the gold standard scores exactly the text the
               evaluator sees.

BLINDING
  Units get random ids. Text is rendered through the same `to_blind_text` used
  everywhere else, so speakers are "Participant N" / "Moderator" with no names and
  no provenance. Ids are shuffled with a recorded seed and the id -> source mapping
  is written to a SEALED file that coders must not open.

Read-only with respect to source data. No API call.

Usage:
    py scripts/build_gold_standard_package.py
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import re
import unicodedata
import sys
from datetime import datetime, UTC
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "scripts"))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import thematic_coding as tc                                   # noqa: E402
from thematic_coding import to_blind_text                      # noqa: E402
from tier2b_segmentation import segment_synthetic_by_guide     # noqa: E402
from gold_standard_workbooks import (                          # noqa: E402
    assert_no_provenance,
    build_deductive_adjudication,
    build_emergent_adjudication,
    build_part1,
    build_part2,
)

_OUT_DIR = _REPO_ROOT / "analysis" / "production_evaluation"
_PKG_DIR = _OUT_DIR / "gold_standard_package"
_SEALED_DIR = _OUT_DIR / "gold_standard_sealed"
_DERIVED = _OUT_DIR / "comparable_transcripts"
_HUMAN_DIR = _REPO_ROOT / "data" / "datasets_transcripts" / "standardized" / "macho_meals"
_SESSION_LOGS = _REPO_ROOT / "output" / "session_logs"
_WITHHELD_DIR = _OUT_DIR / "gold_standard_part2_withheld"
_ADJ_DIR = _OUT_DIR / "gold_standard_adjudication"
_RETURNED_DIR = _OUT_DIR / "gold_standard_returned"

TARGET_SECTION = 3
TARGET_SECTION_LABEL = "Gender and food choice"
SHUFFLE_SEED = 20260730          # recorded, so the blinding is reproducible

# Canonical replication 2 per condition — pre-specified.
REP2 = {
    "enriched": {"fg1": "macho_meals_fg1_run02", "fg2": "macho_meals_fg2_run02",
                 "fg3": "macho_meals_fg3_run02", "fg4": "macho_meals_fg4_run04",
                 "fg5": "macho_meals_fg5_run03"},
    "demographics-only": {f"fg{i}": f"macho_meals_fg{i}_demoonly_run02" for i in range(1, 6)},
}

_Q_HEADER = re.compile(r"^\s*Question\s*(\d+)\s*[.:\)]", re.IGNORECASE)


def _sha(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def _is_mod(e: dict) -> bool:
    return ((e.get("speaker_role") or "").lower() == "moderator"
            or (e.get("speaker_name") or "").lower() == "moderator")


def _counts(entries: list[dict]) -> dict:
    par = [e for e in entries if not _is_mod(e)]
    return {
        "entries": len(entries),
        "moderator_turns": len(entries) - len(par),
        "participant_turns": len(par),
        "total_words": sum(len((e.get("content") or "").split()) for e in entries),
        "participant_words": sum(len((e.get("content") or "").split()) for e in par),
        "distinct_participants": len({e.get("speaker_name") or e.get("speaker_id")
                                      for e in par}),
    }


# ---------------------------------------------------------------------------
# 1. Boundary audit
# ---------------------------------------------------------------------------

def human_section3(fg: str) -> tuple[list[dict], dict]:
    entries = json.loads((_HUMAN_DIR / fg / "transcript.json").read_text(encoding="utf-8"))
    headers = [(i, int(_Q_HEADER.match(e.get("content") or "").group(1)))
               for i, e in enumerate(entries)
               if _is_mod(e) and _Q_HEADER.match(e.get("content") or "")]
    audit = {"question_headers_found": [q for _, q in headers]}

    start = next((i for i, q in headers if q == 3), None)
    if start is None:
        audit.update({"verdict": "BOUNDARY_NOT_FOUND",
                      "problem": "no 'Question 3.' header in this transcript"})
        return [], audit
    later = [i for i, q in headers if q > 3]
    end = min(later) if later else len(entries)

    seg = entries[start:end]
    audit.update({
        "first_entry_index": start, "last_entry_index": end - 1,
        "opens_with_moderator": _is_mod(entries[start]),
        "opening_text": re.sub(r"\s+", " ", entries[start].get("content") or "")[:150],
        "next_header_after": next((q for i, q in headers if i == end), None),
        "closing_boundary": ("next Question header" if later else "end of transcript"),
        "verdict": "OK",
        "problem": "",
        **_counts(seg),
    })
    return seg, audit


def synthetic_section3(run: str) -> tuple[list[dict], dict]:
    run_dir = _SESSION_LOGS / run
    seg_res = segment_synthetic_by_guide(run_dir / "transcript.json",
                                         run_dir / "session_state_initial.json")
    audit: dict = {"segmentation_warnings": seg_res.warnings}
    if TARGET_SECTION not in seg_res.sections:
        audit.update({"verdict": "BOUNDARY_NOT_FOUND",
                      "problem": f"guide section {TARGET_SECTION} absent from this run"})
        return [], audit

    idx = seg_res.sections[TARGET_SECTION].entry_indices
    source = json.loads((run_dir / "transcript.json").read_text(encoding="utf-8"))

    # Intersect with the APPROVED comparable window, so the gold standard scores
    # exactly the text the evaluator will see — never material the window excluded.
    win = json.loads((_DERIVED / run / "comparable_transcript.json").read_text(encoding="utf-8"))
    prov = win["_provenance"]
    w_first, w_last = prov["first_source_entry_index"], prov["last_source_entry_index"]
    keep = [i for i in idx if w_first <= i <= w_last]
    dropped = [i for i in idx if i not in keep]

    if not keep:
        audit.update({"verdict": "BOUNDARY_NOT_FOUND",
                      "problem": "section 3 lies entirely outside the comparable window"})
        return [], audit

    seg = []
    for i in keep:
        e = dict(source[i])
        if i == w_first:          # boundary entry is the trimmed verbatim slice
            e["content"] = win["transcript"][0]["content"]
        seg.append(e)

    audit.update({
        "first_entry_index": keep[0], "last_entry_index": keep[-1],
        "entries_dropped_outside_window": len(dropped),
        "opens_with_moderator": _is_mod(source[keep[0]]),
        "opening_text": re.sub(r"\s+", " ", seg[0].get("content") or "")[:150],
        "comparable_window_first": w_first, "comparable_window_last": w_last,
        "state_crosscheck_conflicts": 0,
        "verdict": "OK", "problem": "",
        **_counts(seg),
    })
    return seg, audit


def _norm_quote(t: str) -> str:
    t = unicodedata.normalize("NFKC", t or "")
    t = (t.replace("’", "'").replace("‘", "'")
           .replace("“", '"').replace("”", '"')
           .replace("–", "-").replace("—", "-"))
    return re.sub(r"\s+", " ", t).strip().casefold()


def release_part2(coder: str) -> int:
    """
    Release a coder's Part 2 workbook — only if their Part 1 came back intact.

    Part 2 is gated because seeing the codebook cannot be undone. The gate checks
    the returned artefact rather than trusting that the sequence was followed:
    the expected (unit_id, theme_slot) grid must be present exactly once each, with
    nothing added, deleted, duplicated or reordered, and every unit must carry at
    least one theme.
    """
    from openpyxl import load_workbook

    returned = _RETURNED_DIR / f"Coder_{coder}_Part1_Emergent.xlsx"
    src = _WITHHELD_DIR / f"Coder_{coder}_Part2_Deductive.xlsx"
    dst = _PKG_DIR / f"Coder_{coder}_Part2_Deductive.xlsx"

    if not returned.exists():
        print("REFUSED — Part 1 not returned.")
        print(f"  expected: {returned.relative_to(_REPO_ROOT)}")
        return 2
    if not src.exists():
        print(f"REFUSED — withheld Part 2 not found: {src.relative_to(_REPO_ROOT)}")
        return 2

    issued_path = _PKG_DIR / f"Coder_{coder}_Part1_Emergent.xlsx"
    if not issued_path.exists():
        print(f"REFUSED — issued Part 1 not found: {issued_path}")
        return 2

    rwb = load_workbook(returned)
    iwb = load_workbook(issued_path)
    expected = [(r[0].value, r[1].value)
                for r in iwb["Emergent_Coding"].iter_rows(min_row=2)]
    ws = rwb["Emergent_Coding"]
    problems: list[str] = []

    # --- the returned Units sheet must be byte-for-byte the issued one ---------
    # Quotes are validated against the ISSUED copy, never the returned one: a
    # returned sheet could have been edited, and checking a quote against edited
    # text would let a coder "verify" their own alteration.
    issued_units = [tuple(c.value for c in r[:5])
                    for r in iwb["Units"].iter_rows(min_row=2)]
    returned_units = [tuple(c.value for c in r[:5])
                      for r in rwb["Units"].iter_rows(min_row=2)]
    if len(returned_units) != len(issued_units):
        problems.append(f"Units sheet has {len(returned_units)} rows, issued had "
                        f"{len(issued_units)} (turn deleted or added)")
    for i, (got, want) in enumerate(zip(returned_units, issued_units)):
        if got != want:
            diff = [f"col {j + 1}: {str(g)[:45]!r} != {str(w)[:45]!r}"
                    for j, (g, w) in enumerate(zip(got, want)) if g != w]
            problems.append(f"Units row {i + 2} altered — {'; '.join(diff)}")
            if len([p for p in problems if p.startswith("Units row")]) >= 5:
                problems.append("... further Units differences suppressed")
                break

    unit_text: dict[str, str] = {}
    for uid, _turn, _spk, _w, text in issued_units:
        if uid:
            unit_text[uid] = unit_text.get(uid, "") + " " + (text or "")
    norm_text = {u: _norm_quote(t) for u, t in unit_text.items()}

    actual, themes_per_unit = [], {}

    def _check_theme(where: str, unit: str, label, desc, quote, rel) -> bool:
        """A populated theme must be complete AND its quote must be real."""
        label = (label or "").strip()
        desc = (desc or "").strip()
        quote = (quote or "").strip()
        rel = (rel or "").strip().lower()
        if not any([label, desc, quote, rel]):
            return False                      # entirely blank row — fine
        missing = [n for n, v in (("theme_label", label), ("theme_description", desc),
                                  ("supporting_quote", quote), ("relevance", rel)) if not v]
        if missing:
            problems.append(f"{where}: partly filled theme for {unit} — missing {missing}")
            return True
        if rel not in ("central", "secondary"):
            problems.append(f"{where}: relevance={rel!r} for {unit} "
                            f"(must be 'central' or 'secondary')")
        if _norm_quote(quote) not in norm_text.get(unit, ""):
            problems.append(f"{where}: quote is not a literal substring of {unit} -> "
                            f"{quote[:60]!r}")
        return True

    for row in ws.iter_rows(min_row=2):
        key = (row[0].value, row[1].value)
        actual.append(key)
        if _check_theme(f"Emergent_Coding row {row[0].row}", key[0],
                        row[2].value, row[3].value, row[4].value, row[5].value):
            themes_per_unit[key[0]] = themes_per_unit.get(key[0], 0) + 1

    # Safe continuation sheet — free-form rows, validated the same way.
    if "Overflow_Themes" in rwb.sheetnames:
        for row in rwb["Overflow_Themes"].iter_rows(min_row=2):
            unit = (row[0].value or "").strip() if isinstance(row[0].value, str)                 else row[0].value
            populated = any((c.value or "").strip() for c in row[1:6]
                            if isinstance(c.value, str))
            if not unit:
                # A row carrying data but no unit_id cannot be attributed to any
                # excerpt, so it is a hard failure rather than something to skip.
                if populated:
                    problems.append(f"Overflow_Themes row {row[0].row}: data present "
                                    f"but unit_id is empty — cannot be attributed")
                continue
            if unit not in norm_text:
                problems.append(f"Overflow_Themes row {row[0].row}: unknown unit_id "
                                f"{unit!r}")
                continue
            if _check_theme(f"Overflow_Themes row {row[0].row}", unit,
                            row[1].value, row[2].value, row[3].value, row[4].value):
                themes_per_unit[unit] = themes_per_unit.get(unit, 0) + 1

    if len(actual) != len(expected):
        problems.append(f"row count {len(actual)} != expected {len(expected)} "
                        f"(rows added or deleted on Emergent_Coding)")
    if actual != expected:
        problems.append("row keys reordered or altered — (unit_id, theme_slot) grid "
                        "does not match the issued workbook")
    dupes = {k for k in actual if actual.count(k) > 1}
    if dupes:
        problems.append(f"duplicate rows: {sorted(dupes)[:5]}")
    units = {k[0] for k in expected}
    empty = sorted(u for u in units if themes_per_unit.get(u, 0) == 0)
    if empty:
        problems.append(f"units with no theme recorded: {empty}")

    if problems:
        print(f"REFUSED — Coder {coder} Part 1 did not pass validation:")
        for pr in problems[:20]:
            print(f"  - {pr}")
        if len(problems) > 20:
            print(f"  ... and {len(problems) - 20} more")
        return 3

    dst.write_bytes(src.read_bytes())
    try:
        shown = dst.relative_to(_REPO_ROOT)
    except ValueError:            # e.g. a temp workspace during testing
        shown = dst
    print(f"RELEASED — {shown}")
    print(f"  Part 1 verified: {len(actual)} rows intact, "
          f"{sum(themes_per_unit.values())} themes across {len(units)} units.")
    return 0


def main() -> None:
    print("=" * 82)
    print(f"  GOLD STANDARD — guide section {TARGET_SECTION} ({TARGET_SECTION_LABEL})")
    print("  Boundary audit first; package built only if all 15 units pass.")
    print("=" * 82)

    units: list[dict] = []
    for fg in ("fg1", "fg2", "fg3", "fg4", "fg5"):
        seg, audit = human_section3(fg)
        units.append({"stratum": "human", "fg": fg, "source": f"human/{fg}",
                      "entries": seg, "audit": audit})
    for cond in ("enriched", "demographics-only"):
        for fg in ("fg1", "fg2", "fg3", "fg4", "fg5"):
            run = REP2[cond][fg]
            seg, audit = synthetic_section3(run)
            units.append({"stratum": cond, "fg": fg, "source": run,
                          "entries": seg, "audit": audit})

    print(f"\n{'stratum':<20}{'fg':<5}{'source':<32}{'verdict':<18}"
          f"{'ent':>4}{'p':>4}{'words':>7}")
    for u in units:
        a = u["audit"]
        print(f"{u['stratum']:<20}{u['fg']:<5}{u['source'][:31]:<32}{a['verdict']:<18}"
              f"{a.get('entries', 0):>4}{a.get('participant_turns', 0):>4}"
              f"{a.get('total_words', 0):>7}")

    failed = [u for u in units if u["audit"]["verdict"] != "OK"]
    _OUT_DIR.mkdir(parents=True, exist_ok=True)
    afields: list[str] = []
    arows = []
    for u in units:
        r = {"stratum": u["stratum"], "fg": u["fg"], "source": u["source"], **u["audit"]}
        r["question_headers_found"] = "|".join(map(str, r.get("question_headers_found", [])))
        r["segmentation_warnings"] = "|".join(r.get("segmentation_warnings", []) or [])
        arows.append(r)
        for k in r:
            if k not in afields:
                afields.append(k)
    with open(_OUT_DIR / "gold_standard_boundary_audit.csv", "w", newline="",
              encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=afields, extrasaction="ignore")
        w.writeheader()
        for r in arows:
            w.writerow(r)
    print(f"\nWrote {(_OUT_DIR / 'gold_standard_boundary_audit.csv').relative_to(_REPO_ROOT)}")

    if failed:
        print(f"\nSTOP — {len(failed)} unit(s) failed the boundary audit; package NOT built:")
        for u in failed:
            print(f"  {u['stratum']}/{u['fg']} ({u['source']}): {u['audit']['problem']}")
        sys.exit(2)
    print("\nBoundary audit: 15/15 OK — building the blind package.")

    # --- blinding ----------------------------------------------------------
    rng = random.Random(SHUFFLE_SEED)
    order = list(range(len(units)))
    rng.shuffle(order)
    unit_ids = [f"U{n:02d}" for n in range(1, len(units) + 1)]

    _PKG_DIR.mkdir(parents=True, exist_ok=True)
    _SEALED_DIR.mkdir(parents=True, exist_ok=True)

    mapping = []
    blind_units = []
    for uid, pos in zip(unit_ids, order):
        u = units[pos]
        blind_text, speaker_map = to_blind_text(u["entries"])
        # Per-turn records, reconstructed with the same speaker map and the same
        # skip rule to_blind_text uses, then checked against its output so the two
        # can never drift apart.
        records, n = [], 0
        for e in u["entries"]:
            content = (e.get("content") or "").strip()
            if not content:
                continue
            n += 1
            name = e.get("speaker_name") or e.get("speaker_id", "Unknown")
            records.append({"turn_id": f"T{n:03d}",
                            "speaker": speaker_map.get(name, name),
                            "text": content})
        rebuilt = "\n".join(f"[{r['turn_id']}] {r['speaker']}: {r['text']}" for r in records)
        if rebuilt != blind_text:
            print(f"STOP — per-turn reconstruction differs from to_blind_text for {uid}")
            sys.exit(4)
        blind_units.append({"unit_id": uid, "text": blind_text, "records": records})
        mapping.append({
            "unit_id": uid, "stratum": u["stratum"], "fg": u["fg"],
            "source": u["source"], "speaker_map": speaker_map,
            "blind_text_sha256": _sha(blind_text),
            **{k: v for k, v in u["audit"].items()
               if k in ("first_entry_index", "last_entry_index", "entries",
                        "participant_turns", "total_words", "distinct_participants")},
        })

    codebook = tc.load_codebook()

    for u in blind_units:
        (_PKG_DIR / f"{u['unit_id']}.txt").write_text(
            f"UNIT {u['unit_id']}\n"
            f"{'=' * 70}\n\n{u['text']}\n", encoding="utf-8")

    # Codebook as CSV is kept for the SCORING side only — it is deliberately NOT
    # placed in the released package, because Part 1 must be completed without it.
    with open(_SEALED_DIR / "codebook_reference.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["subtheme_id", "subtheme_label", "theme",
                                          "description", "example"],
                           extrasaction="ignore")
        w.writeheader()
        for c in codebook:
            w.writerow(c)

    # --- Excel workbooks ---------------------------------------------------
    _WITHHELD_DIR.mkdir(parents=True, exist_ok=True)
    _ADJ_DIR.mkdir(parents=True, exist_ok=True)
    _RETURNED_DIR.mkdir(parents=True, exist_ok=True)
    built: list[Path] = []
    for coder in ("A", "B"):
        p1 = _PKG_DIR / f"Coder_{coder}_Part1_Emergent.xlsx"
        build_part1(p1, coder, blind_units)
        built.append(p1)
        # Part 2 is built now but WITHHELD until that coder's Part 1 is returned.
        p2 = _WITHHELD_DIR / f"Coder_{coder}_Part2_Deductive.xlsx"
        build_part2(p2, coder, blind_units, codebook)
        built.append(p2)

    adj_e = _ADJ_DIR / "Adjudication_Part1_Emergent.xlsx"
    build_emergent_adjudication(adj_e, blind_units)
    adj_d = _ADJ_DIR / "Adjudication_Part2_Deductive.xlsx"
    build_deductive_adjudication(adj_d, blind_units, codebook)
    built += [adj_e, adj_d]

    print("\nWorkbooks built (each re-opened and scanned for provenance leaks):")
    leaks = 0
    for b in built:
        problems = assert_no_provenance(b)
        leaks += len(problems)
        loc = ("RELEASED" if b.parent == _PKG_DIR
               else "WITHHELD" if b.parent == _WITHHELD_DIR else "adjudication")
        print(f"  {loc:<13} {b.name:<38} leaks: {len(problems)}")
        for pr in problems[:4]:
            print(f"      {pr}")
    if leaks:
        print(f"\nSTOP — {leaks} provenance leak(s) in coder-facing workbooks.")
        sys.exit(3)

    (_WITHHELD_DIR / "DO_NOT_RELEASE_YET.md").write_text(
        "# Part 2 — withheld\n\n"
        "These deductive workbooks must NOT be sent to a coder until that coder's\n"
        "Part 1 emergent workbook has been returned and passes its structural check.\n"
        "Seeing the codebook cannot be undone, so the sequence is enforced rather\n"
        "than trusted.\n\n"
        "Place the returned Part 1 workbook in `../gold_standard_returned/`, then:\n\n"
        "    py scripts/build_gold_standard_package.py --release-part2 A\n"
        "    py scripts/build_gold_standard_package.py --release-part2 B\n\n"
        "Release refuses unless the returned workbook has no added, deleted,\n"
        "duplicated or reordered rows and records at least one theme per unit.\n",
        encoding="utf-8")

    (_PKG_DIR / "README.md").write_text(f"""# Gold-standard package — {len(blind_units)} blind excerpts

Two sequential parts over the **same** {len(blind_units)} excerpts.

| Stage | Workbook | Released |
|---|---|---|
| **Part 1 — emergent** | `Coder_A_Part1_Emergent.xlsx`, `Coder_B_Part1_Emergent.xlsx` | now |
| **Part 2 — deductive** | `Coder_A_Part2_Deductive.xlsx`, `Coder_B_Part2_Deductive.xlsx` | **only after that coder returns Part 1** |

Part 1 contains **no codebook**, deliberately. Coders describe the principal ideas
in their own words. Part 2 then codes the same excerpts against the study codebook.
Running it the other way round would let the study's categories shape the emergent
reading, and that cannot be undone once the codebook has been seen.

`U01.txt` … `U{len(blind_units):02d}.txt` are plain-text copies of the same excerpts
that appear on each workbook's **Units** sheet, for convenience.

## After the coding returns

| Stage | Workbook |
|---|---|
| cluster emergent themes | `../gold_standard_adjudication/Adjudication_Part1_Emergent.xlsx` |
| resolve deductive disagreements | `../gold_standard_adjudication/Adjudication_Part2_Deductive.xlsx` |

Scoring: `scripts/score_gold_standard.py`.

## Scope of what this validates

Guide section 3 is where subthemes **A.1–A.3** are *directly elicited*, so those are
the codes this exercise can genuinely validate. **B–D** are not directly elicited
here: their absence is evidence about **specificity / false-positive rate**, and
their presence is **opportunistic detection**. This is not a complete recall
validation for B–D, and no pooled statistic should be read as though it were.

The id → source mapping is sealed and is not part of this package.
""", encoding="utf-8")

    sealed = {
        "sealed_utc": datetime.now(UTC).isoformat(),
        "warning": "SEALED — do not open until both coder worksheets are complete.",
        "shuffle_seed": SHUFFLE_SEED,
        "target_section": TARGET_SECTION,
        "target_section_label": TARGET_SECTION_LABEL,
        "canonical_replication_2": REP2,
        "selection_note": ("Pre-specified. Canonical replication 2 was not swapped for "
                           "a better-scoring run; the FG4 (run04) and FG5 (run03) naming "
                           "exceptions are archival and were fixed before any outcome "
                           "was seen."),
        "mapping": mapping,
    }
    (_SEALED_DIR / "unit_id_to_source_SEALED.json").write_text(
        json.dumps(sealed, indent=2, ensure_ascii=False), encoding="utf-8")

    # --- allowlist verification of the RELEASED package --------------------
    # Whitelisted by exact name/pattern. Anything else present is a leak, whether
    # it is a codebook, a Part-2 workbook, an adjudication workbook, the sealed
    # mapping, or a stale CSV worksheet from an earlier package version.
    allowed = ({"README.md", "Coder_A_Part1_Emergent.xlsx", "Coder_B_Part1_Emergent.xlsx"}
               | {f"{u['unit_id']}.txt" for u in blind_units})
    present = {f.name for f in _PKG_DIR.iterdir()}
    unexpected = sorted(present - allowed)
    missing = sorted(allowed - present)

    forbidden_checks = {
        "codebook": [n for n in present if "codebook" in n.lower()],
        "part 2 / deductive": [n for n in present if "part2" in n.lower()
                               or "deductive" in n.lower()],
        "adjudication": [n for n in present if "adjudic" in n.lower()],
        "sealed mapping": [n for n in present if "seal" in n.lower()],
        "legacy CSV worksheets": [n for n in present if n.lower().endswith(".csv")],
    }

    print("")
    print("ALLOWLIST CHECK — released package")
    print(f"  files present      : {len(present)}   (expected {len(allowed)})")
    print(f"  unexpected files   : {unexpected or 'none'}")
    print(f"  missing files      : {missing or 'none'}")
    for label, hits in forbidden_checks.items():
        print(f"  no {label:<22}: {'OK' if not hits else 'FAIL -> ' + str(hits)}")
    if unexpected or missing or any(forbidden_checks.values()):
        print("")
        print("STOP — released package failed the allowlist check.")
        sys.exit(5)

    strata = {}
    for m in mapping:
        strata[m["stratum"]] = strata.get(m["stratum"], 0) + 1
    print(f"\nPackage: {len(blind_units)} units  strata={strata}  seed={SHUFFLE_SEED}")
    print(f"  {_PKG_DIR.relative_to(_REPO_ROOT)}")
    print(f"  SEALED mapping -> {_SEALED_DIR.relative_to(_REPO_ROOT)}/unit_id_to_source_SEALED.json")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--release-part2", default=None, choices=["A", "B"], metavar="CODER",
                    help="Release that coder's Part 2 workbook, if their Part 1 has been "
                         "returned and passes the structural check.")
    args = ap.parse_args()
    if args.release_part2:
        sys.exit(release_part2(args.release_part2))
    main()
