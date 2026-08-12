"""Rebuild OCA-001 with full turns, minimal context, roster redaction and A.1/A.3."""
import hashlib, json, re, sys, pathlib
from datetime import datetime, UTC
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(".").resolve()
sys.path.insert(0, str(ROOT / "scripts"))
import production_eval_pipeline as pep
import aggregate_production_results as agg
from thematic_coding import to_blind_text, load_codebook

OUT = ROOT / "analysis" / "production_evaluation"
DIR = OUT / "open_coding_adjudication"
SEALED = OUT / "gold_standard_sealed" / "open_coding_item_mapping.json"
ITEM_ID, INTERNAL_ID = "OCA-001", "FG4-DEMO-R01-A1"
RUN = "macho_meals_fg4_demoonly_run01"

# Turns shown: the three cited turns plus the minimum preceding context needed to
# read each speaker's stance. T020 poses the question; T022 is the moderator's
# restatement that T023 answers; T026 is the prompt T027 answers.
SHOW = [20, 21, 22, 23, 26, 27]
CITED = {21, 23, 27}

r = next(x for x in agg.load_results() if x["input"].get("physical_run") == RUN)
entries = pep._entries_for({"path": r["input"]["path"], "side": "synthetic"})
_, smap = to_blind_text(entries)
ROSTER = {name for name in smap if name.lower() != "moderator"}

recs, n = {}, 0
for e in entries:
    c = (e.get("content") or "").strip()
    if not c:
        continue
    n += 1
    nm = e.get("speaker_name") or e.get("speaker_id", "")
    recs[n] = (smap.get(nm, nm), c)

def redact(text):
    """
    Replace each roster name with that speaker's OWN blind label.

    Collapsing every name to "another participant" blinds correctly but destroys
    reference: a moderator turn naming three people becomes three identical phrases
    and the reader can no longer tell who agrees with whom, which is precisely what
    the stance judgement depends on. The speaker labels are already generic, so
    mapping each name to its own label keeps the exchange legible while revealing
    nothing about identity.
    """
    out, hits = text, []
    for name in sorted(ROSTER, key=len, reverse=True):
        label = smap.get(name, "another participant")
        pat = re.compile(r"\b" + re.escape(name) + r"'s\b")
        if pat.search(out):
            hits.append(name + "'s -> " + label + "'s")
            out = pat.sub(label + "'s", out)
        pat = re.compile(r"\b" + re.escape(name) + r"\b")
        if pat.search(out):
            hits.append(name + " -> " + label)
            out = pat.sub(label, out)
    return out, hits

shown, sealed_turns = [], []
for t in SHOW:
    spk, txt = recs[t]
    red, hits = redact(txt)
    shown.append({"turn": f"T{t:03d}", "speaker": spk, "text": red,
                  "cited": t in CITED, "redacted": bool(hits)})
    sealed_turns.append({
        "turn": f"T{t:03d}", "speaker_blind": spk, "cited_in_support_of_A1": t in CITED,
        "original_text": txt, "presented_text": red,
        "redactions": hits,
        "original_sha256": hashlib.sha256(txt.encode()).hexdigest(),
        "presented_sha256": hashlib.sha256(red.encode()).hexdigest(),
    })

cb = {c["subtheme_id"]: c for c in load_codebook()}
A1, A3 = cb["A.1"], cb["A.3"]

HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(bold=True, color="FFFFFF", size=10)
NEED = PatternFill("solid", fgColor="FFF2CC")
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Adjudication"
ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 104
row = 1

def put(label, text, bold=False, fill=None, height=None):
    global row
    c1 = ws.cell(row=row, column=1, value=label); c1.font = Font(bold=True, size=10)
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=row, column=2, value=text); c2.font = Font(bold=bold, size=10)
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    if fill: c2.fill = fill
    if height: ws.row_dimensions[row].height = height
    row += 1
    return row - 1

put("Item", ITEM_ID, True)
put("Task", "Decide whether the evidence below supports subtheme A.1 as the codebook "
            "defines it.", False)
put("", "")
put("Blinding", "You are not told which discussion this is, whether it is human or "
                "machine-generated, or what depends on your answer. Speaker labels are "
                "generic. Participant names inside the text have been replaced with "
                "the same generic label used for that speaker (e.g. Participant 2) "
                "— a minimal redaction for blinding. The passages below are therefore "
                "NOT verbatim quotations.", False)
put("", "")

put("CODEBOOK", "", True)
put("A.1 label", str(A1.get("subtheme_label", "")), True)
put("A.1 definition", str(A1.get("description", "")), True)
put("A.1 example", str(A1.get("example", "")).strip())
put("", "")
put("A.3 label", str(A3.get("subtheme_label", "")), True)
put("A.3 definition", str(A3.get("description", "")), True)
put("A.3 example", str(A3.get("example", "")).strip())
put("", "")
put("Why A.3 is here", "A.1 and A.3 are the neighbouring boundary. They differ in "
                       "STANCE: A.1 acknowledges an influence of gender; A.3 rejects "
                       "or is unsure of one and then goes on to describe a gendered "
                       "food context. The same domestic material can appear in either. "
                       "The question remains whether A.1 is supported; A.3 is shown "
                       "only so the boundary is visible.", False)
put("", "")

put("TRANSCRIPT EXTRACT", "Turns flagged ► were cited in support of A.1. The others are "
                          "the minimum context needed to read each speaker's stance.",
    True)
put("", "")
for s in shown:
    mark = "►" if s["cited"] else " "
    put(f"{mark} {s['turn']} · {s['speaker']}", s["text"],
        height=min(240, 15 + 7.5 * (len(s["text"]) // 100 + s["text"].count("\n") + 1)))
put("", "")
put("Note", "The cited passages were verified as attributed to the turns shown. The "
            "question is not whether they were quoted accurately, but whether they "
            "support A.1.", False)
put("", "")

put("YOUR JUDGEMENT", "", True)
put("options", "SUPPORTS_A1  |  DOES_NOT_SUPPORT_A1  |  UNCERTAIN   "
               "(also available as a dropdown in the cell below)", False)
vr = put("verdict", "", fill=NEED)
dv = DataValidation(type="list",
                    formula1='"SUPPORTS_A1,DOES_NOT_SUPPORT_A1,UNCERTAIN"',
                    allow_blank=True, showDropDown=False)
ws.add_data_validation(dv); dv.add(ws.cell(row=vr, column=2))
put("reasoning (required)", "", fill=NEED, height=110)
put("alternative subtheme (optional)", "", fill=NEED)
put("reviewer", "", fill=NEED)
put("date (UTC)", "", fill=NEED)
put("", "")
put("Reasoning is required", "A verdict without reasoning cannot be used. UNCERTAIN is "
                             "legitimate and also needs reasoning — say what would "
                             "resolve it.", False)
put("No code is changed yet", "Your verdict is recorded against the item. Any change to "
                              "the coded data is a separate, documented step.", False)

DIR.mkdir(parents=True, exist_ok=True)
path = DIR / f"{ITEM_ID}_adjudication.xlsx"
wb.save(path); wb.close()

SEALED.parent.mkdir(parents=True, exist_ok=True)
SEALED.write_text(json.dumps({
    "sealed_utc": datetime.now(UTC).isoformat(),
    "warning": "SEALED — do not supply to the reviewer.",
    "redaction_rule": ("Every roster participant name (and its possessive) is replaced "
                       "with that speaker's own blind label from to_blind_text (e.g. "
                       "'Gregor' -> 'Participant 1'). Reference between speakers is "
                       "preserved; identity is not. Moderator label retained. No other "
                       "alteration. Presented text is therefore NOT a verbatim quote."),
    "roster_names_redacted": sorted(ROSTER),
    "items": [{
        "form_item_id": ITEM_ID, "internal_id": INTERNAL_ID,
        "physical_run": RUN, "fg": "fg4", "condition": "demographics-only",
        "side": "synthetic", "subtheme_under_review": "A.1",
        "boundary_subtheme_shown": "A.3",
        "turns_shown": [f"T{t:03d}" for t in SHOW],
        "turns_cited": [f"T{t:03d}" for t in sorted(CITED)],
        "turn_provenance": sealed_turns,
        "verdict_recorded_in": ("fg4_demographics_only_qualitative_report.json -> "
                                "referred_to_human_review[0].human_review_verdict"),
    }],
}, indent=1, ensure_ascii=False), encoding="utf-8")
print(f"form   : {path}")
print(f"roster : {sorted(ROSTER)}")
print(f"redactions: {sum(len(t['redactions']) for t in sealed_turns)}")
