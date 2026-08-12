"""
Post-hoc thematic-coding library for the Macho Meals fidelity experiment.

LLM-ASSISTED, EVIDENCE-CONSTRAINED THEMATIC FIDELITY ASSESSMENT
This is deductive codebook coding + inductive open-theme comparison via Gemini.
It is NOT "thematic analysis" in the full reflexive qualitative sense.

QUARANTINE: this module is the ONLY place the coding frame
(analysis/coding_frame/CodeBook_Macho Meals.xlsx) is read. It is never
imported by core/, prompts/, or run_session.py, and runs strictly AFTER
synthetic transcripts already exist.

Coder: Gemini (gemini-2.5-flash), a different model family from the Anthropic
models that generate synthetic transcripts — avoids a single model both
producing and grading the same content.

Blind, symmetric: both real and synthetic transcripts are rendered to the same
generic [T001] Participant N format (every provenance field stripped) before
coding. The coding prompt never mentions "real" or "synthetic". Identical
model, temperature, and prompt for every transcript.

Quote-grounded: every Tier 1 "present" decision must be backed by an exact
substring of the blind transcript. Python verifies each quote; codes whose
evidence cannot be verified are demoted to present=False and excluded from
the present set. This defends against fabricated evidence.

Cost discipline: exactly one Tier-1 call and one Tier-2 call per transcript.
They are never bundled — bundling would let the codebook (Tier-1 context)
contaminate the codebook-free Tier-2 open extraction.
All call usage is logged to analysis/coding_frame/gemini_calls.jsonl.
"""

from __future__ import annotations

import json
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from google import genai
from google.genai import errors as _genai_errors

_REPO_ROOT = Path(__file__).resolve().parent.parent
_CODEBOOK_PATH = _REPO_ROOT / "analysis" / "coding_frame" / "CodeBook_Macho Meals.xlsx"
_CALL_LOG_PATH = _REPO_ROOT / "analysis" / "coding_frame" / "gemini_calls.jsonl"
_AUDIT_LOG_PATH = _REPO_ROOT / "analysis" / "coding_frame" / "quote_match_audit.jsonl"

_MODEL = "gemini-2.5-flash"
_TEMPERATURE = 0.0  # deterministic-as-possible coding

# Evaluator configs for the model-comparison experiment.
# Each entry maps an evaluator label to: model id, which .env key to read,
# temperature (None = model does not support it — omit from request), and
# the thinking level the model uses by default (for logging only).
EVALUATOR_CONFIGS: dict[str, dict] = {
    "gemini25": {
        "model":          "gemini-2.5-flash",
        "key_env":        "GEMINI_API_KEY_25",
        "temperature":    0.0,
        "thinking_level": None,   # n/a; temperature controls determinism
    },
    "gemininext": {
        "model":          "gemini-3.5-flash",
        "key_env":        "GEMINI_API_KEY_NEXT",
        "temperature":    None,   # not supported — omitted from request
        "thinking_level": "medium",  # model default; logged but not set via API
    },
}


# ---------------------------------------------------------------------------
# API setup
# ---------------------------------------------------------------------------

def _resolve_api_key() -> str:
    key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
    if not key:
        raise RuntimeError(
            "No Gemini API key found. Set GOOGLE_API_KEY (or GEMINI_API_KEY) "
            "in a .env file at the repo root (see .env.example) or as a "
            "system environment variable."
        )
    return key


def _resolve_backup_api_key() -> str | None:
    return (
        os.environ.get("GOOGLE_API_KEY_BACKUP")
        or os.environ.get("GEMINI_API_KEY_BACKUP")
        or None
    )


def _resolve_backup_model() -> str:
    """Model to use with the backup key — may differ from _MODEL if the backup
    project doesn't have access to the primary model (e.g. new-user restrictions)."""
    return os.environ.get("GOOGLE_API_KEY_BACKUP_MODEL") or _MODEL


def _client() -> genai.Client:
    return genai.Client(api_key=_resolve_api_key())


def _client_for_evaluator(cfg: dict) -> genai.Client:
    """Return a client keyed by cfg['key_env'], falling back to the primary key."""
    key = os.environ.get(cfg.get("key_env", "")) or _resolve_api_key()
    return genai.Client(api_key=key)


def _generate_with_fallback(
    client: genai.Client,
    no_fallback: bool = False,
    **kwargs,
) -> object:
    """
    Calls client.models.generate_content(**kwargs).
    On 429 quota exhaustion, transparently retries once with the backup API key
    (GOOGLE_API_KEY_BACKUP in .env) if one is configured.  Any other error re-raises.
    Pass no_fallback=True (evaluator-comparison context) to suppress the backup
    and keep each evaluator's calls strictly on its own key.
    """
    try:
        return client.models.generate_content(**kwargs)
    except _genai_errors.ClientError as exc:
        if "RESOURCE_EXHAUSTED" not in str(exc):
            raise
        if no_fallback:
            raise
        backup_key = _resolve_backup_api_key()
        if backup_key is None:
            raise
        print("  [quota] Primary key exhausted (429) — retrying with backup key ...", flush=True)
        backup_client = genai.Client(api_key=backup_key)
        backup_kwargs = {**kwargs, "model": _resolve_backup_model()}
        return backup_client.models.generate_content(**backup_kwargs)


def _log_call(event_type: str, metadata: dict) -> None:
    _CALL_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    entry = {"timestamp": datetime.now(UTC).isoformat(), "event_type": event_type, **metadata}
    with open(_CALL_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")


# ---------------------------------------------------------------------------
# Blind transcript format
# ---------------------------------------------------------------------------

def to_blind_text(transcript_entries: list[dict]) -> tuple[str, dict[str, str]]:
    """
    Renders a transcript as:
        [T001] Moderator: <content>
        [T002] Participant 1: <content>

    Speaker names → generic labels (same function for real and synthetic;
    no provenance reaches the coder). Skips empty turns.

    Returns:
        blind_text  — the formatted string to pass to Gemini
        speaker_map — {original_speaker_name: generic_label} for audit trail
    """
    speaker_map: dict[str, str] = {}
    participant_counter = 0
    lines: list[str] = []
    for entry in transcript_entries:
        content = (entry.get("content") or "").strip()
        if not content:
            continue
        speaker_name = entry.get("speaker_name") or entry.get("speaker_id", "Unknown")
        role = (entry.get("speaker_role") or "").lower()
        if speaker_name not in speaker_map:
            if role == "moderator" or speaker_name.lower() == "moderator":
                speaker_map[speaker_name] = "Moderator"
            else:
                participant_counter += 1
                speaker_map[speaker_name] = f"Participant {participant_counter}"
        turn_id = f"T{len(lines) + 1:03d}"
        lines.append(f"[{turn_id}] {speaker_map[speaker_name]}: {content}")
    return "\n".join(lines), speaker_map


def _count_participants(blind_text: str) -> int:
    """Count distinct non-Moderator speaker labels in a blind transcript."""
    speakers: set[str] = set()
    for line in blind_text.splitlines():
        m = re.match(r"\[T\d+\]\s+(.+?):", line)
        if m:
            sp = m.group(1).strip()
            if sp.lower() != "moderator":
                speakers.add(sp)
    return len(speakers)


# ---------------------------------------------------------------------------
# Codebook loading (the ONLY place this file is read)
# ---------------------------------------------------------------------------

def load_codebook() -> list[dict]:
    """Returns [{theme, subtheme_id, subtheme_label, description, example}, ...]."""
    import openpyxl
    wb = openpyxl.load_workbook(_CODEBOOK_PATH, data_only=True)
    ws = wb["Consolidated Table (2)"]
    entries = []
    current_theme = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        _, theme, subtheme, desc, example = row
        if theme:
            current_theme = theme
        if not subtheme and not desc:
            continue
        if subtheme:
            match = re.match(r"([A-Z]\.\d+)\.\s*(.*)", subtheme)
            subtheme_id, subtheme_label = (
                (match.group(1), match.group(2)) if match else (subtheme, subtheme)
            )
        else:
            letter_match = re.match(r"([A-Z])\)", current_theme or "")
            subtheme_id = letter_match.group(1) if letter_match else current_theme
            subtheme_label = current_theme
        entries.append({
            "theme": current_theme,
            "subtheme_id": subtheme_id,
            "subtheme_label": subtheme_label,
            "description": desc,
            "example": example,
        })
    return entries


def _codebook_reference_text(codebook: list[dict]) -> str:
    lines = []
    for c in codebook:
        lines.append(f"- {c['subtheme_id']} ({c['subtheme_label']}): {c['description']}")
    return "\n".join(lines)


# Lazy stable-prefix cache for Tier-1 prompts (Part C — prompt caching).
# Computed once per unique codebook rendering; reused byte-for-byte so that
# Gemini's implicit context caching can engage on the prefix.
_TIER1_PREFIX_CACHE: dict[int, str] = {}


def _get_tier1_stable_prefix(codebook: list[dict]) -> str:
    """Return the stable Tier-1 prompt prefix, building it once per codebook."""
    reference = _codebook_reference_text(codebook)
    key = hash(reference)
    if key not in _TIER1_PREFIX_CACHE:
        _TIER1_PREFIX_CACHE[key] = (
            "CODING FRAME:\n"
            + reference
            + "\n\nCode the transcript below against every entry in the coding "
            "frame above. Quotes must be exact substrings of the transcript text.\n\n"
        )
    return _TIER1_PREFIX_CACHE[key]


# ---------------------------------------------------------------------------
# Pydantic models — Tier 1
# ---------------------------------------------------------------------------

class SupportingQuote(BaseModel):
    turn_id: str = Field(..., description="Turn label, e.g. 'T034' (no brackets).")
    speaker: str = Field(..., description="Generic speaker label, e.g. 'Participant 4'.")
    quote: str = Field(..., description="Exact verbatim substring copied from the transcript.")


class SubthemeCode(BaseModel):
    subtheme_id: str
    present: bool
    confidence: Literal["high", "medium", "low"] | None = None
    supporting_quotes: list[SupportingQuote] = Field(default_factory=list)
    evidence_note: str | None = None
    # Set by Python after substring verification; not in Gemini output
    quote_verified: bool = True
    unverified_quote_count: int = 0
    # Evidence-constrained reach — computed in Python from verified quotes only
    voiced_by: list[str] = Field(default_factory=list)  # distinct non-Moderator speakers with ≥1 verified quote
    reach: float = 0.0  # len(voiced_by) / n_participants in the transcript


class Tier1Result(BaseModel):
    codes: list[SubthemeCode]


@dataclass
class QuoteValidityStats:
    """Counts from substring verification of Tier 1 supporting quotes."""
    total_quotes: int           # quotes submitted for present=True codes
    verified_quotes: int        # quotes confirmed via normalization-robust check
    total_present_codes: int    # codes Gemini marked present (before verification)
    verified_codes: int         # codes with ≥1 verified quote (remain present)
    demoted_codes: int          # codes demoted to present=False (0 verified quotes)
    raw_exact_quotes: int = 0           # quotes that passed the raw (pre-normalization) check
    normalized_recovered_quotes: int = 0  # failed raw but passed normalized check

    @property
    def quote_verification_rate(self) -> float:
        return self.verified_quotes / self.total_quotes if self.total_quotes else 1.0

    @property
    def code_preservation_rate(self) -> float:
        return self.verified_codes / self.total_present_codes if self.total_present_codes else 1.0


# ---------------------------------------------------------------------------
# Scoring data class — Tier 1
# ---------------------------------------------------------------------------

@dataclass
class TierOneScores:
    """
    Fidelity scores comparing a real transcript's code profile to a synthetic
    (or any other) transcript's verified code profile.

    At both subtheme level (stricter) and theme level (easier — the gap is informative).
    """
    subtheme_recall: float      # real codes reproduced in synthetic / real codes present
    subtheme_precision: float   # synthetic codes also in real / synthetic codes present
    subtheme_f1: float
    subtheme_jaccard: float
    theme_recall: float
    theme_precision: float
    theme_f1: float
    theme_jaccard: float

    # Code sets for audit
    real_present: frozenset[str]
    synthetic_present: frozenset[str]
    shared_subthemes: frozenset[str]
    real_themes: frozenset[str]
    synthetic_themes: frozenset[str]
    shared_themes: frozenset[str]

    def interpretation(self) -> str:
        r, p = self.subtheme_recall, self.subtheme_precision
        if r >= 0.8 and p >= 0.8:
            return "strong fidelity: reproduces real themes without adding many extras"
        if r >= 0.8 and p < 0.6:
            return "high-recall / low-precision: reproduces real themes but adds extras not in real"
        if r < 0.6 and p >= 0.8:
            return "low-recall / high-precision: conservative — misses real themes but extras are on-topic"
        return "low-recall / low-precision: poor thematic fidelity"


# ---------------------------------------------------------------------------
# Quote verification
# ---------------------------------------------------------------------------

_CURLY_QUOTE_MAP = str.maketrans({
    "“": '"', "”": '"',   # " "
    "‘": "'", "’": "'",   # ' '
    "‚": "'", "‛": "'",   # ‚ ‛ (rare)
    "`": "'",                   # grave accent used as quote
})

_DASH_ELLIPSIS_MAP = str.maketrans({
    "–": "-",   # en dash
    "—": "-",   # em dash
    "…": "...", # ellipsis character
})


def _normalize_for_match(s: str) -> str:
    """Canonical form for near-verbatim quote matching (exact after normalization)."""
    s = unicodedata.normalize("NFKC", s)
    s = s.translate(_CURLY_QUOTE_MAP)
    s = s.translate(_DASH_ELLIPSIS_MAP)
    s = s.replace(" ", " ")          # non-breaking space → space
    s = re.sub(r"\s+", " ", s)            # collapse whitespace runs
    s = s.casefold()
    s = s.strip()
    s = s.strip("\"'`.,;:!?()-–—…")      # strip leading/trailing punctuation
    return s


def _is_verified_quote(quote: str, blind_text: str) -> bool:
    """True iff the normalized quote appears as a normalized substring of blind_text."""
    nq = _normalize_for_match(quote)
    return bool(nq) and nq in _normalize_for_match(blind_text)


def _append_audit_entry(entry: dict) -> None:
    _AUDIT_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def verify_codes(
    result: Tier1Result,
    blind_text: str,
    transcript_label: str = "",
    n_participants: int = 0,
) -> tuple[Tier1Result, QuoteValidityStats]:
    """
    For every present=True code, check each supporting_quote against blind_text.
    - Individual non-matching quotes are dropped.
    - Codes with zero remaining quotes are demoted to present=False (quote_verified=False).
    - Tracks raw-exact vs normalization-recovered quotes; writes an audit entry for
      each normalization-recovered quote to analysis/coding_frame/quote_match_audit.jsonl.
    Returns the updated Tier1Result and a QuoteValidityStats summary.
    """
    total_quotes = 0
    verified_quotes = 0
    raw_exact_count = 0
    norm_recovered_count = 0
    total_present = 0
    count_verified_codes = 0
    count_demoted = 0
    updated_codes: list[SubthemeCode] = []

    for code in result.codes:
        if not code.present:
            updated_codes.append(code)
            continue

        total_present += 1
        good: list[SupportingQuote] = []
        bad_count = 0
        for q in code.supporting_quotes:
            total_quotes += 1
            raw_q = q.quote.strip()
            if raw_q and raw_q in blind_text:
                # Passed raw exact check
                good.append(q)
                verified_quotes += 1
                raw_exact_count += 1
            elif _is_verified_quote(q.quote, blind_text):
                # Failed raw but passed normalization-robust check — recoverable quote
                good.append(q)
                verified_quotes += 1
                norm_recovered_count += 1
                _append_audit_entry({
                    "transcript_label": transcript_label,
                    "subtheme_id": code.subtheme_id,
                    "turn_id": q.turn_id,
                    "raw_quote": q.quote,
                    "normalized_quote": _normalize_for_match(q.quote),
                })
            else:
                bad_count += 1

        if not good:
            # No verifiable evidence — demote to absent
            updated_codes.append(code.model_copy(update={
                "present": False,
                "quote_verified": False,
                "supporting_quotes": [],
                "unverified_quote_count": bad_count,
                "voiced_by": [],
                "reach": 0.0,
            }))
            count_demoted += 1
        else:
            voiced = sorted({q.speaker for q in good if q.speaker.lower() != "moderator"})
            reach = len(voiced) / n_participants if n_participants > 0 else 0.0
            updated_codes.append(code.model_copy(update={
                "supporting_quotes": good,
                "quote_verified": True,
                "unverified_quote_count": bad_count,
                "voiced_by": voiced,
                "reach": reach,
            }))
            count_verified_codes += 1

    stats = QuoteValidityStats(
        total_quotes=total_quotes,
        verified_quotes=verified_quotes,
        total_present_codes=total_present,
        verified_codes=count_verified_codes,
        demoted_codes=count_demoted,
        raw_exact_quotes=raw_exact_count,
        normalized_recovered_quotes=norm_recovered_count,
    )
    return Tier1Result(codes=updated_codes), stats


# ---------------------------------------------------------------------------
# Tier 1: deductive codebook coding, quote-grounded
# ---------------------------------------------------------------------------

_TIER1_SYSTEM = (
    "You are a qualitative research coding assistant performing deductive thematic coding.\n\n"
    "INPUT: a focus-group transcript with turns labelled [T001], [T002], etc., and a coding "
    "frame (list of subtheme codes with definitions).\n\n"
    "TASK: For EACH code in the frame, determine whether it is evidenced anywhere in the "
    "transcript.\n\n"
    "RULES:\n"
    "1. Apply codes exactly as defined. Do not apply a code unless at least one participant "
    "turn explicitly matches the code's definition.\n"
    "2. For every code marked present=true, provide ONE supporting_quotes entry FOR EACH "
    "DISTINCT PARTICIPANT who voices this subtheme — do not stop at one participant. "
    "Scan the whole transcript and list every participant who clearly expresses this subtheme. "
    "For each:\n"
    "   - turn_id: the exact turn label from the transcript (e.g. 'T034') — no brackets\n"
    "   - speaker: the generic speaker label exactly as shown (e.g. 'Participant 4' or 'Moderator')\n"
    "   - quote: COPY A SHORT VERBATIM PHRASE from that turn — an EXACT substring, NOT a "
    "paraphrase. Copy enough words to identify the moment (≤25 words is sufficient).\n"
    "3. If you cannot find a verbatim phrase that fits the code, set present=false and "
    "supporting_quotes=[].\n"
    "4. confidence: 'high' = direct explicit statement; 'medium' = clear inference; 'low' = "
    "marginal or indirect.\n"
    "5. evidence_note: one sentence explaining why the quote fits the code's definition.\n"
    "6. For present=false codes: confidence=null, supporting_quotes=[], evidence_note=null.\n\n"
    "RESPOND with ONLY a JSON object of this exact shape:\n"
    "{\"codes\": [\n"
    "  {\"subtheme_id\": \"A.1\", \"present\": true, \"confidence\": \"high\",\n"
    "   \"supporting_quotes\": [\n"
    "     {\"turn_id\": \"T034\", \"speaker\": \"Participant 4\", \"quote\": \"exact text here\"},\n"
    "     {\"turn_id\": \"T057\", \"speaker\": \"Participant 2\", \"quote\": \"another participant quote\"}\n"
    "   ],\n"
    "   \"evidence_note\": \"Participants 4 and 2 both explicitly describe X matching A.1.\"},\n"
    "  {\"subtheme_id\": \"A.2\", \"present\": false, \"confidence\": null,\n"
    "   \"supporting_quotes\": [], \"evidence_note\": null},\n"
    "  ...\n"
    "]}\n"
    "One entry per code in the provided frame, in the exact order given."
)


# The Tier-1 output cap. Overridable per evaluator_cfg; this stays the default.
TIER1_DEFAULT_MAX_OUTPUT_TOKENS = 32768

LAST_TIER1_CALL_TELEMETRY: dict = {}


def code_transcript_tier1(
    blind_text: str,
    codebook: list[dict],
    run_label: str,
    evaluator_cfg: dict | None = None,
) -> tuple[Tier1Result, QuoteValidityStats]:
    """
    Codes blind_text against the codebook via Gemini (Tier 1, deductive).
    Automatically verifies all supporting quotes as exact substrings.
    Returns (verified_result, quote_validity_stats).

    evaluator_cfg: entry from EVALUATOR_CONFIGS (model, key_env, temperature).
    When provided the call uses its own key and suppresses the backup fallback so
    evaluator-comparison runs stay on their designated key.
    """
    _ecfg = evaluator_cfg or {}
    _model = _ecfg.get("model", _MODEL)
    _cli = _client_for_evaluator(_ecfg) if evaluator_cfg else _client()

    # Default preserved at 32768. An evaluator_cfg may override it (preflight_v2 uses
    # 16384); the value actually used is transmitted, recorded in
    # effective_request_config and keyed into the cache, so a 16384 result can never
    # be served from a 32768 entry.
    _gen_cfg: dict = {
        "system_instruction": _TIER1_SYSTEM,
        "response_mime_type": "application/json",
        "max_output_tokens":  _ecfg.get("max_output_tokens", TIER1_DEFAULT_MAX_OUTPUT_TOKENS),
    }
    _temp = _ecfg.get("temperature", _TEMPERATURE) if evaluator_cfg else _TEMPERATURE
    if _temp is not None:
        _gen_cfg["temperature"] = _temp
    # Gemini 2.5-class models have a default thinking budget (~8192 tokens) that
    # competes with the output token cap.  Disable thinking for this mechanical
    # substring-matching task so all tokens are available for JSON output.
    if "2.5" in _model:
        _gen_cfg["thinking_config"] = {"thinking_budget": 0}

    n_participants = _count_participants(blind_text)
    # Stable prefix (codebook + task framing) is computed once and reused verbatim
    # so that Gemini's implicit context caching engages on the shared prefix.
    stable_prefix = _get_tier1_stable_prefix(codebook)
    prompt = stable_prefix + f"TRANSCRIPT:\n{blind_text}"
    _MAX_PARSE_RETRIES = 2
    _last_exc: Exception | None = None
    result: Tier1Result | None = None
    for _attempt in range(_MAX_PARSE_RETRIES + 1):
        response = _generate_with_fallback(
            _cli,
            no_fallback=evaluator_cfg is not None,
            model=_model,
            contents=prompt,
            config=_gen_cfg,
        )
        _log_call("tier1_codebook_coding", {
            "run_label": run_label,
            "model": _model,
            "usage": _usage_dict(response),
        })
        # Telemetry for truncation detection. A MAX_TOKENS finish reason must never be
        # allowed to look like a code that was legitimately absent.
        _cands = getattr(response, "candidates", None) or []
        _um = getattr(response, "usage_metadata", None)
        LAST_TIER1_CALL_TELEMETRY.clear()
        LAST_TIER1_CALL_TELEMETRY.update({
            "run_label": run_label,
            "model": _model,
            "parse_attempt": _attempt + 1,
            "max_output_tokens_requested": _gen_cfg.get("max_output_tokens"),
            "finish_reasons": [str(getattr(c, "finish_reason", None)) for c in _cands],
            "n_candidates": len(_cands),
            "prompt_tokens": getattr(_um, "prompt_token_count", None),
            "candidates_tokens": getattr(_um, "candidates_token_count", None),
            "total_tokens": getattr(_um, "total_token_count", None),
            "thoughts_tokens": getattr(_um, "thoughts_token_count", None),
            "cached_tokens": getattr(_um, "cached_content_token_count", None),
            "raw_text_chars": len(response.text or ""),
        })
        raw = _strip_fences(response.text)
        try:
            result = Tier1Result.model_validate(json.loads(raw))
            _last_exc = None
            break
        except (json.JSONDecodeError, ValueError) as exc:
            _last_exc = exc
            if _attempt < _MAX_PARSE_RETRIES:
                print(
                    f"  [retry] JSON parse error on attempt {_attempt + 1}: "
                    f"{str(exc)[:80]} — retrying ...",
                    flush=True,
                )
    if _last_exc is not None:
        raise _last_exc
    assert result is not None
    return verify_codes(result, blind_text, transcript_label=run_label, n_participants=n_participants)


# ---------------------------------------------------------------------------
# Tier 1 scoring
# ---------------------------------------------------------------------------

def _extract_theme_letter(subtheme_id: str) -> str:
    """'A.1' → 'A', 'B.3' → 'B', 'D' → 'D'."""
    return subtheme_id.split(".")[0] if "." in subtheme_id else subtheme_id[:1]


def compute_tier1_scores(
    real_result: Tier1Result,
    synthetic_result: Tier1Result,
) -> TierOneScores:
    """
    Compares verified present-code sets from a real and a synthetic transcript.
    Uses ONLY codes with quote_verified=True (or present=False) from both sides.

    Recall = what fraction of the real's codes the synthetic reproduces.
    Precision = what fraction of the synthetic's codes are also in the real.
    """
    real_present = frozenset(
        c.subtheme_id for c in real_result.codes if c.present and c.quote_verified
    )
    synth_present = frozenset(
        c.subtheme_id for c in synthetic_result.codes if c.present and c.quote_verified
    )

    shared_sub = real_present & synth_present
    union_sub = real_present | synth_present
    sub_recall = len(shared_sub) / len(real_present) if real_present else 0.0
    sub_prec = len(shared_sub) / len(synth_present) if synth_present else 0.0
    sub_f1 = (
        2 * sub_recall * sub_prec / (sub_recall + sub_prec)
        if (sub_recall + sub_prec) else 0.0
    )
    sub_jac = len(shared_sub) / len(union_sub) if union_sub else 0.0

    real_themes = frozenset(_extract_theme_letter(s) for s in real_present)
    synth_themes = frozenset(_extract_theme_letter(s) for s in synth_present)
    shared_th = real_themes & synth_themes
    union_th = real_themes | synth_themes
    th_recall = len(shared_th) / len(real_themes) if real_themes else 0.0
    th_prec = len(shared_th) / len(synth_themes) if synth_themes else 0.0
    th_f1 = (
        2 * th_recall * th_prec / (th_recall + th_prec)
        if (th_recall + th_prec) else 0.0
    )
    th_jac = len(shared_th) / len(union_th) if union_th else 0.0

    return TierOneScores(
        subtheme_recall=sub_recall,
        subtheme_precision=sub_prec,
        subtheme_f1=sub_f1,
        subtheme_jaccard=sub_jac,
        theme_recall=th_recall,
        theme_precision=th_prec,
        theme_f1=th_f1,
        theme_jaccard=th_jac,
        real_present=real_present,
        synthetic_present=synth_present,
        shared_subthemes=shared_sub,
        real_themes=real_themes,
        synthetic_themes=synth_themes,
        shared_themes=shared_th,
    )


# ---------------------------------------------------------------------------
# Tier 2: open theme extraction (codebook-free — deliberately separate call)
# ---------------------------------------------------------------------------

class Tier2Theme(BaseModel):
    theme_label: str = Field(..., description="3–8 word theme phrase.")
    theme_definition: str = Field(..., description="One sentence: what this theme captures.")
    supporting_quotes: list[SupportingQuote] = Field(default_factory=list)
    # Python-computed after substring verification — not from model
    participant_count: int = 0
    verified_quotes: list[SupportingQuote] = Field(default_factory=list)
    position_thirds: dict[str, int] = Field(default_factory=dict)  # early/middle/final → count


class Tier2Result(BaseModel):
    themes: list[Tier2Theme]


class Tier2ThemeMatch(BaseModel):
    real_theme_id: str        # e.g. "R0"
    synthetic_theme_id: str   # e.g. "S2"
    semantic_match: bool
    confidence: Literal["high", "medium", "low"]
    reason: str


class Tier2MatchBatch(BaseModel):
    matches: list[Tier2ThemeMatch]


@dataclass
class Tier2Scores:
    recall: float                          # matched real themes / all real themes
    precision: float                       # matched synth themes / all synth themes
    matched_pairs: list[tuple[int, int]]   # (real_idx, synth_idx)
    emergent_themes: list[Tier2Theme]      # synth-only (no human match)
    missed_themes: list[Tier2Theme]        # human-only (no synth match)
    disagreements: list[dict]              # Gemini ↔ embedding disagree


_TIER2_SYSTEM = (
    "You are a qualitative research assistant performing inductive open coding on a "
    "focus-group transcript. You have NO predefined codebook — work from the transcript only.\n\n"
    "TASK: Identify the most salient themes the participants' discussion centred on.\n\n"
    "RULES:\n"
    "- Extract ONLY themes strongly supported by MULTIPLE participant turns. "
    "Do not invent themes mentioned by only one participant or only once.\n"
    "- No fixed minimum: if only 3 themes are strongly supported, report 3. Up to 8 maximum.\n"
    "- For each theme, provide ONE supporting_quotes entry per DISTINCT PARTICIPANT who voices "
    "it — do not stop at one. Scan the whole transcript.\n"
    "- Quotes must be EXACT verbatim substrings of the transcript (≤25 words each).\n"
    "- Do not assume any external framework or codebook.\n\n"
    "Respond with ONLY a JSON object of this exact shape:\n"
    "{\"themes\": [\n"
    "  {\"theme_label\": \"3-8 word phrase\",\n"
    "   \"theme_definition\": \"One sentence describing what this theme captures.\",\n"
    "   \"supporting_quotes\": [\n"
    "     {\"turn_id\": \"T042\", \"speaker\": \"Participant 3\", \"quote\": \"exact text\"},\n"
    "     {\"turn_id\": \"T067\", \"speaker\": \"Participant 1\", \"quote\": \"exact text\"}\n"
    "   ]},\n"
    "  ...\n"
    "]}"
)

_TIER2_JUDGE_SYSTEM = (
    "You are a qualitative research expert judging whether pairs of focus-group themes "
    "are semantically equivalent — the same underlying topic, even if worded differently. "
    "Substantial overlap in meaning is required; superficial word overlap alone is not. "
    "Respond ONLY with valid JSON."
)


def _turn_third(turn_id: str, total_turns: int) -> str:
    """Return 'early', 'middle', or 'final' for a turn relative to total transcript length."""
    try:
        n = int(turn_id.lstrip("T"))
    except ValueError:
        return "unknown"
    if total_turns == 0:
        return "unknown"
    if n <= total_turns // 3:
        return "early"
    if n <= 2 * total_turns // 3:
        return "middle"
    return "final"


def verify_tier2_themes(
    result: Tier2Result,
    blind_text: str,
    n_participants: int,
) -> Tier2Result:
    """
    Verify Tier-2 supporting quotes as exact/normalized substrings.
    Compute participant_count from verified quotes only (evidence-constrained).
    Compute position_thirds for position-bias reporting.
    """
    total_turns = len(blind_text.splitlines())
    updated: list[Tier2Theme] = []
    for theme in result.themes:
        verified: list[SupportingQuote] = []
        thirds: dict[str, int] = {"early": 0, "middle": 0, "final": 0}
        for q in theme.supporting_quotes:
            if _is_verified_quote(q.quote, blind_text):
                verified.append(q)
                third = _turn_third(q.turn_id, total_turns)
                thirds[third] = thirds.get(third, 0) + 1
        speakers = {q.speaker for q in verified if q.speaker.lower() != "moderator"}
        updated.append(theme.model_copy(update={
            "verified_quotes": verified,
            "participant_count": len(speakers),
            "position_thirds": thirds,
        }))
    return Tier2Result(themes=updated)


def extract_themes_tier2(
    blind_text: str,
    run_label: str,
    evaluator_cfg: dict | None = None,
) -> Tier2Result:
    """
    Extracts open themes from blind_text via Gemini (Tier 2, inductive).
    Receives NO codebook — separate call prevents codebook contamination.
    Verifies quotes and computes participant_count in Python after the call.
    """
    _ecfg = evaluator_cfg or {}
    _model = _ecfg.get("model", _MODEL)
    _cli = _client_for_evaluator(_ecfg) if evaluator_cfg else _client()

    _gen_cfg: dict = {
        "system_instruction": _TIER2_SYSTEM,
        "response_mime_type": "application/json",
        "max_output_tokens":  32768,
    }
    _temp = _ecfg.get("temperature", _TEMPERATURE) if evaluator_cfg else _TEMPERATURE
    if _temp is not None:
        _gen_cfg["temperature"] = _temp
    if "2.5" in _model:
        _gen_cfg["thinking_config"] = {"thinking_budget": 0}

    n_participants = _count_participants(blind_text)

    _MAX_PARSE_RETRIES = 2
    _last_exc: Exception | None = None
    result: Tier2Result | None = None
    for _attempt in range(_MAX_PARSE_RETRIES + 1):
        response = _generate_with_fallback(
            _cli,
            no_fallback=evaluator_cfg is not None,
            model=_model,
            contents=f"TRANSCRIPT:\n{blind_text}",
            config=_gen_cfg,
        )
        _log_call("tier2_open_extraction", {
            "run_label": run_label,
            "model": _model,
            "usage": _usage_dict(response),
        })
        raw = _strip_fences(response.text)
        try:
            result = Tier2Result.model_validate(json.loads(raw))
            _last_exc = None
            break
        except (json.JSONDecodeError, ValueError) as exc:
            _last_exc = exc
            if _attempt < _MAX_PARSE_RETRIES:
                print(
                    f"  [retry] Tier-2 JSON parse error (attempt {_attempt + 1}): "
                    f"{str(exc)[:80]} — retrying ...",
                    flush=True,
                )
    if _last_exc is not None:
        raise _last_exc
    assert result is not None
    return verify_tier2_themes(result, blind_text, n_participants)


def _build_judge_prompt(
    real_themes: list[Tier2Theme],
    synth_themes: list[Tier2Theme],
) -> str:
    lines = ["REAL TRANSCRIPT THEMES:"]
    for i, t in enumerate(real_themes):
        lines.append(f"R{i}: {t.theme_label} — {t.theme_definition}")
    lines += ["", "SYNTHETIC TRANSCRIPT THEMES:"]
    for j, t in enumerate(synth_themes):
        lines.append(f"S{j}: {t.theme_label} — {t.theme_definition}")
    lines += [
        "",
        "For EVERY combination of (real theme Ri, synthetic theme Sj) determine whether "
        "they refer to the same underlying focus-group theme.",
        "",
        'Respond with ONLY: {"matches": ['
        '{"real_theme_id":"R0","synthetic_theme_id":"S1","semantic_match":true,'
        '"confidence":"high","reason":"brief reason"},'
        " ...]} — include exactly one entry per (Ri, Sj) pair, covering all combinations.",
    ]
    return "\n".join(lines)


def _embedding_similarities(
    real_themes: list[Tier2Theme],
    synth_themes: list[Tier2Theme],
) -> list[list[float]]:
    """
    Cosine similarity matrix [real_i][synth_j] using multilingual sentence embeddings.
    Returns zeros on import failure (sentence-transformers optional — diagnostic only).
    """
    try:
        from sentence_transformers import SentenceTransformer  # type: ignore
        import numpy as np  # type: ignore

        _st_model = SentenceTransformer("paraphrase-multilingual-mpnet-base-v2")

        def _text(t: Tier2Theme) -> str:
            return f"{t.theme_label}: {t.theme_definition}"

        real_emb  = _st_model.encode([_text(t) for t in real_themes],  normalize_embeddings=True)
        synth_emb = _st_model.encode([_text(t) for t in synth_themes], normalize_embeddings=True)
        return (np.array(real_emb) @ np.array(synth_emb).T).tolist()
    except ImportError:
        print("  [info] sentence-transformers not installed — embedding cross-check skipped.")
        return [[0.0] * len(synth_themes) for _ in real_themes]


def match_tier2_themes(
    real_result: Tier2Result,
    synth_result: Tier2Result,
    run_label: str,
    evaluator_cfg: dict | None = None,
) -> Tier2Scores:
    """
    Match real vs synthetic open themes using Gemini semantic judgement (primary)
    and multilingual embeddings (diagnostic cross-check).

    Returns Tier2Scores with recall, precision, emergent/missed themes, disagreements.
    """
    real_themes  = real_result.themes
    synth_themes = synth_result.themes

    if not real_themes or not synth_themes:
        return Tier2Scores(
            recall=0.0 if not real_themes else 1.0,
            precision=0.0,
            matched_pairs=[],
            emergent_themes=list(synth_themes),
            missed_themes=list(real_themes),
            disagreements=[],
        )

    # --- Gemini semantic judge (all pairs in one call) ---
    _ecfg = evaluator_cfg or {}
    _model = _ecfg.get("model", _MODEL)
    _cli = _client_for_evaluator(_ecfg) if evaluator_cfg else _client()

    _gen_cfg: dict = {"system_instruction": _TIER2_JUDGE_SYSTEM, "response_mime_type": "application/json"}
    _temp = _ecfg.get("temperature", _TEMPERATURE) if evaluator_cfg else _TEMPERATURE
    if _temp is not None:
        _gen_cfg["temperature"] = _temp

    judge_prompt = _build_judge_prompt(real_themes, synth_themes)

    _MAX_PARSE_RETRIES = 2
    _last_exc: Exception | None = None
    batch: Tier2MatchBatch | None = None
    for _attempt in range(_MAX_PARSE_RETRIES + 1):
        response = _generate_with_fallback(
            _cli,
            no_fallback=evaluator_cfg is not None,
            model=_model,
            contents=judge_prompt,
            config=_gen_cfg,
        )
        _log_call("tier2_theme_matching", {
            "run_label": run_label,
            "model": _model,
            "usage": _usage_dict(response),
        })
        raw = _strip_fences(response.text)
        try:
            batch = Tier2MatchBatch.model_validate(json.loads(raw))
            _last_exc = None
            break
        except (json.JSONDecodeError, ValueError) as exc:
            _last_exc = exc
            if _attempt < _MAX_PARSE_RETRIES:
                print(f"  [retry] Theme-match JSON error (attempt {_attempt + 1}): {str(exc)[:80]} — retrying ...", flush=True)
    if _last_exc is not None:
        raise _last_exc
    assert batch is not None

    # --- Embedding cross-check ---
    sims = _embedding_similarities(real_themes, synth_themes)

    # Build judgment lookup: (real_idx, synth_idx) → match result
    judge_map: dict[tuple[int, int], Tier2ThemeMatch] = {}
    for m in batch.matches:
        try:
            ri = int(m.real_theme_id.lstrip("R"))
            si = int(m.synthetic_theme_id.lstrip("S"))
            judge_map[(ri, si)] = m
        except (ValueError, AttributeError):
            pass

    # Flag disagreements (Gemini-yes + low embedding OR Gemini-no + high embedding)
    disagreements: list[dict] = []
    HIGH_SIM, LOW_SIM = 0.65, 0.35
    for ri in range(len(real_themes)):
        for si in range(len(synth_themes)):
            jm = judge_map.get((ri, si))
            sim = sims[ri][si] if sims else 0.0
            if jm and jm.semantic_match and sim < LOW_SIM:
                disagreements.append({"type": "gemini_yes_embed_low", "ri": ri, "si": si,
                                      "sim": round(sim, 3), "reason": jm.reason})
            elif jm and not jm.semantic_match and sim > HIGH_SIM:
                disagreements.append({"type": "gemini_no_embed_high", "ri": ri, "si": si,
                                      "sim": round(sim, 3), "reason": jm.reason})

    # Greedy matching: confidence-ordered, each theme matched at most once
    _conf_order = {"high": 0, "medium": 1, "low": 2}
    candidates = [
        (m, sims[int(m.real_theme_id.lstrip("R"))][int(m.synthetic_theme_id.lstrip("S"))])
        for m in batch.matches
        if m.semantic_match
        and m.real_theme_id.lstrip("R").isdigit()
        and m.synthetic_theme_id.lstrip("S").isdigit()
    ]
    candidates.sort(key=lambda x: (_conf_order.get(x[0].confidence, 9), -x[1]))

    matched_real:  set[int] = set()
    matched_synth: set[int] = set()
    matched_pairs: list[tuple[int, int]] = []
    for m, _sim in candidates:
        ri = int(m.real_theme_id.lstrip("R"))
        si = int(m.synthetic_theme_id.lstrip("S"))
        if ri not in matched_real and si not in matched_synth:
            matched_pairs.append((ri, si))
            matched_real.add(ri)
            matched_synth.add(si)

    emergent = [synth_themes[j] for j in range(len(synth_themes)) if j not in matched_synth]
    missed   = [real_themes[i]  for i in range(len(real_themes))  if i not in matched_real]

    recall    = len(matched_pairs) / len(real_themes)  if real_themes  else 0.0
    precision = len(matched_pairs) / len(synth_themes) if synth_themes else 0.0

    return Tier2Scores(
        recall=recall,
        precision=precision,
        matched_pairs=matched_pairs,
        emergent_themes=emergent,
        missed_themes=missed,
        disagreements=disagreements,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _strip_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = text[text.index("\n") + 1:] if "\n" in text else text[3:]
    if text.endswith("```"):
        text = text[: text.rfind("```")]
    return text.strip()


def _usage_dict(response) -> dict:
    usage = getattr(response, "usage_metadata", None)
    if usage is None:
        return {}
    return {
        "prompt_tokens":    getattr(usage, "prompt_token_count", None),
        "candidates_tokens": getattr(usage, "candidates_token_count", None),
        "total_tokens":     getattr(usage, "total_token_count", None),
        "cached_tokens":    getattr(usage, "cached_content_token_count", None),
    }
