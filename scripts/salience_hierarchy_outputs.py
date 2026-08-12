"""
Emit the salience-hierarchy deliverables. Offline; reads only existing Tier-1 results.

Order is deliberate: CSVs and the report are written and verified BEFORE the workbook
sheet is added, so the workbook is only touched once the numbers have been checked
against `thematic_reach_long.csv`.

    py scripts/salience_hierarchy_outputs.py
"""
from __future__ import annotations

import csv
import json
import os
import statistics
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))

import salience_hierarchy as sh   # noqa: E402

_RES = _ROOT / "analysis/production_evaluation/results"
_OUT = _ROOT / "analysis/production_evaluation/final"
_XL = _OUT / "FINAL_RESULTS_TABLES.xlsx"
SHEET = "10_Salience_Hierarchy"


def _write_csv(path: Path, rows, cols):
    tmp = path.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    os.replace(tmp, path)
    return len(rows)


# ------------------------------------------------------------ verification
def verify(b) -> dict:
    """Checks that must pass before the workbook is touched."""
    problems = []
    reach = list(csv.DictReader((_RES / "thematic_reach_long.csv").open(encoding="utf-8")))
    pres = list(csv.DictReader(
        (_RES / "thematic_code_presence_long.csv").open(encoding="utf-8")))

    # denominators and keys reconcile against the source
    src = {(r["condition"], r["fg"],
            r["canonical_replication_index"] if r["side"] == "synthetic" else "human",
            r["subtheme_id"]): r for r in reach}
    long = {(r["condition"], r["fg"],
             r["canonical_replication_index"] or "human", r["subtheme_id"]): r
            for r in b["theme_scores_long"]}
    if len(long) != len(b["theme_scores_long"]):
        problems.append("duplicate keys in theme_scores_long")
    n_checked = 0
    for k, r in src.items():
        got = long.get(k)
        if got is None:
            problems.append(f"reach row missing from long output: {k}")
            continue
        if abs(float(got["reach"]) - float(r["reach"])) > 1e-9:
            problems.append(f"reach mismatch at {k}: {got['reach']} vs {r['reach']}")
        den = int(r["participants_n"])
        if den <= 0:
            problems.append(f"non-positive denominator at {k}")
        if abs(float(r["reach"]) - int(r["voiced_by_n"]) / den) > 1e-6:
            problems.append(f"reach != voiced_by_n/participants_n at {k}")
        n_checked += 1

    # true absence -> 0 ; unmeasured -> null ; never swapped
    n_abs = sum(1 for r in b["theme_scores_long"] if r["status"] == "TRUE_ABSENCE")
    n_pres_false = sum(1 for r in pres if r["present"] != "True")
    if n_abs != n_pres_false:
        problems.append(f"true absences {n_abs} != presence rows false {n_pres_false}")
    for r in b["theme_scores_long"]:
        if r["status"] == "TRUE_ABSENCE" and r["reach"] != 0.0:
            problems.append(f"true absence not scored 0: {r}")
        if r["is_unmeasured_null"] and r["reach"] != "":
            problems.append(f"unmeasured coerced to a value: {r}")
        if r["reach"] == 0.0 and r["is_unmeasured_null"]:
            problems.append(f"null became 0: {r}")

    # the primary result covers every human theme
    for row in b["per_run"]:
        if row["n_scored"] + row["n_unmeasured_excluded"] != row["n_human_present"]:
            problems.append(f"primary universe incomplete for {row['fg']}/"
                            f"{row['condition']}/R{row['canonical_replication_index']}")
        if row["n_synthetic_recovered"] + row["n_human_themes_assigned_zero"] \
                != row["n_scored"]:
            problems.append(f"recovered+zero != scored for {row['fg']}")

    return {"problems": problems, "pass": not problems,
            "n_reach_rows_checked": n_checked,
            "n_true_absences_scored_zero": n_abs,
            "n_unmeasured_nulls": sum(1 for r in b["theme_scores_long"]
                                      if r["is_unmeasured_null"])}


# ------------------------------------------------------------------ PNG
def heatmap(b, path: Path):
    """Publication-ready recurrence heatmap with full English subtheme labels."""
    from PIL import Image, ImageDraw, ImageFont

    codebook_path = (_ROOT / "analysis/production_evaluation/gold_standard_sealed/"
                     "codebook_reference.csv")
    with codebook_path.open(encoding="utf-8") as f:
        labels = {r["subtheme_id"]: r["subtheme_label"] for r in csv.DictReader(f)}

    def font(size, bold=False):
        candidates = ([Path("C:/Windows/Fonts/arialbd.ttf")] if bold else []) + [
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("C:/Windows/Fonts/calibri.ttf"),
        ]
        for candidate in candidates:
            if candidate.exists():
                return ImageFont.truetype(str(candidate), size=size)
        return ImageFont.load_default()

    title_font = font(32, bold=True)
    subtitle_font = font(21)
    header_font = font(20, bold=True)
    label_font = font(20)
    value_font = font(22, bold=True)
    note_font = font(17)
    codes = b["codes"]
    cols = [("human", "human", "Human study")]
    for cond, lbl in (("enriched", "Enriched"),
                      ("demographics-only", "Demographics-\nonly")):
        for rep in ("1", "2", "3"):
            cols.append((cond, rep, f"{lbl} R{rep}"))
    hp = b["human_study_profile"]["n_fgs_present"]
    prof = {("human", "human"): hp}
    for s in b["study_replicates"]:
        prof[(s["condition"], s["canonical_replication_index"])] = \
            s["synthetic_n_fgs_present"]

    cw, ch, lm, tm = 175, 54, 410, 170
    W, H = lm + cw * len(cols) + 55, tm + ch * len(codes) + 155
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    d.text((28, 20), "Thematic salience: across-group recurrence",
           fill="#172033", font=title_font)
    d.text((28, 66),
           "Number of focus groups (0–5) in which each deductive subtheme was present",
           fill="#475569", font=subtitle_font)
    d.text((28, tm - 38), "Subtheme", fill="#172033", font=header_font)
    for j, (cond, rep, lbl) in enumerate(cols):
        box = d.multiline_textbbox((0, 0), lbl, font=header_font,
                                   spacing=2, align="center")
        tw = box[2] - box[0]
        d.multiline_text((lm + j * cw + (cw - tw) / 2, tm - 58), lbl,
                         fill="#172033", font=header_font, spacing=2,
                         align="center")
    for i, c in enumerate(codes):
        subtheme_label = labels.get(c, c)
        if c == "D" and subtheme_label.startswith("D) "):
            subtheme_label = subtheme_label[3:]
        label = f"{c} — {subtheme_label}"
        d.text((28, tm + i * ch + 15), label, fill="#172033", font=label_font)
        for j, (cond, rep, _lbl) in enumerate(cols):
            v = prof[(cond, rep)][c]
            f = v / 5.0
            # Pale blue -> dark navy, with zero kept visibly distinct from missing data.
            low, high = (241, 245, 249), (30, 64, 175)
            col = tuple(round(low[k] + (high[k] - low[k]) * f) for k in range(3))
            x0, y0 = lm + j * cw, tm + i * ch
            d.rectangle([x0, y0, x0 + cw - 5, y0 + ch - 5], fill=col,
                        outline="#CBD5E1", width=2)
            value = str(v)
            box = d.textbbox((0, 0), value, font=value_font)
            tw, th = box[2] - box[0], box[3] - box[1]
            d.text((x0 + (cw - 5 - tw) / 2, y0 + (ch - 5 - th) / 2 - 2), value,
                   fill="white" if f >= 0.6 else "#172033", font=value_font)

    note_y = tm + ch * len(codes) + 25
    d.text((28, note_y),
           "0 indicates measured absence from all focus groups in that study realisation; "
           "it is not missing data.", fill="#334155", font=note_font)
    d.text((28, note_y + 30),
           "The human column represents the single human study. Each synthetic column is "
           "one complete five-group study realisation; sessions are never pooled.",
           fill="#334155", font=note_font)
    d.text((28, note_y + 60),
           "Salience is operationalised as recurrence across focus groups, not as validated "
           "interpretive importance or centrality.", fill="#334155", font=note_font)
    tmp = path.with_suffix(".tmp.png")
    img.save(tmp)
    os.replace(tmp, path)
    return path


# ------------------------------------------------------------- workbook
def add_sheet(b):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    wb = load_workbook(_XL)
    before = {ws.title: {c.coordinate: c.value for r in ws.iter_rows() for c in r}
              for ws in wb}
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    bold = Font(bold=True)
    rows = [
        ["PARTICIPANT_BREADTH_AND_RECURRENCE_HIERARCHY_SIMILARITY"], [],
        ["Status", "EXPLORATORY — post-result operationalisation"],
        ["", sh.MANDATORY_STATEMENT], [],
        ["Legacy metric", "tier1_salience_hierarchy"],
        ["  reclassified as", "LEGACY_SHARED-ONLY_AUTOMATIC_DIAGNOSTIC"],
        ["  retained", "yes — but never a primary result"], [],
        ["Reach-based automated salience is SEPARATE from the researcher decision"],
        ["CENTRALITY_NOT_ASSESSED, which this analysis neither uses nor supersedes."], [],
        ["Primary universe", "every subtheme the human FG expressed"],
        ["True absence", "scored 0 — never deleted from the comparison"],
        ["Unmeasured", "null — never coerced to 0"], [],
        ["FG x condition — Kendall tau-b over all human themes"], [],
        ["fg", "condition", "median", "min", "max", "n_defined", "R1", "R2", "R3"],
    ]
    for c in b["by_fg_condition"]:
        rv = c["replicate_values"]
        rows.append([c["fg"], c["condition"], c["median_kendall_tau_b"],
                     c["min_kendall_tau_b"], c["max_kendall_tau_b"], c["n_defined"],
                     rv.get("1"), rv.get("2"), rv.get("3")])
    p = b["paired_summary"]
    rows += [[], ["Paired enriched - demographics-only, medians, n = 5 focus groups"], [],
             ["fg", "enriched_median", "demo_median", "difference", "direction"]]
    for r in p["paired_differences"]:
        rows.append([r["fg"], r["enriched_median"], r["demographics_only_median"],
                     r["difference_enriched_minus_demo"], r["direction"]])
    rows += [[], ["direction counts", json.dumps(p["direction_counts"])],
             ["median difference", p["median_difference"]],
             ["min / max difference", f"{p['min_difference']} / {p['max_difference']}"],
             ["unit of analysis", p["unit_of_analysis"]],
             ["inference", p["inference"]], [],
             ["Study-level recurrence — one row per complete study realisation"], [],
             ["condition", "replicate", "tau_b n_FGs_present", "tau_b mean_reach",
              "top3 overlap n_FGs", "top3 overlap mean_reach"]]
    for s in b["study_replicates"]:
        rows.append([s["condition"], s["canonical_replication_index"],
                     s["kendall_tau_b_n_fgs_present"], s["kendall_tau_b_mean_reach"],
                     s["top3_overlap_tie_aware_n_fgs"],
                     s["top3_overlap_tie_aware_mean_reach"]])
    rows += [[], ["The 15 sessions of a condition are NEVER treated as 15 independent "
                  "focus groups."]]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13)
    for cell in ("A17", "A17"):
        ws[cell].font = bold
    for col, w in (("A", 34), ("B", 26), ("C", 20), ("D", 14), ("E", 14), ("F", 12),
                   ("G", 10), ("H", 10), ("I", 10)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=False)

    tmp = _XL.with_suffix(".tmp")
    wb.save(tmp)
    os.replace(tmp, _XL)
    wb2 = load_workbook(_XL)
    after = {ws.title: {c.coordinate: c.value for r in ws.iter_rows() for c in r}
             for ws in wb2}
    changed = [n for n in before if before[n] != after.get(n)]
    return {"sheet": SHEET, "pre_existing_sheets_changed": changed,
            "n_sheets_before": len(before), "n_sheets_after": len(after)}


def main() -> int:
    b = sh.build()
    v = verify(b)
    print("=== verification (before any workbook write) ===")
    print(f"  reach rows checked        : {v['n_reach_rows_checked']}")
    print(f"  true absences scored 0    : {v['n_true_absences_scored_zero']}")
    print(f"  unmeasured nulls          : {v['n_unmeasured_nulls']}")
    print(f"  PASS                      : {v['pass']}")
    for p in v["problems"][:8]:
        print("     PROBLEM:", p)
    if not v["pass"]:
        print("\nrefusing to write outputs while verification fails")
        return 1

    n1 = _write_csv(_OUT / "salience_hierarchy_per_run.csv", b["per_run"], [
        "fg", "condition", "canonical_replication_index", "n_human_present", "n_scored",
        "n_synthetic_recovered", "n_human_themes_assigned_zero",
        "n_unmeasured_excluded", "n_ties_human", "n_ties_synthetic", "kendall_tau_b",
        "undefined_reason", "spearman_avg_ranks", "normalized_mean_abs_reach_diff",
        "top_theme_overlap_tie_aware", "union_n_themes", "union_kendall_tau_b",
        "union_undefined_reason"])
    n2 = _write_csv(_OUT / "salience_hierarchy_by_fg_condition.csv",
                    [{**c, "R1": c["replicate_values"].get("1"),
                      "R2": c["replicate_values"].get("2"),
                      "R3": c["replicate_values"].get("3")} for c in b["by_fg_condition"]],
                    ["fg", "condition", "n_replicates", "n_defined", "R1", "R2", "R3",
                     "median_kendall_tau_b", "min_kendall_tau_b", "max_kendall_tau_b"])
    n3 = _write_csv(_OUT / "salience_hierarchy_study_replicates.csv",
                    b["study_replicates"],
                    ["condition", "canonical_replication_index", "n_subthemes",
                     "kendall_tau_b_n_fgs_present", "undefined_reason_n_fgs",
                     "kendall_tau_b_mean_reach", "undefined_reason_mean_reach",
                     "top3_overlap_tie_aware_n_fgs", "top3_overlap_tie_aware_mean_reach",
                     "n_unmeasured_cells"])
    n4 = _write_csv(_OUT / "salience_hierarchy_theme_scores_long.csv",
                    b["theme_scores_long"],
                    ["condition", "fg", "canonical_replication_index", "subtheme_id",
                     "reach", "status", "is_true_absence_scored_zero",
                     "is_unmeasured_null"])
    hp = heatmap(b, _OUT / "salience_recurrence_heatmap.png")
    (_OUT / "salience_hierarchy.json").write_text(
        json.dumps({**b, "verification": v}, indent=1, ensure_ascii=False),
        encoding="utf-8")
    print(f"\nCSV rows: per_run {n1}, by_fg_condition {n2}, study_replicates {n3}, "
          f"theme_long {n4}")
    print(f"heatmap : {hp.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
