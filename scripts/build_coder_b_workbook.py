"""
Build the second coder's workbook for the N1 reliability subset.

Takes the units pre-registered for double coding (N1-001..N1-020) straight out
of the coder-A workbook, so coder B sees a byte-identical stimulus -- same
blinded text, same order, same instructions. Rebuilding the text from source
would risk a presentation difference, and a reliability coefficient computed
over two different stimuli measures nothing.

Coder A's labels are not carried over and are not visible anywhere in the file.

Usage:
    py scripts/build_coder_b_workbook.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
_OUT = _REPO_ROOT / "analysis" / "production_evaluation" / "consensus_dynamics" / "N1_triage"

SUBSET = ["N1-%03d" % i for i in range(1, 21)]
LABELS = ["divergencia", "alineacion", "ninguna"]


def main() -> None:
    src = load_workbook(_OUT / "CODING_WORKBOOK.xlsx")
    ws_a = src["Codificacion"]
    head = [c.value for c in ws_a[1]]
    i_uid, i_prev, i_resp, i_w = (head.index(k) for k in
                                  ("unit_id", "turno_previo", "turno_respuesta", "palabras"))

    wanted = set(SUBSET)
    picked = []
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        uid = str(row[i_uid]).strip() if row[i_uid] else ""
        if uid in wanted:
            picked.append((uid, row[i_prev], row[i_resp], row[i_w]))
    picked.sort(key=lambda r: SUBSET.index(r[0]))
    missing = wanted - {p[0] for p in picked}
    if missing:
        raise SystemExit(f"faltan unidades en el cuaderno origen: {sorted(missing)}")

    # keep coder A's sheet objects out of the new file entirely
    wb = load_workbook(_OUT / "CODING_WORKBOOK.xlsx")
    for name in wb.sheetnames:
        if name != "Instrucciones":
            del wb[name]
    ws = wb.create_sheet("Codificacion", 0)

    headers = ["unit_id", "turno_previo", "turno_respuesta", "palabras",
               "ETIQUETA", "nota (opcional)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for uid, prev, resp, w in picked:
        ws.append([uid, prev, resp, w, "", ""])

    for col, width in {"A": 10, "B": 70, "C": 90, "D": 9, "E": 16, "F": 30}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(LABELS), allow_blank=True)
    dv.error = "Elige una de las tres etiquetas."
    ws.add_data_validation(dv)
    dv.add(f"E2:E{ws.max_row}")

    ins = wb["Instrucciones"]
    ins.append([""])
    for line in [
        "SEGUNDA CODIFICADORA — subconjunto de fiabilidad",
        "",
        "Estas 20 unidades ya fueron codificadas por otra persona. No consultes sus",
        "etiquetas ni comentes los casos con ella antes de terminar: el objetivo es",
        "medir cuanto coinciden dos lecturas independientes, y cualquier consulta",
        "previa destruye esa medida.",
        "",
        "Mismas tres etiquetas y mismas reglas de borde de la hoja anterior.",
        f"{sum(p[3] for p in picked):,} palabras de respuesta. Unos 15-20 minutos.",
    ]:
        ins.append([line])
        if line.startswith("SEGUNDA"):
            ins.cell(row=ins.max_row, column=1).font = Font(bold=True)

    out = _OUT / "CODING_WORKBOOK_CODER_B.xlsx"
    wb.save(out)
    print(f"{len(picked)} unidades, {sum(p[3] for p in picked):,} palabras de respuesta")
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
