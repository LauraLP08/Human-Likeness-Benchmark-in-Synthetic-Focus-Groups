"""
Score the two-part gold standard.

STAGES
    --stage structural       check returned workbooks came back intact
    --stage emergent-pool    pool + author-blind Part-1 themes into the adjudication workbook
    --stage deductive-pool   fill coder columns and flag disagreements for adjudication
    --stage score            agreement, alpha and prevalence; evaluator comparison

WHAT THIS REFUSES TO DO
  * Fabricate. A missing or incomplete workbook is reported, never imputed.
  * Accept a damaged workbook. Deleted, added, duplicated or reordered rows FAIL.
  * Accept a positive deductive code without a quote that is a literal substring of
    that unit's excerpt.
  * Compare emergent themes to the evaluator before the coders' free-text themes
    have been clustered and adjudicated. Raw label matching would understate
    agreement, and any automatic similarity cut-off would manufacture it.
  * Report one pooled statistic as if it validated all 11 subthemes equally.

WHAT THE SAMPLE CAN AND CANNOT VALIDATE
Guide section 3 directly elicits subthemes A.1-A.3, so those are the codes this
exercise genuinely validates. B-D are not directly elicited by this question:
  * ABSENT  -> evidence about specificity / false-positive rate;
  * PRESENT -> opportunistic detection.
For B-D this is NOT a complete recall validation, and a code without enough
positive observations is reported as NOT_FULLY_VALIDATED rather than given a
reassuring-looking alpha.

Usage:
    py scripts/score_gold_standard.py --stage structural
    py scripts/score_gold_standard.py --stage emergent-pool
    py scripts/score_gold_standard.py --stage deductive-pool
    py scripts/score_gold_standard.py --stage score [--evaluator-codes p.json]
"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, UTC
from pathlib import Path

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_PKG = _OUT / "gold_standard_package"
_RETURNED = _OUT / "gold_standard_returned"
_ADJ = _OUT / "gold_standard_adjudication"
_SEALED = _OUT / "gold_standard_sealed" / "unit_id_to_source_SEALED.json"

# A subtheme needs at least this many positive observations, across both coders,
# before an agreement statistic on it is worth reporting as validation.
MIN_POSITIVE_OBSERVATIONS = 3

DIRECTLY_ELICITED_PREFIXES = ("A.",)      # section 3 elicits theme A directly
POOL_SEED = 20260730


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKC", s or "")
    s = (s.replace("’", "'").replace("‘", "'")
           .replace("“", '"').replace("”", '"')
           .replace("–", "-").replace("—", "-"))
    return re.sub(r"\s+", " ", s).strip().casefold()


def _unit_texts() -> dict[str, str]:
    """
    Full excerpt text per unit, concatenated from the ISSUED per-turn Units sheet.

    Read from the issued workbook, never a returned one: a quote must be checked
    against the text as sent, or a coder could paste a sentence into their own copy
    and then "verify" their own edit. The Units sheet is one row per TURN, so the
    turns are joined back together here.
    """
    wb = load_workbook(_PKG / "Coder_A_Part1_Emergent.xlsx")
    texts: dict[str, str] = {}
    for r in wb["Units"].iter_rows(min_row=2):
        if r[0].value:
            texts[r[0].value] = texts.get(r[0].value, "") + " " + (r[4].value or "")
    return texts


def _issued_keys(sheet: str, workbook: Path, ncols: int) -> list[tuple]:
    ws = load_workbook(workbook)[sheet]
    return [tuple(c.value for c in row[:ncols]) for row in ws.iter_rows(min_row=2)]


def _check_structure(returned: Path, template: Path, sheet: str,
                     key_cols: int) -> tuple[list[list], list[str]]:
    """Row-grid integrity. Deleted, added, duplicated or reordered rows all FAIL."""
    problems: list[str] = []
    if not returned.exists():
        return [], [f"not returned: {returned.name}"]
    expected = _issued_keys(sheet, template, key_cols)
    ws = load_workbook(returned)[sheet]
    rows = [list(r) for r in ws.iter_rows(min_row=2)]
    actual = [tuple(c.value for c in r[:key_cols]) for r in rows]

    if len(actual) != len(expected):
        problems.append(f"{returned.name}: {len(actual)} rows, expected {len(expected)} "
                        f"({'rows deleted' if len(actual) < len(expected) else 'rows added'})")
    dupes = [k for k, n in Counter(actual).items() if n > 1]
    if dupes:
        problems.append(f"{returned.name}: {len(dupes)} duplicated key(s), e.g. {dupes[:3]}")
    if actual != expected:
        first = next((i for i, (a, e) in enumerate(zip(actual, expected)) if a != e), None)
        if first is not None:
            problems.append(f"{returned.name}: row order/keys altered from row {first + 2} "
                            f"(got {actual[first]}, expected {expected[first]})")
    return rows, problems


# ---------------------------------------------------------------------------

def stage_structural() -> int:
    print("STRUCTURAL CHECK — returned workbooks")
    any_problem = False
    for coder in ("A", "B"):
        for part, sheet, cols, tmpl_dir in (
            ("Part1_Emergent", "Emergent_Coding", 2, _PKG),
            ("Part2_Deductive", "Deductive_Coding", 2, _PKG),
        ):
            name = f"Coder_{coder}_{part}.xlsx"
            tmpl = tmpl_dir / name
            if not tmpl.exists():
                print(f"  {name:<34} template not released yet — skipped")
                continue
            _, problems = _check_structure(_RETURNED / name, tmpl, sheet, cols)
            if problems and problems[0].startswith("not returned"):
                print(f"  {name:<34} NOT RETURNED")
                continue
            any_problem |= bool(problems)
            print(f"  {name:<34} {'OK' if not problems else 'FAIL'}")
            for p in problems:
                print(f"      {p}")
    return 1 if any_problem else 0


def stage_import() -> int:
    """
    Import returned coder workbooks into flat, reviewable CSVs.

    Runs the same structural checks as the release gate, so a workbook that would
    have been refused is not silently imported. Exports are a record of what was
    received; they are never edited in place of the workbook.
    """
    _OUT_IMPORT = _OUT / "gold_standard_imported"
    _OUT_IMPORT.mkdir(parents=True, exist_ok=True)
    any_problem = False
    summary = []

    for coder in ("A", "B"):
        name = f"Coder_{coder}_Part1_Emergent.xlsx"
        if (_RETURNED / name).exists():
            rows, problems = _check_structure(_RETURNED / name, _PKG / name,
                                              "Emergent_Coding", 2)
            any_problem |= bool(problems)
            out = []
            if not problems:
                for r in rows:
                    if (r[2].value or "").strip():
                        out.append({"coder": coder, "unit_id": r[0].value,
                                    "theme_slot": r[1].value,
                                    "theme_label": r[2].value,
                                    "theme_description": r[3].value,
                                    "supporting_quote": r[4].value,
                                    "relevance": r[5].value,
                                    "source_sheet": "Emergent_Coding"})
                wb = load_workbook(_RETURNED / name)
                if "Overflow_Themes" in wb.sheetnames:
                    for r in wb["Overflow_Themes"].iter_rows(min_row=2):
                        if r[0].value and (r[1].value or "").strip():
                            out.append({"coder": coder, "unit_id": r[0].value,
                                        "theme_slot": "overflow",
                                        "theme_label": r[1].value,
                                        "theme_description": r[2].value,
                                        "supporting_quote": r[3].value,
                                        "relevance": r[4].value,
                                        "source_sheet": "Overflow_Themes"})
                _write_csv(_OUT_IMPORT / f"part1_emergent_coder_{coder}.csv", out)
            summary.append((name, len(out), problems))
        else:
            summary.append((name, 0, ["not returned"]))

        name = f"Coder_{coder}_Part2_Deductive.xlsx"
        if (_RETURNED / name).exists():
            codes, problems = _load_deductive(coder)
            any_problem |= bool(problems)
            out = [{"coder": coder, "unit_id": u, "subtheme_id": sub, "present": v}
                   for (u, sub), v in sorted(codes.items())]
            if not problems:
                _write_csv(_OUT_IMPORT / f"part2_deductive_coder_{coder}.csv", out)
            summary.append((name, len(out), problems))
        else:
            summary.append((name, 0, ["not returned"]))

    print("IMPORT — returned workbooks")
    for name, n, problems in summary:
        status = "OK" if not problems else ("NOT RETURNED"
                                            if problems == ["not returned"] else "FAIL")
        print(f"  {name:<34} {status:<14} rows imported: {n}")
        for pr in problems[:5]:
            if pr != "not returned":
                print(f"      {pr}")
    return 1 if any_problem else 0


def _write_csv(path, rows: list[dict]) -> None:
    import csv as _csv
    if not rows:
        return
    fields = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def stage_emergent_pool() -> int:
    """Pool both coders' Part-1 themes, strip authorship, randomise, write to the
    adjudication workbook. Authorship is stripped so clustering cannot become a
    judgement about which coder was right."""
    rows_all = []
    for coder in ("A", "B"):
        name = f"Coder_{coder}_Part1_Emergent.xlsx"
        rows, problems = _check_structure(_RETURNED / name, _PKG / name,
                                          "Emergent_Coding", 2)
        if problems:
            print(f"CANNOT POOL — {name}")
            for p in problems:
                print(f"  {p}")
            return 2
        for r in rows:
            label = (r[2].value or "").strip()
            if not label:
                continue
            rows_all.append({
                "unit_id": r[0].value, "theme_label": label,
                "theme_description": (r[3].value or "").strip(),
                "supporting_quote": (r[4].value or "").strip(),
                "relevance": (r[5].value or "").strip(),
            })
    if not rows_all:
        print("CANNOT POOL — no themes recorded in either workbook.")
        return 2

    rng = random.Random(POOL_SEED)
    by_unit = defaultdict(list)
    for r in rows_all:
        by_unit[r["unit_id"]].append(r)

    wb = load_workbook(_ADJ / "Adjudication_Part1_Emergent.xlsx")
    ws = wb["Pooled_Themes"]
    ws.delete_rows(2, ws.max_row)
    n = 0
    for unit in sorted(by_unit):
        pool = by_unit[unit]
        rng.shuffle(pool)                       # authorship already dropped
        for i, r in enumerate(pool, start=1):
            ws.append([unit, f"{unit}-P{i:02d}", r["theme_label"],
                       r["theme_description"], r["supporting_quote"],
                       r["relevance"], "", "", "", ""])
            n += 1
    wb.save(_ADJ / "Adjudication_Part1_Emergent.xlsx")
    print(f"Pooled {n} themes across {len(by_unit)} units into "
          f"{(_ADJ / 'Adjudication_Part1_Emergent.xlsx').relative_to(_REPO_ROOT)}")
    print("Authorship stripped and order randomised. Cluster them, then re-run --stage score.")
    return 0


def _load_deductive(coder: str) -> tuple[dict, list[str]]:
    name = f"Coder_{coder}_Part2_Deductive.xlsx"
    rows, problems = _check_structure(_RETURNED / name, _PKG / name, "Deductive_Coding", 2)
    if problems:
        return {}, problems
    texts = _unit_texts()
    codes, quotes = {}, {}
    for r in rows:
        unit, sub = r[0].value, r[1].value
        raw = str(r[3].value).strip() if r[3].value is not None else ""
        if raw not in ("0", "1"):
            problems.append(f"{name}: {unit}/{sub} present_0_or_1={raw!r} (must be 0 or 1)")
            continue
        v = int(raw)
        codes[(unit, sub)] = v
        q = (r[4].value or "").strip()
        if v == 1:
            if not q:
                problems.append(f"{name}: {unit}/{sub} present=1 with no supporting quote")
            elif _norm(q) not in _norm(texts.get(unit, "")):
                problems.append(f"{name}: {unit}/{sub} quote is not a literal substring "
                                f"of the excerpt -> {q[:60]!r}")
        quotes[(unit, sub)] = q
    return codes, problems


def stage_deductive_pool() -> int:
    a, pa = _load_deductive("A")
    b, pb = _load_deductive("B")
    if pa or pb:
        print("CANNOT POOL — deductive workbooks not usable:")
        for p in (pa + pb)[:20]:
            print(f"  {p}")
        return 2
    wb = load_workbook(_ADJ / "Adjudication_Part2_Deductive.xlsx")
    ws = wb["Deductive_Adjudication"]
    for r in ws.iter_rows(min_row=2):
        k = (r[0].value, r[1].value)
        if k in a and k in b:
            r[3].value, r[4].value = a[k], b[k]
            r[5].value = "YES" if a[k] != b[k] else ""
            if a[k] == b[k]:
                r[6].value = a[k]          # agreed values carry through automatically
    wb.save(_ADJ / "Adjudication_Part2_Deductive.xlsx")
    dis = sum(1 for k in a if a[k] != b.get(k))
    print(f"Filled {len(a)} rows; {dis} disagreement(s) flagged for adjudication.")
    return 0


def krippendorff_alpha_binary(pairs: list[tuple[int, int]]) -> float | None:
    if len(pairs) < 2:
        return None
    co: Counter = Counter()
    for x, y in pairs:
        co[(x, y)] += 1
        co[(y, x)] += 1
    n = sum(co.values())
    marg: Counter = Counter()
    for (x, _y), c in co.items():
        marg[x] += c
    observed = sum(c for (x, y), c in co.items() if x != y) / n
    expected = sum(marg[x] * marg[y] for x in (0, 1) for y in (0, 1) if x != y) / (n * (n - 1))
    if expected == 0:
        return None
    return 1.0 - observed / expected


def stage_score(evaluator_codes: str | None) -> int:
    a, pa = _load_deductive("A")
    b, pb = _load_deductive("B")
    if pa or pb:
        print("CANNOT SCORE — deductive workbooks not usable. Nothing is imputed.\n")
        for p in (pa + pb)[:25]:
            print(f"  {p}")
        return 2

    shared = sorted(set(a) & set(b))
    pairs = [(a[k], b[k]) for k in shared]
    overall_alpha = krippendorff_alpha_binary(pairs)
    overall_agree = sum(1 for x, y in pairs if x == y) / len(pairs)

    print("DEDUCTIVE — global")
    print(f"  cells              : {len(shared)}")
    print(f"  raw agreement      : {overall_agree:.1%}")
    print(f"  Krippendorff alpha : "
          f"{'undefined' if overall_alpha is None else f'{overall_alpha:.3f}'}")
    print("  NOTE: this global figure is a summary only. It is NOT evidence that all "
          "subthemes were equally validated — see the per-subtheme table.")

    by_sub = defaultdict(list)
    for k in shared:
        by_sub[k[1]].append((a[k], b[k]))

    print("\nDEDUCTIVE — per subtheme")
    print(f"  {'sub':<7}{'n':>4}{'pos_A':>7}{'pos_B':>7}{'prev':>8}"
          f"{'agree':>8}{'alpha':>9}  status")
    per_sub = {}
    for sub in sorted(by_sub):
        ps = by_sub[sub]
        pos_a = sum(x for x, _ in ps)
        pos_b = sum(y for _, y in ps)
        prevalence = (pos_a + pos_b) / (2 * len(ps))
        ag = sum(1 for x, y in ps if x == y) / len(ps)
        al = krippendorff_alpha_binary(ps)
        elicited = sub.startswith(DIRECTLY_ELICITED_PREFIXES)
        enough = (pos_a + pos_b) >= MIN_POSITIVE_OBSERVATIONS
        if elicited and enough:
            status = "VALIDATED (directly elicited)"
        elif elicited:
            status = "NOT_FULLY_VALIDATED (too few positives)"
        elif enough:
            status = "OPPORTUNISTIC DETECTION (not recall validation)"
        else:
            status = "SPECIFICITY EVIDENCE ONLY (no/near-zero positives)"
        per_sub[sub] = {
            "n_cells": len(ps), "positives_coder_A": pos_a, "positives_coder_B": pos_b,
            "prevalence": round(prevalence, 4), "raw_agreement": round(ag, 4),
            "krippendorff_alpha": None if al is None else round(al, 4),
            "directly_elicited_by_section_3": elicited,
            "sufficient_positive_prevalence": enough,
            "validation_status": status,
        }
        print(f"  {sub:<7}{len(ps):>4}{pos_a:>7}{pos_b:>7}{prevalence:>8.2f}"
              f"{ag:>8.1%}{'   n/a' if al is None else f'{al:>9.3f}'}  {status}")

    result = {
        "scored_utc": datetime.now(UTC).isoformat(),
        "deductive_global": {
            "cells": len(shared), "raw_agreement": round(overall_agree, 4),
            "krippendorff_alpha": None if overall_alpha is None else round(overall_alpha, 4),
            "caveat": ("A single pooled statistic must not be read as evidence that all "
                       "11 subthemes were equally validated. See per_subtheme."),
        },
        "deductive_per_subtheme": per_sub,
        "validation_scope": {
            "directly_elicited": "A.1-A.3 — guide section 3 elicits theme A directly",
            "b_to_d_absent": "evidence about specificity / false-positive rate",
            "b_to_d_present": "opportunistic detection, not recall validation",
            "min_positive_observations": MIN_POSITIVE_OBSERVATIONS,
        },
    }

    # --- emergent: gated behind adjudication --------------------------------
    adj = _ADJ / "Adjudication_Part1_Emergent.xlsx"
    clusters = defaultdict(set)
    unclustered = 0
    if adj.exists():
        ws = load_workbook(adj)["Pooled_Themes"]
        for r in ws.iter_rows(min_row=2):
            if not r[0].value:
                continue
            cid = (r[6].value or "").strip()
            if cid:
                clusters[r[0].value].add(cid)
            else:
                unclustered += 1

    if not clusters or unclustered:
        reason = ("no pooled themes have been clustered" if not clusters
                  else f"{unclustered} pooled theme(s) still have no cluster_id")
        print(f"\nEMERGENT — comparison WITHHELD: {reason}.")
        print("  Emergent themes are not compared with the evaluator until the coders'")
        print("  free-text themes have been clustered and adjudicated. Comparing raw")
        print("  labels would understate agreement, and an automatic similarity cut-off")
        print("  would manufacture a number rather than measure one.")
        result["emergent"] = {"status": "WITHHELD_PENDING_ADJUDICATION", "reason": reason}
    else:
        result["emergent"] = {
            "status": "ADJUDICATED",
            "clusters_per_unit": {u: len(c) for u, c in sorted(clusters.items())},
            "total_clusters": sum(len(c) for c in clusters.values()),
            "note": ("Compare this adjudicated cluster set with the LLM Tier-2 themes. "
                     "Agreement is never computed from unmatched free-text labels."),
        }
        print(f"\nEMERGENT — adjudicated: {sum(len(c) for c in clusters.values())} "
              f"clusters across {len(clusters)} units, ready for Tier-2 comparison.")

    # --- evaluator vs adjudicated deductive ---------------------------------
    adj_d = _ADJ / "Adjudication_Part2_Deductive.xlsx"
    adjudicated = {}
    if adj_d.exists():
        for r in load_workbook(adj_d)["Deductive_Adjudication"].iter_rows(min_row=2):
            v = str(r[6].value).strip() if r[6].value is not None else ""
            if v in ("0", "1"):
                adjudicated[(r[0].value, r[1].value)] = int(v)

    if evaluator_codes and adjudicated:
        ev = json.loads(Path(evaluator_codes).read_text(encoding="utf-8"))
        ev_codes = {tuple(k.split("/")): int(v) for k, v in ev.items()}
        common = sorted(set(ev_codes) & set(adjudicated))
        tp = sum(1 for k in common if ev_codes[k] == 1 and adjudicated[k] == 1)
        fp = sum(1 for k in common if ev_codes[k] == 1 and adjudicated[k] == 0)
        fn = sum(1 for k in common if ev_codes[k] == 0 and adjudicated[k] == 1)
        tn = sum(1 for k in common if ev_codes[k] == 0 and adjudicated[k] == 0)
        prec = tp / (tp + fp) if tp + fp else None
        rec = tp / (tp + fn) if tp + fn else None
        print(f"\nEVALUATOR vs adjudicated reference (n={len(common)})")
        print(f"  TP={tp} FP={fp} FN={fn} TN={tn}  "
              f"precision={prec if prec is None else f'{prec:.3f}'}  "
              f"recall={rec if rec is None else f'{rec:.3f}'}")
        print("  Read per subtheme: recall is meaningful only for directly elicited codes.")
        result["evaluator_vs_reference"] = {
            "cells": len(common), "confusion_matrix": {"TP": tp, "FP": fp, "FN": fn, "TN": tn},
            "precision": prec, "recall": rec,
            "caveat": ("Recall is interpretable only for A.1-A.3. For B-D these cells "
                       "measure specificity when absent and opportunistic detection when "
                       "present."),
        }
    else:
        why = []
        if not evaluator_codes:
            why.append("--evaluator-codes not supplied")
        if not adjudicated:
            why.append("deductive adjudication not completed")
        print(f"\nEVALUATOR comparison skipped: {'; '.join(why)}")
        result["evaluator_vs_reference"] = {"skipped_because": why}

    out = _OUT / "gold_standard_results.json"
    out.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nWrote {out.relative_to(_REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--stage", required=True,
                    choices=["structural", "import", "emergent-pool",
                             "deductive-pool", "score"])
    ap.add_argument("--evaluator-codes", default=None)
    args = ap.parse_args()
    sys.exit({
        "structural": stage_structural,
        "import": stage_import,
        "emergent-pool": stage_emergent_pool,
        "deductive-pool": stage_deductive_pool,
        "score": lambda: stage_score(args.evaluator_codes),
    }[args.stage]())
