"""
Build all focus-group agent JSON files from the manifest spreadsheet.
One-shot script — run once, then verify output.

Usage: py scripts/build_fg_agents.py
"""
import sys, io, json, re, unicodedata
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

wb = openpyxl.load_workbook(
    "data/manifests/focus_group_dataset_manifest_v5.xlsx", read_only=True
)
BASE = Path("agents")


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )


def fg_int(val):
    s = str(val).strip()
    if s.startswith("FG"):
        return int(s[2:])
    return int(s)


def score_label(v):
    if v < 2.5:
        return "low"
    if v < 3.5:
        return "below-average"
    if v < 4.5:
        return "moderate"
    if v < 5.5:
        return "above-average"
    return "high"


# ── Psychometric score metadata (from DS03 manifest rows A30-A34) ────────

PSYCH = {
    "masculine_norms": {
        "construct": (
            "Traditional masculinity norms measured with the Male Role Norms "
            "Inventory (MRNI; Levant et al., 1992), subscales: avoidance of "
            "femininity, restrictive emotionality, aggression, achievement "
            "status, self-reliance, and attitudes toward sex (54 items; "
            "homophobia subscale excluded)."
        ),
        "direction": "Higher scores indicate stronger endorsement of traditional masculine norms.",
        "scale": "1–7 (strongly disagree to strongly agree)",
        "native_scale": "1–7",
    },
    "masculinity_of_meat": {
        "construct": (
            "Perceived masculinity of meat measured with the 7-item "
            "Masculinity of Meat scale (Lax & Mertig, 2020; e.g. "
            "'Men need to eat meat more than women do')."
        ),
        "direction": "Higher scores indicate stronger perceived masculinity-meat link and anti-vegetarian bias.",
        "scale": "1–7 (strongly disagree to strongly agree)",
        "native_scale": "1–7",
    },
    "meat_attachment": {
        "construct": (
            "Meat attachment measured with a 5-item shortened form of the "
            "Meat Attachment Questionnaire (Graça et al., 2015; e.g. "
            "'Meat is irreplaceable in my diet')."
        ),
        "direction": "Higher scores indicate stronger emotional and habitual attachment to eating meat.",
        "scale": "1–7 (strongly disagree to strongly agree)",
        "native_scale": "1–7",
    },
    "dairy_attachment": {
        "construct": (
            "Dairy attachment measured on an identical scale to meat "
            "attachment, with 'dairy' replacing 'meat' "
            "(Graça et al., 2015 adaptation)."
        ),
        "direction": "Higher scores indicate stronger emotional and habitual attachment to dairy consumption.",
        "scale": "1–7 (strongly disagree to strongly agree)",
        "native_scale": "1–7",
    },
    "vegetarianism_threat": {
        "construct": (
            "Perceived threat from the rise of vegetarianism measured with "
            "the 8-item Vegetarianism Threat Scale (Dhont & Hodson, 2014; "
            "e.g. 'The rise of vegetarianism poses a threat to our "
            "country’s cultural customs')."
        ),
        "direction": "Higher scores indicate stronger perceived societal or cultural threat from vegetarianism.",
        "scale": "1–7 (rescaled for comparability; see rescaling_note)",
        "native_scale": "1–5",
        "rescaling_note": (
            "Values in the dataset appear rescaled from the original 1–5 "
            "to 1–7 range, likely for comparability with other constructs."
        ),
    },
}

PK = list(PSYCH.keys())
FN = ["red_meat", "poultry", "fish", "egg", "dairy"]
SN = [
    "traditional masculinity norms",
    "perceived masculinity-meat link",
    "meat attachment",
    "dairy attachment",
    "vegetarianism threat",
]


# ═════════════════════════════════════════════════════════════════════════
# DS03 MACHO MEALS UK
# ═════════════════════════════════════════════════════════════════════════

ws = wb["DS03_MACHO_MEALS_UK"]
rows = list(ws.iter_rows(values_only=True))
d = BASE / "macho_meals"
built = ie = excl = 0

for r in rows[1:]:
    if r[0] is None or r[1] is None:
        continue
    try:
        fg = int(r[0])
    except (ValueError, TypeError):
        continue
    if r[2] is None:
        continue
    if fg == 3:
        excl += 1
        continue

    name = str(r[1]).strip()
    aid = f"mm_fg{fg}_{slugify(name)}"

    dem = {"name": name}
    if r[2] is not None:
        dem["age"] = int(r[2])
    if r[3]:
        dem["gender"] = str(r[3]).strip()
    if r[6]:
        dem["diet"] = str(r[6]).strip()
    loc = {}
    if r[5]:
        loc["urban_rural"] = str(r[5]).strip()
    if r[4]:
        loc["region"] = str(r[4]).strip()
    loc["country"] = "UK"
    dem["location"] = loc

    fc = {}
    for i, fn in enumerate(FN):
        v = r[12 + i]
        if v:
            fc[fn] = str(v).strip()

    ps = {}
    for i, pk in enumerate(PK):
        v = r[7 + i]
        if v is not None:
            e = dict(PSYCH[pk])
            e["value"] = float(v)
            e["provenance"] = "observed"
            ps[pk] = e

    # Notes: recruitment context only — psychometric scores are held out
    diet_s = str(r[6]).strip() if r[6] else "Meat eater"
    if diet_s.lower() != "meat eater":
        notes_text = f"Self-identified {diet_s.lower()}."
    else:
        notes_text = "Recruited as a regular meat eater."

    a = {
        "schema_version": "fg_agents_v1",
        "agent_id": aid,
        "language": "en",
        "field_provenance": {
            "language": "observed",
            "persona.demographics.name": "observed",
            "persona.demographics.age": "observed",
            "persona.demographics.gender": "observed",
            "persona.demographics.diet": "observed",
            "persona.demographics.location.urban_rural": "observed",
            "persona.demographics.location.region": "observed",
            "persona.demographics.location.country": "derived",
            "persona.food_consumption": "observed",
            "psychometric_scores": "observed",
            "simulation_config.notes": "derived",
        },
        "persona": {"demographics": dem, "food_consumption": fc},
        "psychometric_scores": ps,
        "study_context": {
            "dataset": "DS03_MACHO_MEALS_UK",
            "focus_group": f"FG{fg}",
        },
        "simulation_config": {
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
            "notes": notes_text,
        },
    }

    intro = str(r[17]).strip() if r[17] else None
    if intro and intro != "None":
        a["opening_intro"] = {
            "intro_eligible": True,
            "text": intro,
            "provenance": "observed_transcript_intro",
        }
        a["field_provenance"]["opening_intro.text"] = "observed_transcript_intro"
        a["field_provenance"]["opening_intro.intro_eligible"] = "derived"
        ie += 1

    write_json(d / f"{aid}.json", a)
    built += 1

write_json(
    d / "_manifest.json",
    {
        "dataset": "DS03_MACHO_MEALS_UK",
        "schema_version": "fg_agents_v1",
        "agents_built": built,
        "intro_eligible": ie,
        "excluded": {
            "fg3_dash_pid_participants": excl,
            "reason": (
                "Researcher confirmed these 5 participants (pseudonyms A–E) "
                "are not part of the study. Survey data cannot be matched to "
                "individual FG3 participants due to a PID recording error."
            ),
        },
    },
)
print(f"DS03 Macho Meals: {built} built, {ie} intro-eligible, {excl} FG3 excluded")


# ═════════════════════════════════════════════════════════════════════════
# DS01 SUSTAINABLE FASHION SPAIN
# ═════════════════════════════════════════════════════════════════════════

ws = wb["DS01_SUSTAINABLE_FASHION_SPAIN"]
rows = list(ws.iter_rows(values_only=True))
d = BASE / "sustainable_fashion"
built = ie = 0
SFG = {1: "never-bought-SF", 2: "never-bought-SF", 3: "has-bought-SF"}

for r in rows[1:]:
    if r[0] is None:
        continue
    try:
        fg = fg_int(r[0])
    except (ValueError, TypeError):
        continue
    name = str(r[2]).strip() if r[2] else None
    if not name:
        continue

    aid = f"sf_fg{fg}_{slugify(name)}"

    dem = {"name": name}
    if r[3] is not None:
        dem["age"] = int(r[3])
    if r[4]:
        dem["gender"] = str(r[4]).strip()

    a = {
        "schema_version": "fg_agents_v1",
        "agent_id": aid,
        "language": "es",
        "field_provenance": {
            "language": "observed",
            "persona.demographics.name": "observed",
            "persona.demographics.age": "observed",
            "persona.demographics.gender": "observed",
            "study_context.sf_group": "observed_paper_context",
        },
        "persona": {"demographics": dem},
        "study_context": {
            "dataset": "DS01_SUSTAINABLE_FASHION_SPAIN",
            "focus_group": f"FG{fg}",
            "participant_label": str(r[1]).strip() if r[1] else None,
            "sf_group": SFG.get(fg, "underspecified"),
        },
        "simulation_config": {
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
        },
    }

    intro = str(r[5]).strip() if r[5] else None
    if intro and intro != "None":
        a["opening_intro"] = {
            "intro_eligible": True,
            "text": intro,
            "provenance": "observed_transcript_intro",
        }
        a["field_provenance"]["opening_intro.text"] = "observed_transcript_intro"
        a["field_provenance"]["opening_intro.intro_eligible"] = "derived"
        ie += 1

    write_json(d / f"{aid}.json", a)
    built += 1

write_json(
    d / "_manifest.json",
    {
        "dataset": "DS01_SUSTAINABLE_FASHION_SPAIN",
        "schema_version": "fg_agents_v1",
        "agents_built": built,
        "intro_eligible": ie,
    },
)
print(f"DS01 Sustainable Fashion: {built} built, {ie} intro-eligible")


# ═════════════════════════════════════════════════════════════════════════
# DS04 DEEPFAKES US
# ═════════════════════════════════════════════════════════════════════════

ws = wb["DS04_DEEPFAKES_US"]
rows = list(ws.iter_rows(values_only=True))
d = BASE / "deepfakes"
built = ie = 0

for r in rows[1:]:
    if r[0] is None:
        continue
    try:
        fg = int(r[0])
    except (ValueError, TypeError):
        continue
    pnum = str(r[1]).strip() if r[1] else None
    if not pnum:
        continue
    ptype = str(r[2]).strip() if r[2] else None
    prof = str(r[3]).strip() if r[3] else None

    slug = re.sub(r"[^a-z0-9]+", "_", pnum.lower()).strip("_")
    aid = f"df_fg{fg}_{slug}"

    sp = {}
    if prof:
        sp["profession"] = prof
    if ptype:
        if ptype.lower() == "student":
            sp["participant_type"] = "Workshop participant (Make a Fake class, MIT)"
        elif ptype.lower() == "expert":
            sp["participant_type"] = "Expert in deepfake/AI technology"
        else:
            sp["participant_type"] = ptype

    a = {
        "schema_version": "fg_agents_v1",
        "agent_id": aid,
        "language": "en",
        "field_provenance": {
            "language": "observed",
            "persona.demographics.name": "derived",
        },
        "persona": {"demographics": {"name": pnum}},
        "study_context": {
            "dataset": "DS04_DEEPFAKES_US",
            "focus_group": f"FG{fg}",
            "participant_number": pnum,
        },
        "simulation_config": {
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
        },
    }

    if sp:
        a["persona"]["study_profile"] = sp
        if "profession" in sp:
            a["field_provenance"]["persona.study_profile.profession"] = "observed"
        if "participant_type" in sp:
            a["field_provenance"]["persona.study_profile.participant_type"] = "observed"

    intro = str(r[4]).strip() if r[4] else None
    if intro and intro != "None":
        a["opening_intro"] = {
            "intro_eligible": True,
            "text": intro,
            "provenance": "observed_transcript_intro",
        }
        a["field_provenance"]["opening_intro.text"] = "observed_transcript_intro"
        a["field_provenance"]["opening_intro.intro_eligible"] = "derived"
        ie += 1

    write_json(d / f"{aid}.json", a)
    built += 1

write_json(
    d / "_manifest.json",
    {
        "dataset": "DS04_DEEPFAKES_US",
        "schema_version": "fg_agents_v1",
        "agents_built": built,
        "intro_eligible": ie,
    },
)
print(f"DS04 Deepfakes: {built} built, {ie} intro-eligible")


# ═════════════════════════════════════════════════════════════════════════
# DS05 SAM MINDFULNESS
# ═════════════════════════════════════════════════════════════════════════

ws = wb["DS05_SAM_MINDFULNESS"]
rows = list(ws.iter_rows(values_only=True))
d = BASE / "mindfulness"
built = ie = 0

for i, r in enumerate(rows[1:], 1):
    rn = str(r[0]).strip() if r[0] else None
    if not rn:
        continue

    anon = f"MF_P{i}"
    aid = f"mf_p{i}"

    dem = {"name": anon}
    if r[1]:
        dem["gender"] = str(r[1]).strip()

    pp = {}
    if r[2]:
        pp["certified_mindfulness_instructor"] = str(r[2]).strip()
    if r[3]:
        pp["certificate"] = str(r[3]).strip()
    if r[4]:
        pp["has_taught_mbsr"] = str(r[4]).strip()
    if r[5] is not None:
        pp["years_experience"] = (
            int(r[5]) if isinstance(r[5], (int, float)) else r[5]
        )
    if r[6]:
        pp["country_of_residence"] = str(r[6]).strip()

    a = {
        "schema_version": "fg_agents_v1",
        "agent_id": aid,
        "language": "en",
        "provisional": True,
        "field_provenance": {
            "language": "observed",
            "persona.demographics.name": "derived",
            "persona.demographics.gender": "observed",
            "persona.professional_profile": "observed",
        },
        "persona": {"demographics": dem},
        "study_context": {
            "dataset": "DS05_SAM_MINDFULNESS",
            "participant_label": anon,
            "anonymisation_note": (
                "Real name withheld — participant identifiable from "
                "their certificate and public LinkedIn profile."
            ),
        },
        "simulation_config": {
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
        },
    }

    if pp:
        a["persona"]["professional_profile"] = pp

    li = str(r[7]).strip() if r[7] else None
    if li and li != "None":
        a["opening_intro"] = {
            "intro_eligible": True,
            "text": li,
            "provenance": "observed_external_profile",
        }
        a["field_provenance"]["opening_intro.text"] = "observed_external_profile"
        a["field_provenance"]["opening_intro.intro_eligible"] = "derived"
        ie += 1

    write_json(d / f"{aid}.json", a)
    built += 1

write_json(
    d / "_manifest.json",
    {
        "dataset": "DS05_SAM_MINDFULNESS",
        "schema_version": "fg_agents_v1",
        "agents_built": built,
        "intro_eligible": ie,
        "provisional": True,
        "provisional_note": (
            "Replica set may change — participant pool is small (5) "
            "and agent selection is provisional."
        ),
    },
)
print(f"DS05 Mindfulness: {built} built, {ie} intro-eligible (provisional)")

print("\nDONE")
