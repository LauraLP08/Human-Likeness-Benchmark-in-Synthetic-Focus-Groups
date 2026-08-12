"""
Excel workbook construction for the two-part gold standard.

Every coder-facing workbook carries: an Instructions sheet, a Units sheet holding
the full blind text, and one Coding sheet with data validation, autofilter, frozen
headers and wrapped text.

NOTHING IDENTIFYING IS EMBEDDED. No condition, no FG, no run label, no origin, no
sealed mapping, no hidden sheet and no defined name that could betray provenance.
`assert_no_provenance` re-opens each finished file and inspects every cell of every
sheet, so the guarantee is verified against the artefact rather than asserted.
"""

from __future__ import annotations

import math
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
_LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")
_ENTRY_FILL = PatternFill("solid", fgColor="FFFDE7")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")

# Excel auto-fits wrapped rows up to 409.5pt (~27 lines at 15pt). The Units text
# column is sized so that the longest turn in this corpus stays well inside that.
UNITS_TEXT_WIDTH = 135
MAX_AUTOFIT_LINES = 27

# Any of these appearing in a coder-facing workbook is a provenance leak.
_FORBIDDEN_PATTERNS = [
    r"\bmacho_meals\b", r"\bdemoonly\b", r"\brun0\d\b", r"\benriched\b",
    r"\bdemographics[- ]only\b", r"\bfg[1-5]\b", r"\bsynthetic\b",
    r"\bhuman\b", r"\bcanonical\b", r"\bsealed\b", r"\bcondition\b",
]


def _style_header(ws, ncols: int, freeze: str) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    ws.row_dimensions[1].height = 30


def _widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_instructions(wb: Workbook, title: str, lines: list[str]) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 118
    r = 3
    for line in lines:
        ws.cell(row=r, column=1, value=line).alignment = Alignment(
            wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = None if len(line) < 90 else 30
        r += 1
    ws.sheet_view.showGridLines = False


def add_units(wb: Workbook, units: list[dict]) -> None:
    """
    One row per TURN, not one cell per excerpt.

    A whole excerpt crammed into a single wrapped cell is unreadable and cannot be
    filtered, and a turn's own line breaks collapse inside it. Splitting by turn
    gives coders a readable, filterable transcript and lets them cite a turn id in
    their notes.
    """
    ws = wb.create_sheet("Units")
    ws.append(["unit_id", "turn_id", "speaker", "words", "turn_text"])
    for u in units:
        for rec in u["records"]:
            ws.append([u["unit_id"], rec["turn_id"], rec["speaker"],
                       len(rec["text"].split()), rec["text"]])
    _style_header(ws, 5, "C2")
    _widths(ws, {"A": 11, "B": 10, "C": 15, "D": 8, "E": UNITS_TEXT_WIDTH})

    # Row heights are deliberately NOT set. Excel auto-fits a wrapped cell only
    # while its height is left automatic; assigning an explicit height is exactly
    # what clips long turns. Leaving it unset guarantees every turn renders in
    # full, with the text itself untouched.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP

    # Auto-fit still cannot exceed Excel's 409.5pt ceiling, so verify at build
    # time that no turn is long enough to hit it rather than discovering a clipped
    # cell by eye later.
    overflowing = []
    for u in units:
        for rec in u["records"]:
            lines = (math.ceil(len(rec["text"]) / UNITS_TEXT_WIDTH)
                     + rec["text"].count(chr(10)))
            if lines > MAX_AUTOFIT_LINES:
                overflowing.append((u["unit_id"], rec["turn_id"], lines))
    if overflowing:
        raise ValueError(
            "Turns too long to render fully at the current column width — widen "
            f"UNITS_TEXT_WIDTH or split the Units sheet: {overflowing[:5]}")


def build_part1(path: Path, coder: str, units: list[dict], rows_per_unit: int = 12) -> None:
    """Emergent coding. The codebook is deliberately absent from this workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions(wb, f"Part 1 — Emergent coding  (Coder {coder})", [
        "WHAT THIS IS",
        "You will read 15 short focus-group excerpts and describe, in your own words, the "
        "principal ideas each one contains. There is no codebook in this workbook and you "
        "should not consult one. Work only from what is in the excerpt.",
        "",
        "ORDER MATTERS",
        "This is Part 1 of two. Part 2 uses a predefined codebook and will be sent to you "
        "only after this workbook is returned. Doing Part 1 without the codebook is the "
        "point: it keeps your emergent reading independent of the study's categories.",
        "",
        "WHAT TO DO",
        "1. Open the 'Units' sheet and read a unit in full before coding it.",
        "2. On the 'Emergent_Coding' sheet, find that unit's block of rows.",
        "3. For each principal idea you identify, complete one row:",
        "     theme_label            a short phrase, roughly 3-8 words",
        "     theme_description      one sentence saying what the idea captures",
        "     supporting_quote       VERBATIM text copied from the excerpt - not paraphrased",
        "     relevance              'central' or 'secondary'",
        "4. Use as many rows as you need, up to the 12 provided per unit. Leave unused "
        "rows blank.",
        "5. Record at least one theme for every unit. Every row you DO fill must have all "
        "four of: theme_label, theme_description, supporting_quote and relevance.",
        "6. If 12 rows are genuinely not enough for a unit, use the 'Overflow_Themes' "
        "sheet — add rows there freely and put the unit_id in the first column. Never "
        "add rows to 'Emergent_Coding'.",
        "",
        "WHAT 'CENTRAL' AND 'SECONDARY' MEAN",
        "central    - the excerpt substantially turns on this idea; removing it would change "
        "what the excerpt is about.",
        "secondary  - genuinely present and worth recording, but peripheral to the main thrust.",
        "",
        "RULES",
        "- Code independently. Do not discuss units with the other coder until both Part 1 "
        "workbooks are returned.",
        "- Quotes must be exact substrings of the excerpt. Copy and paste them.",
        "- Describe what is said, not what you infer the speaker believes.",
        "- Judge participant turns. Moderator turns are context, not evidence.",
        "- Do not add, delete, reorder or re-sort rows on 'Emergent_Coding'. The "
        "(unit_id, theme_slot) grid is fixed and is used to check the workbook came back "
        "intact. Extra themes go on 'Overflow_Themes' instead.",
        "- A partly filled row will be rejected. Either complete all four fields or leave "
        "the whole row blank.",
        "",
        "WHAT YOU ARE NOT TOLD, DELIBERATELY",
        "The excerpts come from more than one source, in unknown proportion and unknown "
        "order, and the ordering is randomised. This is intentional so that your reading is "
        "not shaped by beliefs about where an excerpt came from.",
    ])
    add_units(wb, units)

    ws = wb.create_sheet("Emergent_Coding")
    ws.append(["unit_id", "theme_slot", "theme_label", "theme_description",
               "supporting_quote", "relevance", "coder_notes"])
    for u in units:
        for slot in range(1, rows_per_unit + 1):
            ws.append([u["unit_id"], slot, "", "", "", "", ""])
    _style_header(ws, 7, "C2")
    _widths(ws, {"A": 11, "B": 11, "C": 34, "D": 52, "E": 60, "F": 13, "G": 30})

    dv = DataValidation(type="list", formula1='"central,secondary"', allow_blank=True,
                        showErrorMessage=True, errorTitle="Invalid relevance",
                        error="Choose 'central' or 'secondary'.")
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP
        row[0].fill = _LOCKED_FILL
        row[1].fill = _LOCKED_FILL
        for c in row[2:]:
            c.fill = _ENTRY_FILL

    # Safe continuation: extra themes go here, so the fixed grid on
    # 'Emergent_Coding' is never disturbed and the integrity check still holds.
    ov = wb.create_sheet("Overflow_Themes")
    ov.append(["unit_id", "theme_label", "theme_description", "supporting_quote",
               "relevance", "coder_notes"])
    _style_header(ov, 6, "B2")
    # Guidance sits OUTSIDE the six-column data grid. A hint placed in a data row
    # would itself be a populated row with no unit_id — precisely what the release
    # gate rejects.
    ov["H1"] = ("Use this sheet ONLY if a unit needs more than 12 themes. "
                "Add rows freely, but every row you fill MUST have the unit_id in "
                "column A, plus label, description, verbatim quote and relevance. "
                "Rows carrying data without a unit_id are rejected.")
    ov["H1"].alignment = _WRAP_TOP
    ov.column_dimensions["H"].width = 60
    ov.row_dimensions[1].height = 60
    _widths(ov, {"A": 11, "B": 34, "C": 52, "D": 60, "E": 13, "F": 34})
    dv2 = DataValidation(type="list", formula1='"central,secondary"', allow_blank=True)
    ov.add_data_validation(dv2)
    dv2.add("E2:E400")
    for row in ov.iter_rows(min_row=2, max_row=ov.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP
    wb.save(path)


def build_part2(path: Path, coder: str, units: list[dict], codebook: list[dict]) -> None:
    """Deductive coding against the 11-subtheme study codebook."""
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions(wb, f"Part 2 — Deductive coding  (Coder {coder})", [
        "WHAT THIS IS",
        "The same 15 excerpts you read in Part 1, now coded against a predefined codebook of "
        f"{len(codebook)} subthemes. The codebook is on the 'Codebook' sheet.",
        "",
        "DO PART 1 FIRST",
        "This workbook is released only after your Part 1 emergent workbook has been "
        "returned. Do not revise Part 1 in light of what you see here.",
        "",
        "WHAT TO DO",
        "1. Read the 'Codebook' sheet in full before coding.",
        "2. For each row on 'Deductive_Coding', decide whether that subtheme is evidenced in "
        "that unit.",
        "     present = 1  at least one participant turn clearly evidences the subtheme",
        "     present = 0  it does not",
        "3. Every present = 1 REQUIRES a verbatim supporting_quote copied from the excerpt. "
        "A positive code without a quote that matches the excerpt exactly will be rejected "
        "by the scoring script.",
        "4. Leave no present_0_or_1 cell blank.",
        "",
        "AN IMPORTANT ASYMMETRY",
        "The codebook was built for the study as a whole, not to cover this one question "
        "exhaustively. Expect many subthemes to be absent from many units. Absence is a real "
        "and useful finding - it is not a failure to find something, and you should not "
        "stretch a code to fit.",
        "",
        "RULES",
        "- Code independently. Do not discuss units with the other coder.",
        "- Code only what is in the excerpt. Do not infer from what you expect.",
        "- Judge participant turns. Moderator turns are context, not evidence.",
        "- Do not add, delete, reorder or re-sort rows. unit_id and subtheme_id are fixed and "
        "are used to check the workbook came back intact.",
    ])
    add_units(wb, units)

    cb = wb.create_sheet("Codebook")
    cb.append(["subtheme_id", "subtheme_label", "theme", "description", "example"])
    for c in codebook:
        cb.append([c.get("subtheme_id"), c.get("subtheme_label"), c.get("theme"),
                   c.get("description"), c.get("example")])
    _style_header(cb, 5, "A2")
    _widths(cb, {"A": 13, "B": 26, "C": 30, "D": 74, "E": 74})
    for row in cb.iter_rows(min_row=2, max_row=cb.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP
        cb.row_dimensions[row[0].row].height = 58

    ws = wb.create_sheet("Deductive_Coding")
    ws.append(["unit_id", "subtheme_id", "subtheme_label", "present_0_or_1",
               "supporting_quote", "coder_notes"])
    for u in units:
        for c in codebook:
            ws.append([u["unit_id"], c.get("subtheme_id"), c.get("subtheme_label"),
                       "", "", ""])
    _style_header(ws, 6, "D2")
    _widths(ws, {"A": 11, "B": 13, "C": 26, "D": 15, "E": 66, "F": 30})

    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False,
                        showErrorMessage=True, errorTitle="Invalid value",
                        error="Enter 0 (absent) or 1 (present).")
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP
        for c in row[:3]:
            c.fill = _LOCKED_FILL
        for c in row[3:]:
            c.fill = _ENTRY_FILL
    wb.save(path)


def build_emergent_adjudication(path: Path, units: list[dict]) -> None:
    """
    Blinded clustering of the two coders' emergent themes.

    Free-text labels cannot be compared directly — two coders will name the same
    idea differently, so raw label matching would understate agreement and inventing
    a similarity threshold would manufacture it. The coders' themes are therefore
    pooled per unit, stripped of any indication of who wrote them, and clustered by
    human judgement into semantically equivalent groups. Only the adjudicated
    cluster set is compared with the LLM Tier-2 themes.
    """
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions(wb, "Emergent adjudication — theme clustering", [
        "WHEN TO USE THIS",
        "After BOTH Part 1 emergent workbooks have been returned. Not before.",
        "",
        "WHY THIS STEP EXISTS",
        "Two coders describing the same idea will rarely choose the same words. Comparing "
        "their free-text labels directly would understate agreement, and picking a string- "
        "or embedding-similarity cut-off would manufacture a number rather than measure one. "
        "So equivalence is decided by coder judgement, once, and recorded.",
        "",
        "HOW TO DO IT",
        "1. Run: py scripts/score_gold_standard.py --stage emergent-pool",
        "   This fills the 'Pooled_Themes' sheet with every theme both coders recorded, "
        "   per unit, in randomised order and with the author hidden.",
        "2. For each unit, group themes that express the same underlying idea.",
        "3. Give each group a cluster_id that is unique WITHIN that unit: C1, C2, C3 ...",
        "4. Give each cluster a short adjudicated_label and a one-sentence "
        "adjudicated_description.",
        "5. A theme that matches nothing else still forms its own cluster.",
        "6. Set cluster_relevance to 'central' or 'secondary'.",
        "",
        "IMPORTANT",
        "- You are clustering ideas, not judging who was right.",
        "- Do not look up which coder wrote which theme; the sheet is deliberately blinded.",
        "- Do not add or delete rows. Fill cluster_id for every populated row.",
        "- Only after this is complete is the adjudicated set compared with the automated "
        "  Tier-2 themes.",
    ])
    add_units(wb, units)

    ws = wb.create_sheet("Pooled_Themes")
    ws.append(["unit_id", "pooled_row_id", "theme_label", "theme_description",
               "supporting_quote", "relevance", "cluster_id", "adjudicated_label",
               "adjudicated_description", "cluster_relevance"])
    _style_header(ws, 10, "C2")
    _widths(ws, {"A": 11, "B": 14, "C": 32, "D": 48, "E": 52, "F": 12,
                 "G": 12, "H": 32, "I": 48, "J": 18})
    note = wb.create_sheet("HOW_TO_POPULATE")
    note["A1"] = ("Run:  py scripts/score_gold_standard.py --stage emergent-pool\n"
                  "It writes the pooled, author-blinded themes into 'Pooled_Themes'.\n"
                  "Do not paste them in by hand — the pooling step randomises order and "
                  "strips authorship, and doing it manually would defeat both.")
    note["A1"].alignment = _WRAP_TOP
    note.column_dimensions["A"].width = 110
    note.row_dimensions[1].height = 70
    wb.save(path)


def build_deductive_adjudication(path: Path, units: list[dict], codebook: list[dict]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb, "Deductive adjudication — resolving disagreements", [
        "WHEN TO USE THIS",
        "After BOTH Part 2 deductive workbooks have been returned.",
        "",
        "HOW TO DO IT",
        "1. Run: py scripts/score_gold_standard.py --stage deductive-pool",
        "   This fills coder_A_present and coder_B_present and flags disagreements.",
        "2. For every flagged row, agree a final adjudicated_present (0 or 1).",
        "3. Record a one-line adjudication_rationale.",
        "4. Rows where both coders agree may be left as they are; the script carries the "
        "   agreed value through automatically.",
        "",
        "WHAT THIS IS FOR",
        "The adjudicated codes become the reference standard against which the automated "
        "evaluator is scored. Agreement statistics themselves are computed from the two "
        "INDEPENDENT coder sets, not from these adjudicated values.",
    ])
    add_units(wb, units)
    ws = wb.create_sheet("Deductive_Adjudication")
    ws.append(["unit_id", "subtheme_id", "subtheme_label", "coder_A_present",
               "coder_B_present", "disagreement", "adjudicated_present",
               "adjudication_rationale"])
    for u in units:
        for c in codebook:
            ws.append([u["unit_id"], c.get("subtheme_id"), c.get("subtheme_label"),
                       "", "", "", "", ""])
    _style_header(ws, 8, "D2")
    _widths(ws, {"A": 11, "B": 13, "C": 26, "D": 16, "E": 16, "F": 14, "G": 19, "H": 60})
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _WRAP_TOP
    wb.save(path)


def assert_no_provenance(path: Path) -> list[str]:
    """Re-open a finished workbook and scan every cell of every sheet."""
    problems: list[str] = []
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            problems.append(f"{path.name}/{ws.title}: sheet is not visible ({ws.sheet_state})")
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                low = cell.value.lower()
                for pat in _FORBIDDEN_PATTERNS:
                    if re.search(pat, low):
                        problems.append(
                            f"{path.name}/{ws.title}!{cell.coordinate}: matches {pat!r} "
                            f"-> {cell.value[:60]!r}")
    if wb.defined_names:
        problems.append(f"{path.name}: defined names present: {list(wb.defined_names)}")
    return problems
