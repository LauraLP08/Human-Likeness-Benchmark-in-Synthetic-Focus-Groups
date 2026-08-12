"""
Score the N1 triage: per-side error of the D1 divergence detector.

Reads the filled CODING_WORKBOOK.xlsx, unseals the unit map, and reports the
one number that decides whether the raw D1 human-vs-synthetic contrast is
interpretable at all: the gap between per-side recall.

Estimation. The sample is stratified with known inclusion probabilities, so
every rate is a Horvitz-Thompson estimate: each coded unit stands for
N_stratum / n_drawn units of the corpus. Reading the sample counts directly
would badly overstate divergence, because the flagged strata are deliberately
oversampled by 7-24x.

Intervals are Wilson intervals on the EFFECTIVE sample size (Kish), not on the
weighted count -- weighting buys unbiasedness, not precision. With n=40 per
side these are wide by design; N1 is a go/no-go, not an estimate to quote.

Usage:
    py scripts/score_consensus_triage.py
    py scripts/score_consensus_triage.py --coder-b path/to/CODER_B.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
_OUT = _REPO_ROOT / "analysis" / "production_evaluation" / "consensus_dynamics" / "N1_triage"

POSITIVE = "divergencia"


def _read_labels(path: Path) -> dict[str, str]:
    ws = load_workbook(path, data_only=True)["Codificacion"]
    head = [c.value for c in ws[1]]
    i_uid, i_lab = head.index("unit_id"), head.index("ETIQUETA")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_uid] and row[i_lab]:
            out[str(row[i_uid]).strip()] = str(row[i_lab]).strip().lower()
    return out


def _wilson(k: float, n: float, z: float = 1.96) -> tuple[float, float]:
    if n <= 0:
        return (float("nan"), float("nan"))
    p = k / n
    d = 1 + z * z / n
    c = (p + z * z / (2 * n)) / d
    h = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return (max(0.0, c - h), min(1.0, c + h))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(_OUT / "CODING_WORKBOOK.xlsx"))
    ap.add_argument("--coder-b", default=None)
    args = ap.parse_args()

    sealed = json.loads((_OUT / "unit_map_SEALED.json").read_text(encoding="utf-8"))["units"]
    labels = _read_labels(Path(args.workbook))
    if not labels:
        raise SystemExit("No labels found in ETIQUETA. Code the workbook first.")
    missing = [u for u in sealed if u not in labels]
    print(f"coded {len(labels)}/{len(sealed)} units"
          + (f"  ({len(missing)} pending)" if missing else ""))

    # ---- weights on what was ACTUALLY coded --------------------------------
    # Coding may stop early. Unit order was randomised under a fixed seed
    # independent of stratum and side, so a prefix of the workbook is a random
    # subsample of the sample -- but the inclusion probability is then
    # n_coded/N_stratum, not n_drawn/N_stratum. Using n_drawn here would
    # silently overweight whichever strata happen to be under-coded.
    coded_per_cell: dict[tuple[str, str], int] = defaultdict(int)
    for uid in labels:
        m = sealed[uid]
        coded_per_cell[(m["side"], m["stratum"])] += 1

    print(f"\n{'side':<11}{'stratum':<18}{'N':>6}{'sorteadas':>11}{'codificadas':>13}{'peso':>9}")
    for uid_meta in sorted({(m["side"], m["stratum"], m["N_stratum"], m["n_drawn"])
                            for m in sealed.values()}):
        side, stratum, N, drawn = uid_meta
        c = coded_per_cell[(side, stratum)]
        w = N / c if c else float("nan")
        print(f"{side:<11}{stratum:<18}{N:>6}{drawn:>11}{c:>13}{w:>9.2f}")
        if c == 0:
            print(f"{'':<11}  ^ estrato sin codificar: no contribuye a la estimacion")

    # ---- Horvitz-Thompson accumulation ------------------------------------
    # per side: weighted TP / FP / FN / TN of D1(opening) against the human label
    acc: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    eff: dict[str, list[float]] = defaultdict(list)

    for uid, lab in labels.items():
        meta = sealed[uid]
        w = meta["N_stratum"] / coded_per_cell[(meta["side"], meta["stratum"])]
        side = meta["side"]
        pred_pos = meta["d1_label"] in ("divergence", "mixed")
        true_pos = lab == POSITIVE
        cell = ("tp" if true_pos else "fp") if pred_pos else ("fn" if true_pos else "tn")
        acc[side][cell] += w
        acc[side]["n_true_pos"] += w if true_pos else 0.0
        acc[side]["N"] += w
        eff[side].append(w)

    print(f"\n{'side':<11} {'prevalencia HT':>15} {'recall D1':>22} {'precision D1':>22}")
    summary = {}
    for side in ("human", "synthetic"):
        a = acc[side]
        ws_ = eff[side]
        n_eff = (sum(ws_) ** 2) / sum(w * w for w in ws_) if ws_ else 0.0
        prev = a["n_true_pos"] / a["N"] if a["N"] else float("nan")
        rec = a["tp"] / (a["tp"] + a["fn"]) if (a["tp"] + a["fn"]) else float("nan")
        prc = a["tp"] / (a["tp"] + a["fp"]) if (a["tp"] + a["fp"]) else float("nan")
        lo, hi = _wilson(prev * n_eff, n_eff)
        summary[side] = {"prevalence_ht": prev, "prevalence_ci": [lo, hi],
                         "recall": rec, "precision": prc, "n_eff": n_eff,
                         "n_coded": len(ws_)}
        print(f"{side:<11} {prev:>14.3f} [{lo:.2f},{hi:.2f}] {rec:>21.3f} {prc:>21.3f}")

    # ---- the go/no-go -----------------------------------------------------
    rh, rs = summary["human"]["recall"], summary["synthetic"]["recall"]
    gap = abs(rh - rs) if not (math.isnan(rh) or math.isnan(rs)) else float("nan")
    print("\n" + "-" * 72)
    print(f"GAP DE RECALL ENTRE LADOS: {gap:.3f}")
    if math.isnan(gap):
        verdict = "INDETERMINADO — un lado no tiene positivos verdaderos suficientes"
    elif gap < 0.15:
        verdict = ("PASA — el sesgo diferencial es pequeno; el contraste crudo de D1 "
                   "es interpretable (con su banda)")
    elif gap < 0.30:
        verdict = ("MARGINAL — reportar D1 solo junto a un segundo detector; "
                   "no sostener una conclusion direccional con D1 solo")
    else:
        verdict = ("NO PASA — D1 mide registro, no postura. El contraste crudo no es "
                   "interpretable; escalar a N2 o basar la conclusion en D2/D4/D5")
    print(f"VEREDICTO: {verdict}")
    print("-" * 72)

    # ---- reliability ------------------------------------------------------
    # Cohen's kappa alone is misleading here. With marginals as skewed as these
    # (divergence is rare), kappa is depressed even when raw agreement is high
    # -- the well-known kappa paradox. Gwet's AC1 is reported alongside because
    # it is built to be stable under skew. Where the two disagree, the skew is
    # the explanation, and both belong in the write-up.
    reliability = None
    if args.coder_b:
        b = _read_labels(Path(args.coder_b))
        shared = sorted(set(labels) & set(b))
        if shared:
            cats = sorted({labels[u] for u in shared} | {b[u] for u in shared})
            n = len(shared)
            po = sum(labels[u] == b[u] for u in shared) / n

            pa = {c: sum(labels[u] == c for u in shared) / n for c in cats}
            pb = {c: sum(b[u] == c for u in shared) / n for c in cats}
            pe_k = sum(pa[c] * pb[c] for c in cats)
            kappa = (po - pe_k) / (1 - pe_k) if pe_k < 1 else float("nan")

            q = len(cats)
            pi = {c: (pa[c] + pb[c]) / 2 for c in cats}
            pe_g = sum(pi[c] * (1 - pi[c]) for c in cats) / (q - 1) if q > 1 else 0.0
            ac1 = (po - pe_g) / (1 - pe_g) if pe_g < 1 else float("nan")

            print(f"\nfiabilidad sobre {n} unidades dobles")
            print(f"  acuerdo bruto      {po:.3f}")
            print(f"  kappa de Cohen     {kappa:.3f}")
            print(f"  AC1 de Gwet        {ac1:.3f}")
            print(f"  {'categoria':<14}{'A':>4}{'B':>4}{'coinciden':>11}")
            for c in cats:
                both = sum(1 for u in shared if labels[u] == c and b[u] == c)
                print(f"  {c:<14}{int(pa[c]*n):>4}{int(pb[c]*n):>4}{both:>11}")
            disagreements = [(u, labels[u], b[u]) for u in shared if labels[u] != b[u]]
            if disagreements:
                print("  desacuerdos:", ", ".join(f"{u}({a}/{bb})"
                                                  for u, a, bb in disagreements))
            reliability = {"n": n, "raw_agreement": po, "cohen_kappa": kappa,
                           "gwet_ac1": ac1,
                           "disagreements": [list(d) for d in disagreements]}

    (_OUT / "N1_RESULTS.json").write_text(json.dumps({
        "n_coded": len(labels), "n_units": len(sealed),
        "per_side": summary, "recall_gap": gap, "verdict": verdict,
        "reliability": reliability,
        "estimator": "Horvitz-Thompson; Wilson CI on Kish effective n",
        "detector": "D1 opening-window (OPENING_CLAUSES=2)",
    }, indent=2), encoding="utf-8")
    print(f"\nwrote {_OUT / 'N1_RESULTS.json'}")


if __name__ == "__main__":
    main()
