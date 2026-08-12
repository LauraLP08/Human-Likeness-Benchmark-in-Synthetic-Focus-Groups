"""
N1 triage sample — human validation of the D1 divergence detector.

Purpose (and its limit): N1 does NOT correct the corpus rate. It answers one
question, which is the only one that can invalidate the whole comparison:
    does D1 detect divergence at a materially different rate on the human side
    than on the synthetic side?
If it does, the raw D1 contrast is uninterpretable regardless of its direction.
N2 (240 units, Horvitz-Thompson + prediction-powered correction) is what buys
corrected rates; N1 buys the go/no-go.

Design
------
80 units, 40 per side, stratified on the PRIMARY D1 label (opening window) with
known inclusion probabilities, so per-side recall and precision are estimable
rather than eyeballed. Divergence-flagged human acts are taken as a census
(N=10) because that stratum is too small to sample.

Blinding: speaker names masked, side / fg / run / D1 label withheld, unit order
shuffled under a fixed seed. Blinding is imperfect by construction -- a 250-word
turn is recognisable next to a 37-word one -- which is exactly why the estimand
is per-side detector error rather than an assumed-symmetric one.

Outputs
-------
    consensus_dynamics/N1_triage/CODING_WORKBOOK.xlsx     <- the coder's file
    consensus_dynamics/N1_triage/sample_manifest.csv      <- strata + weights
    consensus_dynamics/N1_triage/unit_map_SEALED.json     <- do not open before coding

Usage:
    py scripts/build_consensus_triage_sample.py
"""

from __future__ import annotations

import csv
import hashlib
import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
_CD = _REPO_ROOT / "analysis" / "production_evaluation" / "consensus_dynamics"
_ACTS = _CD / "response_acts.csv"
_OUT = _CD / "N1_triage"

SEED = 20260803
BLIND_SALT = "consensus_n1_triage_v1"

# stratum -> {side: n to draw}. "census" draws the whole stratum.
PLAN: dict[str, dict[str, int | str]] = {
    "A_d1_divergence": {"human": "census", "synthetic": 14},
    "B_d1_alignment":  {"human": 10, "synthetic": 10},
    "C_d1_none":       {"human": 20, "synthetic": 16},
}

LABELS = ["divergencia", "alineacion", "ninguna"]


def _stratum(row: dict) -> str:
    lab = row["d1_label"]
    if lab in ("divergence", "mixed"):
        return "A_d1_divergence"
    if lab == "alignment":
        return "B_d1_alignment"
    return "C_d1_none"


def _mask(text: str, names: list[str]) -> str:
    out = text
    for n in sorted({n for n in names if n and len(n) > 2}, key=len, reverse=True):
        out = out.replace(n, "[nombre]")
    return out


def main() -> None:
    _OUT.mkdir(parents=True, exist_ok=True)
    rows = list(csv.DictReader(_ACTS.open(encoding="utf-8")))
    all_names = [r["resp_speaker"] for r in rows] + [r["prev_speaker"] for r in rows]

    rng = random.Random(SEED)
    picked: list[dict] = []
    manifest: list[dict] = []

    for stratum, per_side in PLAN.items():
        for side in ("human", "synthetic"):
            pool = sorted(
                (r for r in rows if r["side"] == side and _stratum(r) == stratum),
                key=lambda r: r["act_id"],
            )
            want = per_side[side]
            n = len(pool) if want == "census" else min(int(want), len(pool))
            draw = pool if want == "census" else rng.sample(pool, n)
            for r in draw:
                picked.append({**r, "_stratum": stratum,
                               "_N_stratum": len(pool), "_n_drawn": n})
            manifest.append({
                "stratum": stratum, "side": side,
                "N_in_corpus": len(pool), "n_sampled": n,
                "inclusion_prob": round(n / len(pool), 6) if pool else 0.0,
                "ht_weight": round(len(pool) / n, 4) if n else "",
            })

    rng.shuffle(picked)

    sealed, sheet_rows = {}, []
    for i, r in enumerate(picked, start=1):
        uid = "N1-%03d" % i
        sealed[uid] = {
            "act_id": r["act_id"], "side": r["side"], "fg": r["fg"], "run": r["run"],
            "condition": r["condition"], "section_index": int(r["section_index"]),
            "stratum": r["_stratum"], "N_stratum": r["_N_stratum"], "n_drawn": r["_n_drawn"],
            "d1_label": r["d1_label"], "d1_label_full": r["d1_label_full"],
            "d1_div_hits": r["d1_div_hits"], "d1_align_hits": r["d1_align_hits"],
            "resp_words": int(r["resp_words"]),
            "blind_check": hashlib.sha256(
                f"{BLIND_SALT}|{r['act_id']}".encode()).hexdigest()[:8],
        }
        sheet_rows.append({
            "unit_id": uid,
            "turno_previo": _mask(r["prev_text"], all_names),
            "turno_respuesta": _mask(r["resp_text"], all_names),
            "palabras": int(r["resp_words"]),
        })

    # ---- workbook ---------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Codificacion"
    headers = ["unit_id", "turno_previo", "turno_respuesta", "palabras",
               "ETIQUETA", "nota (opcional)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for r in sheet_rows:
        ws.append([r["unit_id"], r["turno_previo"], r["turno_respuesta"],
                   r["palabras"], "", ""])

    widths = {"A": 10, "B": 70, "C": 90, "D": 9, "E": 16, "F": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(LABELS), allow_blank=True)
    dv.error = "Elige una de las tres etiquetas."
    ws.add_data_validation(dv)
    dv.add(f"E2:E{ws.max_row}")

    # ---- instructions -----------------------------------------------------
    ins = wb.create_sheet("Instrucciones")
    total_words = sum(r["palabras"] for r in sheet_rows)
    for line in [
        ("Tarea de triaje N1 — divergencia en actos de respuesta", True),
        ("", False),
        ("Una sola decisión por unidad. No juzgues intensidad, ni tipo, ni si se resolvió.", False),
        ("", False),
        ("Pregunta: en el TURNO DE RESPUESTA, ¿la persona se posiciona en contra de algo", False),
        ("dicho en el TURNO PREVIO?", False),
        ("", False),
        ("  divergencia  = expresa desacuerdo, matiza en contra, corrige, o marca contraste", False),
        ("                 con lo dicho antes (aunque sea de forma cortés o indirecta).", False),
        ("  alineacion   = expresa acuerdo, confirma, o construye sobre lo dicho antes", False),
        ("                 sin oponerse.", False),
        ("  ninguna      = ni una ni otra: cambia de tema, responde al moderador, aporta", False),
        ("                 algo paralelo, o no hay postura identificable frente al turno previo.", False),
        ("", False),
        ("Reglas de borde:", True),
        ("  · Si contrasta con SU PROPIA idea y no con la del turno previo -> ninguna.", False),
        ("  · Si primero acuerda y luego se opone ('sí, pero...') -> divergencia.", False),
        ("  · Si sólo dice que su experiencia es distinta, sin oponerse -> ninguna.", False),
        ("  · Si dudas entre dos, elige la que describa el efecto sobre la conversación.", False),
        ("  · No dejes celdas vacías. Si es ilegible, escribe la nota y marca ninguna.", False),
        ("", False),
        (f"Unidades: {len(sheet_rows)}. Palabras de respuesta a leer: {total_words:,}.", False),
        ("Tiempo estimado: 60–75 minutos. No hace falta terminar de una sentada;", False),
        ("el orden ya está aleatorizado, así que parar a mitad no sesga nada.", False),
        ("", False),
        ("Los turnos están ciegos: nombres enmascarados, y no se indica si el extracto", False),
        ("es humano o sintético. La longitud puede delatarlo — no dejes que eso te guíe;", False),
        ("codifica lo que dice el texto, no de dónde crees que viene.", False),
        ("", False),
        ("Doble codificación: las unidades N1-001 a N1-020 las codifica también la", False),
        ("segunda persona, en una copia limpia de este archivo (para κ).", False),
    ]:
        ins.append([line[0]])
        if line[1]:
            ins.cell(row=ins.max_row, column=1).font = Font(bold=True)
    ins.column_dimensions["A"].width = 100

    wb.save(_OUT / "CODING_WORKBOOK.xlsx")

    with (_OUT / "sample_manifest.csv").open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(manifest[0]))
        w.writeheader()
        w.writerows(manifest)

    (_OUT / "unit_map_SEALED.json").write_text(json.dumps({
        "_warning": "SEALED. Do not open until CODING_WORKBOOK.xlsx is filled in.",
        "seed": SEED, "blind_salt": BLIND_SALT,
        "source": str(_ACTS.relative_to(_REPO_ROOT)),
        "source_sha256": hashlib.sha256(_ACTS.read_bytes()).hexdigest(),
        "units": sealed,
    }, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"N1 sample: {len(sheet_rows)} units  ({total_words:,} response words to read)")
    print(f"{'stratum':<18} {'side':<10} {'N':>5} {'n':>4} {'p_incl':>8} {'weight':>8}")
    for m in manifest:
        print(f"{m['stratum']:<18} {m['side']:<10} {m['N_in_corpus']:>5} {m['n_sampled']:>4} "
              f"{m['inclusion_prob']:>8.4f} {str(m['ht_weight']):>8}")
    print(f"\nwrote {_OUT / 'CODING_WORKBOOK.xlsx'}")


if __name__ == "__main__":
    main()
