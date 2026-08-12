"""
Consolidation scaffold for the single-coder sample — EMPTY BY DESIGN.

WHY CONSOLIDATION EXISTS
A coder's raw rows are not a reference set. One theme may be entered twice in
different words, and two rows may look alike while making different claims. Using raw
rows as the recall denominator would let a coder's duplication inflate or deflate
every downstream number without anyone touching the extractor.

So a human consolidates within each unit first, and the CONSOLIDATED themes become
`human_reference_themes`. That set is frozen before the extractor runs.

Classification:
    SINGLE_CODER_HUMAN_REFERENCE_WITH_POST_CODING_CONSOLIDATION

This is NOT inter-coder agreement and is NOT calibration-grade. One coder produced
the themes; one person consolidated them. It cannot become either by relabelling.

THE SCAFFOLD IS LEFT EMPTY. Nothing here merges, splits or judges a theme. Doing that
automatically would be the same error as auto-clustering the primary sample: the
judgement is the point.

No LLM call of any kind.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_DIR = _OUT / "transportability_sample"
_WB = _DIR / "Transportability_Consolidation.xlsx"

CLASSIFICATION = "SINGLE_CODER_HUMAN_REFERENCE_WITH_POST_CODING_CONSOLIDATION"
UNITS = [f"S{i:02d}" for i in range(1, 7)]

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
NEED = PatternFill("solid", fgColor="FFF2CC")


def _rel(p):
    """Display path, tolerant of a redirected output workspace."""
    try:
        return p.relative_to(_REPO_ROOT)
    except ValueError:
        return p


class ConsolidationNotReady(RuntimeError):
    pass


def build() -> Path:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    for i, (t, b) in enumerate([
        ("Consolidation — single-coder reference themes", True),
        ("", False),
        ("WHY THIS STEP EXISTS", True),
        ("The coder's raw rows are not the reference set. The same theme may appear", False),
        ("twice in different words, and two rows may look similar while making", False),
        ("different claims. If raw rows were used as the denominator, duplication", False),
        ("alone would move recall without the extractor changing at all.", False),
        ("", False),
        ("WHAT YOU DO", True),
        ("1. Work one unit at a time. Read every raw row for that unit first.", False),
        ("2. Give each raw row a consolidated_theme_id on the Raw_To_Consolidated", False),
        ("   sheet. Rows you judge to be the SAME theme share an id.", False),
        ("   IDS MUST BE UNIT-PREFIXED: S01_T1, S01_T2, S02_T1, ... An id belongs to", False),
        ("   exactly one unit and must never appear under another. This is checked.", False),
        ("3. MERGE only when the rows make the same claim about the same thing in", False),
        ("   different words.", False),
        ("4. KEEP SEPARATE when they differ in the claim made, in who or what is the", False),
        ("   agent, or in the position taken — even if the wording is close.", False),
        ("5. On Consolidated_Themes, write one row per consolidated theme: a label, a", False),
        ("   one-sentence description, and its centrality FOR THAT UNIT.", False),
        ("6. Centrality is judged per consolidated theme x unit. A theme may be", False),
        ("   central in one unit and secondary in another.", False),
        ("", False),
        ("WHAT IS NOT ALLOWED", True),
        ("Do not delete or edit any raw row. They stay exactly as the coder wrote them;", False),
        ("this sheet records how they group, not what they should have said.", False),
        ("Every consolidated theme must list the source_row_ids it came from.", False),
        ("", False),
        ("WHAT THIS IS NOT", True),
        ("This is not inter-coder agreement. One coder produced these themes and one", False),
        ("person is consolidating them. No agreement statistic can come from it, and", False),
        ("it is not calibration-grade.", False),
    ], start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=b, size=11 if b else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Raw_To_Consolidated")
    for j, (h, w) in enumerate(zip(
            ["blind_unit_id", "source_row_id", "raw_theme_label",
             "raw_theme_description", "raw_supporting_quote", "raw_relevance",
             "consolidated_theme_id", "consolidator_note"],
            [14, 16, 32, 46, 52, 13, 22, 32]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(row=2, column=1, value=(
        "POPULATED AUTOMATICALLY from the returned coder workbook once it validates. "
        "Left empty on purpose — filling it now would require inventing the coder's rows."))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C2"

    ws = wb.create_sheet("Consolidated_Themes")
    for j, (h, w) in enumerate(zip(
            ["blind_unit_id", "consolidated_theme_id", "consolidated_label",
             "consolidated_description", "is_central_in_this_unit",
             "source_row_ids", "consolidator_note"],
            [14, 22, 34, 50, 24, 30, 32]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    dv = DataValidation(type="list", formula1='"central,secondary"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("E2:E400")
    for i, uid in enumerate(UNITS, start=2):
        ws.cell(row=i, column=1, value=uid)
        for j in range(2, 8):
            ws.cell(row=i, column=j).fill = NEED
    ws.cell(row=len(UNITS) + 3, column=1, value=(
        "One row per consolidated theme. Add rows as needed — the six seeded rows are "
        "only a reminder that every unit must end with at least one consolidated theme."))
    ws.cell(row=len(UNITS) + 3, column=1).alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "C2"

    ws = wb.create_sheet("Scope")
    for i, (a, b) in enumerate([
        ("classification", CLASSIFICATION),
        ("coders", "one"),
        ("consolidation", "one person, after coding"),
        ("agreement statistics", "NOT POSSIBLE — one coder"),
        ("calibration-grade", "NO"),
        ("primary sample", "U01-U07 (Q3), two coders — separate, never pooled"),
        ("recall denominator", "consolidated themes, never raw rows"),
    ], start=1):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70

    _DIR.mkdir(parents=True, exist_ok=True)
    wb.save(_WB)
    wb.close()
    return _WB


# ---------------------------------------------------------------------------
# Import raw rows from a VALIDATED coder workbook, and seal them
# ---------------------------------------------------------------------------

_RAW_SEAL = _OUT / "gold_standard_sealed" / "transportability_raw_rows_seal.json"

RAW_COLS = ("blind_unit_id", "source_row_id", "raw_theme_label",
            "raw_theme_description", "raw_supporting_quote", "raw_relevance")
EDITABLE_COLS = ("consolidated_theme_id", "consolidator_note")
VALID_CENTRALITY = ("central", "secondary")


def _row_content_sha(r: dict) -> str:
    return hashlib.sha256("␟".join(
        str(r[c]) for c in RAW_COLS).encode("utf-8")).hexdigest()


THEME_ID_RULE = ("consolidated_theme_id must be unit-prefixed, e.g. S01_T1. An id "
                 "belongs to exactly one unit and never appears under another.")


def import_raw_rows(coder_workbook: Path, consolidation_workbook: Path = None,
                    seal_path: Path = None) -> dict:
    """
    Populate Raw_To_Consolidated from a RETURNED, VALIDATED coder workbook.

    Nothing is invented. This runs only after the coder returns their file and the
    return gate passes; until then there are no raw rows to import and the
    consolidation sheet stays empty.

    Each raw theme gets a stable, unique `source_row_id`:
        S01_slot_03      a numbered slot on Emergent_Coding
        S01_ovf_01       an Overflow_Themes row, numbered in sheet order
    The id encodes the unit, so a row can never be silently reassigned to another
    unit later.
    """
    import build_transportability_package as _pkg

    consolidation_workbook = consolidation_workbook or _WB
    seal_path = seal_path or _RAW_SEAL

    problems = _pkg.validate(coder_workbook)
    if problems:
        raise ConsolidationNotReady(
            "the coder workbook has not passed the return gate; import refused:\n  - "
            + "\n  - ".join(problems[:20]))

    wbc = openpyxl.load_workbook(coder_workbook, read_only=True, data_only=True)

    def _rows(sheet):
        rows = list(wbc[sheet].iter_rows(values_only=True))
        hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [{h: ("" if v is None else str(v).strip())
                 for h, v in zip(hdr, r)} for r in rows[1:]]

    raw = []
    for r in _rows("Emergent_Coding"):
        # relevance is NOT_ASSESSED by decision and is not required for a row to be
        # substantively complete.
        if not all(r.get(f) for f in _pkg.REQUIRED_THEME_FIELDS):
            continue
        raw.append({
            "blind_unit_id": r["blind_unit_id"],
            "source_row_id": f"{r['blind_unit_id']}_slot_{int(r['slot']):02d}",
            "raw_theme_label": r["theme_label"],
            "raw_theme_description": r["theme_description"],
            "raw_supporting_quote": r["supporting_quote"],
            "raw_relevance": r.get("relevance") or _pkg.RELEVANCE_STATUS,
            "coder_note": r.get("coder_note", ""),
        })
    ovf_n = {}
    for r in _rows("Overflow_Themes"):
        # Exactly the same completeness rule as Emergent_Coding above. relevance is
        # NOT_ASSESSED by decision, so an overflow row carrying label, description and
        # quote is substantively complete and must be imported, not skipped. This
        # branch previously still demanded relevance, which would have silently dropped
        # every overflow theme the coder wrote.
        if not all(r.get(f) for f in _pkg.REQUIRED_THEME_FIELDS):
            continue
        uid = r["blind_unit_id"]
        ovf_n[uid] = ovf_n.get(uid, 0) + 1
        raw.append({
            "blind_unit_id": uid,
            "source_row_id": f"{uid}_ovf_{ovf_n[uid]:02d}",
            "raw_theme_label": r["theme_label"],
            "raw_theme_description": r["theme_description"],
            "raw_supporting_quote": r["supporting_quote"],
            "raw_relevance": r.get("relevance") or _pkg.RELEVANCE_STATUS,
            "coder_note": r.get("coder_note", ""),
        })
    wbc.close()

    ids = [r["source_row_id"] for r in raw]
    if len(set(ids)) != len(ids):
        dupes = sorted({i for i in ids if ids.count(i) > 1})
        raise ConsolidationNotReady(f"source_row_id collision: {dupes}")
    if not raw:
        raise ConsolidationNotReady("the validated workbook contained no complete themes")

    # --- write the mapping sheet, protecting the imported columns ------------
    wb = openpyxl.load_workbook(consolidation_workbook)
    ws = wb["Raw_To_Consolidated"]
    ws.delete_rows(2, ws.max_row)
    for i, r in enumerate(sorted(raw, key=lambda x: x["source_row_id"]), start=2):
        for j, col in enumerate(RAW_COLS, start=1):
            c = ws.cell(row=i, column=j, value=r[col])
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.protection = Protection(locked=True)
        for j, col in enumerate(EDITABLE_COLS, start=len(RAW_COLS) + 1):
            c = ws.cell(row=i, column=j)
            c.fill = NEED
            c.protection = Protection(locked=False)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.protection.sheet = True
    ws.protection.enable()
    wb.save(consolidation_workbook)
    wb.close()

    seal = {
        "sealed_utc": datetime.now(UTC).isoformat(),
        "classification": CLASSIFICATION,
        "warning": "SEALED — the authority on which raw rows exist.",
        "source_workbook": str(coder_workbook),
        "n_raw_rows": len(raw),
        "protected_columns": list(RAW_COLS),
        "editable_columns": list(EDITABLE_COLS),
        "rows": {r["source_row_id"]: {"blind_unit_id": r["blind_unit_id"],
                                      "content_sha256": _row_content_sha(r)}
                 for r in raw},
    }
    seal_path.parent.mkdir(parents=True, exist_ok=True)
    seal_path.write_text(json.dumps(seal, indent=1, ensure_ascii=False),
                         encoding="utf-8")
    return seal


# ---------------------------------------------------------------------------
# Freeze — the sealed mapping is the authority
# ---------------------------------------------------------------------------

def freeze_reference(path: Path = None, seal_path: Path = None,
                     force: bool = False) -> dict:
    """
    Freeze `human_reference_themes` from a COMPLETED consolidation workbook.

    THE SEALED RAW-ROW SET IS THE SOURCE OF TRUTH. `source_row_ids` typed on
    Consolidated_Themes is a human-entered restatement and may be wrong; where the two
    disagree the workbook is rejected rather than one being preferred. Silently
    trusting either would let a mistyped id drop a theme from the denominator.
    """
    path = path or _WB
    seal_path = seal_path or _RAW_SEAL
    out_path = _DIR / "human_reference_themes.json"

    # A frozen reference is a commitment: it fixes the recall denominator. Replacing
    # it silently would let the denominator move after results existed, so an
    # existing reference is refused unless the caller says so explicitly.
    if out_path.exists() and not force:
        raise ConsolidationNotReady(
            f"a frozen reference already exists at {out_path.name}; refusing to "
            f"overwrite it. Pass force=True only if you intend to discard the "
            f"existing denominator.")
    if not seal_path.exists():
        raise ConsolidationNotReady(
            "no raw-row seal — import the returned coder workbook first")
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    sealed_rows = seal["rows"]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def _rows(sheet):
        rows = list(wb[sheet].iter_rows(values_only=True))
        hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [{h: ("" if v is None else str(v).strip())
                 for h, v in zip(hdr, r)} for r in rows[1:]]

    mapping = [r for r in _rows("Raw_To_Consolidated") if r.get("source_row_id")]
    themes_raw = _rows("Consolidated_Themes")
    wb.close()

    problems: list[str] = []

    # --- 1. the raw-row set must match the seal exactly ---------------------
    ids = [r["source_row_id"] for r in mapping]
    seen = set(ids)
    missing = sorted(set(sealed_rows) - seen)
    added = sorted(seen - set(sealed_rows))
    dupes = sorted({i for i in ids if ids.count(i) > 1})
    if missing:
        problems.append(f"raw rows missing from the mapping: {missing}")
    if added:
        problems.append(f"raw rows not present in the seal: {added}")
    if dupes:
        problems.append(f"duplicated raw rows in the mapping: {dupes}")
    for r in mapping:
        sid = r["source_row_id"]
        if sid not in sealed_rows:
            continue
        if r["blind_unit_id"] != sealed_rows[sid]["blind_unit_id"]:
            problems.append(
                f"{sid}: unit changed from {sealed_rows[sid]['blind_unit_id']} "
                f"to {r['blind_unit_id']}")
        # Recompute the content hash over exactly RAW_COLS. Sheet protection in Excel
        # is a convenience, not a guarantee: it can be turned off, and a file can be
        # edited by anything that reads xlsx. The hash is what actually holds.
        observed = _row_content_sha(r)
        expected = sealed_rows[sid]["content_sha256"]
        if observed != expected:
            problems.append(
                f"{sid}: raw content was modified after import "
                f"(expected {expected[:16]}..., observed {observed[:16]}...). "
                f"The imported columns {RAW_COLS} are the coder's words and must not "
                f"change during consolidation.")

    # --- 2. every raw row assigned exactly once -----------------------------
    assignment: dict[str, str] = {}
    for r in mapping:
        tid = r.get("consolidated_theme_id", "")
        if not tid:
            problems.append(f"{r['source_row_id']}: no consolidated_theme_id")
            continue
        assignment[r["source_row_id"]] = f"{r['blind_unit_id']}::{tid}"

    # --- 3. consolidated themes ---------------------------------------------
    themes, by_key = [], {}
    for r in themes_raw:
        uid = r.get("blind_unit_id", "")
        vals = [r.get(k, "") for k in ("consolidated_theme_id", "consolidated_label",
                                       "consolidated_description",
                                       "is_central_in_this_unit", "source_row_ids")]
        if not uid and not any(vals):
            continue
        if uid not in UNITS:
            problems.append(f"row with blind_unit_id {uid!r} is not a known unit")
            continue
        if not any(vals):
            continue                       # seeded blank row: reported as a gap below
        tid, label, desc, central, src = vals
        missing_f = [n for n, v in (("consolidated_theme_id", tid),
                                    ("consolidated_label", label),
                                    ("consolidated_description", desc),
                                    ("is_central_in_this_unit", central),
                                    ("source_row_ids", src)) if not v]
        if missing_f:
            problems.append(f"{uid} {tid or '(no id)'}: partially completed, "
                            f"missing {missing_f}")
            continue
        if central not in VALID_CENTRALITY:
            problems.append(f"{uid} {tid}: is_central_in_this_unit is {central!r}, "
                            f"expected one of {VALID_CENTRALITY}")
        if not str(tid).startswith(uid + "_"):
            problems.append(f"{uid} {tid}: consolidated_theme_id is not unit-prefixed "
                            f"(expected {uid}_...). {THEME_ID_RULE}")
        key = f"{uid}::{tid}"
        if key in by_key:
            problems.append(f"{uid}: consolidated_theme_id {tid} appears more than once")
        declared = [s.strip() for s in str(src).replace(",", "|").split("|") if s.strip()]
        by_key[key] = declared
        themes.append({"blind_unit_id": uid, "consolidated_theme_id": tid,
                       "label": label, "description": desc,
                       "is_central_in_this_unit": central,
                       "declared_source_row_ids": declared})

    # --- 4. an id must not cross units --------------------------------------
    per_id: dict[str, set] = {}
    for t in themes:
        per_id.setdefault(t["consolidated_theme_id"], set()).add(t["blind_unit_id"])
    for tid, units_ in sorted(per_id.items()):
        if len(units_) > 1:
            problems.append(
                f"consolidated_theme_id {tid} is used in more than one unit "
                f"({sorted(units_)}) — ids must be unique within a unit and never "
                f"cross units")

    # --- 5. declared ids must reconcile with the mapping, both ways ---------
    for t in themes:
        key = f"{t['blind_unit_id']}::{t['consolidated_theme_id']}"
        declared = set(t["declared_source_row_ids"])
        for sid in sorted(declared):
            if sid not in sealed_rows:
                problems.append(f"{key}: declared source_row_id {sid} does not exist")
            elif sealed_rows[sid]["blind_unit_id"] != t["blind_unit_id"]:
                problems.append(
                    f"{key}: declared source_row_id {sid} belongs to "
                    f"{sealed_rows[sid]['blind_unit_id']}, not {t['blind_unit_id']}")
        from_mapping = {sid for sid, k in assignment.items() if k == key}
        if declared != from_mapping:
            problems.append(
                f"{key}: source_row_ids disagree with the mapping — declared "
                f"{sorted(declared)}, mapping says {sorted(from_mapping)}. The sealed "
                f"mapping is the authority.")
        if not declared:
            problems.append(f"{key}: consolidated theme with no raw rows")

    # --- 5b. mapping keys and theme keys must correspond exactly ------------
    # Without this, a mapping row pointing at a consolidated_theme_id that has no row
    # in Consolidated_Themes is invisible: its key never appears among the themes, so
    # nothing compares against it and the raw row silently leaves the denominator.
    mapping_keys = set(assignment.values())
    theme_keys = set(by_key)
    orphan_mapping = sorted(mapping_keys - theme_keys)
    orphan_themes = sorted(theme_keys - mapping_keys)
    if orphan_mapping:
        problems.append(
            f"mapping points at consolidated_theme_id(s) with no row in "
            f"Consolidated_Themes: {orphan_mapping}")
    if orphan_themes:
        problems.append(
            f"consolidated theme(s) with no raw row assigned in the mapping: "
            f"{orphan_themes}")

    # --- 5c. three-way set equality, checked BEFORE anything is written -----
    sealed_ids = set(sealed_rows)
    declared_ids = {sid for t in themes for sid in t["declared_source_row_ids"]}
    assigned_ids = set(assignment)
    if not (sealed_ids == declared_ids == assigned_ids):
        problems.append(
            "the three raw-row sets are not identical — "
            f"sealed({len(sealed_ids)}) vs declared({len(declared_ids)}) vs "
            f"assigned({len(assigned_ids)}); "
            f"sealed-but-not-declared={sorted(sealed_ids - declared_ids)}, "
            f"declared-but-not-sealed={sorted(declared_ids - sealed_ids)}, "
            f"sealed-but-not-assigned={sorted(sealed_ids - assigned_ids)}, "
            f"assigned-but-not-sealed={sorted(assigned_ids - sealed_ids)}")

    # --- 6. no raw row in two themes ----------------------------------------
    claim_count: dict[str, list] = {}
    for t in themes:
        for sid in t["declared_source_row_ids"]:
            claim_count.setdefault(sid, []).append(
                f"{t['blind_unit_id']}::{t['consolidated_theme_id']}")
    for sid, owners in sorted(claim_count.items()):
        if len(owners) > 1:
            problems.append(f"raw row {sid} is claimed by {owners}")

    # --- 7. coverage ---------------------------------------------------------
    covered = {t["blind_unit_id"] for t in themes}
    for uid in UNITS:
        if uid not in covered:
            problems.append(f"{uid}: no consolidated theme — consolidation incomplete")
    stray = covered - set(UNITS)
    if stray:
        problems.append(f"themes for unknown units: {sorted(stray)}")

    if problems:
        raise ConsolidationNotReady(
            "human_reference_themes NOT frozen:\n  - " + "\n  - ".join(problems))

    out = {
        "frozen_utc": datetime.now(UTC).isoformat(),
        "classification": CLASSIFICATION,
        "is_inter_coder_agreement": False,
        "calibration_grade": False,
        "denominator_rule": ("recall denominator = consolidated themes per unit; raw "
                             "coder rows are NEVER a denominator"),
        "raw_rows_seal": str(seal_path),
        "n_raw_rows": len(sealed_rows),
        "n_consolidated_themes": len(themes),
        "human_reference_themes": [
            {**t, "source_row_ids": t.pop("declared_source_row_ids")} for t in themes],
    }
    # Write to a temporary file in the same directory, then rename. A crash or a
    # disk error partway through must not leave a half-written reference that later
    # looks authoritative.
    tmp = out_path.with_suffix(".json.tmp")
    try:
        tmp.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")
        os.replace(tmp, out_path)
    finally:
        if tmp.exists():
            tmp.unlink()
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--build", action="store_true")
    ap.add_argument("--freeze", action="store_true")
    ap.add_argument("--import-raw", dest="import_raw", default=None,
                    help="path to the RETURNED, validated coder workbook")
    a = ap.parse_args()
    if a.build:
        p = build()
        print(f"scaffold: {_rel(p)}  (EMPTY — for a human)")
        return 0
    if a.import_raw:
        try:
            seal = import_raw_rows(Path(a.import_raw))
        except ConsolidationNotReady as exc:
            print(f"REFUSED: {exc}")
            return 2
        print(f"imported {seal['n_raw_rows']} raw rows; seal written")
        return 0
    try:
        out = freeze_reference()
    except ConsolidationNotReady as exc:
        print(f"REFUSED: {exc}")
        return 1
    print(f"frozen {out['n_consolidated_themes']} consolidated themes "
          f"from {out['n_raw_rows_assigned']} raw rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


# ---------------------------------------------------------------------------
# Direct freeze — consolidation was audited and found unnecessary
# ---------------------------------------------------------------------------

SUPPLEMENTARY_CLASSIFICATION = "SUPPLEMENTARY_SINGLE_CODER_TRANSPORTABILITY_SAMPLE"
CONSOLIDATION_DECISION = "CONSOLIDATION_NOT_REQUIRED — CODER_ROWS_ALREADY_DISTINCT"


def freeze_supplementary_reference(coder_workbook: Path = None,
                                   out_path: Path = None,
                                   force: bool = False) -> dict:
    """
    Freeze the supplementary human reference straight from the returned workbook.

    No consolidation workbook is involved: the duplicate audit found no exact,
    near-exact or ambiguous duplicates within any unit, so every complete coder row is
    already a distinct human theme and merging would invent structure that is not
    there. Each row becomes one theme, carrying its own source_row_id.

    relevance is written as NOT_ASSESSED — never as `secondary`, `false`, `0`, or an
    absent theme. The coder workbook is opened READ-ONLY and never modified.
    """
    import build_transportability_package as _pkg
    import hashlib as _h

    coder_workbook = coder_workbook or _pkg._WB
    out_path = out_path or (_DIR / "supplementary_human_reference.json")

    problems = _pkg.validate(coder_workbook)
    if problems:
        raise ConsolidationNotReady(
            "the coder workbook has not passed the return gate; freeze refused:\n  - "
            + "\n  - ".join(problems[:20]))
    if out_path.exists() and not force:
        raise ConsolidationNotReady(
            f"a frozen supplementary reference already exists at {out_path.name}; "
            f"refusing to overwrite it. Pass force=True only to discard it.")

    wbc = openpyxl.load_workbook(coder_workbook, read_only=True, data_only=True)

    def _rows(sheet):
        rows = list(wbc[sheet].iter_rows(values_only=True))
        hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [{h: ("" if v is None else str(v).strip())
                 for h, v in zip(hdr, r)} for r in rows[1:]]

    coding, overflow = _rows("Emergent_Coding"), _rows("Overflow_Themes")
    units_rows = _rows("Units")
    wbc.close()

    _sealed_dir = _RAW_SEAL.parent
    man = json.loads((_sealed_dir / "transportability_sample_manifest.json")
                     .read_text(encoding="utf-8"))
    qof = {u["blind_unit_id"]: u["question_id"] for u in man["units"]}
    sof = {u["blind_unit_id"]: u["stratum"] for u in man["units"]}

    themes, n_empty_slots = [], 0
    for r in coding:
        uid = r.get("blind_unit_id", "")
        if not uid:
            continue
        if not any(r.get(f) for f in _pkg.REQUIRED_THEME_FIELDS):
            n_empty_slots += 1
            continue
        sid = f"{uid}_slot_{int(r['slot']):02d}"
        themes.append({
            "supplementary_key": f"{uid}::{sid}",
            "blind_unit_id": uid,
            "question_id": qof.get(uid),
            "stratum": sof.get(uid),
            "source_row_id": sid,
            "theme_label": r["theme_label"],
            "theme_description": r["theme_description"],
            "supporting_quote": r["supporting_quote"],
            "relevance_status": _pkg.RELEVANCE_STATUS,
            "relevance_value_as_recorded": r.get("relevance") or None,
            "coder_note": r.get("coder_note", ""),
            "content_sha256": _h.sha256("\u241f".join(
                [uid, sid, r["theme_label"], r["theme_description"],
                 r["supporting_quote"]]).encode("utf-8")).hexdigest(),
        })
    for i, r in enumerate(overflow, start=1):
        uid = r.get("blind_unit_id", "")
        if uid and all(r.get(f) for f in _pkg.REQUIRED_THEME_FIELDS):
            sid = f"{uid}_ovf_{i:02d}"
            themes.append({
                "supplementary_key": f"{uid}::{sid}", "blind_unit_id": uid,
                "question_id": qof.get(uid), "stratum": sof.get(uid),
                "source_row_id": sid, "theme_label": r["theme_label"],
                "theme_description": r["theme_description"],
                "supporting_quote": r["supporting_quote"],
                "relevance_status": _pkg.RELEVANCE_STATUS,
                "relevance_value_as_recorded": r.get("relevance") or None,
                "coder_note": r.get("coder_note", ""),
                "content_sha256": _h.sha256("\u241f".join(
                    [uid, sid, r["theme_label"], r["theme_description"],
                     r["supporting_quote"]]).encode("utf-8")).hexdigest(),
            })

    keys = [t["supplementary_key"] for t in themes]
    if len(keys) != len(set(keys)):
        raise ConsolidationNotReady("duplicate supplementary keys")

    per_unit = {}
    for t in themes:
        per_unit.setdefault(t["blind_unit_id"], 0)
        per_unit[t["blind_unit_id"]] += 1
    if any(n == 0 for n in per_unit.values()) or len(per_unit) != 6:
        raise ConsolidationNotReady(f"expected 6 units with themes, got {per_unit}")

    out = {
        "frozen_utc": datetime.now(UTC).isoformat(),
        "classification": SUPPLEMENTARY_CLASSIFICATION,
        "consolidation_decision": CONSOLIDATION_DECISION,
        "consolidation_rule": (
            "the within-unit duplicate audit found 0 exact, 0 near-exact and 0 "
            "interpretively ambiguous pairs, so every complete coder row stands as a "
            "distinct human theme; nothing was merged and no LLM rewrote any theme"),
        "relevance_status": _pkg.RELEVANCE_STATUS,
        "relevance_note": (
            "The researcher reviewed all six units and decided not to adjudicate "
            "central vs secondary: it could not be determined reliably and is not "
            "needed for the supplementary transportability objective. This is a "
            "methodological decision, not missing data. It must never be read as "
            "`secondary`, `false`, `0` or as an absent theme. No result about "
            "centrality, relevance, salience or thematic hierarchy may be reported "
            "from this sample."),
        "coder_workbook": coder_workbook.name,
        "coder_workbook_sha256": _h.sha256(coder_workbook.read_bytes()).hexdigest(),
        "n_units": len(per_unit),
        "n_themes": len(themes),
        "n_empty_slots": n_empty_slots,
        "denominator_note": (
            "The denominator for any future transportability comparison is the number "
            "of HUMAN THEMES per unit below. Empty slots are unused capacity, not "
            "absent themes, and are never a denominator."),
        "denominators_per_unit": {u: {"n_human_themes": per_unit[u],
                                      "question_id": qof.get(u),
                                      "stratum": sof.get(u)}
                                  for u in sorted(per_unit)},
        "never_combine_numerically_with": (
            "U01-U07 / Q3 emergent calibration — different questions, different "
            "design, different denominators. These are never pooled."),
        "limitations": [
            "one coder", "six units", "four different guide questions",
            "no inter-coder agreement", "not a gold standard",
            "no generalisation to all guide questions",
        ],
        "themes": themes,
    }

    tmp = out_path.with_suffix(".json.tmp")
    try:
        tmp.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")
        os.replace(tmp, out_path)
    finally:
        if tmp.exists():
            tmp.unlink()
    return out
