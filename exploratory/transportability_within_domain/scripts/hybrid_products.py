"""
Emit the hybrid products from the COMPLETE 93-pair universe: the combined cross-model
record, the cost record, the results narrative, the tables workbook and the traceability
index.

Reads only from the sealed hybrid artefacts. Writes only inside hybrid_evaluation/.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
from collections import Counter
from datetime import datetime, UTC
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))

import hybrid_transportability as hy   # noqa: E402
import hybrid_complement as hc         # noqa: E402

_HY = hy._HY
_L = lambda n: json.loads((_HY / n).read_text(encoding="utf-8"))   # noqa: E731

CLAUDE_BATCH_IN, CLAUDE_BATCH_OUT = 2.50, 12.50   # verified 2026-08-02, list Batch
ROUNDS = ("claude_round1_results.json", "claude_round2_results.json",
          "claude_complement_results.json")


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


# ---------------------------------------------------------------- product 4
def cross_model_record():
    uni = _L("hybrid_universe.json")
    gate = {r["case_id"]: r for r in uni["rows"]}
    der = _L("hybrid_matching_derivation.json")
    gate2 = {x["case_id"]: x for x in der["machine_only"] + der["granularity"]}

    srcs = [(_L(n), lbl) for n, lbl in zip(ROUNDS, (1, 2, "complement"))]
    cases = {}
    for src, rnd in srcs:
        for r in src["results"]:
            c = cases.setdefault(r["case_id"], {
                "case_id": r["case_id"], "round": rnd, "task": r["task"],
                "source_round": (hc.SOURCE_COMPLEMENT if rnd == "complement"
                                 else hc.SOURCE_ORIGINAL),
                "blind_unit_id": r["blind_unit_id"], "question_id": r["question_id"],
                "provenance": r["provenance"], "repetitions": {}})
            c["repetitions"][r["repetition_index"]] = {
                "custom_id": r["custom_id"], "status": r["status"],
                "cache_key": r["cache_key"], "attempt": r.get("attempt", 1),
                "retry_reason": r.get("retry_reason"),
                "category": (r.get("judgement") or {}).get("category"),
                "confidence": (r.get("judgement") or {}).get("confidence"),
                "rationale": (r.get("judgement") or {}).get("rationale"),
                "quotations": (r.get("judgement") or {}).get("quotations"),
                "usage": r.get("usage")}
    for cid, c in cases.items():
        g = gate.get(cid) or gate2.get(cid)
        c["gate_outcome"] = {"status": g["status"], "category": g.get("category"),
                             "reasons": g.get("reasons", [])} if g else None

    rec = {
        "built_utc": datetime.now(UTC).isoformat(),
        "classification": hy.CLASSIFICATION,
        "auditor_model": hy.AUDITOR_MODEL, "execution_mode": "batch",
        "effort": hy.AUDITOR_EFFORT,
        "correspondence_space": "complete — all 93 within-unit pairs adjudicated",
        "blinding": ("the auditor saw a blinded extract id, the reference theme and the "
                     "candidate theme, with no prior result, no metric, no "
                     "classification, no Q3, no experimental condition and no "
                     "human/model provenance"),
        "repetition_policy": ("two repetitions per case with block and evidence order "
                              "reversed in repetition 1; a case is only settled when "
                              "both agree, neither is LOW, and every cited quotation is "
                              "literal and not the moderator's"),
        "retry_policy": _L("hybrid_complement_manifest.json")["retry_policy"],
        "jobs": [{"round": lbl, "job_id": s["job_id"], "n_requests": s["n_results"],
                  "n_complete": s["n_complete"],
                  "failures": [r["custom_id"] for r in s["results"]
                               if r["status"] != "COMPLETE"],
                  "retry": s.get("retry"), "usage": s["total_usage"]}
                 for s, lbl in srcs],
        "n_cases": len(cases),
        "cases": [cases[k] for k in sorted(cases)],
    }
    hy._atomic(_HY / "claude_cross_model_results.json", rec)
    return rec


# ---------------------------------------------------------------- product 5
def cost_record():
    g = _L("gemini_extraction_results.json")
    per = []
    ci = co = 0
    for n, lbl in zip(ROUNDS, ("audit round 1", "audit round 2",
                               "complementary audit (32 omitted pairs)")):
        s = _L(n)
        ci += s["total_usage"]["input_tokens"]
        co += s["total_usage"]["output_tokens"]
        per.append({"stage": lbl, "job_id": s["job_id"],
                    "n_requests": s["n_results"], "n_complete": s["n_complete"],
                    "input_tokens": s["total_usage"]["input_tokens"],
                    "output_tokens": s["total_usage"]["output_tokens"],
                    "retry": s.get("retry")})
    cost = ci / 1e6 * CLAUDE_BATCH_IN + co / 1e6 * CLAUDE_BATCH_OUT
    rec = {
        "record_type": "POST_RUN_MEASURED_USAGE",
        "scope": hy.CLASSIFICATION,
        "kept_separate_from": "any pre-run estimate; this file contains no forecast",
        "claude": {
            "per_stage": per,
            "actual_input_tokens": ci, "actual_output_tokens": co,
            "batch_input_rate_per_mtok_usd": CLAUDE_BATCH_IN,
            "batch_output_rate_per_mtok_usd": CLAUDE_BATCH_OUT,
            "formula": "(input/1e6 x in_rate) + (output/1e6 x out_rate)",
            "worked": (f"({ci}/1e6 x {CLAUDE_BATCH_IN}) + ({co}/1e6 x "
                       f"{CLAUDE_BATCH_OUT}) = {ci/1e6*CLAUDE_BATCH_IN:.4f} + "
                       f"{co/1e6*CLAUDE_BATCH_OUT:.4f}"),
            "calculated_list_batch_cost_usd": round(cost, 2),
            "rate_source": "https://platform.claude.com/docs/en/about-claude/pricing",
            "rate_verified_utc": "2026-08-02"},
        "gemini": {
            "job_name": _L("gemini_job.json")["job_name"], "model": hy.GEMINI_MODEL,
            "actual_input_tokens": g["total_usage"]["input_tokens"],
            "actual_output_tokens": g["total_usage"]["output_tokens"],
            "re_run_for_the_complement": False,
            "calculated_cost_usd": None,
            "cost_status": "NOT_CALCULATED_RATE_NOT_VERIFIED",
            "why": ("no published Batch rate for this model was verified during this "
                    "study, and inventing one would put an unsourced number into the "
                    "record. The measured token counts above are the factual part.")},
        "IMPORTANT": ("costs CALCULATED at published list rates from measured tokens. "
                      "Not necessarily what the organisation is billed. The Console "
                      "invoice is authoritative."),
    }
    hy._atomic(_HY / "hybrid_cost_actual.json", rec)
    return rec


# ---------------------------------------------------------------- product 6
def results_md(m, cost):
    o, pq, pu = m["overall_within_check"], m["per_question"], m["per_unit"]
    uni = _L("hybrid_universe.json")
    ev = m["evidence"]["literal_evidence_attachment_rate"]
    fr, bal = m["FROZEN_RULE_CLASSIFICATION"], m["BALANCED_INTERPRETATION"]
    L = []
    A = L.append
    A("# Exploratory out-of-Q3 transportability check — results\n")
    A(f"**Classification — `{hy.CLASSIFICATION}`**\n")
    A("This is not a formal validation and its numbers are never pooled with the "
      "U01–U07/Q3 calibration or with the deductive analysis. It asks one narrow "
      "question: when the same emergent extraction and the same blinded cross-model "
      "adjudication are pointed at six supplementary units drawn from four *other* "
      "guide questions, does the procedure behave in a way that is descriptively "
      "compatible with what Q3 showed?\n")
    A(f"Built {m['built_utc']}. Supersedes an earlier version computed over an "
      "incomplete correspondence space — see §2 and `PROTOCOL_DEVIATIONS.md`.\n")

    A("## 1. What was analysed\n")
    A("| Unit | Question | Human reference themes | Candidate themes | Pairs |")
    A("|---|---|---:|---:|---:|")
    for u, v in pu.items():
        A(f"| {u} | {v['question_id']} | {v['n_human_themes']} | "
          f"{v['n_machine_themes']} | {v['n_human_themes'] * v['n_machine_themes']} |")
    A(f"| **total** | 4 questions | **{o['n_human_themes']}** | "
      f"**{o['n_machine_themes']}** | **{uni['n_pairs']}** |\n")
    A("The 18 human themes are the single-coder supplementary reference, frozen before "
      "any model ran and never edited by this check. Centrality and relevance remain "
      "`NOT_ASSESSED` throughout.\n")

    A("## 2. The correspondence space is now complete\n")
    A("An earlier version of this document computed recall and precision from **61 of "
      "the 93** possible within-unit pairs. The other 32 had been dropped by a "
      "deterministic similarity screener whose documented role is to *propose* pairs, "
      "never to decide one. Treating its exclusions as settled non-correspondences "
      "promoted a heuristic into an adjudicator, and made the recall band, the claim of "
      "zero unresolved human themes, and the closure of the classification unsupported "
      "at the time they were published.\n")
    A("All 32 omitted pairs have since been adjudicated under the same model, mode, "
      "effort, prompt, schema, blinding, categories and gates as the original 61. No "
      "historical decision was re-run or re-interpreted.\n")
    A("| Source | Confirmed match | Confirmed non-correspondence | Unresolved | Total |")
    A("|---|---:|---:|---:|---:|")
    for s, lbl in ((hc.SOURCE_ORIGINAL, "`ORIGINAL_SCREENED_61`"),
                   (hc.SOURCE_COMPLEMENT, "`COMPLEMENT_32`")):
        d = m["pair_status_by_source"][s]
        A(f"| {lbl} | {d.get('HYBRID_CONFIRMED_MATCH', 0)} | "
          f"{d.get('HYBRID_CONFIRMED_NON_CORRESPONDENCE', 0)} | "
          f"{d.get('HYBRID_UNRESOLVED', 0)} | {sum(d.values())} |")
    tot = m["pair_status_counts"]
    A(f"| **all 93** | **{tot.get('HYBRID_CONFIRMED_MATCH', 0)}** | "
      f"**{tot.get('HYBRID_CONFIRMED_NON_CORRESPONDENCE', 0)}** | "
      f"**{tot.get('HYBRID_UNRESOLVED', 0)}** | **93** |\n")
    A("**What the 32 changed.** They produced no new confirmed matches, so recall did "
      "not move. They produced 25 further confirmed non-correspondences and 7 further "
      "unresolved pairs, and those unresolved pairs changed the precision picture "
      "substantially: the number of candidate themes that might yet correspond to a "
      "human theme rose from 1 to 5, widening the precision band from "
      f"[0.6000, 0.6333] to [{o['strict_confirmed_precision']:.4f}, "
      f"{o['possible_precision_upper_bound']:.4f}].\n")
    A("**What they also did is make the recall claim legitimate.** Under the screened "
      "set, the two unrecovered human themes had been judged against only 2 of 4 and 4 "
      "of 6 candidates in their units. Each has now been judged against its complete "
      "local universe, and every pair came back a confirmed non-correspondence. The "
      "same figure that was previously asserted is now earned.\n")

    A("## 3. Headline figures\n")
    A("| Metric | Value |")
    A("|---|---|")
    A(f"| Confirmed recall (lower bound) | **{o['confirmed_recall_lower_bound']:.4f}** "
      f"({o['n_recovered']}/{o['n_human_themes']}) |")
    A(f"| Possible recall (upper bound) | **{o['possible_recall_upper_bound']:.4f}** |")
    A(f"| Strict confirmed precision | **{o['strict_confirmed_precision']:.4f}** "
      f"({o['n_machine_matched']}/{o['n_machine_themes']}) |")
    A(f"| Possible precision (upper bound) | "
      f"**{o['possible_precision_upper_bound']:.4f}** "
      f"(+{o['n_machine_unresolved_possibly_matched']} unresolved) |")
    A(f"| Exploratory adjusted precision, counting corroborated novelty | "
      f"{o['exploratory_adjusted_precision_including_corroborated_novelty']:.4f} |")
    A(f"| Literal evidence attachment | {ev['value']:.4f} "
      f"({ev['numerator']}/{ev['denominator']}) |\n")
    A("Theme-level states, from the complete universe:\n")
    A("| Side | State | n |")
    A("|---|---|---:|")
    for k, v in m["human_theme_states"].items():
        A(f"| human | {k} | {v} |")
    for k, v in m["machine_theme_states"].items():
        A(f"| candidate | {k} | {v} |")
    A("")
    A(f"**The recall band is zero-width and this time that is a finding.** All "
      f"{o['n_human_themes']} human themes were adjudicated against every candidate in "
      f"their unit. {o['n_recovered']} have at least one confirmed match; "
      f"{o['n_confirmed_not_recovered']} — "
      f"{', '.join('`' + k + '`' for k in o['human_themes_confirmed_not_recovered'])} — "
      "have a confirmed non-correspondence with every single candidate in their unit "
      "and nothing unresolved. No human theme sits in between, so there is no "
      "uncertainty for the band to express. Every unresolved pair in the study attaches "
      "to a human theme that is already recovered through another candidate.\n")
    A(f"**Precision is the weaker axis and carries a real band.** "
      f"{o['n_machine_matched']} of {o['n_machine_themes']} candidates correspond to a "
      f"human theme, {o['n_machine_confirmed_unmatched']} are confirmed to match "
      f"nothing, and {o['n_machine_unresolved_possibly_matched']} hold at least one "
      "unsettled pair and could fall either way. The honest statement is the interval "
      f"[{o['strict_confirmed_precision']:.4f}, "
      f"{o['possible_precision_upper_bound']:.4f}], not a point.\n")
    both = m["machine_only"]["corroborated_novel_but_pairwise_unresolved"]
    if both:
        A(f"Four of those open candidates — "
          f"{', '.join('`' + k + '`' for k in both)} — were separately judged "
          "`VALID_NOVEL_THEME` when shown the unit's complete reference inventory, "
          "while their pairwise correspondence against one specific reference theme "
          "stayed unsettled. Those are different questions and the tension is left "
          "standing rather than resolved by fiat: a corroborated novel theme is never "
          "converted into a human correspondence.\n")
    A("The adjusted figure of "
      f"{o['exploratory_adjusted_precision_including_corroborated_novelty']:.4f} counts "
      f"the {o['n_corroborated_novel']} candidates a blinded auditor twice judged to be "
      "valid themes the human coder did not record. That is a claim about *the "
      "auditor's* reading, not evidence that the coder missed something — a single "
      "coder working to a defined scope is entitled to leave material uncoded. Read it "
      "as an upper envelope, not a correction to the reference.\n")

    A("## 4. By question\n")
    A("| Q | Units | Human | Cand. | Recall lower | Recall upper | Prec. lower | "
      "Prec. upper | Prec. band | Corrob. novel |")
    A("|---|---|---:|---:|---:|---:|---:|---:|---:|---:|")
    for q, v in pq.items():
        A(f"| {q} | {', '.join(v['units'])} | {v['n_human_themes']} | "
          f"{v['n_machine_themes']} | {v['confirmed_recall_lower_bound']:.4f} | "
          f"{v['possible_recall_upper_bound']:.4f} | "
          f"{v['strict_confirmed_precision']:.4f} | "
          f"{v['possible_precision_upper_bound']:.4f} | "
          f"{v['precision_band_width']:.3f} | {v['n_corroborated_novel']} |")
    A("")
    A("Recall is at or above 0.75 in every question and reaches 1.00 in Q2 and Q4. "
      "Precision is the axis that moves, and it moves downward everywhere except Q2. "
      "The pattern has one straightforward reading: the extractor is generous. It "
      "recovers nearly everything the coder recorded and then proposes more besides — "
      "most visibly in Q4, where eleven candidates stand against four human themes.\n")
    A("**These four rows must not be compared with one another statistically.** Each "
      "rests on one or two units and between three and six human themes; a difference "
      "of one theme moves a per-question rate by 0.17 to 0.33.\n")

    A("## 5. Adjudication outcomes\n")
    A("| Stage | Outcome | n |")
    A("|---|---|---:|")
    for k, v in sorted(tot.items(), key=lambda x: -x[1]):
        A(f"| correspondence, all 93 pairs | {k} | {v} |")
    for k, v in sorted(m["machine_only"]["counts"].items(), key=lambda x: -x[1]):
        A(f"| candidate-only status | {k} | {v} |")
    for k, v in m["granularity"]["counts"].items():
        A(f"| granularity | {k.replace('|', ' → ')} | {v} |")
    A("")
    A("Granularity and candidate-only status were **re-derived from the complete 93-pair "
      "universe**. Because all 19 confirmed matches come from the original 61 pairs, the "
      "fragmentation and fusion multiplicities are unchanged and the set of candidates "
      "with no confirmed match is unchanged, so the complement created no new cases of "
      "either kind. Nothing already corroborated was re-audited.\n")

    def _why(r):
        j = " ".join(r["reasons"])
        return ("disagree" if "disagree" in j else "errored" if "no judgement" in j
                else "quote" if "not literal" in j else "LOW" if "LOW" in j else "other")
    w = Counter(_why(r) for r in m["unresolved_pairs"])
    n_un = len(m["unresolved_pairs"])
    A(f"{n_un} of the 93 pairs remain `HYBRID_UNRESOLVED`: {w['disagree']} because the "
      f"two repetitions disagreed — typically `RELATED_BUT_DISTINCT` against a partial "
      f"overlap, the boundary the rubric is least sharp at — {w['quote']} because a "
      f"cited quotation was not literal, and {w['errored']} because a request errored "
      "and was retained rather than resent. They are listed individually in the tables "
      "workbook and none is resolved by inference.\n")

    A("## 6. Comparison with Q3 — two conclusions\n")
    A(f"Q3 reference, descriptive only: recall {hy.Q3_REFERENCE['recall']:.4f}, strict "
      f"precision {hy.Q3_REFERENCE['strict_precision']:.4f}. Those figures come from a "
      "different question, a different denominator (44 theme × unit instances) and a "
      "two-coder reference. They are a landmark, not a control condition, and no test "
      "is run against them.\n")
    pc = m["pre_complement_classification"]
    A(f"Superseded historical result: `{pc['value']}` — **{pc['status']}**. "
      f"{pc['why_superseded'].capitalize()}. It is retained for the record and must "
      "never be cited as a current figure.\n")
    A("### 6.1 Frozen-rule classification\n")
    A(f"> **`{fr['value']}`** — {fr['reason']}.\n")
    A("The rule was fixed before any result existed and keys on **recall only**. It has "
      "not been retrofitted to include precision, because rewriting a predefined rule "
      "after seeing the data would destroy the reason for freezing it. Applied to the "
      f"corrected figures: unresolved share of human themes "
      f"{o['n_unresolved_possibly_recovered']}/{o['n_human_themes']} = 0%, well under "
      "the 40% ceiling; mean per-question recall band 0.000, far under 0.35; every "
      f"question's band at or above {hy.Q3_REFERENCE['recall']}; no unsupported theme "
      "corroborated in two or more units.\n")
    A("### 6.2 Balanced interpretation\n")
    A(f"> {bal['statement']}\n")
    d = bal["dimensions_weighed"]
    A("| Dimension | Outside Q3 | Q3 landmark |")
    A("|---|---|---|")
    A(f"| Recall band | [{d['recall_outside_q3'][0]:.4f}, "
      f"{d['recall_outside_q3'][1]:.4f}] | {hy.Q3_REFERENCE['recall']:.4f} |")
    A(f"| Strict precision | {d['strict_precision_outside_q3']:.4f} | "
      f"{hy.Q3_REFERENCE['strict_precision']:.4f} |")
    A(f"| Precision band | [{d['precision_band_outside_q3'][0]:.4f}, "
      f"{d['precision_band_outside_q3'][1]:.4f}] | — |")
    A(f"| Candidates per human theme | "
      f"{d['thematic_proliferation']['n_machine_themes_per_human_theme']} | — |")
    A(f"| Corroborated novel candidates | "
      f"{d['thematic_proliferation']['n_corroborated_novel']} | — |")
    A(f"| Unresolved pairs | {d['n_unresolved_pairs']} of 93 | — |")
    A("")
    A(f"Strict precision outside Q3 sits {d['strict_precision_gap_vs_q3']:.4f} below "
      "the Q3 landmark, and even the optimistic end of the precision band "
      f"({o['possible_precision_upper_bound']:.4f}) does not reach it. The frozen rule "
      "returns compatibility because it measures recall; it is not a summary of overall "
      "fidelity, and reporting it alone would overstate the finding. Both conclusions "
      "are reported together for that reason.\n")
    A("This check does **not** establish transportability, validate the procedure, or "
      "show the two settings to be equivalent. Six units, 18 human themes, one coder, "
      "no second human adjudicator, and an auditor rated "
      "`USABLE_FOR_CORROBORATION_ONLY` in the Q3 phase after producing non-verbatim "
      "quotations of its own.\n")

    A("### 6.3 Substantive conclusion\n")
    A(f"Across the six supplementary units, the automatic extractor recovered "
      f"**{o['n_recovered']} of {o['n_human_themes']}** human themes. The "
      f"{o['n_confirmed_not_recovered']} it did not recover — "
      f"{', '.join('`' + k + '`' for k in o['human_themes_confirmed_not_recovered'])} — "
      "were each adjudicated against **every** candidate theme in their own unit, and "
      "every one of those pairs returned a confirmed non-correspondence. Their absence "
      "is a measured result, not a gap in the adjudication.\n")
    A(f"Precision is the weaker axis. **Strict confirmed precision is "
      f"{o['n_machine_matched']}/{o['n_machine_themes']} = "
      f"{o['strict_confirmed_precision']:.4f}** — this is the primary precision "
      f"estimate. **{o['n_machine_unresolved_possibly_matched']} candidate themes hold "
      "a correspondence that is still uncertain**, so the possible upper bound is "
      f"{o['n_machine_matched'] + o['n_machine_unresolved_possibly_matched']}/"
      f"{o['n_machine_themes']} = {o['possible_precision_upper_bound']:.4f}. Both ends "
      "of that interval are below the Q3 landmark of "
      f"{hy.Q3_REFERENCE['strict_precision']:.4f}.\n")
    A(f"A further **{o['n_corroborated_novel']} candidates were corroborated by Claude "
      "as novel themes** — valid, distinct, and absent from the coder's record for that "
      "unit. That is **automated corroboration, not human validation**: no researcher "
      "adjudicated them, and a single coder working to a defined scope is entitled to "
      "leave material uncoded. The resulting **adjusted precision of "
      f"{o['n_machine_matched'] + o['n_corroborated_novel']}/{o['n_machine_themes']} = "
      f"{o['exploratory_adjusted_precision_including_corroborated_novelty']:.4f} is an "
      "optimistic exploratory ceiling, not the headline estimate** — the headline "
      f"estimate is {o['strict_confirmed_precision']:.4f}.\n")
    q2, q4 = pq["Q2"], pq["Q4"]
    A(f"By question, **Q2 was the cleanest case**: {q2['n_human_themes']} human themes, "
      f"{q2['n_machine_themes']} candidates, recall "
      f"{q2['confirmed_recall_lower_bound']:.4f} and precision "
      f"{q2['strict_confirmed_precision']:.4f}, with no unresolved pair and no surplus "
      f"candidate. **Q4 showed the greatest thematic proliferation**: "
      f"{q4['n_machine_themes']} candidates against {q4['n_human_themes']} human "
      f"themes, precision {q4['strict_confirmed_precision']:.4f} and "
      f"{q4['n_corroborated_novel']} corroborated novel themes. The extractor's "
      "characteristic behaviour outside Q3 is generosity: it recovers nearly everything "
      "the coder recorded, and proposes a good deal more.\n")
    A("Taken together, the two conclusions in §6.1 and §6.2 stand as the finding. The "
      "frozen rule returns compatibility because it measures recall, and recall is "
      "genuinely comparable to Q3. Precision is not. **Neither equivalence nor "
      "established transportability is demonstrated**, and this check is not a "
      "validation of the procedure.\n")

    A("## 7. Cost\n")
    c = cost["claude"]
    for s in c["per_stage"]:
        A(f"- {s['stage']}: {s['n_requests']} requests, {s['input_tokens']:,} in / "
          f"{s['output_tokens']:,} out — `{s['job_id']}`")
    A(f"- **Total Claude: {c['actual_input_tokens']:,} in / "
      f"{c['actual_output_tokens']:,} out → "
      f"${c['calculated_list_batch_cost_usd']:.2f}** at the verified list Batch rate.")
    A(f"- Gemini extraction, {cost['gemini']['actual_input_tokens']:,} in / "
      f"{cost['gemini']['actual_output_tokens']:,} out → cost **not calculated**: no "
      f"published Batch rate for `{hy.GEMINI_MODEL}` was verified, and an unsourced "
      "rate would be worse than none. Gemini was **not** re-run for the complement.")
    A("- Calculated figures, not invoices. The Console is authoritative.\n")

    A("## 8. Limitations\n")
    for s in [
        "Six units, 18 human themes, one coder. Every rate here has an integer "
        "numerator small enough that a single reclassification moves it visibly.",
        "The human reference is single-coded. There is no inter-coder agreement figure "
        "for these units and none can be produced without new human work.",
        "Relevance and centrality are `NOT_ASSESSED` by methodological decision.",
        "The adjudicator is an LLM. It was blinded and required to agree with itself "
        f"across two order-reversed repetitions with literal evidence, which is why "
        f"{n_un} pairs and 1 candidate remain unresolved rather than forced.",
        "One round-1 request errored and was retained as unresolved rather than "
        "stopping the run, contrary to the stopping rule as written. Recorded as "
        "`PROTOCOL_DEVIATION_01`.",
        "The first published version of these metrics used an incomplete "
        "correspondence space. Recorded as `PROTOCOL_DEVIATION_02`.",
        "`literal_evidence_attachment_rate` = 1.0000 means every candidate carries a "
        "quotation verbatim in its own unit and not the moderator's. It says nothing "
        "about whether the theme is a warranted reading of that quotation.",
        "The frozen rule keys on recall only. A precision-keyed rule would return a "
        "different class on identical data — hence the balanced interpretation in §6.2.",
    ]:
        A(f"- {s}")
    A("")
    p = _HY / "HYBRID_TRANSPORTABILITY_RESULTS.md"
    p.write_text("\n".join(L), encoding="utf-8")
    return p


# ---------------------------------------------------------------- product 7
def tables_xlsx(m, cost):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    der = _L("hybrid_matching_derivation.json")
    uni = _L("hybrid_universe.json")
    wb = Workbook()
    bold = Font(bold=True)

    def sheet(title, headers, rows, widths=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)
    ws = wb.create_sheet("README")
    for row in [
        ["EXPLORATORY OUT-OF-Q3 TRANSPORTABILITY CHECK — TABLES"], [],
        ["Classification", hy.CLASSIFICATION],
        ["Built", m["built_utc"]],
        ["Correspondence space", "COMPLETE — all 93 within-unit pairs adjudicated"],
        ["Pre-complement classification", m["pre_complement_classification"]["value"]],
        ["  status", m["pre_complement_classification"]["status"]],
        ["FROZEN_RULE_CLASSIFICATION", m["FROZEN_RULE_CLASSIFICATION"]["value"]],
        ["  reason", m["FROZEN_RULE_CLASSIFICATION"]["reason"]],
        ["BALANCED_INTERPRETATION", m["BALANCED_INTERPRETATION"]["statement"]], [],
        ["These figures describe six supplementary units from questions Q1, Q2, Q4 and"],
        ["Q5. They are NOT a validation, NOT pooled with U01-U07/Q3, and NOT pooled"],
        ["with the deductive analysis. Q3 appears only as a descriptive landmark; no"],
        ["test is run against it."], [],
        ["Per-question rows rest on 3-6 human themes each and must not be compared"],
        ["with one another statistically."], [],
        ["See PROTOCOL_DEVIATIONS.md for the two recorded defects."],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 78
    ws.column_dimensions["B"].width = 56

    o = m["overall_within_check"]
    sheet("Headline", ["metric", "value", "numerator", "denominator", "note"], [
        ["confirmed_recall_lower_bound", o["confirmed_recall_lower_bound"],
         o["n_recovered"], o["n_human_themes"], "human themes with >=1 confirmed match"],
        ["possible_recall_upper_bound", o["possible_recall_upper_bound"],
         o["n_recovered"] + o["n_unresolved_possibly_recovered"], o["n_human_themes"],
         "adds human themes with no match but >=1 unresolved pair"],
        ["n_confirmed_not_recovered", o["n_confirmed_not_recovered"], "", "",
         "every candidate in the unit adjudicated, all confirmed non-correspondences"],
        ["strict_confirmed_precision", o["strict_confirmed_precision"],
         o["n_machine_matched"], o["n_machine_themes"],
         "candidates with >=1 confirmed match"],
        ["possible_precision_upper_bound", o["possible_precision_upper_bound"],
         o["n_machine_matched"] + o["n_machine_unresolved_possibly_matched"],
         o["n_machine_themes"], "adds candidates with >=1 unresolved pair"],
        ["exploratory_adjusted_precision_including_corroborated_novelty",
         o["exploratory_adjusted_precision_including_corroborated_novelty"],
         o["n_machine_matched"] + o["n_corroborated_novel"], o["n_machine_themes"],
         "an envelope on the auditor's reading, NOT a correction to the coder"],
        ["literal_evidence_attachment_rate",
         m["evidence"]["literal_evidence_attachment_rate"]["value"],
         m["evidence"]["literal_evidence_attachment_rate"]["numerator"],
         m["evidence"]["literal_evidence_attachment_rate"]["denominator"],
         m["evidence"]["literal_evidence_attachment_rate"]["does_NOT_measure"]],
    ], [58, 12, 11, 12, 76])

    cols = ["n_human_themes", "n_machine_themes", "confirmed_recall_lower_bound",
            "possible_recall_upper_bound", "recall_band_width",
            "strict_confirmed_precision", "possible_precision_upper_bound",
            "precision_band_width",
            "exploratory_adjusted_precision_including_corroborated_novelty",
            "n_recovered", "n_confirmed_not_recovered",
            "n_machine_matched", "n_machine_unresolved_possibly_matched",
            "n_machine_confirmed_unmatched", "n_corroborated_novel"]
    sheet("Per_question", ["question_id", "units"] + cols,
          [[q, ", ".join(v["units"])] + [v[c] for c in cols]
           for q, v in m["per_question"].items()], [11, 12] + [15] * len(cols))
    sheet("Per_unit", ["unit_id", "question_id"] + cols,
          [[u, v["question_id"]] + [v[c] for c in cols]
           for u, v in m["per_unit"].items()], [9, 11] + [15] * len(cols))

    sheet("All_93_pairs",
          ["case_id", "source_round", "unit_id", "question_id", "human_key",
           "machine_key", "category", "status", "evidence_verified",
           "gate_failure_reasons"],
          [[r["case_id"], r["source_round"], r["blind_unit_id"], r["question_id"],
            r["human_key"], r["machine_key"], r["category"] or "", r["status"],
            r.get("evidence_verified"), "; ".join(r["reasons"])]
           for r in uni["rows"]],
          [34, 22, 9, 11, 20, 12, 40, 36, 14, 62])

    sheet("Human_theme_states",
          ["human_key", "unit", "state", "n_pairs_adjudicated", "n_pairs_in_unit",
           "local_universe_complete", "confirmed_matches", "unresolved_pairs"],
          [[k, v["unit"], v["state"], v["n_pairs_adjudicated"], v["n_pairs_in_unit"],
            v["local_universe_complete"], ", ".join(v["confirmed_matches"]),
            ", ".join(v["unresolved_pairs"])]
           for k, v in sorted(uni["human_state"].items())],
          [20, 8, 34, 14, 13, 14, 30, 26])

    sheet("Machine_theme_states",
          ["machine_key", "unit", "state", "n_pairs_adjudicated", "n_pairs_in_unit",
           "local_universe_complete", "confirmed_matches", "unresolved_pairs",
           "candidate_only_status"],
          [[k, v["unit"], v["state"], v["n_pairs_adjudicated"], v["n_pairs_in_unit"],
            v["local_universe_complete"], ", ".join(v["confirmed_matches"]),
            ", ".join(v["unresolved_pairs"]),
            next((x["status"] for x in der["machine_only"]
                  if x["machine_key"] == k), "")]
           for k, v in sorted(uni["machine_state"].items())],
          [14, 8, 32, 14, 13, 14, 34, 26, 36])

    sheet("Unresolved",
          ["case_id", "source_round", "unit_id", "human_key", "machine_key",
           "why_unresolved", "moves_recall", "moves_precision"],
          [[r["case_id"], r["source_round"], r["blind_unit_id"], r["human_key"],
            r["machine_key"], "; ".join(r["reasons"]),
            ("no — this human theme is already recovered via another candidate"
             if uni["human_state"][r["human_key"]]["state"] == "RECOVERED" else "YES"),
            ("no — this candidate already has a confirmed match"
             if uni["machine_state"][r["machine_key"]]["state"] == "MATCHED"
             else "YES — inside the precision band")]
           for r in m["unresolved_pairs"]],
          [34, 22, 9, 20, 12, 62, 56, 44])

    sheet("Candidate_only",
          ["case_id", "machine_key", "unit_id", "question_id", "category", "status",
           "also_pairwise_unresolved", "gate_failure_reasons"],
          [[x["case_id"], x["machine_key"], x["blind_unit_id"], x["question_id"],
            x["category"] or "", x["status"],
            x["machine_key"] in m["machine_only"]
            ["corroborated_novel_but_pairwise_unresolved"],
            "; ".join(x["reasons"])] for x in der["machine_only"]],
          [24, 14, 9, 11, 26, 36, 22, 62])

    sheet("Granularity",
          ["case_id", "unit_id", "question_id", "relation", "category", "status",
           "keys"],
          [[x["case_id"], x["blind_unit_id"], x["question_id"], x["relation"] or "",
            x["category"] or "", x["status"],
            json.dumps({k: v for k, v in x["provenance"].items() if k != "relation"})]
           for x in der["granularity"]],
          [34, 9, 11, 24, 34, 16, 52])

    c = cost["claude"]
    rows = [[f"{s['stage']} — input_tokens", s["input_tokens"], s["job_id"]]
            for s in c["per_stage"]]
    rows += [[f"{s['stage']} — output_tokens", s["output_tokens"], s["job_id"]]
             for s in c["per_stage"]]
    rows += [
        ["claude_total_input_tokens", c["actual_input_tokens"], "measured"],
        ["claude_total_output_tokens", c["actual_output_tokens"], "measured"],
        ["claude_batch_input_rate_per_mtok_usd", c["batch_input_rate_per_mtok_usd"],
         "verified 2026-08-02"],
        ["claude_batch_output_rate_per_mtok_usd", c["batch_output_rate_per_mtok_usd"],
         "verified 2026-08-02"],
        ["claude_calculated_list_batch_cost_usd", c["calculated_list_batch_cost_usd"],
         "calculated, not an invoice"],
        ["gemini_input_tokens", cost["gemini"]["actual_input_tokens"],
         "measured; not re-run for the complement"],
        ["gemini_output_tokens", cost["gemini"]["actual_output_tokens"], "measured"],
        ["gemini_calculated_cost_usd", "NOT_CALCULATED", cost["gemini"]["why"]]]
    sheet("Cost", ["item", "value", "note"], rows, [46, 16, 76])

    p = _HY / "HYBRID_TRANSPORTABILITY_TABLES.xlsx"
    tmp = p.with_suffix(".tmp")
    wb.save(tmp)
    os.replace(tmp, p)
    return p


# ---------------------------------------------------------------- product 8
def traceability_md(m, cost):
    man, cman = _L("hybrid_manifest.json"), _L("hybrid_complement_manifest.json")
    gj = _L("gemini_job.json")
    uni = _L("hybrid_universe.json")
    L = []
    A = L.append
    A("# Exploratory out-of-Q3 transportability check — traceability\n")
    A(f"**`{hy.CLASSIFICATION}`** · built {datetime.now(UTC).isoformat()}\n")
    A("Every figure in the results document can be reached from this index. Artefacts "
      "are listed with the SHA-256 of the file as it stands now, so any later edit is "
      "detectable.\n")

    A("## Provenance chain\n")
    A("| # | Stage | Produced by | Artefact | Gate |")
    A("|---|---|---|---|---|")
    for row in [
        ("0", "Freeze protocol and rules", "scripts/hybrid_transportability.py",
         "hybrid_manifest.json", "rules frozen before any model ran"),
        ("1", "Validate inputs", "scripts/hybrid_transportability.py --validate",
         "manifest.input_validation",
         "18/18 themes, 6/6 unit hashes, boundaries clean, quotes literal"),
        ("2", "Emergent extraction", "scripts/hybrid_gemini_extract.py",
         "gemini_extraction_results.json", "6/6 COMPLETE or stop"),
        ("3", "Candidate proposal", "scripts/hybrid_candidates.py",
         "hybrid_candidates.json",
         "both sides covered — NOT the whole pair space, see DEVIATION_02"),
        ("4a", "Blinded adjudication, round 1", "scripts/hybrid_claude_audit.py",
         "claude_round1_results.json", "blinding + schema + repetition gates"),
        ("4b", "Blinded adjudication, round 2", "scripts/hybrid_claude_audit.py",
         "claude_round2_results.json", "same gates; candidate-only and granularity"),
        ("4c", "Complement: the 32 omitted pairs", "scripts/hybrid_complement.py",
         "hybrid_complement_manifest.json, claude_complement_results.json",
         "set equality 61+32=93 proved before submission; same gates"),
        ("5", "Universe integration", "scripts/hybrid_universe.py",
         "hybrid_universe.json",
         "93/93 present, no duplicates, no cross-unit pair, history unchanged"),
        ("6", "Metrics", "scripts/hybrid_metrics.py",
         "hybrid_metrics.json, hybrid_matching_derivation.json",
         "recomputed from the complete universe"),
        ("7", "Products", "scripts/hybrid_products.py",
         "results, tables, traceability, cost",
         "frozen rule + balanced interpretation, both reported"),
    ]:
        A("| " + " | ".join(f"`{x}`" if i in (2, 3) else x
                            for i, x in enumerate(row)) + " |")
    A("")

    A("## The correspondence space\n")
    q = cman["arithmetic"]
    A(f"| Quantity | n |")
    A("|---|---:|")
    A(f"| human themes | {q['n_human_themes']} |")
    A(f"| candidate themes | {q['n_machine_themes']} |")
    A(f"| within-unit pairs possible | {q['n_cartesian_within_unit']} |")
    A(f"| adjudicated in the original run | {q['n_historical_audited']} |")
    A(f"| adjudicated in the complement | {q['n_complement']} |")
    A(f"| duplicates | {q['duplicates']} |")
    A("")
    A("Per unit: " + ", ".join(f"{u} {v['n_human']}×{v['n_machine']}={v['n_pairs']}"
                               for u, v in q["per_unit"].items()) + ".\n")
    A(f"61 + 32 reconstitutes the cartesian exactly: "
      f"`{q['reconstitutes_cartesian']}`. The screener's own rejection list was checked "
      "independently and is exactly the complement. The 61 historical pairs were "
      "carried forward read-only; their SHA-256 record is in "
      "`hybrid_complement_manifest.json → historical_pairs_read_only`.\n")

    A("## Artefact hashes\n")
    A("| Artefact | SHA-256 | bytes |")
    A("|---|---|---:|")
    me = "HYBRID_TRANSPORTABILITY_TRACEABILITY.md"
    for f in sorted(_HY.glob("*")):
        if f.is_file() and f.name != me:
            A(f"| `{f.name}` | `{_sha(f)}` | {f.stat().st_size:,} |")
    A("")
    A(f"`{me}` is not listed: a file cannot carry its own hash. Verify it by "
      "regenerating it — `py scripts/hybrid_products.py` is deterministic apart from "
      "the build timestamp in its header.\n")

    A("## Model calls\n")
    A("| Stage | Model | Mode | Job id | Requests | Complete |")
    A("|---|---|---|---|---:|---:|")
    A(f"| extraction | `{hy.GEMINI_MODEL}` | batch | `{gj['job_name']}` | 6 | 6 |")
    for s in cost["claude"]["per_stage"]:
        A(f"| {s['stage']} | `{hy.AUDITOR_MODEL}` | batch | `{s['job_id']}` | "
          f"{s['n_requests']} | {s['n_complete']} |")
    A("")
    A(f"Extraction prompt SHA-256 `{gj['prompt_sha256']}` — byte-identical to the "
      "prompt frozen for U01–U07/Q3. Response schema SHA-256 "
      f"`{gj['response_schema_sha256']}`. **Gemini was not re-run for the complement**; "
      "the extraction is the original sealed one.\n")
    a = cman["audit_configuration"]
    A(f"The complementary audit used the same adjudication configuration as the "
      f"original: `{a['model']}`, {a['execution_mode']}, effort `{a['effort']}`, "
      f"structured output, {a['repetitions_per_pair']} repetitions per pair, prompt "
      f"SHA-256 `{a['prompt_sha256'][:16]}…`, schema SHA-256 "
      f"`{a['schema_sha256'][:16]}…`, identical categories and identical gates. Cache "
      "keys incorporate the pair case id, the rendered content hash and the repetition "
      "index; all 64 were unique and none collided with the 152 historical keys.\n")
    A("Every response was matched to its case by `custom_id`, never by position. "
      "custom_ids are allocated per batch, so they are unique within a job and repeat "
      "across jobs; matching is always through that job's own id map.\n")

    A("## Where each headline number comes from\n")
    o = m["overall_within_check"]
    A("| Figure | Value | Source |")
    A("|---|---|---|")
    for lbl, val, src in [
        ("confirmed_recall_lower_bound", f"{o['confirmed_recall_lower_bound']:.4f}",
         "hybrid_universe.json → human_state, count of state == RECOVERED"),
        ("possible_recall_upper_bound", f"{o['possible_recall_upper_bound']:.4f}",
         "adds state == UNRESOLVED_POSSIBLY_RECOVERED (there are none)"),
        ("n_confirmed_not_recovered", str(o["n_confirmed_not_recovered"]),
         "state == CONFIRMED_NOT_RECOVERED; each verified local_universe_complete"),
        ("strict_confirmed_precision", f"{o['strict_confirmed_precision']:.4f}",
         "machine_state, count of state == MATCHED"),
        ("possible_precision_upper_bound",
         f"{o['possible_precision_upper_bound']:.4f}",
         f"adds {o['n_machine_unresolved_possibly_matched']} with state == "
         f"UNRESOLVED_POSSIBLY_MATCHED: "
         f"{', '.join(o['machine_themes_unresolved_possibly_matched'])}"),
        ("exploratory_adjusted_precision",
         f"{o['exploratory_adjusted_precision_including_corroborated_novelty']:.4f}",
         "adds candidate-only cases with status HYBRID_CORROBORATED_NOVEL"),
        ("literal_evidence_attachment_rate",
         f"{m['evidence']['literal_evidence_attachment_rate']['value']:.4f}",
         "gemini_extraction_results.json; verified at acceptance"),
        ("FROZEN_RULE_CLASSIFICATION", m["FROZEN_RULE_CLASSIFICATION"]["value"],
         "hybrid_transportability.FINAL_RULE, frozen in hybrid_manifest.json before "
         "any result existed; unmodified"),
        ("BALANCED_INTERPRETATION", "see §6.2",
         "hybrid_metrics.json → BALANCED_INTERPRETATION.dimensions_weighed"),
        ("PRE_COMPLEMENT_CLASSIFICATION",
         m["pre_complement_classification"]["value"],
         m["pre_complement_classification"]["status"]),
    ]:
        A(f"| `{lbl}` | {val} | {src} |")
    A("")

    A("## Closure figures\n")
    A("| Quantity | Value |")
    A("|---|---|")
    t = m["pair_status_counts"]
    for lbl, val in [
        ("total pairs adjudicated", uni["n_pairs"]),
        ("  historical (`ORIGINAL_SCREENED_61`)", uni["n_historical"]),
        ("  complementary (`COMPLEMENT_32`)", uni["n_complement"]),
        ("confirmed matches", t.get("HYBRID_CONFIRMED_MATCH", 0)),
        ("confirmed non-correspondences",
         t.get("HYBRID_CONFIRMED_NON_CORRESPONDENCE", 0)),
        ("unresolved pairs", t.get("HYBRID_UNRESOLVED", 0)),
        ("recall (confirmed)",
         f"{o['n_recovered']}/{o['n_human_themes']} = "
         f"{o['confirmed_recall_lower_bound']:.4f}"),
        ("strict confirmed precision",
         f"{o['n_machine_matched']}/{o['n_machine_themes']} = "
         f"{o['strict_confirmed_precision']:.4f}"),
        ("possible precision upper bound",
         f"{o['n_machine_matched'] + o['n_machine_unresolved_possibly_matched']}/"
         f"{o['n_machine_themes']} = {o['possible_precision_upper_bound']:.4f}"),
        ("machine themes possibly matched (uncertain)",
         o["n_machine_unresolved_possibly_matched"]),
        ("corroborated novel (automated, not human-validated)",
         o["n_corroborated_novel"]),
        ("adjusted precision (optimistic exploratory ceiling)",
         f"{o['n_machine_matched'] + o['n_corroborated_novel']}/"
         f"{o['n_machine_themes']} = "
         f"{o['exploratory_adjusted_precision_including_corroborated_novelty']:.4f}"),
        ("cumulative Claude cost",
         f"USD {cost['claude']['calculated_list_batch_cost_usd']:.2f} at the list Batch "
         "rate (calculated, not an invoice)"),
    ]:
        A(f"| {lbl} | {val} |")
    A("")
    A(f"{t.get('HYBRID_CONFIRMED_MATCH', 0)} + "
      f"{t.get('HYBRID_CONFIRMED_NON_CORRESPONDENCE', 0)} + "
      f"{t.get('HYBRID_UNRESOLVED', 0)} = {uni['n_pairs']}.\n")
    A("**Recorded deviation affecting these figures.** One round-1 request errored and "
      "was never resent (`PROTOCOL_DEVIATION_01`). That pair belonged to "
      "`ORIGINAL_SCREENED_61`, was not re-examined in the complementary audit, and "
      "remains `HYBRID_UNRESOLVED` with one completed repetition. It is one of the "
      f"{o['n_machine_unresolved_possibly_matched']} sources of precision uncertainty, "
      "via `S06::M6`.\n")

    A("## Integrity checks\n")
    A(f"- 93/93 pairs present; {uni['n_historical']} + {uni['n_complement']} = "
      f"{uni['n_pairs']}; zero duplicate pair keys; every combination appears exactly "
      "once; no pair crosses a unit boundary.")
    A("- Every human and candidate theme was adjudicated against its complete local "
      "universe; a theme is called `CONFIRMED_NOT_RECOVERED` only on that basis.")
    A("- The 61 historical decisions are re-derived from the sealed round-1 results and "
      "compared cell for cell; any alteration fails the build.")
    A("- Mutation tests in `tests/test_hybrid_universe.py` plant a missing pair, a "
      "duplicated pair, a cross-unit pair, an altered historical decision, an "
      "unresolved counted as a match, and a theme declared unrecovered on an incomplete "
      "universe — each is proved to fail the corresponding guard.")
    A("")

    A("## What this check did not touch\n")
    A("Read-only throughout, and re-verified after the run by "
      "`tests/test_hybrid_transportability.py`:\n")
    for s in ("Transportability_Emergent_SingleCoder.xlsx — the single-coder workbook",
              "supplementary_human_reference.json — the frozen supplementary reference",
              "the U01–U07/Q3 calibration and every artefact belonging to it",
              "the clustering and matching workbooks",
              "the deductive results",
              "the human transcripts and the comparable windows",
              "the frozen evaluation specification and the metric registry",
              "output/session_logs/",
              "the historical Claude round-1 and round-2 responses"):
        A(f"- {s}")
    A("")
    A("Constraints held for the duration:\n")
    for s in man.get("protections", []):
        A(f"- {s}")
    A("")
    A("No new researcher task was created. Cases that could not be resolved "
      "automatically remain `HYBRID_UNRESOLVED`.\n")
    A("Deviations from the frozen protocol are recorded in `PROTOCOL_DEVIATIONS.md`. The "
      "protocol defined **four** stopping points: **1–3 passed; stopping point 4 was not "
      "applied as written** when one request errored. That pair belonged to "
      "`ORIGINAL_SCREENED_61`, was not re-examined in the complementary audit (which "
      "covered only the 32 omitted pairs), and remains `HYBRID_UNRESOLVED` with one "
      "completed repetition, contributing to the precision uncertainty around "
      "`S06::M6`.\n")

    A("## Reproduction\n")
    A("```bash")
    A("py scripts/hybrid_transportability.py --validate")
    A("py scripts/hybrid_complement.py --manifest")
    A("py scripts/hybrid_universe.py")
    A("py scripts/hybrid_metrics.py")
    A("py scripts/hybrid_products.py")
    A("```")
    A("")
    A("Derivation and products are pure functions of the sealed batch results and "
      "reproduce exactly. The three batch submissions are not re-runnable without new "
      "API calls and are guarded: every submit path refuses to run when a job record "
      "already exists.\n")

    p = _HY / "HYBRID_TRANSPORTABILITY_TRACEABILITY.md"
    p.write_text("\n".join(L), encoding="utf-8")
    return p


def main() -> int:
    m = _L("hybrid_metrics.json")
    rec = cross_model_record()
    cost = cost_record()
    p6 = results_md(m, cost)
    p7 = tables_xlsx(m, cost)
    p8 = traceability_md(m, cost)
    print(f"cross-model record : {rec['n_cases']} cases, "
          f"{sum(j['n_requests'] for j in rec['jobs'])} requests across "
          f"{len(rec['jobs'])} jobs")
    print(f"cost               : claude "
          f"${cost['claude']['calculated_list_batch_cost_usd']:.2f}, gemini "
          f"{cost['gemini']['cost_status']}")
    for p in (p6, p7, p8):
        print(f"wrote              : {p.name}  ({p.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
