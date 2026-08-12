"""
Coder package + return validator for the supplementary transportability sample.

Design follows the Part-1 emergent workbook: Instructions, Units (one row per turn),
Emergent_Coding (12 slots per unit), Overflow_Themes. No codebook, no provenance.

SINGLE CODER, SUPPLEMENTARY. No agreement statistic can or will be computed from it.

Run modes:
    --build     write the workbook and the sealed blinding map
    --validate  check a returned workbook and refuse it if incomplete or tampered

No LLM call of any kind.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_DIR = _OUT / "transportability_sample"
_SEALED = _OUT / "gold_standard_sealed"
_WB = _DIR / "Transportability_Emergent_SingleCoder.xlsx"
_SEAL = _SEALED / "transportability_package_seal.json"

CLASSIFICATION = "SUPPLEMENTARY_SINGLE_CODER_TRANSPORTABILITY_SAMPLE"
SLOTS = 12
# RELEVANCE WAS DELIBERATELY NOT ASSESSED.
#
# The researcher reviewed all six units and decided not to adjudicate central vs
# secondary: the distinction could not be determined reliably from this material and is
# not needed for the supplementary transportability objective. This is a methodological
# decision, not missing data, and it is recorded as NOT_ASSESSED.
#
# An empty relevance must never be rendered as `secondary`, `false`, `0`, or as an
# absent theme. What remains REQUIRED is the thematic substance: a label, a description
# and a literal quote. Those are what the transportability comparison uses.
RELEVANCE_STATUS = "NOT_ASSESSED"

# Required of every theme row. A row missing any of these is genuinely incomplete.
REQUIRED_THEME_FIELDS = ("theme_label", "theme_description", "supporting_quote")
# Recorded if present, never required, never imputed, never used analytically.
OPTIONAL_THEME_FIELDS = ("relevance",)
# Retained for callers that still ask for the full column set.
THEME_FIELDS = REQUIRED_THEME_FIELDS + OPTIONAL_THEME_FIELDS
OPTIONAL_FIELDS = ("coder_note",)

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
NEED = PatternFill("solid", fgColor="FFF2CC")


def _rel(p):
    """Display path, tolerant of a redirected output workspace."""
    try:
        return p.relative_to(_REPO_ROOT)
    except ValueError:
        return p


class PackageError(RuntimeError):
    pass


def _sha(t: str) -> str:
    return hashlib.sha256(t.encode("utf-8")).hexdigest()


def _norm(s: str) -> str:
    return " ".join(re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split())


def _units() -> list[dict]:
    p = _DIR / "_units_for_packaging.json"
    if not p.exists():
        raise PackageError("run build_transportability_sample.py first")
    return json.loads(p.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build() -> Path:
    units = _units()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    for i, (t, b) in enumerate([
        ("Emergent coding — supplementary sample", True),
        ("", False),
        ("WHAT THIS IS", True),
        ("Six short discussion extracts. For each one, write down the themes you see.", False),
        ("Work inductively: read the extract and describe what is there. There is no", False),
        ("codebook and you are not matching against one.", False),
        ("", False),
        ("WHAT YOU ARE NOT TOLD", True),
        ("Where each extract comes from, or anything about the others. Speakers are", False),
        ("labelled Participant 1, Participant 2 and so on. That is deliberate.", False),
        ("", False),
        ("HOW TO CODE", True),
        ("1. Read the whole extract on the Units sheet before writing anything.", False),
        ("2. On Emergent_Coding, use the rows for that unit. Twelve slots are provided;", False),
        ("   leave unused slots blank. Blank rows are fine.", False),
        ("3. For each theme give ALL FOUR of: theme_label, theme_description (one", False),
        ("   sentence), supporting_quote, relevance. A row with some fields filled and", False),
        ("   others empty will be rejected — either complete it or clear it.", False),
        ("4. supporting_quote must be copied VERBATIM from that unit's text. It is", False),
        ("   checked as a literal substring.", False),
        ("5. relevance: 'central' if it is a main idea of the extract, 'secondary' if", False),
        ("   it is present but minor.", False),
        ("6. Every unit needs at least one complete theme.", False),
        ("7. More than twelve themes for one unit? Use Overflow_Themes, and put the", False),
        ("   blind_unit_id on every overflow row.", False),
        ("", False),
        ("THIS IS A SUPPLEMENTARY SAMPLE", True),
        ("One coder, six units. It is not a gold standard and no agreement statistic", False),
        ("will be computed from it.", False),
    ], start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=b, size=11 if b else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    # --- Units: one row per turn ----------------------------------------
    ws = wb.create_sheet("Units")
    for j, (h, w) in enumerate(zip(
            ["blind_unit_id", "turn_id", "speaker", "text"], [14, 10, 15, 135]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for u in units:
        for line in u["lines"]:
            m = re.match(r"^\[(T\d+)\]\s+([^:]+):\s*(.*)$", line, re.S)
            tid, spk, txt = m.groups() if m else ("", "", line)
            ws.cell(row=r, column=1, value=u["blind_unit_id"])
            ws.cell(row=r, column=2, value=tid)
            ws.cell(row=r, column=3, value=spk.strip())
            c = ws.cell(row=r, column=4, value=txt.strip())
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{r - 1}"

    # --- Emergent_Coding: fixed grid ------------------------------------
    ws = wb.create_sheet("Emergent_Coding")
    cols = ["blind_unit_id", "slot", "theme_label", "theme_description",
            "supporting_quote", "relevance", "coder_note"]
    widths = [14, 7, 34, 50, 60, 13, 30]
    for j, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    r = 2
    for u in units:
        for slot in range(1, SLOTS + 1):
            ws.cell(row=r, column=1, value=u["blind_unit_id"])
            ws.cell(row=r, column=2, value=slot)
            for j in range(3, 8):
                cell = ws.cell(row=r, column=j)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if j < 7:
                    cell.fill = NEED
            r += 1
    dv = DataValidation(type="list", formula1='"central,secondary"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{r - 1}")
    ws.freeze_panes = "C2"

    # --- Overflow -------------------------------------------------------
    ws = wb.create_sheet("Overflow_Themes")
    for j, (h, w) in enumerate(zip(
            ["blind_unit_id", "theme_label", "theme_description", "supporting_quote",
             "relevance", "coder_note"], [14, 34, 50, 60, 13, 30]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(row=1, column=8, value=(
        "Use ONLY if a unit needs more than 12 themes. Every row must carry its "
        "blind_unit_id."))
    dv2 = DataValidation(type="list", formula1='"central,secondary"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("E2:E400")
    ws.freeze_panes = "B2"

    _DIR.mkdir(parents=True, exist_ok=True)
    wb.save(_WB)
    wb.close()

    _SEAL.write_text(json.dumps({
        "sealed_utc": datetime.now(UTC).isoformat(),
        "classification": CLASSIFICATION,
        "warning": "SEALED — do not give to the coder.",
        "workbook": str(_rel(_WB)),
        "slots_per_unit": SLOTS,
        "units": [{"blind_unit_id": u["blind_unit_id"], "question_id": u["question"],
                   "n_turns": len(u["turn_ids"]), "turn_ids": u["turn_ids"],
                   "unit_text_sha256": _sha(u["text"])} for u in units],
        "grid_fingerprint": _sha("|".join(
            f"{u['blind_unit_id']}:{s}" for u in units for s in range(1, SLOTS + 1))),
        "units_fingerprint": {u["blind_unit_id"]: _sha(u["text"]) for u in units},
    }, indent=1, ensure_ascii=False), encoding="utf-8")
    return _WB


# ---------------------------------------------------------------------------
# Validate a returned workbook
# ---------------------------------------------------------------------------

def _read(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = [{h: ("" if v is None else str(v).strip()) for h, v in zip(hdr, r)}
           for r in rows[1:]]
    wb.close()
    return out


def validate(path: Path = _WB) -> list[str]:
    problems: list[str] = []
    if not _SEAL.exists():
        return ["no package seal found — build the package before validating"]
    seal = json.loads(_SEAL.read_text(encoding="utf-8"))
    known = {u["blind_unit_id"]: u for u in seal["units"]}

    units_rows = _read(path, "Units")
    coding = _read(path, "Emergent_Coding")
    overflow = [r for r in _read(path, "Overflow_Themes")
                if any(v for k, v in r.items() if k)]

    # --- unit text unchanged ---------------------------------------------
    by_unit: dict[str, list[str]] = {}
    for r in units_rows:
        uid = r.get("blind_unit_id")
        if uid:
            by_unit.setdefault(uid, []).append(
                f"[{r.get('turn_id')}] {r.get('speaker')}: {r.get('text')}")
    for uid, exp in seal["units_fingerprint"].items():
        got = _sha("\n".join(by_unit.get(uid, [])))
        if uid not in by_unit:
            problems.append(f"{uid}: unit missing from the Units sheet")
        elif got != exp:
            problems.append(f"{uid}: unit text was modified")
    for uid in by_unit:
        if uid not in known:
            problems.append(f"{uid}: unknown blind_unit_id on the Units sheet")

    # --- fixed grid intact ------------------------------------------------
    grid = _sha("|".join(f"{r.get('blind_unit_id')}:{r.get('slot')}" for r in coding))
    if grid != seal["grid_fingerprint"]:
        problems.append(
            "the Emergent_Coding grid changed — rows were added, deleted or reordered")

    # --- theme completeness ----------------------------------------------
    complete_per_unit: dict[str, int] = {u: 0 for u in known}
    for r in coding + overflow:
        uid = r.get("blind_unit_id", "")
        slot = r.get("slot", "overflow")
        filled = [f for f in REQUIRED_THEME_FIELDS if r.get(f)]
        if not filled and not r.get("coder_note") and not r.get("relevance"):
            continue
        if uid not in known:
            problems.append(f"row with blind_unit_id {uid!r} is not a known unit")
            continue
        if len(filled) != len(REQUIRED_THEME_FIELDS):
            missing = [f for f in REQUIRED_THEME_FIELDS if not r.get(f)]
            problems.append(f"{uid} slot {slot}: partially completed, missing {missing}")
            continue
        # relevance is NOT checked here: it was not assessed by decision, and a row
        # carrying label, description and a literal quote is substantively complete.
        src = "\n".join(by_unit.get(uid, []))
        if _norm(r["supporting_quote"]) not in _norm(src):
            problems.append(f"{uid} slot {slot}: supporting_quote is not a literal "
                            f"substring of the unit")
        complete_per_unit[uid] = complete_per_unit.get(uid, 0) + 1

    for r in overflow:
        if not r.get("blind_unit_id") and any(r.get(f) for f in THEME_FIELDS):
            problems.append("Overflow_Themes row carries data but no blind_unit_id")

    for uid, n in sorted(complete_per_unit.items()):
        if n == 0:
            problems.append(f"{uid}: no complete theme — every unit needs at least one")

    # --- provenance leak --------------------------------------------------
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    blob = " ".join(str(c.value or "") for s in wb.sheetnames for row in wb[s].iter_rows()
                    for c in row).lower()
    wb.close()
    for leak in ("enriched", "demographics-only", "macho_meals", "session_logs",
                 "comparable_transcript", "fg1", "fg2", "fg3", "fg4", "fg5",
                 "synthetic", "run01", "run02", "run03"):
        if leak in blob:
            problems.append(f"provenance leak in the workbook: {leak!r}")
    return problems


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--build", action="store_true")
    ap.add_argument("--validate", action="store_true")
    ap.add_argument("--workbook", default=str(_WB))
    a = ap.parse_args()
    if a.build:
        p = build()
        print(f"workbook: {_rel(p)}")
        print(f"seal    : {_rel(_SEAL)}")
        return 0
    problems = validate(Path(a.workbook))
    print("=" * 74)
    print(f"  {CLASSIFICATION} — return validation")
    print("=" * 74)
    if problems:
        print(f"\nNOT READY — {len(problems)} problem(s):")
        for p in problems[:40]:
            print("  -", p)
        if len(problems) > 40:
            print(f"  ... and {len(problems) - 40} more")
        return 1
    print("\nREADY — all six units coded, quotes literal, grid intact.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except PackageError as exc:
        print(f"REFUSED: {exc}")
        raise SystemExit(2)


def review_flags(path: Path = _WB) -> list[str]:
    """
    Non-blocking observations. Relevance never gates; if a value survives in the sheet
    it is preserved verbatim and reported here, never used analytically.
    """
    flags: list[str] = []
    rows = _read(path, "Emergent_Coding") + _read(path, "Overflow_Themes")
    present = [(r.get("blind_unit_id"), r.get("slot"), r.get("relevance"))
               for r in rows if str(r.get("relevance") or "").strip()]
    for uid, slot, val in present:
        if val not in ("central", "secondary"):
            flags.append(f"{uid} slot {slot}: relevance is {val!r}, which is neither "
                         f"'central' nor 'secondary'. Not blocking — relevance is "
                         f"{RELEVANCE_STATUS} for this sample.")
    if present:
        flags.append(f"{len(present)} row(s) carry a relevance value. Preserved "
                     f"verbatim and NOT used analytically: relevance is "
                     f"{RELEVANCE_STATUS} for this sample.")
    return flags
