"""
Supplementary transportability sample: sampling constraints, blinding, return gate.

This sample is SINGLE-CODER and supplementary. It must never be merged with the
U01-U07 primary calibration sample, and no agreement statistic can come from it.

Every mutation test works on a COPY. The issued workbook is never modified.
No API calls.
"""

import json
import re
import pathlib
import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import build_transportability_package as pkg   # noqa: E402
import build_transportability_sample as smp    # noqa: E402

OUT = ROOT / "analysis" / "production_evaluation"
SEALED = OUT / "gold_standard_sealed"
WB = OUT / "transportability_sample" / "Transportability_Emergent_SingleCoder.xlsx"



@pytest.fixture(scope="module")
def issued(tmp_path_factory):
    """
    A freshly BUILT issued workbook — empty, deterministic, no human content.

    These tests used to treat the live returned file as a pristine template. Once the
    coder completed the sample that stopped being true: her 18 rows survived into the
    fixture and two tests failed. A file that legitimately changes cannot be a
    template.
    """
    import build_transportability_package as _pkg
    import shutil as _sh
    d = tmp_path_factory.mktemp("issued")
    sample, sealed = d / "s", d / "z"
    sample.mkdir(); sealed.mkdir()
    old = (_pkg._DIR, _pkg._WB, _pkg._SEAL)
    _sh.copy2(OUT / "transportability_sample" / "_units_for_packaging.json",
              sample / "_units_for_packaging.json")
    _pkg._DIR, _pkg._WB, _pkg._SEAL = sample, sample / "coder.xlsx", sealed / "s.json"
    try:
        _pkg.build()
    finally:
        _pkg._DIR, _pkg._WB, _pkg._SEAL = old
    return sample / "coder.xlsx"

@pytest.fixture(scope="module")
def manifest():
    p = SEALED / "transportability_sample_manifest.json"
    if not p.exists():
        pytest.skip("sample not built")
    return json.loads(p.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def audit():
    return json.loads(
        (SEALED / "transportability_boundary_audit.json").read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Sampling design
# ---------------------------------------------------------------------------

def test_exactly_six_units_two_per_stratum(manifest):
    units = manifest["units"]
    assert len(units) == 6
    assert Counter(u["stratum"] for u in units) == {
        "human": 2, "enriched": 2, "demographics-only": 2}


def test_all_four_questions_and_four_fgs(manifest):
    units = manifest["units"]
    assert {u["question_id"] for u in units} == {"Q1", "Q2", "Q4", "Q5"}
    assert len({u["fg"] for u in units}) >= 4


def test_q3_is_never_sampled(manifest):
    """Q3 is the primary calibration question; mixing it in would confound them."""
    assert all(u["question_id"] != "Q3" for u in manifest["units"])
    assert all(u["section_index"] != 3 for u in manifest["units"])
    assert 3 in smp.EXCLUDED_SECTIONS


def test_synthetic_units_are_replication_index_two(manifest):
    for u in manifest["units"]:
        if u["stratum"] == "human":
            assert u["canonical_replication_index"] is None
        else:
            assert u["canonical_replication_index"] == 2


def test_synthetic_sources_are_comparable_windows_only(manifest):
    for u in manifest["units"]:
        path = u["source_path"].replace("\\", "/")
        if u["stratum"] == "human":
            assert "standardized" in path
        else:
            assert "comparable_transcript.json" in path
            assert "output/session_logs" not in path


def test_archived_runs_are_absent(manifest):
    runs = {u["physical_run"] for u in manifest["units"]}
    assert "macho_meals_fg4_run02" not in runs
    assert "macho_meals_fg5_run02" not in runs


def test_selection_is_deterministic_and_seeded(manifest):
    assert manifest["selection_seed"] == smp.SELECTION_SEED
    assert manifest["selection_algorithm"]
    assert manifest["accepted_on_attempt"] >= 1
    frame, _ = smp.build_frame()
    again, attempt = smp.select(frame)
    got = sorted((p["stratum"], p["fg"], p["question"]) for p in again)
    want = sorted((u["stratum"], u["fg"], u["question_id"]) for u in manifest["units"])
    assert got == want, "re-running the seed changed the sample"
    assert attempt == manifest["accepted_on_attempt"]


def test_selection_did_not_consult_results(manifest):
    assert "Tier-1 metrics" in manifest["not_consulted_during_selection"]
    assert "thematic content" in manifest["not_consulted_during_selection"]


def test_manifest_records_full_provenance(manifest):
    for u in manifest["units"]:
        for k in ("blind_unit_id", "question_id", "stratum", "fg", "source_path",
                  "source_sha256", "section_text_sha256", "word_count",
                  "boundary_method"):
            assert u.get(k) not in (None, ""), f"{u['blind_unit_id']} missing {k}"
        assert len(u["source_sha256"]) == 64


# ---------------------------------------------------------------------------
# Boundary audit
# ---------------------------------------------------------------------------

def test_every_unit_passed_the_boundary_audit(audit):
    assert audit["all_clear"]
    for u in audit["units"]:
        assert u["problems"] == [], f"{u['blind_unit_id']}: {u['problems']}"
        assert u["contiguous"], f"{u['blind_unit_id']}: turns were dropped"
        assert u["first_included_turn"] and u["last_included_turn"]


def test_boundary_check_is_stratum_appropriate(audit, manifest):
    """Synthetic moderators paraphrase; a 'Question N' header check would be wrong."""
    strat = {u["blind_unit_id"]: u["stratum"] for u in manifest["units"]}
    for u in audit["units"]:
        if strat[u["blind_unit_id"]] == "human":
            assert "Question N" in u["boundary_check"]
        else:
            assert "sub-entry suffix slice" in u["boundary_check"]


def test_prior_question_recap_is_removed_not_merely_advised(audit):
    """
    An earlier version RETAINED the Q3 recap and flagged it in an advisory. It is now
    excised by a verbatim suffix slice, with the dropped prefix kept in the sealed
    audit. Reporting residue is not the same as removing it.
    """
    sliced = [u for u in audit["units"] if u["opening_boundary"].get("applied")]
    assert sliced, "no unit needed a slice — check the detector"
    for u in sliced:
        ob = u["opening_boundary"]
        assert ob["dropped_prefix"], "the dropped prefix must be recorded, not discarded"
        assert ob["dropped_words"] > 0
        assert ob["dropped_prefix"] not in u["first_included_turn"]
    assert all(not u.get("advisories") for u in audit["units"]), (
        "residue should now be removed, so no advisory should remain")


# ---------------------------------------------------------------------------
# Blinding
# ---------------------------------------------------------------------------

def test_workbook_carries_no_provenance():
    if not WB.exists():
        pytest.skip("package not built")
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    blob = " ".join(str(c.value or "") for s in wb.sheetnames
                    for row in wb[s].iter_rows() for c in row).lower()
    wb.close()
    # "codebook" is NOT provenance — the instructions deliberately say there is
    # none. Only real provenance terms are scanned.
    for leak in ("enriched", "demographics", "macho_meals", "session_logs",
                 "comparable_transcript", "synthetic", "fg1", "fg2", "fg3",
                 "fg4", "fg5"):
        assert leak not in blob, f"{leak!r} leaked into the coder workbook"


def test_instructions_state_there_is_no_codebook():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    blob = " ".join(str(c.value or "") for row in wb["Instructions"].iter_rows()
                    for c in row).lower()
    wb.close()
    assert "there is no codebook" in blob
    assert "inductively" in blob


def test_speakers_are_blinded():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    speakers = {str(r[2].value) for r in wb["Units"].iter_rows(min_row=2) if r[2].value}
    wb.close()
    assert speakers
    assert all(s == "Moderator" or re.fullmatch(r"Participant \d+", s)
               for s in speakers), speakers


def test_sheet_set_is_exactly_the_part1_design():
    wb = openpyxl.load_workbook(WB, read_only=True)
    names = wb.sheetnames
    wb.close()
    assert names == ["Instructions", "Units", "Emergent_Coding", "Overflow_Themes"]


def test_twelve_slots_per_unit():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    slots = Counter(r[0].value for r in wb["Emergent_Coding"].iter_rows(min_row=2)
                    if r[0].value)
    wb.close()
    assert len(slots) == 6
    assert set(slots.values()) == {pkg.SLOTS} == {12}


# ---------------------------------------------------------------------------
# Return gate
# ---------------------------------------------------------------------------

@pytest.fixture
def filled(tmp_path, issued):
    """A valid return: one complete theme per unit, quote copied from the unit."""
    dst = tmp_path / "returned.xlsx"
    shutil.copy2(issued, dst)
    wb = openpyxl.load_workbook(dst)
    units = {}
    for r in wb["Units"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            units.setdefault(r[0], []).append(str(r[3]))
    ws = wb["Emergent_Coding"]
    done = set()
    for row in range(2, ws.max_row + 1):
        uid = ws.cell(row=row, column=1).value
        if not uid or uid in done:
            continue
        done.add(uid)
        ws.cell(row=row, column=3).value = "theme for " + str(uid)
        ws.cell(row=row, column=4).value = "One sentence describing the theme."
        ws.cell(row=row, column=5).value = " ".join(units[uid][0].split()[:8])
        ws.cell(row=row, column=6).value = "central"
    wb.save(dst)
    wb.close()
    return dst


def _edit(path, sheet, row, col, value):
    wb = openpyxl.load_workbook(path)
    wb[sheet].cell(row=row, column=col).value = value
    wb.save(path)
    wb.close()


def _first_coded_row(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Emergent_Coding"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value:
            wb.close()
            return row
    wb.close()
    pytest.skip("fixture produced no coded row")


def test_issued_workbook_is_rejected(issued):
    """An empty issued workbook has no complete theme anywhere."""
    problems = pkg.validate(issued)
    assert problems
    assert all("no complete theme" in p for p in problems)


def test_the_live_return_is_not_used_as_a_template():
    """
    The live coder file may be referenced, never copied into a fixture. This is the
    guard for the defect that broke these two tests when the coder finished.
    """
    import re as _re
    src = pathlib.Path(__file__).read_text(encoding="utf-8")
    hits = [ln.strip() for ln in src.splitlines()
            if _re.search(r"copy2?\s*\(\s*WB\b", ln)
            and not ln.strip().startswith("#")]
    assert hits == [], f"the live return is being copied: {hits}"


def test_a_valid_return_is_accepted(filled):
    problems = pkg.validate(filled)
    assert problems == [], problems


def test_a_genuinely_incomplete_theme_is_rejected(filled):
    """
    A theme missing one of the three REQUIRED fields is incomplete and must be
    rejected. Column 4 is theme_description.
    """
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 4, None)
    problems = pkg.validate(filled)
    assert any("partially completed" in p for p in problems), problems
    assert any("theme_description" in p for p in problems), problems


@pytest.mark.parametrize("column,field", [(3, "theme_label"),
                                          (4, "theme_description"),
                                          (5, "supporting_quote")])
def test_each_required_field_is_still_required(filled, column, field):
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), column, None)
    assert any(field in p and "partially completed" in p
               for p in pkg.validate(filled))


def test_an_empty_relevance_does_NOT_reject_a_row(filled):
    """
    THE DISTINCTION THAT MATTERS.

    relevance was deliberately not assessed: the researcher reviewed all six units and
    declined the central/secondary judgement. An empty relevance is a recorded decision,
    not an omission, and must not block a row that carries label, description and a
    literal quote.
    """
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 6, None)
    problems = pkg.validate(filled)
    assert problems == [], problems
    assert not any("relevance" in p for p in problems)


def test_clearing_relevance_everywhere_still_validates(filled):
    """Not just one sampled row — every row in the workbook."""
    wb = openpyxl.load_workbook(filled)
    ws = wb["Emergent_Coding"]
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=6).value = None
    wb.save(filled); wb.close()
    assert pkg.validate(filled) == []


def test_relevance_is_optional_not_required():
    assert pkg.REQUIRED_THEME_FIELDS == ("theme_label", "theme_description",
                                         "supporting_quote")
    assert pkg.OPTIONAL_THEME_FIELDS == ("relevance",)
    assert "relevance" not in pkg.REQUIRED_THEME_FIELDS
    assert pkg.RELEVANCE_STATUS == "NOT_ASSESSED"


def test_at_least_one_theme_per_unit_is_still_required(filled):
    """Relaxing relevance must not relax this."""
    wb = openpyxl.load_workbook(filled)
    ws = wb["Emergent_Coding"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "S01":
            for c in (3, 4, 5, 6):
                ws.cell(row=r, column=c).value = None
    wb.save(filled); wb.close()
    assert any("S01: no complete theme" in p for p in pkg.validate(filled))


def test_a_non_literal_quote_is_still_rejected(filled):
    """Relaxing relevance must not relax quote verification."""
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 5,
          "a sentence that appears nowhere in this unit")
    assert any("not a literal" in p for p in pkg.validate(filled))

def test_non_literal_quote_is_rejected(filled):
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 5,
          "a quote that is nowhere in the unit")
    assert any("not a literal substring" in p for p in pkg.validate(filled))


def test_modified_unit_text_is_rejected(filled):
    _edit(filled, "Units", 2, 4, "the coder rewrote this turn")
    assert any("unit text was modified" in p for p in pkg.validate(filled))


def test_deleted_grid_row_is_rejected(filled):
    wb = openpyxl.load_workbook(filled)
    wb["Emergent_Coding"].delete_rows(5)
    wb.save(filled)
    wb.close()
    assert any("grid changed" in p for p in pkg.validate(filled))


def test_reordered_grid_is_rejected(filled):
    wb = openpyxl.load_workbook(filled)
    ws = wb["Emergent_Coding"]
    a = ws.cell(row=2, column=2).value
    ws.cell(row=2, column=2).value = ws.cell(row=3, column=2).value
    ws.cell(row=3, column=2).value = a
    wb.save(filled)
    wb.close()
    assert any("grid changed" in p for p in pkg.validate(filled))


def test_unknown_blind_unit_id_is_rejected(filled):
    _edit(filled, "Overflow_Themes", 2, 1, "S99")
    _edit(filled, "Overflow_Themes", 2, 2, "stray theme")
    assert any("S99" in p for p in pkg.validate(filled))


def test_overflow_row_without_unit_id_is_rejected(filled):
    _edit(filled, "Overflow_Themes", 2, 2, "a theme with no unit id")
    assert any("no blind_unit_id" in p for p in pkg.validate(filled))


def test_unit_with_no_themes_is_rejected(filled):
    wb = openpyxl.load_workbook(filled)
    ws = wb["Emergent_Coding"]
    target = ws.cell(row=2, column=1).value
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == target:
            for col in range(3, 8):
                ws.cell(row=row, column=col).value = None
    wb.save(filled)
    wb.close()
    assert any(str(target) + ": no complete theme" in p for p in pkg.validate(filled))


def test_an_odd_relevance_value_is_flagged_not_gated(filled):
    """
    relevance is out of the gate entirely, so a stray value cannot block a valid
    workbook. It is preserved verbatim and surfaced as a review flag instead.
    """
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 6, "quite important")
    assert pkg.validate(filled) == []
    flags = pkg.review_flags(filled)
    assert any("quite important" in f for f in flags), flags
    assert any("NOT used analytically" in f for f in flags), flags

def test_provenance_pasted_into_the_workbook_is_rejected(filled):
    _edit(filled, "Emergent_Coding", _first_coded_row(filled), 7,
          "note: this looked like macho_meals fg3 enriched")
    assert any("provenance leak" in p for p in pkg.validate(filled))


# ---------------------------------------------------------------------------
# Separation from the primary sample
# ---------------------------------------------------------------------------

def test_classification_is_supplementary_everywhere(manifest, audit):
    assert manifest["classification"] == smp.CLASSIFICATION
    assert "SUPPLEMENTARY_SINGLE_CODER" in smp.CLASSIFICATION
    assert "PRIMARY EMERGENT CALIBRATION SAMPLE" in manifest["relationship_to_primary"]
    assert "never pooled" in manifest["relationship_to_primary"]
    assert audit["classification"] == smp.CLASSIFICATION


def test_blind_ids_cannot_collide_with_the_primary_sample(manifest):
    """Primary units are U01-U15; these are S01-S06. No overlap is possible."""
    ids = {u["blind_unit_id"] for u in manifest["units"]}
    assert all(re.fullmatch(r"S\d{2}", i) for i in ids)
    assert not any(i.startswith("U") for i in ids)
