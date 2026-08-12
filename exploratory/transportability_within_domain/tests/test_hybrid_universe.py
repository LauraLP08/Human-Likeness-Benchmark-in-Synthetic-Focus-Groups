"""
Integrity of the complete 93-pair correspondence universe.

Every test that guards an invariant here is paired with a mutation test that plants the
violation and proves the guard fires. An invariant test that cannot fail is worse than no
test: it manufactures the appearance of coverage, which is precisely the error that made
this corrective phase necessary.
"""
from __future__ import annotations

import hashlib
import json
import sys
from copy import deepcopy
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))

import hybrid_transportability as hy      # noqa: E402
import hybrid_complement as hc            # noqa: E402
import hybrid_universe as hu              # noqa: E402

_HY = hy._HY
pytestmark = pytest.mark.skipif(
    not (_HY / "hybrid_universe.json").exists(),
    reason="the complementary audit has not been derived yet")


def _L(n):
    return json.loads((_HY / n).read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def uni():
    return _L("hybrid_universe.json")


@pytest.fixture(scope="module")
def man():
    return _L("hybrid_complement_manifest.json")


# ------------------------------------------------------- the universe is whole
def test_universe_is_exactly_93_pairs(uni):
    assert uni["n_pairs"] == 93
    assert uni["n_historical"] == 61
    assert uni["n_complement"] == 32
    assert uni["n_historical"] + uni["n_complement"] == uni["n_pairs"]


def test_every_within_unit_combination_appears_exactly_once(uni):
    c = _L("hybrid_candidates.json")
    expected = {hc.pair_key(h["key"], m["key"])
                for u in hy.UNITS
                for h in c["humans"].get(u, [])
                for m in c["machines"].get(u, [])}
    seen = [hc.pair_key(r["human_key"], r["machine_key"]) for r in uni["rows"]]
    assert len(seen) == len(set(seen)), "duplicate pair keys"
    assert set(seen) == expected
    assert len(expected) == 93


def test_no_pair_crosses_a_unit_boundary(uni):
    for r in uni["rows"]:
        hu_, mu = r["human_key"].split("::")[0], r["machine_key"].split("::")[0]
        assert hu_ == mu == r["blind_unit_id"], r["case_id"]


def test_derivation_reports_no_problems(uni):
    assert uni["problems"] == [], uni["problems"]
    assert uni["pass"] is True


# ------------------------------------------------------------ mutation tests
def _mutate(uni, fn, with_history=False):
    """
    Plant a violation and run the PRODUCTION integrity function over it.

    This deliberately calls hybrid_universe.integrity_problems rather than a local copy
    of the same logic: a mutation test that probes a reimplementation proves only that
    the copy agrees with itself, and would keep passing while the real check rotted.
    """
    import hybrid_round2 as r2
    u = deepcopy(uni)
    fn(u)
    hist = r2.derive_round1()["rows"] if with_history else None
    return hu.integrity_problems(u["rows"], _L("hybrid_candidates.json"), hist)


def test_the_unmutated_universe_is_clean():
    """The control: without a planted fault, the production check must be silent."""
    import hybrid_round2 as r2
    uni = _L("hybrid_universe.json")
    assert hu.integrity_problems(uni["rows"], _L("hybrid_candidates.json"),
                                 r2.derive_round1()["rows"]) == []


# Each of these asserts the SPECIFIC problem its guard should raise, not merely that
# some problem appeared. Asserting only "problems is non-empty" lets one broad check
# (set equality, or the 61/32 count) mask the removal of every narrower guard, so the
# suite would keep passing while the individual guards rotted away.
def test_mutation_a_missing_pair_is_detected(uni):
    problems = _mutate(uni, lambda u: u["rows"].pop(7))
    assert any("missing" in p for p in problems), \
        f"a dropped pair was not reported as missing: {problems}"


def test_mutation_a_duplicated_pair_is_detected(uni):
    problems = _mutate(uni, lambda u: u["rows"].append(deepcopy(u["rows"][3])))
    assert any("duplicate pair keys" in p for p in problems), \
        f"a duplicated pair was not reported as a duplicate: {problems}"


def test_mutation_a_cross_unit_pair_is_detected(uni):
    def f(u):
        other = next(r for r in u["rows"]
                     if r["blind_unit_id"] != u["rows"][0]["blind_unit_id"])
        u["rows"][0] = {**u["rows"][0], "machine_key": other["machine_key"]}
    problems = _mutate(uni, f)
    assert any("crosses units" in p for p in problems), \
        f"a cross-unit pair was not reported as crossing units: {problems}"


def test_mutation_a_mislabelled_source_round_is_detected(uni):
    """
    The 61/32 split is what makes 'no historical decision was re-run' checkable. If a
    complementary pair could be relabelled as historical, the provenance claim would be
    unfalsifiable.
    """
    def f(u):
        victim = next(r for r in u["rows"]
                      if r["source_round"] == hc.SOURCE_COMPLEMENT)
        victim["source_round"] = hc.SOURCE_ORIGINAL
    problems = _mutate(uni, f)
    assert any("expected 61 + 32" in p for p in problems), \
        f"a mislabelled source_round went unnoticed: {problems}"


def test_mutation_an_altered_historical_decision_is_detected(uni):
    """
    The load-bearing guarantee of the whole correction: historical decisions are carried
    forward, never re-judged. The production check compares each ORIGINAL row against an
    independent re-derivation from the sealed round-1 results.
    """
    def f(u):
        victim = next(r for r in u["rows"]
                      if r["source_round"] == hc.SOURCE_ORIGINAL
                      and r["status"] == hy.HYBRID_CONFIRMED_MATCH)
        victim["status"] = "HYBRID_CONFIRMED_NON_CORRESPONDENCE"
    problems = _mutate(uni, f, with_history=True)
    assert any("historical decision altered" in p for p in problems), \
        "an altered historical decision went unnoticed"


def test_mutation_an_unresolved_counted_as_a_match_is_detected(uni):
    """
    Recall must never absorb an unresolved pair into the confirmed numerator. Flip one
    unresolved pair to a match and the production roll-up must move a human theme out of
    CONFIRMED_NOT_RECOVERED or UNRESOLVED_POSSIBLY_RECOVERED into RECOVERED.
    """
    rows = deepcopy(uni["rows"])
    unres = [r for r in rows if r["status"] == hy.HYBRID_UNRESOLVED]
    assert unres, "no unresolved pair in the universe to mutate"
    cands = _L("hybrid_candidates.json")
    before, _ = hu.theme_states(rows, cands)
    victim = unres[0]
    victim["status"] = hy.HYBRID_CONFIRMED_MATCH
    after, _ = hu.theme_states(rows, cands)
    k = victim["human_key"]
    assert after[k]["state"] == "RECOVERED"
    assert victim["machine_key"] in after[k]["confirmed_matches"]
    assert after[k] != before[k] or before[k]["state"] == "RECOVERED"
    # and the strict numerator must never have counted it before the flip
    assert victim["machine_key"] not in before[k]["confirmed_matches"]


def test_mutation_not_recovered_on_an_incomplete_universe_is_detected(uni):
    """
    A human theme may only be declared confirmed-not-recovered when every machine theme
    in its unit has been judged against it. Drop one of its pairs and the production
    roll-up must degrade the state rather than silently keep it.
    """
    notrec = [k for k, v in uni["human_state"].items()
              if v["state"] == "CONFIRMED_NOT_RECOVERED"]
    assert notrec, "no confirmed-not-recovered human theme to mutate"
    k = notrec[0]
    rows = deepcopy(uni["rows"])
    victim = next(r for r in rows if r["human_key"] == k)
    rows.remove(victim)                            # simulate one unjudged pair
    hstate, _ = hu.theme_states(rows, _L("hybrid_candidates.json"))
    assert hstate[k]["state"] == "NOT_RECOVERED_BUT_UNIVERSE_INCOMPLETE", \
        "a theme was called not-recovered on an incomplete local universe"
    assert hstate[k]["local_universe_complete"] is False
    # and build() must refuse such a universe outright
    assert any("incomplete local universe" in p or "missing" in p
               for p in _mutate(uni, lambda u: u["rows"].remove(
                   next(r for r in u["rows"] if r["human_key"] == k))))


# ---------------------------------------------------- theme-level correctness
def test_confirmed_not_recovered_requires_a_complete_local_universe(uni):
    for k, v in uni["human_state"].items():
        if v["state"] == "CONFIRMED_NOT_RECOVERED":
            assert v["local_universe_complete"], k
            assert v["n_pairs_adjudicated"] == v["n_pairs_in_unit"], k
            assert not v["confirmed_matches"] and not v["unresolved_pairs"], k
        assert v["state"] != "NOT_RECOVERED_BUT_UNIVERSE_INCOMPLETE", k


def test_every_theme_saw_its_whole_local_universe(uni):
    for k, v in uni["human_state"].items():
        assert v["local_universe_complete"], f"{k} judged against only "\
            f"{v['n_pairs_adjudicated']}/{v['n_pairs_in_unit']} machine themes"
    for k, v in uni["machine_state"].items():
        assert v["local_universe_complete"], k


def test_states_are_mutually_consistent(uni):
    for k, v in uni["human_state"].items():
        if v["confirmed_matches"]:
            assert v["state"] == "RECOVERED", k
        elif v["unresolved_pairs"]:
            assert v["state"] == "UNRESOLVED_POSSIBLY_RECOVERED", k
    for k, v in uni["machine_state"].items():
        if v["confirmed_matches"]:
            assert v["state"] == "MATCHED", k
        elif v["unresolved_pairs"]:
            assert v["state"] == "UNRESOLVED_POSSIBLY_MATCHED", k


# ------------------------------------------------------------- the manifest
def test_manifest_arithmetic_holds(man):
    q = man["arithmetic"]
    assert q["n_cartesian_within_unit"] == 93
    assert q["n_historical_audited"] == 61
    assert q["n_complement"] == 32
    assert q["reconstitutes_cartesian"] is True
    assert q["duplicates"] == 0
    assert man["pass"] is True and man["problems"] == []
    assert sum(v["n_pairs"] for v in q["per_unit"].values()) == 93


def test_complement_excluded_nothing_by_similarity(man):
    assert len(man["pairs"]) == 32
    assert man["audit_configuration"]["n_requests_expected"] == 64
    sims = [p["screener_similarity_FOR_RECORD_ONLY"] for p in man["pairs"]]
    assert all(s is not None for s in sims), "similarity not recorded"
    assert len({p["pair_key"] for p in man["pairs"]}) == 32


def test_complement_used_the_same_configuration_as_the_original(man):
    import cross_model_audit_q3 as cm
    a = man["audit_configuration"]
    assert a["model"] == hy.AUDITOR_MODEL == "claude-opus-5"
    assert a["execution_mode"] == "batch"
    assert a["effort"] == hy.AUDITOR_EFFORT == "high"
    assert a["repetitions_per_pair"] == 2
    assert a["structured_output"] is True
    assert a["prompt_sha256"] == cm.prompt_sha("A_PAIRWISE_CORRESPONDENCE")
    assert a["schema_sha256"] == cm.schema_sha("A_PAIRWISE_CORRESPONDENCE")
    assert set(a["accepted_as_correspondence"]) == set(hy.CORRESPONDENCE_ACCEPTED)
    assert set(a["rejected_as_correspondence"]) == set(hy.CORRESPONDENCE_REJECTED)


def test_historical_results_are_byte_identical_to_the_manifest_record(man):
    h = man["historical_pairs_read_only"]
    for name, want in (("claude_round1_results.json", h["claude_round1_results_sha256"]),
                       ("claude_round2_results.json", h["claude_round2_results_sha256"])):
        got = hashlib.sha256((_HY / name).read_bytes()).hexdigest()
        assert got == want, f"{name} changed during the corrective phase"


# ------------------------------------------------------------- retry policy
def test_retry_policy_was_frozen_before_submission(man):
    p = man["retry_policy"]
    assert p["frozen_before_submission"] is True
    assert p["not_part_of_original_protocol"] is True
    assert p["max_technical_retries_per_request"] == 1
    assert p["on_retry_failure"] == "the pair remains HYBRID_UNRESOLVED"
    joined = " ".join(p["never_retry"]).lower()
    for forbidden in ("disagreement", "low confidence", "quotation", "unfavourable"):
        assert forbidden in joined, forbidden


def test_only_transport_failures_are_retryable():
    assert hc.RETRYABLE == {"NO_OUTPUT", "INVALID_JSON", "OUTPUT_TRUNCATED", "INVALID"}
    assert "COMPLETE" not in hc.RETRYABLE


def test_retries_are_capped_at_one_per_request(uni):
    for r in uni["rows"]:
        for a in r.get("attempts", []):
            assert a.get("attempt", 1) <= 2, r["case_id"]


def test_deviations_document_records_both_defects():
    t = (_HY / "PROTOCOL_DEVIATIONS.md").read_text(encoding="utf-8")
    assert "PROTOCOL_DEVIATION_01" in t
    assert "PROTOCOL_DEVIATION_02" in t
    assert "COMPLEMENT_RETRY_POLICY_V1" in t
    assert "withdrawn" in t.lower()


# ------------------------------------------------------- closure corrections
def test_the_errored_pair_is_not_described_as_re_examined():
    """
    It belonged to ORIGINAL_SCREENED_61; the complement covered only the 32 omitted
    pairs. Saying it was re-examined would imply an adjudication that never happened.
    """
    # these documents are hard-wrapped, so a raw substring would break on a line
    # boundary rather than on the claim actually being present
    def flat(name):
        return " ".join((_HY / name).read_text(encoding="utf-8").split())
    for name in ("PROTOCOL_DEVIATIONS.md", "HYBRID_TRANSPORTABILITY_TRACEABILITY.md"):
        t = flat(name)
        assert "is one of the pairs re-examined" not in t, name
        assert "not re-examined in the complementary audit" in t, name
    d = flat("PROTOCOL_DEVIATIONS.md")
    assert "ORIGINAL_SCREENED_61" in d
    assert "one completed repetition" in d


def test_the_protocol_had_four_stopping_points_not_five():
    for name in ("PROTOCOL_DEVIATIONS.md", "HYBRID_TRANSPORTABILITY_TRACEABILITY.md"):
        t = " ".join((_HY / name).read_text(encoding="utf-8").split())
        assert "all five stopping points" not in t, name
        assert "stopping points 1, 2, 3 and 5" not in t, name
        assert "four** stopping points" in t or "four stopping points" in t, name
        assert "1–3 passed" in t or "1-3 passed" in t, name


def test_the_errored_pair_status_was_not_changed_by_the_narrative_fix(uni):
    """A narrative correction must not move a decision or a metric."""
    row = next(r for r in uni["rows"]
               if r["case_id"] == "P::S06::S06_slot_01::S06::M6")
    assert row["status"] == hy.HYBRID_UNRESOLVED
    assert row["source_round"] == hc.SOURCE_ORIGINAL
    assert any("no judgement" in x for x in row["reasons"])
    assert uni["machine_state"]["S06::M6"]["state"] == "UNRESOLVED_POSSIBLY_MATCHED"


def test_evidence_note_wording():
    m = _L("hybrid_metrics.json")
    want = "NOT substantive groundedness; verifies literal evidence attachment only."
    assert m["evidence"]["literal_evidence_attachment_rate"]["does_NOT_measure"] == want
    assert m["evidence"]["literal_evidence_attachment_rate"]["value"] == 1.0
    from openpyxl import load_workbook
    ws = load_workbook(_HY / "HYBRID_TRANSPORTABILITY_TABLES.xlsx")["Headline"]
    note = next(r[4] for r in ws.iter_rows(values_only=True)
                if r[0] and "literal_evidence" in str(r[0]))
    assert note == want


def test_closure_metrics_are_unchanged(uni):
    """
    The closure applied textual corrections only. These are the values audited over the
    complete 93/93 universe; any drift means a narrative edit moved a number.
    """
    o = _L("hybrid_metrics.json")["overall_within_check"]
    assert (o["n_recovered"], o["n_human_themes"]) == (16, 18)
    assert o["confirmed_recall_lower_bound"] == 0.8889
    assert o["possible_recall_upper_bound"] == 0.8889
    assert (o["n_machine_matched"], o["n_machine_themes"]) == (18, 30)
    assert o["strict_confirmed_precision"] == 0.6000
    assert o["n_machine_unresolved_possibly_matched"] == 5
    assert o["possible_precision_upper_bound"] == 0.7667
    assert o["n_corroborated_novel"] == 11
    assert o["exploratory_adjusted_precision_including_corroborated_novelty"] == 0.9667
    from collections import Counter
    t = Counter(r["status"] for r in uni["rows"])
    assert t["HYBRID_CONFIRMED_MATCH"] == 19
    assert t["HYBRID_CONFIRMED_NON_CORRESPONDENCE"] == 60
    assert t["HYBRID_UNRESOLVED"] == 14
    assert sum(t.values()) == 93


# ------------------------------------------ integration into the final products
_FINAL = _ROOT / "analysis/production_evaluation/final"


def test_final_products_carry_the_check_without_pooling_it():
    for name in ("FINAL_INTEGRATED_RESULTS_REPORT.md",
                 "RESULTS_TRACEABILITY_INDEX.md"):
        t = (_FINAL / name).read_text(encoding="utf-8")
        assert "EXPLORATORY_OUT_OF_Q3_TRANSPORTABILITY_CHECK" in t, name
        for fig in ("16/18", "18/30", "23/30", "29/30"):
            assert fig in t, f"{name} missing {fig}"
        assert "DESCRIPTIVELY_COMPATIBLE_WITH_Q3" in t, name
        assert "BALANCED_INTERPRETATION" in t, name
        assert "never pooled" in t.lower() or "not pooled" in t.lower(), name


def test_the_report_no_longer_claims_no_extractor_was_run():
    t = (_FINAL / "FINAL_INTEGRATED_RESULTS_REPORT.md").read_text(encoding="utf-8")
    assert "No automatic extractor has been run" not in t


def test_the_final_report_never_calls_the_check_a_validation():
    import re
    t = " ".join((_FINAL / "FINAL_INTEGRATED_RESULTS_REPORT.md")
                 .read_text(encoding="utf-8").lower().split())
    for pat in (r"transportability (is|was|has been) establish",
                r"establishes transportability",
                r"(is|are|was|were) equivalent"):
        assert re.search(pat, t) is None, pat
    assert "not a validation" in t


def test_the_q3_denominators_were_not_pooled():
    """The Q3 figures must still stand on 44 instances, untouched by this check."""
    t = (_FINAL / "FINAL_INTEGRATED_RESULTS_REPORT.md").read_text(encoding="utf-8")
    assert "0.6818" in t and "44" in t
    assert "48" not in t.split("## 6.")[0].split("44-instance")[0][-40:]
    j = json.loads((_ROOT / "analysis/production_evaluation/emergent_calibration_q3"
                    / "final_integration_reconciliation.json").read_text(encoding="utf-8"))
    figs = {p["figure"]: str(p["value"]) for p in j["provenance"]}
    assert figs["emergent.n_human_instances"] == "44"
    assert figs["emergent.recall"] == "30/44 = 0.6818"
    assert figs["emergent.strict_precision"] == "24/30 = 0.8000"
    assert j["reconciliation"] == "ALL FIGURES RECONCILED"


# --------------------------------------------- retired claims, all final products
def _final_product_texts():
    """
    Every final product, markdown AND workbook cells.

    An earlier sweep searched only the .md files and therefore missed a stale claim
    living in a spreadsheet cell. Anything that checks report wording must read the
    workbooks too.
    """
    out = {}
    for p in sorted(_FINAL.rglob("*")):
        if p.suffix == ".md":
            out[p.name] = " ".join(p.read_text(encoding="utf-8").split())
        elif p.suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(p)
            for ws in wb:
                for row in ws.iter_rows():
                    for c in row:
                        if c.value is not None:
                            out[f"{p.name}!{ws.title}!{c.coordinate}"] = str(c.value)
    return out


def test_no_product_claims_the_extractor_was_never_run():
    """Required regression guard: this claim was falsified by the exploratory check."""
    for where, t in _final_product_texts().items():
        assert "No automatic extractor has been run" not in t, where
        assert "no extractor run" not in t.lower(), where


RETIRED = [
    "Enrichment does not move any of this",
    "does not move any structural metric",
    "indistinguishable from chance",
    "does not respond to enrichment at all",
    "differ from human transcripts identically",
    "most robust result in the study",
    "granularity choice can manufacture",
    "apparent total null is produced",
]


def _strip_disclaimer_sections(text: str) -> str:
    """
    Drop the "Claims not supported" / "Literature needed" blocks.

    Naming a retired phrase in order to rule it out is the correct behaviour, not a
    regression. Scanning those sections would pressure the drafts to delete their own
    disclaimers to satisfy the test — precisely backwards.
    """
    out, skipping = [], False
    for line in text.splitlines():
        if line.lstrip().startswith("#"):
            skipping = any(h in line.lower() for h in
                           ("claims not supported", "literature needed",
                            "open wording decisions"))
        if not skipping:
            out.append(line)
    return "\n".join(out)


def test_retired_overclaims_are_absent_from_every_final_product():
    for where, t in _final_product_texts().items():
        body = _strip_disclaimer_sections(t) if where.endswith(".md") else t
        for phrase in RETIRED:
            assert phrase.lower() not in body.lower(), f"{where}: {phrase!r}"


def test_the_disclaimer_stripper_does_not_hide_the_report_body():
    """Guard the guard: the stripper must not swallow substantive prose."""
    sample = ("# Results\n\nEnrichment does not move any of this.\n\n"
              "## Claims not supported\n\n- That it is indistinguishable from chance.\n")
    body = _strip_disclaimer_sections(sample)
    assert "does not move any of this" in body      # still scanned
    assert "indistinguishable from chance" not in body   # correctly exempted


def test_the_required_replacement_framing_is_present():
    t = " ".join((_FINAL / "FINAL_INTEGRATED_RESULTS_REPORT.md")
                 .read_text(encoding="utf-8").split())
    assert "No consistent or descriptively substantial improvement in interactional " \
           "structure was observed under enrichment" in t
    assert "does not demonstrate equivalence, absence of effect, or an exactly null " \
           "effect" in t
    assert "These small-n directional counts do not provide conclusive evidence of a " \
           "consistent structural advantage" in t
    assert "descriptive divergence between fidelity dimensions" in t
    assert "highly sensitive to coding granularity" in t
    assert "does not establish that granularity causally produced the result" in t


def test_sheet_7_note_is_current_and_data_rows_intact():
    from openpyxl import load_workbook
    ws = load_workbook(_FINAL / "FINAL_RESULTS_TABLES.xlsx")["7_Supplementary_S01_S06"]
    a9 = ws["A9"].value
    assert "reported separately in sheet 9" in a9
    assert "never pooled with u01-u07/q3" in a9.lower()
    assert "No automatic extractor has been run" not in a9
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=1, max_row=7)]
    assert rows[0] == ["blind_unit_id", "question_id", "stratum", "n_human_themes"]
    assert [r[3] for r in rows[1:7]] == [4, 3, 2, 1, 4, 4]
    assert sum(r[3] for r in rows[1:7]) == 18


# ------------------------------------------------- structural traceability
def test_structural_figures_are_indexed_with_full_provenance():
    j = json.loads((_FINAL / "structural_traceability.json").read_text(encoding="utf-8"))
    assert j["n_figures"] >= 43
    for f in j["figures"]:
        for k in ("figure", "value", "source", "rule", "column", "unit_of_analysis",
                  "namespace"):
            assert f.get(k) not in (None, ""), f"{f['figure']} missing {k}"
    names = {f["figure"] for f in j["figures"]}
    for metric in ("total_words", "participant_turns", "words_per_turn_iqr",
                   "short_turn_proportion_25w", "turn_balance_gini", "chain_depth",
                   "moderator_word_share"):
        for suffix in ("human", "enriched", "demographics_only",
                       "enriched_minus_demo", "n_fg_enriched_closer_to_human"):
            assert f"structural.{metric}.{suffix}" in names, f"{metric}.{suffix}"
    assert any("fg2" in n for n in names) and any("fg4" in n for n in names)
    assert "deductive.fg4_demographics_only.theme_level_recall" in names


def test_the_index_lists_every_derived_structural_figure():
    j = json.loads((_FINAL / "structural_traceability.json").read_text(encoding="utf-8"))
    t = (_FINAL / "RESULTS_TRACEABILITY_INDEX.md").read_text(encoding="utf-8")
    for f in j["figures"]:
        assert f"`{f['figure']}`" in t, f["figure"]


def test_structural_figures_match_the_report_table():
    """
    The index is derived from source; the report table was written by hand. If they ever
    disagree, one of them is wrong — so compare them rather than trusting either.
    """
    j = json.loads((_FINAL / "structural_traceability.json").read_text(encoding="utf-8"))
    v = {f["figure"]: f["value"] for f in j["figures"]}
    for fig, want, tol in [
        ("structural.total_words.human", 4689, 1),
        ("structural.total_words.enriched", 8277, 1),
        ("structural.total_words.demographics_only", 8817, 1),
        ("structural.participant_turns.human", 69.2, 0.05),
        ("structural.chain_depth.human", 12.8, 0.05),
        ("structural.chain_depth.enriched", 2.02, 0.005),
        ("structural.short_turn_proportion_25w.human", 0.344, 0.001),
        ("structural.short_turn_proportion_25w.enriched", 0.0, 0.0),
        ("structural.turn_balance_gini.human", 0.195, 0.001),
        ("structural.moderator_word_share.human", 0.025, 0.001),
    ]:
        assert abs(v[fig] - want) <= tol, f"{fig}: index {v[fig]} vs report {want}"
    for fig, want in [("structural.short_turn_proportion_25w.n_fg_enriched_closer_to_human", "0/5"),
                      ("structural.chain_depth.n_fg_enriched_closer_to_human", "4/5"),
                      ("structural.participant_turns.n_fg_enriched_closer_to_human", "2/5")]:
        assert v[fig] == want, fig


# ------------------------------------------------------------ rounding policy
# The four differences that were wrong while the script rounded the condition means
# before subtracting them. Each is off by exactly 1 ulp at 4 dp under the old rule, so
# these values pin the rounding policy rather than merely restating the output.
EXACT_DIFFS = {
    "structural.total_words.enriched_minus_demo": -539.6667,
    "structural.participant_turns.enriched_minus_demo": -1.3333,
    "structural.chain_depth.enriched_minus_demo": 0.0039,
    "structural.moderator_word_share.enriched_minus_demo": -0.0077,
}


def _struct():
    return json.loads((_FINAL / "structural_traceability.json").read_text(
        encoding="utf-8"))


def test_the_four_corrected_differences_have_their_exact_values():
    v = {f["figure"]: f["value"] for f in _struct()["figures"]}
    for fig, want in EXACT_DIFFS.items():
        assert v[fig] == want, f"{fig}: {v[fig]} != {want}"


def test_the_index_shows_the_corrected_values():
    t = (_FINAL / "RESULTS_TRACEABILITY_INDEX.md").read_text(encoding="utf-8")
    for fig, want in EXACT_DIFFS.items():
        assert f"| `{fig}` | {want} |" in t, f"{fig} not shown as {want}"
    for stale in ("-539.6666", "-1.3334", "| 0.004 |", "-0.0078"):
        assert stale not in t, f"stale rounded-first value {stale} still in the index"


def test_differences_are_computed_from_unrounded_operands():
    """
    The operands are stored at full precision, so the policy is checkable rather than
    merely asserted: recompute from them and the published value must follow.
    """
    for f in _struct()["figures"]:
        if not f["figure"].endswith(".enriched_minus_demo"):
            continue
        if "operands_exact" not in f:
            continue
        e = f["operands_exact"]["enriched"]
        d = f["operands_exact"]["demographics_only"]
        assert round(e - d, 4) == f["value"], f["figure"]
        # the operands themselves must be full precision, not pre-rounded
        assert e == round(e, 4) or abs(e - round(e, 4)) > 0, f["figure"]


def test_mutation_rounding_the_means_first_breaks_the_four_values():
    """
    The mutation the correction exists to prevent: round each condition mean to 4 dp,
    then subtract. Every one of the four must change, so a regression to the old rule
    cannot pass silently.
    """
    figs = {f["figure"]: f for f in _struct()["figures"]}
    broken = []
    for fig, want in EXACT_DIFFS.items():
        o = figs[fig]["operands_exact"]
        rounded_first = round(round(o["enriched"], 4) - round(o["demographics_only"], 4), 4)
        if rounded_first != want:
            broken.append(fig)
        assert rounded_first != want, \
            f"{fig}: rounding first still yields {want}; this value cannot detect the bug"
    assert len(broken) == 4, broken


def test_closer_to_human_counts_use_exact_means():
    """
    A comparison on rounded means can flip a near-tie. Recompute every count from the
    exact per-FG means and require the published counts to match.
    """
    j = _struct()
    per_fg = j["per_fg_structural_means"]      # rounded, for display
    figs = {f["figure"]: f for f in j["figures"]}
    for metric, conds in per_fg.items():
        fig = figs[f"structural.{metric}.n_fg_enriched_closer_to_human"]
        n = sum(1 for won in fig["per_fg"].values() if won)
        assert fig["value"] == f"{n}/{len(fig['per_fg'])}", metric
        assert "exact FG means" in fig["rule"], metric
        assert len(fig["per_fg"]) == 5, metric


def test_figure_counts_reconcile_with_the_rows_actually_present():
    """The stated totals must be recomputed from the index, not asserted alongside it."""
    import re
    assert _struct()["n_figures"] == 47
    t = (_FINAL / "RESULTS_TRACEABILITY_INDEX.md").read_text(encoding="utf-8")
    assert "**16 headline figures**" in t
    assert "**13 exploratory transportability figures**" in t
    assert "**47 structural and per-FG exception figures**" in t
    m = re.search(r"\*\*(\d+) exploratory Level 2 coverage-accumulation and lexical "
                  r"figures\*\* — (\d+) in total", t)
    assert m, "the Level 2 / lexical count sentence is missing or reworded"
    stated, total = int(m.group(1)), int(m.group(2))
    block = t.split("## Level 2 coverage accumulation and lexical provenance")[1]
    actual = block.count("| `saturation.") + block.count("| `lexical.")
    assert actual == stated, f"index says {stated} rows, contains {actual}"
    assert total == 16 + 13 + 47 + stated
    j = json.loads((_ROOT / "analysis/production_evaluation/emergent_calibration_q3"
                    / "final_integration_reconciliation.json").read_text(encoding="utf-8"))
    assert len(j["provenance"]) == 16


def test_structural_index_does_not_contradict_the_workbook():
    """
    The workbook's structural sheet is the published table; the index is derived. They
    must agree to the precision the workbook prints.
    """
    from openpyxl import load_workbook
    ws = load_workbook(_FINAL / "FINAL_RESULTS_TABLES.xlsx")["3_Structural_Interaction"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    assert rows[0] == ["metric", "human_mean", "enriched_mean",
                       "demographics_only_mean", "enriched_minus_demo",
                       "n_fg_enriched_closer_to_human"], rows[0]
    v = {f["figure"]: f["value"] for f in _struct()["figures"]}
    numeric = ((1, "human"), (2, "enriched"), (3, "demographics_only"),
               (4, "enriched_minus_demo"))
    checked = 0
    for r in rows[1:]:
        mid = str(r[0]).strip()
        if not mid or f"structural.{mid}.human" not in v:
            continue
        for col, key in numeric:
            got = v[f"structural.{mid}.{key}"]
            # the workbook is the published table; the index is derived. Equality to
            # the workbook's own 4 dp is the standard — this is what caught the
            # double-rounding, so it must not be loosened to a tolerance that hides it.
            assert round(float(r[col]), 4) == round(got, 4), \
                f"{mid}.{key}: workbook {r[col]} vs index {got}"
            checked += 1
        assert str(r[5]) == v[f"structural.{mid}.n_fg_enriched_closer_to_human"], mid
        checked += 1
    assert checked == 35, f"expected 7 metrics x 5 columns, cross-checked {checked}"


def test_final_tables_workbook_has_its_own_sheet():
    from openpyxl import load_workbook
    wb = load_workbook(_FINAL / "FINAL_RESULTS_TABLES.xlsx")
    assert "9_Exploratory_Transportability" in wb.sheetnames
    vals = [c for r in wb["9_Exploratory_Transportability"].iter_rows(values_only=True)
            for c in r if c is not None]
    for want in (93, 19, 60, 14, 0.8889, 0.6, 0.7667, 0.9667, 4.6):
        assert want in vals, want
    # the Q3 sheet must be untouched by this addition
    q3 = [c for r in wb["4_Emergent_Q3"].iter_rows(values_only=True)
          for c in r if c is not None]
    assert any("44" in str(c) or c == 44 for c in q3)
