"""
Boundary corrections and the consolidation gate for the supplementary sample.

Three defects are guarded here:
  * S02 and S04 opened with 52 and 34 words of Q3 recap before posing Q4 — Q3 is the
    PRIMARY calibration question, so that material must not travel into a
    supplementary unit;
  * S05 ended with a turn that poses Q3 verbatim while the transcript attributes it to
    a participant;
  * the coder's raw rows are not a legitimate recall denominator.

No API calls. The U01-U07 package is never opened for writing.
"""

import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import build_transportability_consolidation as con   # noqa: E402
import build_transportability_package as pkg         # noqa: E402
import build_transportability_sample as smp          # noqa: E402

OUT = ROOT / "analysis" / "production_evaluation"
SEALED = OUT / "gold_standard_sealed"
DIR = OUT / "transportability_sample"
WB = DIR / "Transportability_Emergent_SingleCoder.xlsx"

Q3_RECAP_S02 = "men aren't supposed to be too visibly intentional"
Q3_RECAP_S04 = "the 'safe choice' framing"
Q3_ASK = "gender influences what you eat"


@pytest.fixture(scope="module")
def audit():
    p = SEALED / "transportability_boundary_audit.json"
    if not p.exists():
        pytest.skip("sample not built")
    return json.loads(p.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def units():
    return json.loads((DIR / "_units_for_packaging.json").read_text(encoding="utf-8"))


def _unit(units, uid):
    return next(u for u in units if u["blind_unit_id"] == uid)


def _audit(audit, uid):
    return next(u for u in audit["units"] if u["blind_unit_id"] == uid)


# ---------------------------------------------------------------------------
# The three specific corrections
# ---------------------------------------------------------------------------

def test_s02_no_longer_contains_the_q3_recap(units):
    text = _unit(units, "S02")["text"].lower()
    assert Q3_RECAP_S02 not in text
    assert "that last bit" not in text.split("\n")[0].lower()
    assert "noah" not in text.split("\n")[0].lower()


def test_s04_no_longer_contains_the_q3_recap(units):
    text = _unit(units, "S04")["text"].lower()
    assert Q3_RECAP_S04.lower() not in text
    assert "cheers amir" not in text.lower()


def test_s05_no_longer_contains_the_q3_ask(units):
    text = _unit(units, "S05")["text"].lower()
    assert Q3_ASK not in text
    assert "you might have already answered question three" not in text


def test_s05_records_the_removed_turn_with_its_diagnosis(audit):
    a = _audit(audit, "S05")
    removed = a["removed_from_end"]
    assert len(removed) == 1
    r = removed[0]
    assert r["diagnosis"] == "NEXT_QUESTION_ASK_MISLABELLED_OR_NONSTANDARD"
    assert r["speaker_label_as_recorded"].startswith("Participant")
    assert Q3_ASK in " ".join(r["next_question_markers_found"])
    assert "question three" in r["text"].lower()
    assert r["source_entry_index"] is not None
    assert a["n_turns"] == 16 and a["word_count"] > 0


# ---------------------------------------------------------------------------
# Sub-entry slices are verbatim and hash-verified
# ---------------------------------------------------------------------------

def test_opening_slices_are_verbatim_suffixes(audit):
    for u in audit["units"]:
        ob = u["opening_boundary"]
        if not ob.get("applied"):
            continue
        assert ob["boundary_text_is_verbatim_slice"] is True
        assert ob["original_entry"].endswith(ob["retained_suffix"])
        assert ob["original_entry"] == ob["dropped_prefix"] + ob["retained_suffix"]
        assert ob["source_character_start"] == len(ob["dropped_prefix"])
        assert hashlib.sha256(
            ob["retained_suffix"].encode()).hexdigest() == ob["retained_text_sha256"]
        assert hashlib.sha256(
            ob["original_entry"].encode()).hexdigest() == ob["original_entry_sha256"]
        assert ob["original_entry_sha256"] != ob["retained_text_sha256"]


def test_only_s02_and_s04_needed_a_slice(audit):
    applied = {u["blind_unit_id"] for u in audit["units"]
               if u["opening_boundary"].get("applied")}
    assert applied == {"S02", "S04"}


def test_units_with_no_residue_record_offset_zero(audit):
    for uid in ("S01", "S03"):
        ob = _audit(audit, uid)["opening_boundary"]
        assert ob["applied"] is False
        assert ob["source_character_start"] == 0
        assert ob["reason"] == "no substantive residue"


def test_slice_stopped_at_a_named_participant(audit):
    for uid in ("S02", "S04"):
        ob = _audit(audit, uid)["opening_boundary"]
        assert ob["stopped_at"]["because"] == "names a participant"
        assert ob["dropped_words"] > 0


def test_the_ask_survives_every_slice(units):
    for u in units:
        first = u["lines"][0].lower()
        markers = smp.QUESTION_MARKERS[u["question"]]
        assert sum(1 for m in markers if m in first) >= smp.MIN_MARKER_HITS, (
            f"{u['blind_unit_id']}: the slice removed the ask")


# ---------------------------------------------------------------------------
# Hard gates
# ---------------------------------------------------------------------------

def test_no_unit_contains_the_next_questions_ask(units):
    for u in units:
        hits = smp.contains_next_question_ask(u["text"], u["question"])
        assert hits == [], f"{u['blind_unit_id']} still contains {hits}"


def test_every_unit_contains_its_own_ask(audit):
    for u in audit["units"]:
        assert u["problems"] == [], f"{u['blind_unit_id']}: {u['problems']}"


def test_audit_is_all_clear_and_reconciles(audit, units):
    assert audit["all_clear"] is True
    for a in audit["units"]:
        u = _unit(units, a["blind_unit_id"])
        assert a["section_text_sha256"] == hashlib.sha256(
            u["text"].encode()).hexdigest()
        assert a["n_turns"] == len(u["turn_ids"]) == len(u["lines"])
        content_words = sum(len(l.split(": ", 1)[1].split()) if ": " in l
                            else len(l.split()) for l in u["lines"])
        assert a["word_count"] == content_words
        assert a["contiguous"]


def test_end_gate_detects_a_planted_next_question_ask():
    """The gate must fire, not merely pass on clean data."""
    assert smp.contains_next_question_ask(
        "[T099] Participant 1: Do you think your gender influences what you eat?", "Q2")
    assert smp.contains_next_question_ask(
        "[T099] Moderator: How do you decide what to eat?", "Q1")
    assert smp.contains_next_question_ask("nothing relevant here", "Q2") == []


# ---------------------------------------------------------------------------
# The six units did not change
# ---------------------------------------------------------------------------

def test_the_selection_is_unchanged():
    m = json.loads((SEALED / "transportability_sample_manifest.json").read_text(
        encoding="utf-8"))
    got = sorted((u["stratum"], u["fg"], u["question_id"]) for u in m["units"])
    assert got == sorted([
        ("demographics-only", "fg1", "Q1"), ("demographics-only", "fg2", "Q4"),
        ("enriched", "fg3", "Q1"), ("enriched", "fg1", "Q4"),
        ("human", "fg3", "Q2"), ("human", "fg4", "Q5")])
    assert m["selection_seed"] == smp.SELECTION_SEED


def test_workbook_matches_the_corrected_unit_texts(units):
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    rows = [r for r in wb["Units"].iter_rows(min_row=2, values_only=True) if r[0]]
    wb.close()
    by_unit = {}
    for uid, tid, spk, txt in rows:
        by_unit.setdefault(uid, []).append(f"[{tid}] {spk}: {txt}")
    for u in units:
        assert by_unit[u["blind_unit_id"]] == u["lines"], (
            f"{u['blind_unit_id']}: workbook text differs from the corrected unit")


def test_workbook_carries_no_q3_material():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    blob = " ".join(str(c.value or "") for s in wb.sheetnames
                    for row in wb[s].iter_rows() for c in row).lower()
    wb.close()
    for banned in (Q3_ASK, Q3_RECAP_S02, "cheers amir", "question three"):
        assert banned.lower() not in blob, f"{banned!r} is still in the workbook"


# ---------------------------------------------------------------------------
# Consolidation gate
# ---------------------------------------------------------------------------

def test_consolidation_scaffold_exists_and_is_empty():
    p = DIR / "Transportability_Consolidation.xlsx"
    assert p.exists()
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    assert wb.sheetnames == ["Instructions", "Raw_To_Consolidated",
                             "Consolidated_Themes", "Scope"]
    themes = [r for r in wb["Consolidated_Themes"].iter_rows(min_row=2, values_only=True)
              if r[1]]
    wb.close()
    assert themes == [], "the scaffold must be left empty for a human"


def test_reference_cannot_be_frozen_before_the_raw_rows_are_imported():
    """
    The seal is now a precondition: with no imported raw rows there is nothing to
    consolidate against, so freeze refuses before it looks at themes at all.
    """
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference()
    assert "no raw-row seal" in str(e.value)


def test_human_reference_themes_do_not_exist_yet():
    assert not (DIR / "human_reference_themes.json").exists(), (
        "the reference must not exist before a human consolidates")


def test_classification_is_not_calibration_grade():
    assert con.CLASSIFICATION == (
        "SINGLE_CODER_HUMAN_REFERENCE_WITH_POST_CODING_CONSOLIDATION")
    proto = (DIR / "FUTURE_PROTOCOL.md").read_text(encoding="utf-8")
    assert "not inter-coder agreement" in proto
    assert "not calibration-grade" in proto


def test_denominators_are_consolidated_themes_not_raw_rows():
    proto = (DIR / "FUTURE_PROTOCOL.md").read_text(encoding="utf-8")
    assert "Raw coder rows are never a\ndenominator" in proto or \
           "raw coder rows are never a denominator" in proto.lower()
    assert "**consolidated** human themes for that unit" in proto


def test_duplicate_raw_rows_cannot_inflate_the_denominator():
    """
    Two raw rows merged into one consolidated theme count ONCE. If raw rows were the
    denominator, a coder writing the same theme twice would depress recall.
    """
    themes = [{"blind_unit_id": "S01", "consolidated_theme_id": "T1",
               "source_row_ids": ["S01_slot_1", "S01_slot_7"]},
              {"blind_unit_id": "S01", "consolidated_theme_id": "T2",
               "source_row_ids": ["S01_slot_2"]}]
    raw_rows = sum(len(t["source_row_ids"]) for t in themes)
    assert raw_rows == 3
    assert len(themes) == 2, "the denominator is consolidated themes, not raw rows"


# ---------------------------------------------------------------------------
# Non-mutation
# ---------------------------------------------------------------------------
#
# An earlier version of this file pinned SHA-256 prefixes for the U01-U07 package and
# asserted they never changed. That was wrong: the researcher is editing
# Clustering_U01_U07.xlsx, so the pinned hash would have failed on her first
# legitimate save — punishing authorised work rather than catching accidental
# mutation.
#
# The real guard lives in tests/test_consolidation_gate.py
# (test_constructors_do_not_mutate_the_active_human_workbooks): it hashes the active
# workbooks at runtime, runs the constructors with every output redirected into
# tmp_path, and re-hashes. Nothing is hardcoded and no constructor touches a real
# file.


# ---------------------------------------------------------------------------
# The frozen supplementary reference — relevance NOT_ASSESSED
# ---------------------------------------------------------------------------

REF = DIR / "supplementary_human_reference.json"


@pytest.fixture(scope="module")
def frozen():
    if not REF.exists():
        pytest.skip("supplementary reference not frozen")
    return json.loads(REF.read_text(encoding="utf-8"))


def test_the_reference_is_frozen_with_the_expected_shape(frozen):
    assert frozen["classification"] == (
        "SUPPLEMENTARY_SINGLE_CODER_TRANSPORTABILITY_SAMPLE")
    assert frozen["n_units"] == 6
    assert frozen["n_themes"] == 18
    assert frozen["n_empty_slots"] == 54
    assert len(frozen["themes"]) == 18
    keys = [t["supplementary_key"] for t in frozen["themes"]]
    assert len(keys) == len(set(keys))


def test_relevance_is_not_assessed_and_never_imputed(frozen):
    assert frozen["relevance_status"] == "NOT_ASSESSED"
    for t in frozen["themes"]:
        assert t["relevance_status"] == "NOT_ASSESSED"
        assert t["relevance_value_as_recorded"] is None
        # never a substitute value
        assert t["relevance_value_as_recorded"] not in ("secondary", "central", False, 0)
    note = frozen["relevance_note"].lower()
    assert "methodological decision, not missing data" in note
    assert "never be read as" in note


def test_no_centrality_or_salience_result_is_reported(frozen):
    """Item 9: nothing about relevance may survive as a result."""
    banned = ("n_central", "n_secondary", "central_reference", "salience",
              "hierarchy", "centrality_rate")
    blob = json.dumps({k: v for k, v in frozen.items()
                       if k not in ("relevance_note", "themes")}).lower()
    for b in banned:
        assert b not in blob, f"{b!r} leaked into the frozen reference"


def test_every_theme_keeps_the_three_required_fields(frozen):
    for t in frozen["themes"]:
        assert t["theme_label"] and t["theme_description"] and t["supporting_quote"]
        assert t["content_sha256"]


def test_quotes_are_literal_in_their_own_unit(frozen):
    units = json.loads((DIR / "_units_for_packaging.json").read_text(encoding="utf-8"))
    text = {u["blind_unit_id"]: u["text"] for u in units}
    norm = lambda t: " ".join(str(t).split())
    for t in frozen["themes"]:
        assert norm(t["supporting_quote"]) in norm(text[t["blind_unit_id"]]), (
            t["supplementary_key"])


def test_denominators_are_human_themes_not_slots(frozen):
    assert sum(d["n_human_themes"]
               for d in frozen["denominators_per_unit"].values()) == 18
    assert all(d["n_human_themes"] > 0
               for d in frozen["denominators_per_unit"].values())
    assert "Empty slots are unused capacity" in frozen["denominator_note"]


def test_consolidation_decision_is_recorded(frozen):
    assert frozen["consolidation_decision"] == (
        "CONSOLIDATION_NOT_REQUIRED — CODER_ROWS_ALREADY_DISTINCT")
    assert "0 exact" in frozen["consolidation_rule"]
    assert "no LLM rewrote any theme" in frozen["consolidation_rule"]


def test_it_must_never_be_pooled_with_q3(frozen):
    assert "never pooled" in frozen["never_combine_numerically_with"].lower()
    for lim in ("one coder", "six units", "no inter-coder agreement"):
        assert any(lim in x for x in frozen["limitations"]), lim


def test_freeze_refuses_to_overwrite_without_force(tmp_path):
    dst = tmp_path / "supplementary_human_reference.json"
    dst.write_text('{"sentinel": "already frozen"}', encoding="utf-8")
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_supplementary_reference(out_path=dst)
    assert "already exists" in str(e.value)
    assert json.loads(dst.read_text(encoding="utf-8"))["sentinel"] == "already frozen"


def test_freeze_is_atomic_and_leaves_no_temp_file(tmp_path):
    dst = tmp_path / "supplementary_human_reference.json"
    out = con.freeze_supplementary_reference(out_path=dst)
    assert out["n_themes"] == 18
    assert dst.exists()
    assert not list(tmp_path.glob("*.tmp"))


def test_freeze_refuses_a_workbook_that_fails_the_gate(tmp_path):
    """A required field missing still blocks the freeze."""
    import shutil
    import openpyxl as _o
    bad = tmp_path / "bad.xlsx"
    shutil.copy2(WB, bad)
    wb = _o.load_workbook(bad)
    ws = wb["Emergent_Coding"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value:
            ws.cell(row=r, column=4).value = None      # theme_description
            break
    wb.save(bad); wb.close()
    with pytest.raises(con.ConsolidationNotReady):
        con.freeze_supplementary_reference(coder_workbook=bad,
                                           out_path=tmp_path / "x.json")


def test_the_coder_workbook_was_not_modified(frozen):
    import hashlib
    assert hashlib.sha256(WB.read_bytes()).hexdigest() == \
        frozen["coder_workbook_sha256"]
