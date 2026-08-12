"""
Blind pooling and clustering workbook for the partial emergent review — U01–U07 only.

WHAT THIS IS
Two coders independently coded the same first seven units. This pools their emergent
themes per unit, strips authorship, randomises order, and issues an Excel workbook in
which a HUMAN decides which labels denote the same theme.

WHAT THIS IS NOT
It is not a gold standard, and it computes NO agreement. Emergent labels are free
text; whether "Eating meat as the normal" and "an existing but unintelligible norm"
are the same theme is a human judgement. Any agreement figure produced before that
judgement would be a string-matching artefact wearing a reliability number's clothes.

BLINDING
Authorship is removed from the workbook and the pooled order is shuffled per unit
with a per-unit deterministic seed, so the ordering is reproducible for audit but
carries no authorship signal. The coder mapping is written to a SEALED file that the
clustering workbook does not contain.

U08 IS OUT OF SCOPE
Coder A also coded U08; Coder B did not. It is excluded from clustering and from any
future agreement computation. U08–U15 are NOT_REVIEWED — not absences, not zeros.

Writes only to analysis/production_evaluation/. Never modifies the returned workbooks.
"""

from __future__ import annotations

import hashlib
import json
import random
import re
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_PKG = _OUT / "gold_standard_package"
_DIR = _OUT / "partial_emergent_clustering"
_SEALED = _OUT / "gold_standard_sealed"

WORKBOOKS = {"Coder_A": _PKG / "Coder_A_Part1_Emergent.xlsx",
             "Coder_B": _PKG / "Coder_B_Part1_Emergent.xlsx"}
SHARED_UNITS = [f"U{i:02d}" for i in range(1, 8)]          # U01..U07
ISSUED_UNITS = [f"U{i:02d}" for i in range(1, 16)]         # U01..U15

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
NEED = PatternFill("solid", fgColor="FFF2CC")


def _norm(s) -> str:
    return " ".join(re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split())


def read_units_text(ws) -> dict[str, str]:
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
    i_unit = header.index("unit_id")
    out: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[i_unit]:
            continue
        for cell in row:
            if cell and str(cell).strip() != str(row[i_unit]).strip():
                out.setdefault(str(row[i_unit]).strip(), []).append(str(cell))
    return {k: "\n".join(v) for k, v in out.items()}


def read_themes(path: Path) -> tuple[list[dict], dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    units_text = read_units_text(wb["Units"])
    ws = wb["Emergent_Coding"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}
    themes = []
    for r in rows[1:]:
        uid = r[idx["unit_id"]]
        if not uid:
            continue
        get = lambda k: (str(r[idx[k]]).strip()
                         if k in idx and idx[k] < len(r) and r[idx[k]] is not None else "")
        if not get("theme_label"):
            continue
        themes.append({
            "unit_id": str(uid).strip(),
            "theme_slot": get("theme_slot"),
            "theme_label": get("theme_label"),
            "theme_description": get("theme_description"),
            "supporting_quote": get("supporting_quote"),
            "relevance": get("relevance"),
            "coder_notes": get("coder_notes"),
        })
    wb.close()
    return themes, units_text


def build() -> dict:
    _DIR.mkdir(parents=True, exist_ok=True)
    _SEALED.mkdir(parents=True, exist_ok=True)

    pooled, units_text, per_coder = [], {}, {}
    for coder, path in WORKBOOKS.items():
        themes, utext = read_themes(path)
        units_text.update(utext)
        per_coder[coder] = themes
        for t in themes:
            t["_coder"] = coder

    # quote validation against the ISSUED unit text
    for coder, themes in per_coder.items():
        for t in themes:
            src = units_text.get(t["unit_id"], "")
            t["_quote_literal"] = bool(t["supporting_quote"]) and \
                _norm(t["supporting_quote"]) in _norm(src)

    # pool the shared units only, blind and shuffled per unit
    sealed, rows = [], []
    n = 0
    for uid in SHARED_UNITS:
        items = [t for c in per_coder for t in per_coder[c] if t["unit_id"] == uid]
        seed = int(hashlib.sha256(f"partial_emergent::{uid}".encode()).hexdigest()[:8], 16)
        random.Random(seed).shuffle(items)
        for t in items:
            n += 1
            pid = f"P{n:03d}"
            rows.append({
                "pooled_id": pid, "unit_id": uid,
                "theme_label": t["theme_label"],
                "theme_description": t["theme_description"],
                "supporting_quote": t["supporting_quote"],
                "relevance": t["relevance"] or "(not provided)",
                "quote_literal_in_unit": "yes" if t["_quote_literal"] else "NO",
            })
            sealed.append({"pooled_id": pid, "unit_id": uid, "coder": t["_coder"],
                           "theme_slot": t["theme_slot"],
                           "relevance_provided": bool(t["relevance"])})
    pooled = rows

    _write_workbook(pooled)
    sealed_path = _SEALED / "partial_emergent_pooled_authorship.json"
    sealed_path.write_text(json.dumps({
        "created_utc": datetime.now(UTC).isoformat(),
        "warning": ("SEALED — authorship map for the blind clustering workbook. Do not "
                    "supply to the person doing the clustering."),
        "scope": SHARED_UNITS,
        "map": sealed,
    }, indent=1, ensure_ascii=False), encoding="utf-8")

    return {"pooled": pooled, "sealed": sealed, "per_coder": per_coder,
            "sealed_path": sealed_path}


def _write_workbook(pooled: list[dict]) -> Path:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("Partial emergent review — clustering and adjudication (U01–U07)", True),
        ("", False),
        ("WHAT THIS IS", True),
        ("Two coders independently coded the same seven units. Their emergent themes are", False),
        ("pooled below with authorship removed and order shuffled. Your task is to decide", False),
        ("which themes denote the SAME underlying theme.", False),
        ("", False),
        ("THIS IS NOT A GOLD STANDARD", True),
        ("Seven of fifteen units were coded. U08–U15 are NOT_REVIEWED — they are not", False),
        ("thematic absences and must never be treated as zeros.", False),
        ("", False),
        ("WHY NO AGREEMENT NUMBER YET", True),
        ("Agreement cannot be computed on free-text labels. Whether two differently worded", False),
        ("labels are the same theme is your judgement. Any figure computed before you", False),
        ("finish would be a string-matching artefact, not a reliability estimate.", False),
        ("", False),
        ("HOW TO COMPLETE", True),
        ("1. Work unit by unit. Read all pooled themes for a unit before deciding.", False),
        ("2. Give every row a cluster_id. Rows you judge to be the same theme share one id.", False),
        ("   Use C01, C02, ... consistently ACROSS units — a theme recurring in U02 and U05", False),
        ("   must carry the SAME cluster_id, or the saturation curve will be wrong.", False),
        ("3. A theme only one coder raised still gets its own cluster_id. Do not drop it.", False),
        ("4. On the Cluster_Definitions sheet, name each cluster once and mark whether it is", False),
        ("   central (a main idea of the unit) or peripheral.", False),
        ("5. Do NOT consult the codebook while clustering. The codebook comparison happens", False),
        ("   afterwards, on the Codebook_Comparison sheet, and only then.", False),
        ("", False),
        ("quote_literal_in_unit flags any supporting quote that is not verbatim in the unit.", False),
        ("relevance shows '(not provided)' where a coder left it blank.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    # --- pooled themes for clustering ---
    ws = wb.create_sheet("Clustering")
    cols = ["pooled_id", "unit_id", "theme_label", "theme_description",
            "supporting_quote", "relevance", "quote_literal_in_unit",
            "cluster_id", "cluster_label", "is_central", "adjudicator_notes"]
    widths = [10, 9, 34, 46, 52, 15, 13, 12, 28, 12, 30]
    for j, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(pooled, start=2):
        for j, h in enumerate(cols, start=1):
            c = ws.cell(row=i, column=j, value=r.get(h, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if h in ("cluster_id", "cluster_label", "is_central"):
                c.fill = NEED
        ws.row_dimensions[i].height = None
    dv = DataValidation(type="list", formula1='"central,peripheral"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"J2:J{len(pooled) + 1}")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(pooled) + 1}"

    # --- downstream sheets, deliberately empty ---
    ws = wb.create_sheet("Cluster_Definitions")
    for j, (h, w) in enumerate(zip(
            ["cluster_id", "cluster_label", "definition", "is_central",
             "first_seen_unit", "n_units_present", "raised_by_both_coders", "notes"],
            [12, 30, 52, 12, 15, 16, 22, 30]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(row=2, column=1, value="(one row per cluster — complete after Clustering)")

    ws = wb.create_sheet("Presence_Matrix")
    ws.cell(row=1, column=1, value=(
        "Built automatically from Clustering + the sealed authorship map once "
        "adjudication is complete. Rows = clusters, columns = unit x coder."))
    ws.cell(row=2, column=1, value=(
        "Left empty on purpose: filling it now would require guessing your clustering."))
    ws.column_dimensions["A"].width = 100
    for r in (1, 2):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Saturation")
    for j, (h, w) in enumerate(zip(
            ["unit_id", "n_pooled_themes", "n_clusters_total_so_far",
             "n_new_clusters_this_unit", "n_new_CENTRAL_clusters_this_unit",
             "notes"], [10, 18, 24, 26, 32, 40]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = w
    counts = {}
    for r in pooled:
        counts[r["unit_id"]] = counts.get(r["unit_id"], 0) + 1
    for i, u in enumerate(SHARED_UNITS, start=2):
        ws.cell(row=i, column=1, value=u)
        ws.cell(row=i, column=2, value=counts.get(u, 0))
        for j in (3, 4, 5):
            ws.cell(row=i, column=j).fill = NEED
    ws.cell(row=len(SHARED_UNITS) + 3, column=1, value=(
        "Saturation is assessed only after the cumulative curve exists. New clusters "
        "in U06–U07 — especially new CENTRAL clusters — argue against saturation. "
        "No claim of complete saturation may be made from seven units."))
    ws.cell(row=len(SHARED_UNITS) + 3, column=1).alignment = Alignment(wrap_text=True)

    ws = wb.create_sheet("Codebook_Comparison")
    for j, (h, w) in enumerate(zip(
            ["cluster_id", "cluster_label", "closest_codebook_subtheme",
             "relationship", "notes"], [12, 30, 28, 24, 46]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = w
    dv2 = DataValidation(
        type="list",
        formula1='"matches,partially overlaps,not in codebook,codebook theme not found"',
        allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("D2:D200")
    ws.cell(row=2, column=1, value=(
        "COMPLETE THIS LAST — only after emergent clustering is finished."))

    ws = wb.create_sheet("Scope")
    rows = [["unit_id", "status"]] + \
           [[u, "SHARED — coded by both coders, in scope"] for u in SHARED_UNITS] + \
           [["U08", "OUT OF SCOPE — one coder only; excluded from agreement"]] + \
           [[u, "NOT_REVIEWED"] for u in ISSUED_UNITS[8:]]
    for i, r in enumerate(rows, start=1):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if i == 1:
                c.fill, c.font = HDR, HDR_FONT
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 56

    path = _DIR / "Clustering_U01_U07.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    res = build()
    print(f"pooled themes: {len(res['pooled'])}")
    print(f"workbook : {(_DIR / 'Clustering_U01_U07.xlsx').relative_to(_REPO_ROOT)}")
    print(f"sealed   : {res['sealed_path'].relative_to(_REPO_ROOT)}")
