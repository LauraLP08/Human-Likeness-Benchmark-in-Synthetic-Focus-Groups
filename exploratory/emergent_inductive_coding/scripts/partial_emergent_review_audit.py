"""
Read-only audit of the RETURNED Part-1 emergent workbooks.

This is NOT the gold standard. The package was issued as 15 units per coder; the
coders completed a subset. Classifying a partial exercise as a completed gold
standard would let interpretive metrics be reported against a foundation that does
not exist, so the material is labelled PARTIAL_EMERGENT_HUMAN_REVIEW throughout and
the original 15-unit gate is NOT forced.

STRICTLY READ-ONLY. The workbooks are opened `read_only=True` and never written.
Nothing is invented for units that were not coded.

Part 2 (deductive) is NOT released and NOT requested.

No agreement statistic is computed here. Emergent free-text labels cannot be
compared until a human clustering pass has decided which labels denote the same
theme; computing agreement on raw strings first would fabricate a number.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_PKG = _OUT / "gold_standard_package"
_ARTIFACT = _OUT / "partial_emergent_human_review_audit.json"

WORKBOOKS = {"Coder_A": _PKG / "Coder_A_Part1_Emergent.xlsx",
             "Coder_B": _PKG / "Coder_B_Part1_Emergent.xlsx"}

REQUIRED_FOR_COMPLETE = ("theme_label", "theme_description", "supporting_quote",
                         "relevance")
ANY_CONTENT = REQUIRED_FOR_COMPLETE + ("coder_notes",)


def _norm(s: str) -> str:
    return " ".join(re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split())


def read_units(ws) -> dict[str, str]:
    """unit_id -> concatenated verbatim text, for literal quote checking."""
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
    try:
        i_unit = header.index("unit_id")
    except ValueError:
        return {}
    text_cols = [i for i, h in enumerate(header) if h in ("text", "turn_text", "content")]
    out: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = row[i_unit]
        if not uid:
            continue
        for i in (text_cols or [len(row) - 1]):
            if i < len(row) and row[i]:
                out.setdefault(str(uid).strip(), []).append(str(row[i]))
    return {k: "\n".join(v) for k, v in out.items()}


def audit_workbook(coder: str, path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    units_text = read_units(wb["Units"])
    ws = wb["Emergent_Coding"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    issued_units, complete_rows, partial_rows = [], [], []
    per_unit: dict[str, list[dict]] = {}
    for r in rows[1:]:
        uid = r[idx["unit_id"]]
        if not uid:
            continue
        uid = str(uid).strip()
        if uid not in issued_units:
            issued_units.append(uid)
        vals = {h: (str(r[i]).strip() if i < len(r) and r[i] is not None else "")
                for h, i in idx.items()}
        filled_required = [f for f in REQUIRED_FOR_COMPLETE if vals.get(f)]
        any_filled = [f for f in ANY_CONTENT if vals.get(f)]
        if len(filled_required) == len(REQUIRED_FOR_COMPLETE):
            complete_rows.append((uid, vals))
            per_unit.setdefault(uid, []).append(vals)
        elif any_filled:
            partial_rows.append({"unit_id": uid,
                                 "theme_slot": vals.get("theme_slot"),
                                 "filled_fields": any_filled,
                                 "missing_required": [f for f in REQUIRED_FOR_COMPLETE
                                                      if not vals.get(f)]})

    # overflow sheet
    ov = wb["Overflow_Themes"]
    ov_rows = [r for r in ov.iter_rows(min_row=2, values_only=True)
               if any(c not in (None, "") for c in r)]

    # literal quote checking against the issued Units text
    quotes_valid, quotes_invalid = 0, []
    for uid, vals in complete_rows:
        q = vals.get("supporting_quote", "")
        src = units_text.get(uid, "")
        if q and _norm(q) in _norm(src):
            quotes_valid += 1
        elif q:
            quotes_invalid.append({"unit_id": uid, "theme_slot": vals.get("theme_slot"),
                                   "quote_excerpt": q[:120]})
    wb.close()

    coded_units = sorted(per_unit)
    return {
        "coder": coder,
        "workbook": str(path.relative_to(_REPO_ROOT)),
        "workbook_modified_utc": datetime.fromtimestamp(
            path.stat().st_mtime, UTC).isoformat(),
        "units_issued": len(issued_units),
        "units_with_at_least_one_complete_theme": len(coded_units),
        "units_coded": coded_units,
        "units_empty": sorted(set(issued_units) - set(coded_units)),
        "n_units_empty": len(issued_units) - len(coded_units),
        "complete_themes_total": len(complete_rows),
        "complete_themes_per_unit": {u: len(v) for u, v in sorted(per_unit.items())},
        "partially_completed_rows": partial_rows,
        "n_partially_completed_rows": len(partial_rows),
        "quotes_literal_valid": quotes_valid,
        "quotes_literal_invalid": quotes_invalid,
        "n_quotes_literal_invalid": len(quotes_invalid),
        "overflow_rows_used": len(ov_rows),
    }


def main() -> int:
    audits = {c: audit_workbook(c, p) for c, p in WORKBOOKS.items()}
    a, b = audits["Coder_A"], audits["Coder_B"]
    sa, sb = set(a["units_coded"]), set(b["units_coded"])
    shared = sorted(sa & sb)

    out = {
        "audited_utc": datetime.now(UTC).isoformat(),
        "classification": "PARTIAL_EMERGENT_HUMAN_REVIEW",
        "explicitly_not": ("NOT a completed gold standard. The 15-unit gate is NOT "
                           "applied and NOT forced; the exercise is reported at the "
                           "size it actually reached."),
        "read_only": True,
        "workbooks_modified_by_this_audit": False,
        "part2_deductive_status": ("NOT_CONDUCTED — prior human anchor available; "
                                   "current coder exercise limited to partial emergent "
                                   "review"),
        "part2_released": False,
        "part2_requested": False,
        "interpretive_metrics": ("WITHHELD — the metrics that required a completed gold "
                                 "standard remain withheld"),
        "agreement_computed": False,
        "agreement_note": ("No agreement statistic is computed. Emergent labels are "
                           "free text; which labels denote the same theme is a human "
                           "clustering decision that has not been made. Computing "
                           "agreement on raw strings first would fabricate a number."),
        "per_coder": audits,
        "overlap": {
            "units_coded_by_both": shared,
            "n_units_coded_by_both": len(shared),
            "units_only_coder_a": sorted(sa - sb),
            "units_only_coder_b": sorted(sb - sa),
            "n_only_coder_a": len(sa - sb),
            "n_only_coder_b": len(sb - sa),
        },
        "units_not_attempted_by_either": sorted(
            set(range(0)) or (set(a["units_empty"]) & set(b["units_empty"]))),
        "no_data_invented": ("Units neither coder completed are reported as empty. "
                             "Nothing is imputed for them."),
    }

    if shared and not (sa - sb) and not (sb - sa):
        out["blind_pooling"] = {
            "eligible": True,
            "scope": shared,
            "n_units": len(shared),
            "instruction": ("Pool the emergent labels for these units blind to coder "
                            "identity, for human clustering. Agreement is computed only "
                            "AFTER clustering, and only on these units."),
        }
    elif shared:
        out["blind_pooling"] = {
            "eligible": True,
            "scope": shared,
            "n_units": len(shared),
            "instruction": ("Pool blind for these shared units only. Units coded by one "
                            "coder alone carry no agreement information and are excluded "
                            "from any agreement computation."),
            "excluded_from_agreement": {"only_coder_a": sorted(sa - sb),
                                        "only_coder_b": sorted(sb - sa)},
        }
    else:
        out["blind_pooling"] = {
            "eligible": False,
            "reason": "no unit was coded by both coders; there is no shared basis",
        }

    _ARTIFACT.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")

    print("=" * 78)
    print("  PARTIAL EMERGENT HUMAN REVIEW — read-only audit")
    print("=" * 78)
    for c in ("Coder_A", "Coder_B"):
        d = audits[c]
        print(f"\n{c}  ({d['workbook_modified_utc'][:16]}Z)")
        print(f"   units issued                     : {d['units_issued']}")
        print(f"   units with >=1 complete theme    : {d['units_with_at_least_one_complete_theme']}")
        print(f"   units empty                      : {d['n_units_empty']}")
        print(f"   complete themes (total)          : {d['complete_themes_total']}")
        print(f"   themes per coded unit            : {d['complete_themes_per_unit']}")
        print(f"   partially completed rows         : {d['n_partially_completed_rows']}")
        print(f"   quotes literal valid / invalid   : {d['quotes_literal_valid']} / {d['n_quotes_literal_invalid']}")
        print(f"   overflow rows used               : {d['overflow_rows_used']}")
        print(f"   coded units                      : {', '.join(d['units_coded'])}")
    o = out["overlap"]
    print(f"\nshared units (both coders): {o['n_units_coded_by_both']} -> {', '.join(o['units_coded_by_both'])}")
    print(f"only Coder_A: {o['units_only_coder_a'] or 'none'}")
    print(f"only Coder_B: {o['units_only_coder_b'] or 'none'}")
    print(f"\nblind pooling eligible: {out['blind_pooling'].get('eligible')}")
    print(f"artifact: {_ARTIFACT.relative_to(_REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
