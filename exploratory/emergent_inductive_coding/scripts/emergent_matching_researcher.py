"""
Researcher-facing matching workbook for the Q3 emergent calibration (V2), plus its gate.

WHY V1 WAS REPLACED
-------------------
Emergent_Matching_Q3_POPULATED.xlsx listed 74 rows with `relation` and `decision` but no
counterpart-key column, so a match could not be expressed at all. The one-sided 44-row
design fixed that. Direct validation of V1 then found three further defects, all fixed
here:

  1. EVIDENCE WAS TRUNCATED. Quotes were cut at 180/200 characters, mid-sentence. The
     researcher was being asked to judge a match against evidence she could not fully
     read. All slicing is gone; every quote is complete. Excel allows 32,767 characters
     per cell, every cell is checked against that before writing, and if any ever
     exceeded it the evidence would move to a one-row-per-quote sheet with an explicit
     reference — never a silent truncation.
  2. IMMUTABLE COLUMNS WERE UNSEALED. Sheet protection is a convenience that any writer
     can switch off. The gate now rebuilds the canonical representation from
     human_reference_q3.json and extraction_results_q3.json and compares every immutable
     cell, including the relation_derived formulas, rejecting any edit, deletion,
     addition or reordering.
  3. UNCERTAIN WAS TREATED AS A MATCH. Keys named on an UNCERTAIN row were flowing into
     the same link set as MATCHED. They are now kept strictly apart: they never count as
     matches, never raise recall, never clear a theme from the adjudication queue, and
     never produce confirmed fusion or fragmentation. While any remain, the overall state
     is UNRESOLVED.

    py scripts/emergent_matching_researcher.py --build
    py scripts/emergent_matching_researcher.py --validate
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "scripts"))

import emergent_calibration_q3 as cal   # noqa: E402

_DIR = cal._DIR
_WB = _DIR / "Emergent_Matching_Q3_RESEARCHER_V2.xlsx"

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")
EDIT_FILL = PatternFill("solid", fgColor="FFF7E6")

EXCEL_CELL_LIMIT = 32767

HUMAN_DECISIONS = ("MATCHED", "NO_MATCH_HUMAN_ONLY", "UNCERTAIN")

IMMUTABLE_COLS = ("unit_id", "human_key", "human_cluster_label", "human_definition",
                  "human_supporting_quotes", "available_machine_keys",
                  "available_machine_labels", "available_machine_quotes",
                  "relation_derived")
EDITABLE = ("human_decision", "matched_machine_keys", "researcher_reasoning")
HUMAN_COLS = ["unit_id", "human_key", "human_cluster_label", "human_definition",
              "human_supporting_quotes", "available_machine_keys",
              "available_machine_labels", "available_machine_quotes",
              "human_decision", "matched_machine_keys", "relation_derived",
              "researcher_reasoning"]

RELEVANCE_CAVEAT = "DESCRIPTIVE_MODEL_METADATA_NOT_HUMAN_VALIDATED"
UNRESOLVED = "UNRESOLVED_PENDING_UNCERTAIN_RESOLUTION"


class MatchingNotReady(RuntimeError):
    pass


def _load():
    ref = json.loads((_DIR / "human_reference_q3.json").read_text(encoding="utf-8"))
    res = json.loads((_DIR / "extraction_results_q3.json").read_text(encoding="utf-8"))
    human = {}
    for r in ref["union_reference"]:
        human.setdefault(r["unit_id"], []).append(r)
    machine = {}
    for r in res["results"]:
        if r.get("status") == "COMPLETE":
            machine[r["unit_id"]] = r.get("themes", [])
    return human, machine


# ---------------------------------------------------------------------------
# The canonical representation — one source of truth for build AND gate
# ---------------------------------------------------------------------------

def canonical_rows() -> list[dict]:
    """
    The immutable cells, rebuilt from the frozen sources. Nothing is truncated.

    build() writes exactly this; validate() rebuilds it and compares. A drift between
    the workbook and the sources is therefore always detectable, whatever Excel's sheet
    protection happens to say.
    """
    human, machine = _load()
    col = {h: i + 1 for i, h in enumerate(HUMAN_COLS)}
    out = []
    excel_row = 2
    for unit in cal.UNITS:
        mach = machine.get(unit, [])
        keys = "; ".join(cal.machine_key(unit, m["machine_theme_id"]) for m in mach)
        labels = "\n\n".join(
            f"{cal.machine_key(unit, m['machine_theme_id'])} — {m.get('label', '')}\n"
            f"    {m.get('one_sentence_description', '')}" for m in mach)
        quotes = "\n\n".join(
            "\n".join(
                f"{cal.machine_key(unit, m['machine_theme_id'])} [{e['turn_id']}] "
                f"{e.get('speaker', '')}: \"{e['quote']}\""
                for e in (m.get("evidence") or []))
            for m in mach)
        for h in human.get(unit, []):
            hq = "\n\n".join(f"• {q['quote']}" for q in h.get("supporting_quotes", []))
            mk = f"{get_column_letter(col['matched_machine_keys'])}{excel_row}"
            dc = f"{get_column_letter(col['human_decision'])}{excel_row}"
            formula = (f'=IF({dc}="","",IF({dc}="UNCERTAIN","candidate only - not a match",'
                       f'IF({dc}<>"MATCHED","n/a",'
                       f'IF({mk}="","MISSING KEY",'
                       f'IF(LEN({mk})-LEN(SUBSTITUTE({mk},";",""))=0,'
                       f'"one_to_one (provisional)",'
                       f'"one_to_many / possible fragmentation")))))')
            out.append({
                "excel_row": excel_row,
                "unit_id": unit,
                "human_key": h["human_key"],
                "human_cluster_label": h["cluster_label"],
                "human_definition": h.get("consolidated_definition", ""),
                "human_supporting_quotes": hq,
                "available_machine_keys": keys,
                "available_machine_labels": labels,
                "available_machine_quotes": quotes,
                "relation_derived": formula,
            })
            excel_row += 1
    return out


def oversized_cells(rows: list[dict] = None) -> list[str]:
    """Any cell that would exceed Excel's limit. Nothing is ever silently cut."""
    rows = rows if rows is not None else canonical_rows()
    bad = []
    for r in rows:
        for c in IMMUTABLE_COLS:
            n = len(str(r[c]))
            if n > EXCEL_CELL_LIMIT:
                bad.append(f"{r['human_key']}.{c}: {n} chars exceeds "
                           f"{EXCEL_CELL_LIMIT}")
    return bad


def _atomic(path: Path, payload: dict) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        tmp.write_text(json.dumps(payload, indent=1, ensure_ascii=False),
                       encoding="utf-8")
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            tmp.unlink()


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build(out_path: Path = None) -> Path:
    out_path = out_path or _WB
    rows = canonical_rows()
    over = oversized_cells(rows)
    if over:
        raise MatchingNotReady(
            "evidence would exceed Excel's cell limit; move it to a one-row-per-quote "
            "sheet rather than truncating:\n  - " + "\n  - ".join(over))

    _, machine = _load()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("Matching the automated themes to your own — U01 to U07", True),
        ("", False),
        ("WHAT YOU ARE DOING", True),
        ("You work through ONE sheet: Human_Matching. It has 44 rows, one for each", False),
        ("theme you and the other coder identified, grouped by unit.", False),
        ("", False),
        ("You never work from the machine side. Which automated themes went", False),
        ("unmatched, and which need a separate judgement, is worked out from what", False),
        ("you enter here. You will not be asked to repeat a relationship twice.", False),
        ("", False),
        ("FOR EACH ROW", True),
        ("1. Read human_cluster_label, its definition, and its full quotes.", False),
        ("2. Read the automated themes for THAT SAME UNIT. Their keys, labels,", False),
        ("   descriptions and every one of their quotes are on the same row, in", False),
        ("   the three 'available_...' columns. Nothing is abbreviated and you do", False),
        ("   not need to leave the sheet.", False),
        ("3. Choose ONE value in human_decision:", False),
        ("      MATCHED              - an automated theme expresses this same idea", False),
        ("      NO_MATCH_HUMAN_ONLY  - no automated theme expresses it", False),
        ("      UNCERTAIN            - the evidence allows more than one reading", False),
        ("4. If MATCHED, put the machine key(s) in matched_machine_keys:", False),
        ("      one key      ->  U01::M2", False),
        ("      several keys ->  U01::M2; U01::M4        (semicolon separated)", False),
        ("   Use only keys from the SAME unit. They are listed for you.", False),
        ("5. researcher_reasoning: required for UNCERTAIN, and whenever a match is", False),
        ("   not obvious.", False),
        ("", False),
        ("ABOUT UNCERTAIN", True),
        ("If you choose UNCERTAIN you MAY still name the key(s) you were weighing.", False),
        ("They are kept as CANDIDATES only. They do not count as a match, they do", False),
        ("not change any score, and they do not remove the automated theme from", False),
        ("the list still needing a judgement. The calibration stays UNRESOLVED", False),
        ("until every UNCERTAIN row is settled, so use it when you mean it and", False),
        ("say why.", False),
        ("", False),
        ("RULES THE SHEET ENFORCES", True),
        ("  MATCHED needs at least one machine key.", False),
        ("  NO_MATCH_HUMAN_ONLY must leave matched_machine_keys empty.", False),
        ("  UNCERTAIN needs a reason; keys are optional.", False),
        ("  A key from another unit is rejected.", False),
        ("", False),
        ("WHAT YOU DO NOT DO", True),
        ("  Do NOT consult the codebook. This is an emergent comparison.", False),
        ("  Do NOT edit any grey column. They are checked character by character", False),
        ("  against the frozen sources and any change is rejected.", False),
        ("  Do NOT fill relation_derived — it fills itself as you type.", False),
        ("  Do NOT add your name or a date. Those are recorded when I import.", False),
        ("", False),
        ("A machine theme that matches nothing of yours is NOT automatically an", False),
        ("error. Those go to a separate short adjudication afterwards, built from", False),
        ("your answers here.", False),
        ("", False),
        ("Grey columns are locked. Only the three cream columns are yours.", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 96

    ws = wb.create_sheet("Human_Matching")
    widths = [8, 13, 38, 38, 62, 20, 58, 70, 22, 24, 28, 44]
    for j, (h, w) in enumerate(zip(HUMAN_COLS, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 30

    col = {h: i + 1 for i, h in enumerate(HUMAN_COLS)}
    for rec in rows:
        r = rec["excel_row"]
        for c in IMMUTABLE_COLS:
            cell = ws.cell(row=r, column=col[c], value=rec[c])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = LOCKED_FILL
            cell.protection = Protection(locked=True)
        for c in EDITABLE:
            cell = ws.cell(row=r, column=col[c])
            cell.fill = EDIT_FILL
            cell.protection = Protection(locked=False)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 150

    dv = DataValidation(type="list", formula1='"' + ",".join(HUMAN_DECISIONS) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    last = rows[-1]["excel_row"]
    dv.add(f"{get_column_letter(col['human_decision'])}2:"
           f"{get_column_letter(col['human_decision'])}{last}")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HUMAN_COLS))}{last}"
    ws.protection.sheet = True
    ws.protection.enable()

    ws = wb.create_sheet("Machine_Themes")
    mcols = ["unit_id", "machine_key", "label", "description", "evidence_quotes",
             "model_relevance", "relevance_caveat"]
    for j, (h, w) in enumerate(zip(mcols, [8, 13, 40, 52, 88, 16, 46]), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    i = 2
    for unit in cal.UNITS:
        for m in machine.get(unit, []):
            ev = "\n\n".join(
                f"[{e['turn_id']}] {e.get('speaker', '')}: \"{e['quote']}\""
                for e in (m.get("evidence") or []))
            for j, v in enumerate([unit, cal.machine_key(unit, m["machine_theme_id"]),
                                   m.get("label", ""),
                                   m.get("one_sentence_description", ""), ev,
                                   m.get("relevance", ""), RELEVANCE_CAVEAT], start=1):
                if len(str(v)) > EXCEL_CELL_LIMIT:
                    raise MatchingNotReady(f"Machine_Themes cell too long: {len(str(v))}")
                c = ws.cell(row=i, column=j, value=v)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.protection = Protection(locked=True)
            ws.row_dimensions[i].height = 120
            i += 1
    ws.freeze_panes = "C2"
    ws.protection.sheet = True
    ws.protection.enable()

    ws = wb.create_sheet("Scope")
    for k, (a, b) in enumerate([
        ("scope", f"{cal.UNITS[0]}-{cal.UNITS[-1]}, {cal.GUIDE_QUESTION} only"),
        ("human instances", "44 (unit_id x cluster_id)"),
        ("machine themes", "30 (unit_id x machine_theme_id)"),
        ("evidence", "COMPLETE — no quote is abbreviated anywhere in this workbook"),
        ("immutable columns", "sealed and compared character-by-character on return"),
        ("UNCERTAIN keys", "candidates only — never a match, never counted"),
        ("coverage reference", cal.COVERAGE_REFERENCE),
        ("centrality", cal.CENTRALITY_NOT_AVAILABLE),
        ("model relevance", RELEVANCE_CAVEAT),
        ("supersedes", "Emergent_Matching_Q3_RESEARCHER.xlsx (truncated evidence)"),
        ("codebook", "do not consult it — this is an emergent comparison"),
    ], start=1):
        ws.cell(row=k, column=1, value=a).font = Font(bold=True)
        ws.cell(row=k, column=2, value=b)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 82

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Return gate
# ---------------------------------------------------------------------------

def read_rows(path: Path = None) -> list[dict]:
    path = path or _WB
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb["Human_Matching"]
    hdr = [c.value for c in ws[1]]
    rows = []
    for n, r in enumerate(ws.iter_rows(min_row=2), start=2):
        rec = dict(zip(hdr, [c.value for c in r]))
        if all(v is None for v in rec.values()):
            continue
        rec["_excel_row"] = n
        rows.append(rec)
    wb.close()
    return rows


def validate(path: Path = None) -> list[str]:
    """Problems that must be fixed before the matching can be imported."""
    path = path or _WB
    problems: list[str] = []
    rows = read_rows(path)
    canon = canonical_rows()
    _, machine = _load()
    valid_keys = {u: {cal.machine_key(u, m["machine_theme_id"])
                      for m in machine.get(u, [])} for u in cal.UNITS}
    all_keys = {k for v in valid_keys.values() for k in v}

    # --- the immutable columns, rebuilt and compared -----------------------
    if len(rows) != len(canon):
        problems.append(f"{len(rows)} human rows, expected {len(canon)} — rows were "
                        f"added or deleted")
    by_row = {r["_excel_row"]: r for r in rows}
    for c in canon:
        got = by_row.get(c["excel_row"])
        if got is None:
            problems.append(f"row {c['excel_row']} ({c['human_key']}) is missing")
            continue
        if str(got.get("human_key") or "").strip() != c["human_key"]:
            problems.append(f"row {c['excel_row']}: human_key is "
                            f"{got.get('human_key')!r}, expected {c['human_key']!r} — "
                            f"rows were reordered or replaced")
            continue
        for field in IMMUTABLE_COLS:
            exp = c[field]
            act = got.get(field)
            act = "" if act is None else str(act)
            if act != exp:
                problems.append(
                    f"{c['human_key']}: {field} was modified "
                    f"(expected {len(exp)} chars, found {len(act)}); immutable columns "
                    f"are compared against the frozen sources")
    extra = sorted(set(by_row) - {c["excel_row"] for c in canon})
    for n in extra:
        problems.append(f"row {n} is not part of the sealed 44 — rows were added")

    # --- the decisions -----------------------------------------------------
    for r in rows:
        n = r["_excel_row"]
        unit = str(r.get("unit_id") or "").strip()
        hk = str(r.get("human_key") or "").strip()
        dec = str(r.get("human_decision") or "").strip()
        raw = str(r.get("matched_machine_keys") or "").strip()
        why = str(r.get("researcher_reasoning") or "").strip()
        keys = [k.strip() for k in raw.split(";") if k.strip()]

        if not dec:
            problems.append(f"row {n} ({hk}): no decision")
            continue
        if dec not in HUMAN_DECISIONS:
            problems.append(f"row {n} ({hk}): decision {dec!r} is not one of "
                            f"{list(HUMAN_DECISIONS)}")
            continue
        if dec == "MATCHED" and not keys:
            problems.append(f"row {n} ({hk}): MATCHED but no machine key given")
        if dec == "NO_MATCH_HUMAN_ONLY" and keys:
            problems.append(f"row {n} ({hk}): NO_MATCH_HUMAN_ONLY must leave "
                            f"matched_machine_keys empty, got {keys}")
        if dec == "UNCERTAIN" and not why:
            problems.append(f"row {n} ({hk}): UNCERTAIN requires a short reason")
        if len(keys) != len(set(keys)):
            problems.append(f"row {n} ({hk}): duplicate machine keys in {raw!r}")
        for k in keys:
            if k in valid_keys.get(unit, set()):
                continue
            if k in all_keys:
                other = next(u for u, ks in valid_keys.items() if k in ks)
                problems.append(f"row {n} ({hk}): {k} belongs to {other}, not {unit} — "
                                f"themes from different units cannot be related")
            else:
                problems.append(f"row {n} ({hk}): unknown machine key {k!r}")
    return problems


# ---------------------------------------------------------------------------
# Derivation — MATCHED and UNCERTAIN are kept strictly apart
# ---------------------------------------------------------------------------

def derive(path: Path = None) -> dict:
    problems = validate(path)
    if problems:
        raise MatchingNotReady("matching is not usable yet:\n  - " +
                               "\n  - ".join(problems))
    rows = read_rows(path)
    _, machine = _load()

    confirmed: dict[str, list[str]] = {}     # machine_key -> human_keys (MATCHED only)
    candidates: dict[str, list[str]] = {}    # machine_key -> human_keys (UNCERTAIN)
    out_rows, uncertain_rows = [], []

    for r in rows:
        unit = str(r["unit_id"]).strip()
        hk = str(r["human_key"]).strip()
        dec = str(r["human_decision"]).strip()
        keys = [k.strip() for k in
                str(r.get("matched_machine_keys") or "").split(";") if k.strip()]
        why = str(r.get("researcher_reasoning") or "").strip()

        if dec == "MATCHED":
            for k in keys:
                confirmed.setdefault(k, []).append(hk)
            rel = "one_to_one" if len(keys) == 1 else "one_to_many"
            out_rows.append({"unit_id": unit, "human_key": hk, "decision": dec,
                             "confirmed_machine_keys": keys, "relation": rel,
                             "reasoning": why})
        elif dec == "NO_MATCH_HUMAN_ONLY":
            out_rows.append({"unit_id": unit, "human_key": hk, "decision": dec,
                             "confirmed_machine_keys": [],
                             "relation": "no_match_human_only", "reasoning": why})
        else:                                   # UNCERTAIN
            for k in keys:
                candidates.setdefault(k, []).append(hk)
            # NOT a link. No relation is asserted, in either direction.
            out_rows.append({"unit_id": unit, "human_key": hk, "decision": dec,
                             "confirmed_machine_keys": [],
                             "candidate_machine_keys": keys,
                             "relation": "unresolved", "reasoning": why})
            uncertain_rows.append({"unit_id": unit, "human_key": hk,
                                   "candidate_machine_keys": keys, "reasoning": why})

    all_machine = [cal.machine_key(u, m["machine_theme_id"])
                   for u in cal.UNITS for m in machine.get(u, [])]
    # A candidate key does NOT clear a theme from the queue.
    queue = [k for k in all_machine if k not in confirmed]

    n_matched = sum(1 for r in out_rows if r["decision"] == "MATCHED")
    resolved = not uncertain_rows

    return {
        "derived_utc": datetime.now(UTC).isoformat(),
        "n_human_rows": len(out_rows),
        "n_machine_themes": len(all_machine),
        "rows": out_rows,

        "confirmed_links": {k: sorted(v) for k, v in sorted(confirmed.items())},
        "candidate_uncertain_links": {k: sorted(v)
                                      for k, v in sorted(candidates.items())},
        "uncertain_rows": uncertain_rows,
        "n_uncertain": len(uncertain_rows),

        "machine_only_requiring_adjudication": queue,
        "machine_only_verdicts": list(cal.MACHINE_ONLY_VERDICTS),

        # confirmed structure only; a candidate never creates either of these
        "possible_fusion_one_machine_many_human": {
            k: v for k, v in confirmed.items() if len(v) > 1},
        "possible_fragmentation_one_human_many_machine": {
            r["human_key"]: r["confirmed_machine_keys"] for r in out_rows
            if len(r["confirmed_machine_keys"]) > 1},

        # recall numerator counts MATCHED rows only
        "recall_numerator_matched_human_instances": n_matched,
        "recall_denominator_union_reference": len(out_rows),
        "resolution_state": ("RESOLVED" if resolved else UNRESOLVED),
        "resolution_note": (
            "Every row is decided." if resolved else
            f"{len(uncertain_rows)} UNCERTAIN row(s) remain. Their candidate keys are "
            f"NOT matches: they do not raise recall, do not clear a machine theme from "
            f"the adjudication queue, and do not create confirmed fusion or "
            f"fragmentation. The calibration state stays {UNRESOLVED} until they are "
            f"settled."),
    }


def main() -> int:
    a = sys.argv[1:]
    if "--build" in a:
        print(f"built {build().name}")
    elif "--validate" in a:
        probs = validate()
        print("READY" if not probs else f"NOT_READY — {len(probs)} problem(s)")
        for x in probs[:40]:
            print("  -", x)
    else:
        print(__doc__)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
