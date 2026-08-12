"""
PRIMARY_EMERGENT_AUTOMATION_CALIBRATION_Q3 — PREPARED, NOTHING EXECUTED.

Calibrates the automated emergent extractor against the human clustering of
U01-U07 (guide question Q3). This module builds the frozen prompt, the output
schema, the cache-key design, the reference export and the human matching workbook.

IT MAKES NO API CALL, AND IT NEVER OPENS THE ACTIVE CLUSTERING WORKBOOK FOR WRITING.
The reference export reads that workbook only after `--validate` reports READY, and
only for reading.

WHAT THIS IS NOT
Not an application to the 35-transcript corpus. Seven units, one guide question. Its
purpose is to find out whether the extractor can be trusted at all, before anyone
asks it to do more.

THE MODEL CHOICE IS AN OPEN DECISION
`gemini-3.5-flash` is recommended for consistency with Tier 1, but it is NOT
automatically validated for open-ended extraction. Tier 1 was codebook-driven
classification against 11 fixed subthemes; this asks the model to invent the
categories. Those are different tasks and the qualification does not transfer.
U01-U07 IS that qualification, which is why it must run before anything else.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime, UTC
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "scripts"))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_OUT = _REPO_ROOT / "analysis" / "production_evaluation"
_DIR = _OUT / "emergent_calibration_q3"
_PKG = _OUT / "gold_standard_package"
_SEALED = _OUT / "gold_standard_sealed"
_CLUSTERING = _OUT / "partial_emergent_clustering" / "Clustering_U01_U07.xlsx"

CLASSIFICATION = "PRIMARY_EMERGENT_AUTOMATION_CALIBRATION_Q3"
UNITS = [f"U{i:02d}" for i in range(1, 8)]
SCHEMA_VERSION = "emergent_theme_v1"
GUIDE_QUESTION = "Q3"

HDR = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
NEED = PatternFill("solid", fgColor="FFF2CC")


class CalibrationNotReady(RuntimeError):
    """Raised when a precondition for the next step is not met."""


def _rel(p: Path):
    try:
        return p.relative_to(_REPO_ROOT)
    except ValueError:
        return p


def _sha(t: str) -> str:
    return hashlib.sha256(t.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# The frozen extraction prompt
# ---------------------------------------------------------------------------
#
# It carries no codebook, no subtheme names or letters, no Tier-1 result, no
# condition, no human/synthetic marker, no focus-group id and no study metadata. It
# does not name the discussion topic either: telling the model what the conversation
# is "about" would seed the categories it is supposed to discover.

EXTRACTION_SYSTEM_PROMPT = """\
You are analysing one short extract from a group discussion.

Read it and identify the distinct substantive themes it contains. Work only from the
extract in front of you. You have no category list and are not matching against one.

WHAT COUNTS AS A THEME

A theme is a distinct substantive claim or idea, not merely a topic. "Cost" is a
topic; "buying differently would cost more than they can justify" is a theme.

  * Merge formulations that say the same thing in different words.
  * Keep opposing positions separate. Asserting something and denying it are two
    themes, not one, however similar the wording.
  * A difference in WHO or WHAT is acting, in what is acted upon, or in the stance
    taken, may itself make two themes distinct. Judge whether the underlying claim
    differs.

EVIDENCE

Every theme must be supported by at least one quotation:

  * copied VERBATIM from the extract, character for character;
  * attributed to the turn_id it appears in;
  * spoken by a PARTICIPANT. Never quote the moderator. The moderator's questions
    and summaries are not evidence of what the group thinks.

If you cannot support a candidate theme with a participant quotation, do not report
it.

RELEVANCE

  * "central"   — a main idea of this extract; the discussion would be
                  misrepresented without it.
  * "secondary" — genuinely present, but minor or raised in passing.

HOW MANY THEMES

Report as many as the extract genuinely contains, and no more. There is no target
number. Do not pad the list to look thorough, and do not compress distinct claims to
look concise.

OUTPUT

Return JSON only, matching the schema you are given. No commentary.
"""

# JSON Schema for one unit's output.
EXTRACTION_SCHEMA = {
    "schema_version": SCHEMA_VERSION,
    "type": "object",
    "required": ["themes"],
    "additionalProperties": False,
    "properties": {
        "themes": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["machine_theme_id", "label", "one_sentence_description",
                             "relevance", "evidence", "voiced_by"],
                "additionalProperties": False,
                "properties": {
                    "machine_theme_id": {"type": "string",
                                         "description": "unique within this unit, e.g. M1"},
                    "label": {"type": "string"},
                    "one_sentence_description": {"type": "string"},
                    "relevance": {"type": "string", "enum": ["central", "secondary"]},
                    "evidence": {
                        "type": "array", "minItems": 1,
                        "items": {
                            "type": "object",
                            "required": ["turn_id", "speaker", "quote"],
                            "additionalProperties": False,
                            "properties": {
                                "turn_id": {"type": "string"},
                                "speaker": {"type": "string"},
                                "quote": {"type": "string",
                                          "description": "verbatim from that turn"},
                            },
                        },
                    },
                    "voiced_by": {"type": "array", "items": {"type": "string"}},
                    "evidence_note": {"type": "string"},
                },
            },
        },
    },
}

# Terms that must never appear in the prompt. Codebook ids, theme names, subtheme
# labels and the study's framing vocabulary.
FORBIDDEN_IN_PROMPT = (
    "a.1", "a.2", "a.3", "b.1", "b.2", "b.3", "b.4", "c.1", "c.2", "c.3",
    "4n", "codebook", "subtheme", "gender", "masculin", "plant-based", "plant based",
    "meat", "vegetarian", "vegan", "food", "eat", "diet", "macho",
    "natural", "normal", "necessary", "unnatural", "insufficient",
    "justification", "extreme case", "focus group", "synthetic", "enriched",
    "demographics", "participant 1", "tier 1", "tier-1", "recall", "precision",
)


# ---------------------------------------------------------------------------
# Effective configuration and cache key
# ---------------------------------------------------------------------------


def transmitted_response_schema() -> dict:
    """
    The schema ACTUALLY SENT to the model.

    The prompt says "matching the schema you are given", so a schema must genuinely be
    transmitted or that instruction dangles. Gemini's structured-output subset does not
    accept `additionalProperties` or a custom `schema_version` key, so those are dropped
    here. What is dropped is only a validator-side restriction; the fields, types,
    enums and required lists are identical, and the FULL schema still governs
    validate_extraction() locally.
    """
    def strip(node):
        if isinstance(node, dict):
            return {k: strip(v) for k, v in node.items()
                    if k not in ("additionalProperties", "schema_version")}
        if isinstance(node, list):
            return [strip(x) for x in node]
        return node
    return strip(EXTRACTION_SCHEMA)


def response_schema_sha() -> str:
    return hashlib.sha256(
        json.dumps(transmitted_response_schema(), sort_keys=True,
                   ensure_ascii=False).encode("utf-8")).hexdigest()


def proposed_effective_config(model: str = "gemini-3.5-flash",
                              max_output_tokens: int = 16384) -> dict:
    """
    PROPOSED — not frozen until approved.

    Mirrors the Tier-1 effective-config discipline: every transmitted parameter is
    recorded, and the whole dict keys the cache.
    """
    return {
        "task": "emergent_extraction",
        "execution_mode": "batch",
        "model": model,
        "response_mime_type": "application/json",
        # The response schema is a TRANSMITTED parameter. It must appear here, and in
        # the cache key: two runs that sent different schemas are not the same run.
        "response_schema_transmitted": True,
        "response_schema_sha256": response_schema_sha(),
        "max_output_tokens": max_output_tokens,
        "temperature_transmitted": False,
        "temperature": None,
        "thinking_config_transmitted": False,
        "thinking_config": None,
        "thinking_level_effective": "model_default_unpinned",
        "schema_version": SCHEMA_VERSION,
    }


def canonical_config(effective: dict) -> str:
    return json.dumps(effective, sort_keys=True, separators=(",", ":"))


def cache_key(unit_text_sha: str, prompt_sha: str, effective: dict) -> str:
    """
    NOTE THE ABSENT COMPONENT.

    The Tier-1 key includes `codebook_sha256`. This one does not, because there is no
    codebook — that absence is the defining property of the task, and a key that
    silently reused the Tier-1 shape would imply a codebook was involved.
    """
    return hashlib.sha256("|".join([
        unit_text_sha, "emergent_q3", prompt_sha,
        SCHEMA_VERSION, canonical_config(effective),
    ]).encode("utf-8")).hexdigest()


def prompt_sha() -> str:
    return _sha(EXTRACTION_SYSTEM_PROMPT)


def prompt_purity_problems(prompt: str = None) -> list[str]:
    """Any forbidden term present in the prompt."""
    text = (prompt if prompt is not None else EXTRACTION_SYSTEM_PROMPT).lower()
    return [t for t in FORBIDDEN_IN_PROMPT if t in text]


# ---------------------------------------------------------------------------
# Human reference export — three views, never mixed
# ---------------------------------------------------------------------------

# union_reference is the PRIMARY AND ONLY reference for coverage. The coder views are
# not alternative references: they exist so each coder's own recall against the SAME
# union can be computed, which is what the B+ coverage benchmark needs.
REFERENCE_VIEWS = ("union_reference", "coder_a_view", "coder_b_view")
PRIMARY_VIEW = "union_reference"
COVERAGE_REFERENCE = "union_reference"

# Centrality was deliberately not assessed. There is no central_reference and there
# must not be one: an empty centrality artefact would otherwise be read as a reference
# containing zero central themes, converting a declined judgement into a finding.
CENTRALITY_STATUS = "NOT_ASSESSED"
CENTRALITY_NOT_AVAILABLE = "NOT_AVAILABLE — CENTRALITY_NOT_ASSESSED"
FORBIDDEN_VIEWS = ("central_reference",)

# Retained for provenance only; these rows are now treated exactly like every other.
CENTRALITY_MISSING_ROWS = ("P034", "P040")


def human_key(unit_id: str, cluster_id: str) -> str:
    """Cluster identity is ALWAYS the pair. A bare cluster_id is not an identity."""
    return f"{unit_id}::{cluster_id}"


def export_human_reference(clustering_workbook: Path = None,
                           out_dir: Path = None) -> dict:
    """
    Export the human reference from the VALIDATED clustering workbook.

    Refuses unless the workbook passes its own return gate. Reads only; the workbook
    is never opened for writing and never re-sealed.
    """
    import partial_emergent_clustering_pipeline as pipe

    wb_path = clustering_workbook or _CLUSTERING
    out_dir = out_dir or _DIR

    problems = pipe.validate(wb_path)
    if problems:
        raise CalibrationNotReady(
            f"the clustering workbook is not READY ({len(problems)} problem(s)); "
            f"the human reference cannot be exported yet. First: "
            f"{problems[0]}")

    rows = pipe._read(wb_path)
    who = {}
    seal = _SEALED / "partial_emergent_pooled_authorship.json"
    if seal.exists():
        who = {m["pooled_id"]: m["coder"]
               for m in json.loads(seal.read_text(encoding="utf-8"))["map"]}

    cells: dict[tuple, dict] = {}
    for r in rows:
        cid, unit, pid = r["cluster_id"], r["unit_id"], r["pooled_id"]
        c = cells.setdefault((unit, cid), {
            "human_key": human_key(unit, cid),
            "unit_id": unit, "cluster_id": cid,
            "cluster_label": r.get("cluster_label", ""),
            "consolidated_definition": r.get("cluster_label", ""),
            "source_pooled_ids": [], "coders": set(),
            "centrality_status": CENTRALITY_STATUS,
            "centrality_value_as_recorded": None,
            "supporting_quotes": [],
        })
        c["source_pooled_ids"].append(pid)
        c["coders"].add(who.get(pid, "?"))
        # Any surviving centrality value is preserved verbatim but never used.
        if (r.get("is_central") or "").strip():
            c["centrality_value_as_recorded"] = r["is_central"]
        if r.get("supporting_quote"):
            c["supporting_quotes"].append(
                {"pooled_id": pid, "quote": r["supporting_quote"]})

    records = []
    for key in sorted(cells):
        c = cells[key]
        coders = c.pop("coders")
        records.append({
            **c,
            "n_source_rows": len(c["source_pooled_ids"]),
            "coder_a_present": int("Coder_A" in coders),
            "coder_b_present": int("Coder_B" in coders),
            "raised_by_both": int({"Coder_A", "Coder_B"} <= coders),
            "n_coders": len({x for x in coders if x in ("Coder_A", "Coder_B")}),
        })

    # The coder views are SUBSETS of the union, keyed identically, so each coder's
    # recall against the union is |view| / |union|.
    views = {
        "union_reference": records,
        "coder_a_view": [r for r in records if r["coder_a_present"]],
        "coder_b_view": [r for r in records if r["coder_b_present"]],
    }
    n_union = len(records)
    out = {
        "exported_utc": datetime.now(UTC).isoformat(),
        "classification": CLASSIFICATION,
        "guide_question": GUIDE_QUESTION,
        "units": UNITS,
        "cluster_identity": "(unit_id, cluster_id)",
        "primary_view": PRIMARY_VIEW,
        "coverage_reference": COVERAGE_REFERENCE,
        "view_rule": (
            "union_reference is the primary AND ONLY reference for coverage. "
            "coder_a_view and coder_b_view are subsets of it, present so that each "
            "coder's own recall against the SAME union can be computed; they are "
            "never used as alternative denominators and are never pooled."),
        "centrality_status": CENTRALITY_STATUS,
        "central_reference": CENTRALITY_NOT_AVAILABLE,
        "centrality_note": (
            "The researcher decided not to classify clusters as central or "
            "peripheral because that distinction could not be determined reliably "
            "from the available material. This is a methodological decision, not "
            "missing data. No central_reference exists. An absent centrality "
            "artefact must NOT be interpreted as a reference with zero central "
            "themes. The human review validates theme identification, description, "
            "textual evidence and grouping; it does NOT validate hierarchy, relative "
            "importance or thematic salience."),
        "counts": {k: len(v) for k, v in views.items()},
        "coder_recall_vs_union": {
            "Coder_A": {"numerator": len(views["coder_a_view"]),
                        "denominator": n_union},
            "Coder_B": {"numerator": len(views["coder_b_view"]),
                        "denominator": n_union},
        },
        **views,
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    # Atomic: a crash partway through must not leave a half-written reference that
    # later looks authoritative.
    dst = out_dir / "human_reference_q3.json"
    tmp = dst.with_suffix(".json.tmp")
    try:
        tmp.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")
        os.replace(tmp, dst)
    finally:
        if tmp.exists():
            tmp.unlink()
    return out


# ---------------------------------------------------------------------------
# Technical validation of extractor output — before any matching
# ---------------------------------------------------------------------------

def validate_extraction(unit_id: str, payload: dict, telemetry: dict,
                        unit_lines: list[str]) -> dict:
    """
    Gate one unit's extractor output. Incomplete output is quarantined, not cached.
    """
    problems: list[str] = []
    turns = {}
    for line in unit_lines:
        m = re.match(r"^\[(T\d+)\]\s+([^:]+):\s*(.*)$", line, re.S)
        if m:
            turns[m.group(1)] = (m.group(2).strip(), m.group(3).strip())

    reasons = telemetry.get("finish_reasons") or []
    if not reasons or any("STOP" not in str(r).upper() for r in reasons):
        problems.append(f"finish_reason is {reasons}, expected STOP")

    themes = (payload or {}).get("themes")
    if themes is None:
        problems.append("no `themes` array — schema invalid")
        themes = []

    ids = [t.get("machine_theme_id") for t in themes]
    if len(set(ids)) != len(ids):
        problems.append(f"duplicate machine_theme_id: "
                        f"{sorted({i for i in ids if ids.count(i) > 1})}")
    if any(not i for i in ids):
        problems.append("a theme has no machine_theme_id")

    norm = lambda s: " ".join(re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split())
    for t in themes:
        tid = t.get("machine_theme_id", "?")
        for field in ("label", "one_sentence_description", "relevance"):
            if not t.get(field):
                problems.append(f"{tid}: missing {field}")
        if t.get("relevance") not in (None, "central", "secondary"):
            problems.append(f"{tid}: relevance is {t['relevance']!r}")
        ev = t.get("evidence") or []
        if not ev:
            problems.append(f"{tid}: no evidence — a theme without a quotation is "
                            f"not reportable")
        for e in ev:
            tur = e.get("turn_id")
            if tur not in turns:
                problems.append(f"{tid}: turn_id {tur!r} is not in this unit")
                continue
            speaker, content = turns[tur]
            if speaker.lower().startswith("moderator"):
                problems.append(f"{tid}: quotes the moderator at {tur} — the "
                                f"moderator is not evidence of what the group thinks")
            if norm(e.get("quote", "")) not in norm(content):
                problems.append(f"{tid}: quote at {tur} is not verbatim in that turn")

    return {"unit_id": unit_id,
            "status": "COMPLETE" if not problems else "QUARANTINE",
            "n_themes": len(themes), "problems": problems}


def assert_corpus_complete(results: dict) -> None:
    """All seven units, all COMPLETE, before any matching begins."""
    problems = []
    missing = [u for u in UNITS if u not in results]
    if missing:
        problems.append(f"missing units: {missing}")
    for u, r in sorted(results.items()):
        if u not in UNITS:
            problems.append(f"unexpected unit: {u}")
        elif r.get("status") != "COMPLETE":
            problems.append(f"{u}: {r.get('status')} — {r.get('problems', [])[:2]}")
    if problems:
        raise CalibrationNotReady(
            "extraction corpus incomplete; matching refused:\n  - "
            + "\n  - ".join(problems))


# ---------------------------------------------------------------------------
# Bipartite matching workbook — human decides, always
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# FROZEN ADJUDICATION RULES — approved by the researcher 2026-07-31, before any
# extraction ran. Rule 6.4 = Alternative 1, the QUALITATIVE gate. Nothing here may be
# rewritten after results exist.
# ---------------------------------------------------------------------------

MACHINE_ONLY_VERDICTS = ("VALID_NOVEL_THEME", "UNSUPPORTED_OR_SPURIOUS",
                         "DUPLICATE_MACHINE_THEME", "UNCERTAIN")

ADJUDICATION_RULES = {
    "UNSUPPORTED_OR_SPURIOUS": (
        "the theme asserts an idea for which there is not sufficient textual evidence "
        "in the unit"),
    "SEVERE_UNSUPPORTED_THEME": (
        "contradicts the text, attributes a position nobody holds, invents "
        "relationships or actors, or turns an incidental mention into a substantive "
        "thematic claim"),
    "RECURRENT_UNSUPPORTED_PATTERN": (
        "the same KIND of error appears in at least two different units. Used as a "
        "signal of systematic failure, NOT as a statistical test"),
    "DUPLICATE_MACHINE_THEME": (
        "two or more machine themes represent substantially the same claim within one "
        "unit"),
    "UNCERTAIN": (
        "the evidence permits more than one reasonable reading; counted automatically "
        "as neither correct nor incorrect"),
    "VALID_NOVEL_THEME": (
        "does not match the human reference, but is clearly supported by the text and "
        "constitutes a distinct thematic idea"),
}

# Every adjudication must retain these, per decision.
ADJUDICATION_REQUIRED_FIELDS = ("unit_id", "machine_theme_id", "verdict", "quote",
                                "human_justification")

# A PASS is NOT earned by clearing the coverage benchmark alone.
PASS_CONDITIONS = (
    "machine recall vs union_reference >= 0.6364 (28/44, the lower coder recall)",
    "no recurrent severe unsupported errors",
    "complete adjudication of every machine-only theme",
    "explicit human review of fragmentation and fusion",
)
COVERAGE_BENCHMARK = 28 / 44

FINAL_STATES = ("PASS_WITH_SAMPLED_HUMAN_VERIFICATION",
                "BORDERLINE — FALL_BACK_TO_ASSISTIVE_REVIEW",
                "FAIL — FALL_BACK_TO_ASSISTIVE_REVIEW",
                "UNRESOLVED_AT_THIS_SAMPLE_SIZE")


# ---------------------------------------------------------------------------
# Unit text — the ONLY input transmitted per call
# ---------------------------------------------------------------------------

_UNIT_TURN = re.compile(r"^\[(T\d+)\]\s+([^:]+):\s*(.*)$")


def unit_lines(unit_id: str) -> list[str]:
    """
    The unit's turns as "[Tnnn] Speaker: text".

    The "UNIT U0n" banner and rule line are stripped: the unit id is provenance and
    must never reach the model. Blank lines inside a turn are paragraph breaks and are
    preserved.
    """
    src = _PKG / f"{unit_id}.txt"
    out: list[list[str]] = []
    for line in src.read_text(encoding="utf-8").splitlines():
        m = _UNIT_TURN.match(line)
        if m:
            out.append([m.group(1), m.group(2).strip(), m.group(3)])
        elif out is not None and out and not line.startswith(("UNIT ", "=====")):
            out[-1][2] += "\n" + line
    return [f"[{t}] {sp}: {tx.rstrip()}" for t, sp, tx in out]


def unit_text(unit_id: str) -> str:
    return "\n".join(unit_lines(unit_id))


def unit_text_problems(unit_id: str) -> list[str]:
    """The transmitted text must carry no provenance."""
    low = " ".join(unit_text(unit_id).split()).lower()
    bad = []
    for leak in ("unit u0", "unit u1", "fg1", "fg2", "fg3", "fg4", "fg5",
                 "enriched", "demographics-only", "macho meals"):
        if leak in low:
            bad.append(f"{unit_id}: transmitted text contains {leak!r}")
    return bad


MATCH_RELATIONS = ("one_to_one", "one_to_many", "many_to_one",
                   "no_match_human_only", "no_match_machine_only")
MATCHED_RELATIONS = ("one_to_one", "one_to_many", "many_to_one")
NO_MATCH_RELATIONS = ("no_match_human_only", "no_match_machine_only")
DECISIONS = ("match", "no_match")

# ENTITY IDENTITY IS ALWAYS THE PAIR.
#   human theme   : (unit_id, human_cluster_id)
#   machine theme : (unit_id, machine_theme_id)
# A bare "M01", "H01" or "C01" is NOT an identity. The same id text recurs across
# units denoting entirely different themes, so adjudicating "M01" without its unit
# would silently complete a different unit's theme.
def machine_key(unit_id: str, machine_theme_id: str) -> str:
    return f"{unit_id}::{machine_theme_id}"


def validate_matching(rows: list[dict], human_keys: set, machine_keys: set) -> list[str]:
    """
    Structural validation of the matching sheet. Returns problems; never repairs.

    `human_keys` and `machine_keys` are sets of "UNIT::ID" strings — the full
    inventory that must be accounted for.
    """
    problems: list[str] = []
    human_by_id: dict[str, set] = {}
    machine_by_id: dict[str, set] = {}
    for k in human_keys:
        u, i = k.split("::", 1)
        human_by_id.setdefault(i, set()).add(u)
    for k in machine_keys:
        u, i = k.split("::", 1)
        machine_by_id.setdefault(i, set()).add(u)

    seen_pairs: dict[tuple, int] = {}
    decided_h: set = set()
    decided_m: set = set()
    matched_h: dict[str, set] = {}
    matched_m: dict[str, set] = {}
    said_match: set = set()
    said_no_match: set = set()

    for n, r in enumerate(rows, start=2):
        where = f"row {n}"
        unit = (r.get("unit_id") or "").strip()
        hid = (r.get("human_cluster_id") or "").strip()
        mid = (r.get("machine_theme_id") or "").strip()
        rel = (r.get("relation") or "").strip()
        dec = (r.get("decision") or "").strip()

        if not unit:
            problems.append(f"{where}: no unit_id — an entity without its unit has no identity")
            continue
        if unit not in UNITS:
            problems.append(f"{where}: unit_id {unit!r} is outside {UNITS[0]}-{UNITS[-1]}")
            continue
        if not hid and not mid:
            problems.append(f"{where}: orphan row — neither a human cluster nor a machine theme")
            continue

        # --- relation / decision presence and coherence --------------------
        if rel and rel not in MATCH_RELATIONS:
            problems.append(f"{where}: relation {rel!r} is not one of {list(MATCH_RELATIONS)}")
        if dec and dec not in DECISIONS:
            problems.append(f"{where}: decision {dec!r} is not one of {list(DECISIONS)}")
        if dec and not rel:
            problems.append(f"{where}: decision {dec!r} with no relation")
        if rel and not dec:
            problems.append(f"{where}: relation {rel!r} with no decision")
        if rel in MATCHED_RELATIONS and dec == "no_match":
            problems.append(f"{where}: relation {rel!r} conflicts with decision 'no_match'")
        if rel in NO_MATCH_RELATIONS and dec == "match":
            problems.append(f"{where}: relation {rel!r} conflicts with decision 'match'")
        if rel == "no_match_human_only" and mid:
            problems.append(f"{where}: no_match_human_only must leave machine_theme_id empty, got {mid!r}")
        if rel == "no_match_machine_only" and hid:
            problems.append(f"{where}: no_match_machine_only must leave human_cluster_id empty, got {hid!r}")
        if rel in MATCHED_RELATIONS and not (hid and mid):
            problems.append(f"{where}: relation {rel!r} needs both a human cluster and a machine theme")

        # --- keys resolve, within this unit --------------------------------
        hk = f"{unit}::{hid}" if hid else None
        mk = machine_key(unit, mid) if mid else None
        if hk and hk not in human_keys:
            elsewhere = sorted(human_by_id.get(hid, set()))
            if elsewhere:
                problems.append(
                    f"{where}: cross-unit relation — human cluster {hid!r} does not exist "
                    f"in {unit}; it exists in {elsewhere}. Identity is (unit_id, cluster_id).")
            else:
                problems.append(f"{where}: unknown human cluster key {hk!r}")
        if mk and mk not in machine_keys:
            elsewhere = sorted(machine_by_id.get(mid, set()))
            if elsewhere:
                problems.append(
                    f"{where}: cross-unit relation — machine theme {mid!r} does not exist "
                    f"in {unit}; it exists in {elsewhere}. Identity is (unit_id, machine_theme_id).")
            else:
                problems.append(f"{where}: unknown machine theme key {mk!r}")

        # --- duplicates within a unit --------------------------------------
        pair = (unit, hid, mid)
        if pair in seen_pairs:
            problems.append(f"{where}: duplicate pairing {pair} (already on row {seen_pairs[pair]})")
        else:
            seen_pairs[pair] = n

        # --- bookkeeping for the cross-row checks --------------------------
        if dec:
            if hk:
                decided_h.add(hk)
            if mk:
                decided_m.add(mk)
        if dec == "match":
            if hk:
                said_match.add(hk)
                matched_h.setdefault(hk, set()).add(mk)
            if mk:
                said_match.add(mk)
                matched_m.setdefault(mk, set()).add(hk)
        if dec == "no_match":
            for k in (hk, mk):
                if k:
                    said_no_match.add(k)

    # --- an entity cannot be both matched and unmatched --------------------
    for k in sorted(said_match & said_no_match):
        problems.append(f"{k}: marked BOTH matched and not matched")

    # --- declared cardinality must match the rows actually present ---------
    for n, r in enumerate(rows, start=2):
        unit = (r.get("unit_id") or "").strip()
        rel = (r.get("relation") or "").strip()
        hid = (r.get("human_cluster_id") or "").strip()
        mid = (r.get("machine_theme_id") or "").strip()
        if not unit or rel not in MATCHED_RELATIONS:
            continue
        hk, mk = (f"{unit}::{hid}" if hid else None), (machine_key(unit, mid) if mid else None)
        n_m = len(matched_h.get(hk, set()))
        n_h = len(matched_m.get(mk, set()))
        if rel == "one_to_many" and n_m < 2:
            problems.append(f"row {n}: declared one_to_many but human {hk} is matched to "
                            f"{n_m} machine theme(s)")
        if rel == "many_to_one" and n_h < 2:
            problems.append(f"row {n}: declared many_to_one but machine {mk} is matched to "
                            f"{n_h} human cluster(s)")
        if rel == "one_to_one" and (n_m > 1 or n_h > 1):
            problems.append(f"row {n}: declared one_to_one but human {hk} has {n_m} machine "
                            f"match(es) and machine {mk} has {n_h} human match(es)")

    # --- nothing may be left undecided -------------------------------------
    for k in sorted(human_keys - decided_h):
        problems.append(f"human cluster {k} has no decision")
    for k in sorted(machine_keys - decided_m):
        problems.append(f"machine theme {k} has no decision")
    for n, r in enumerate(rows, start=2):
        if (r.get("relation") or "").strip() in ("one_to_many", "many_to_one") \
                and not (r.get("reasoning") or "").strip():
            problems.append(f"row {n}: {r['relation']} requires reasoning")
    return problems


def build_matching_workbook(out_path: Path = None) -> Path:
    """
    Per-unit bipartite adjudication sheet. EMPTY: every decision is a human one.

    Candidate orderings by quote overlap and lexical similarity may be filled in as
    a convenience, but nothing is ever auto-accepted and no decision may be inferred
    from a similarity score.
    """
    out_path = out_path or (_DIR / "Emergent_Matching_Q3.xlsx")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    for i, (t, b) in enumerate([
        ("Emergent extractor calibration — matching (U01–U07)", True),
        ("", False),
        ("WHAT YOU ARE DOING", True),
        ("For each unit, decide how the automated themes relate to the human", False),
        ("clusters. Both sides are listed; you decide the correspondence.", False),
        ("", False),
        ("RULES", True),
        ("1. Nothing is pre-accepted. Every human cluster and every machine theme", False),
        ("   needs an explicit decision, including 'no match'.", False),
        ("2. Similarity columns, where present, are a reading aid only. A match is", False),
        ("   never justified by label similarity alone — check the evidence.", False),
        ("3. Do NOT consult the codebook. This is an emergent comparison.", False),
        ("4. Relations allowed: one_to_one, one_to_many, many_to_one, and no-match", False),
        ("   on either side. Record each pairing on its own row.", False),
        ("5. Reasoning is REQUIRED wherever the relation is many-to-one,", False),
        ("   one-to-many, or you consider the boundary doubtful.", False),
        ("6. A machine theme with no human counterpart is MACHINE-ONLY: not", False),
        ("   observed in the human reference. It is not an error, and it is NOT", False),
        ("   automatically a false positive — it goes to novel-theme adjudication.", False),
        ("7. IDENTITY IS (unit_id, id). The same id text recurs in other units", False),
        ("   meaning something different. Always work within one unit's block.", False),
        ("", False),
        ("WHAT HAPPENS NEXT", True),
        ("Metrics are computed only after every row has a decision. The metrics", False),
        ("must never be used to revisit a match.", False),
    ], start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=b, size=11 if b else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Matching")
    cols = ["unit_id", "human_cluster_id", "human_cluster_label",
            "machine_theme_id", "machine_theme_label",
            "quote_overlap", "label_similarity",
            "relation", "decision", "reasoning", "adjudicator", "date_utc"]
    widths = [10, 18, 34, 18, 34, 13, 15, 20, 14, 46, 16, 14]
    for j, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(row=2, column=1, value=(
        "POPULATED after extraction: one row per candidate pair, plus a row for every "
        "human cluster and machine theme so nothing can be skipped. Left empty here."))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(MATCH_RELATIONS) + '"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("H2:H2000")
    dv2 = DataValidation(type="list", formula1='"match,no_match"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("I2:I2000")
    ws.freeze_panes = "C2"

    ws = wb.create_sheet("Coverage")
    for j, h in enumerate(["unit_id", "n_human_clusters", "n_machine_themes",
                           "n_rows_decided", "all_human_decided",
                           "all_machine_decided"], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = 20
    for i, u in enumerate(UNITS, start=2):
        ws.cell(row=i, column=1, value=u)

    ws = wb.create_sheet("Scope")
    for i, (a, b) in enumerate([
        ("classification", CLASSIFICATION),
        ("scope", "U01–U07, guide question Q3 only"),
        ("primary reference", "union_reference"),
        ("sensitivities", "two_coder_reference, central_reference — never pooled"),
        ("codebook", "NOT used in extraction or matching"),
        ("supplementary sample", "S01–S06 — separate, never mixed with this"),
        ("machine-only themes", "not observed in the human reference; NOT errors"),
    ], start=1):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    return out_path


def build_populated_matching_workbook(results: dict = None, out_path: Path = None) -> Path:
    """
    The matching workbook, POPULATED with both inventories and every decision blank.

    One row per human cluster and one row per machine theme, per unit, so nothing can
    be skipped by omission. Candidate pairs are NOT pre-proposed: an ordering by
    similarity would anchor the adjudicator, and a match is never justified by label
    similarity alone.
    """
    out_path = out_path or (_DIR / "Emergent_Matching_Q3_POPULATED.xlsx")
    ref = json.loads((_DIR / "human_reference_q3.json").read_text(encoding="utf-8"))
    if results is None:
        results = json.loads((_DIR / "extraction_results_q3.json").read_text(encoding="utf-8"))

    human = {}
    for r in ref["union_reference"]:
        human.setdefault(r["unit_id"], []).append(r)
    machine = {}
    for r in results["results"]:
        if r.get("status") == "COMPLETE":
            machine[r["unit_id"]] = r.get("themes", [])

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Rules"
    rows = [("FROZEN ADJUDICATION RULES — approved before the extractor ran", True),
            ("", False),
            ("These definitions are frozen. They are not to be reinterpreted in the", False),
            ("light of what the extractor actually produced.", False), ("", False)]
    for k in ("UNSUPPORTED_OR_SPURIOUS", "SEVERE_UNSUPPORTED_THEME",
              "RECURRENT_UNSUPPORTED_PATTERN", "DUPLICATE_MACHINE_THEME",
              "UNCERTAIN", "VALID_NOVEL_THEME"):
        rows.append((k, True))
        rows.append(("    " + ADJUDICATION_RULES[k], False))
    rows += [("", False), ("A PASS REQUIRES ALL OF:", True)]
    rows += [("    " + c, False) for c in PASS_CONDITIONS]
    rows += [("", False), ("Every decision must retain: quote, unit, category and", False),
             ("human justification.", False)]
    for i, (t, b) in enumerate(rows, start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(bold=b, size=11 if b else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 108

    # --- Matching -----------------------------------------------------------
    ws = wb.create_sheet("Matching")
    cols = ["unit_id", "side", "human_key", "human_cluster_id", "human_cluster_label",
            "machine_key", "machine_theme_id", "machine_theme_label",
            "machine_quote", "relation", "decision", "reasoning", "adjudicator",
            "date_utc"]
    widths = [9, 9, 14, 15, 40, 14, 15, 40, 46, 20, 12, 44, 14, 12]
    for j, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w

    i = 2
    for unit in UNITS:
        for h in human.get(unit, []):
            ws.cell(row=i, column=1, value=unit)
            ws.cell(row=i, column=2, value="human")
            ws.cell(row=i, column=3, value=h["human_key"])
            ws.cell(row=i, column=4, value=h["cluster_id"])
            ws.cell(row=i, column=5, value=h["cluster_label"])
            i += 1
        for m in machine.get(unit, []):
            ws.cell(row=i, column=1, value=unit)
            ws.cell(row=i, column=2, value="machine")
            ws.cell(row=i, column=6, value=machine_key(unit, m["machine_theme_id"]))
            ws.cell(row=i, column=7, value=m["machine_theme_id"])
            ws.cell(row=i, column=8, value=m.get("label", ""))
            ev = (m.get("evidence") or [{}])[0]
            ws.cell(row=i, column=9, value=ev.get("quote", ""))
            i += 1
    for row in ws.iter_rows(min_row=2, max_row=i - 1):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"' + ",".join(MATCH_RELATIONS) + '"',
                        allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"J2:J{i + 400}")
    dv2 = DataValidation(type="list", formula1='"match,no_match"', allow_blank=True)
    ws.add_data_validation(dv2); dv2.add(f"K2:K{i + 400}")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:N{i - 1}"

    # --- Machine-only adjudication -----------------------------------------
    ws = wb.create_sheet("Machine_Only_Adjudication")
    cols2 = ["unit_id", "machine_key", "machine_theme_id", "machine_theme_label",
             "quote", "turn_id", "verdict", "is_severe", "error_kind",
             "human_justification", "adjudicator", "date_utc"]
    widths2 = [9, 14, 15, 40, 50, 9, 26, 11, 26, 50, 14, 12]
    for j, (h, w) in enumerate(zip(cols2, widths2), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.cell(row=2, column=1, value=(
        "Populated after matching identifies which machine themes have no human "
        "counterpart. A machine-only theme is NOT automatically a false positive: the "
        "union comes from two coders over seven units and may contain human omissions."))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    dv3 = DataValidation(type="list",
                         formula1='"' + ",".join(MACHINE_ONLY_VERDICTS) + '"',
                         allow_blank=True)
    ws.add_data_validation(dv3); dv3.add("G3:G500")
    dv4 = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv4); dv4.add("H3:H500")
    ws.freeze_panes = "C3"

    # --- Fragmentation / fusion review -------------------------------------
    ws = wb.create_sheet("Fragmentation_Fusion")
    for j, h in enumerate(["unit_id", "pattern", "human_keys", "machine_keys",
                           "human_review_required", "reviewer_note"], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = [9, 18, 30, 30, 22, 60][j - 1]
    ws.cell(row=2, column=1, value=(
        "Explicit human review of fragmentation and fusion is a PASS CONDITION. "
        "It is not inferred from the relation column."))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # --- Coverage -----------------------------------------------------------
    ws = wb.create_sheet("Coverage")
    for j, h in enumerate(["unit_id", "n_human_clusters", "n_machine_themes",
                           "all_human_decided", "all_machine_decided"], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HDR, HDR_FONT
        ws.column_dimensions[get_column_letter(j)].width = 20
    for k, u in enumerate(UNITS, start=2):
        ws.cell(row=k, column=1, value=u)
        ws.cell(row=k, column=2, value=len(human.get(u, [])))
        ws.cell(row=k, column=3, value=len(machine.get(u, [])))

    ws = wb.create_sheet("Scope")
    for k, (a, b) in enumerate([
        ("scope", f"{UNITS[0]}-{UNITS[-1]}, {GUIDE_QUESTION} only"),
        ("cluster identity", "(unit_id, cluster_id)"),
        ("machine identity", "(unit_id, machine_theme_id)"),
        ("coverage reference", COVERAGE_REFERENCE),
        ("centrality", CENTRALITY_NOT_AVAILABLE),
        ("supplementary sample", "S01-S06 are a SEPARATE study and are never mixed in"),
        ("codebook", "do not consult the codebook — this is an emergent comparison"),
    ], start=1):
        ws.cell(row=k, column=1, value=a).font = Font(bold=True)
        ws.cell(row=k, column=2, value=b)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 74

    wb.save(out_path)
    return out_path


def assert_matching_complete(rows: list[dict], human_keys: set, machine_keys: set) -> None:
    """
    Metrics are refused unless the matching is structurally sound AND complete.

    `human_keys` / `machine_keys` are "UNIT::ID" strings. Passing bare ids would let
    a decision recorded in one unit satisfy a same-named theme in another.
    """
    for k in list(human_keys) + list(machine_keys):
        if "::" not in str(k):
            raise CalibrationNotReady(
                f"key {k!r} is not (unit_id, id) — a bare id is not an identity")
    problems = validate_matching(rows, set(human_keys), set(machine_keys))
    if problems:
        raise CalibrationNotReady(
            "matching not usable; metrics refused:\n  - " + "\n  - ".join(problems))


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--build-matching", action="store_true")
    ap.add_argument("--export-reference", action="store_true")
    ap.add_argument("--show-config", action="store_true")
    a = ap.parse_args()

    if a.build_matching:
        p = build_matching_workbook()
        print(f"matching workbook: {_rel(p)}  (EMPTY — every decision is human)")
        return 0
    if a.export_reference:
        try:
            out = export_human_reference()
        except CalibrationNotReady as exc:
            print(f"REFUSED: {exc}")
            return 1
        print(f"exported {out['counts']}")
        return 0

    eff = proposed_effective_config()
    print("=" * 76)
    print(f"  {CLASSIFICATION}  — PREPARED, NOTHING EXECUTED")
    print("=" * 76)
    print(f"\nscope            : {UNITS}  ({GUIDE_QUESTION} only)")
    print(f"schema version   : {SCHEMA_VERSION}")
    print(f"prompt sha256    : {prompt_sha()}")
    print(f"prompt purity    : {prompt_purity_problems() or 'clean'}")
    print(f"proposed config  : {canonical_config(eff)}")
    print(f"example key      : {cache_key('<unit_sha>', prompt_sha(), eff)[:32]}...")
    print(f"\nexpected cost    : exactly 7 extraction calls (one per unit), or one "
          f"Batch job of 7 requests.")
    print("NOT RUN. Approval required before any call.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
