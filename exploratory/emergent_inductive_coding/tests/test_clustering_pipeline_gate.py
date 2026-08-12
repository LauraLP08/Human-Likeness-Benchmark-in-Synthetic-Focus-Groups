"""
The downstream pipeline must refuse an incomplete or tampered clustering workbook.

A half-finished workbook that silently produced a presence matrix would yield a
saturation curve built on partial clustering — and nothing downstream would show it.
So the gate is checked in both directions here: it rejects every incomplete state,
AND it accepts a properly completed one.

Every test works on a COPY in tmp_path. The issued workbook is never modified.

No API calls.
"""

import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import partial_emergent_clustering_pipeline as pipe   # noqa: E402

WB = ROOT / "analysis" / "production_evaluation" / \
     "partial_emergent_clustering" / "Clustering_U01_U07.xlsx"


@pytest.fixture
def filled(tmp_path):
    """A plausibly completed workbook: one cluster per unique theme_label."""
    dst = tmp_path / "returned.xlsx"
    shutil.copy2(WB, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    assigned, n = {}, 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=ci["pooled_id"]).value
        label = str(ws.cell(row=r, column=ci["theme_label"]).value)[:20].lower()
        if label not in assigned:
            n += 1
            assigned[label] = f"C{n:02d}"
        ws.cell(row=r, column=ci["cluster_id"]).value = assigned[label]
        ws.cell(row=r, column=ci["cluster_label"]).value = f"cluster {assigned[label]}"
        # is_central is deliberately left blank: centrality was NOT_ASSESSED.
        _ = pid
    wb.save(dst)
    wb.close()
    return dst


def _multi_member_cluster(path):
    """A cluster_id with at least two rows — a singleton cannot be inconsistent."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    members = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=ci["cluster_id"]).value
        members.setdefault(cid, []).append(
            ws.cell(row=r, column=ci["pooled_id"]).value)
    wb.close()
    for cid, ids in members.items():
        if len(ids) >= 2:
            return cid, ids
    pytest.skip("fixture produced no multi-member cluster")


def _edit(path, pooled_id, column, value):
    wb = openpyxl.load_workbook(path)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=ci["pooled_id"]).value == pooled_id:
            ws.cell(row=r, column=ci[column]).value = value
            break
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# The gate rejects
# ---------------------------------------------------------------------------

def test_the_returned_workbook_is_ready():
    """The researcher's completed workbook passes the gate as returned."""
    assert pipe.validate(WB) == []


def test_a_workbook_with_no_cluster_ids_is_not_ready(tmp_path):
    dst = tmp_path / "blank.xlsx"
    shutil.copy2(WB, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Clustering"]
    ci = {h: i + 1 for i, h in enumerate([c.value for c in ws[1]])}
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=ci["cluster_id"]).value = None
        ws.cell(row=r, column=ci["cluster_label"]).value = None
    wb.save(dst); wb.close()
    problems = pipe.validate(dst)
    assert len(problems) == 152, f"expected 76 ids + 76 labels, got {len(problems)}"
    with pytest.raises(pipe.ClusteringNotReady):
        pipe.analyse(dst)


def _unused_issued_not_ready():
    problems = pipe.validate(WB)
    assert problems, "the unfilled issued workbook must not validate"
    assert any("cluster_id is empty" in p for p in problems)


def test_analyse_refuses_an_incomplete_workbook(tmp_path):
    dst = tmp_path / "incomplete.xlsx"
    shutil.copy2(WB, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Clustering"]
    ci = {h: i + 1 for i, h in enumerate([c.value for c in ws[1]])}
    ws.cell(row=2, column=ci["cluster_id"]).value = None
    wb.save(dst); wb.close()
    with pytest.raises(pipe.ClusteringNotReady):
        pipe.analyse(dst)


def test_missing_cluster_id_is_rejected(filled):
    _edit(filled, "P005", "cluster_id", None)
    assert any("P005: cluster_id is empty" in p for p in pipe.validate(filled))


def test_missing_is_central_is_NOT_rejected(filled):
    """
    Centrality was deliberately not assessed. A blank is a recorded decision, not an
    omission, so no row may fail for it.
    """
    _edit(filled, "P005", "is_central", None)
    problems = pipe.validate(filled)
    assert not any("is_central" in p for p in problems), problems
    assert problems == []


def test_centrality_is_absent_from_the_required_columns():
    assert "is_central" not in pipe.REQUIRED_OF_ADJUDICATOR
    assert "is_central" in pipe.OPTIONAL_OF_ADJUDICATOR
    assert pipe.CENTRALITY_STATUS == "NOT_ASSESSED"


def test_no_row_at_all_fails_for_a_blank_centrality(filled):
    """Every one of the 76 rows, not just a sampled one."""
    wb = openpyxl.load_workbook(filled)
    ws = wb["Clustering"]
    ci = {h: i + 1 for i, h in enumerate([c.value for c in ws[1]])}
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=ci["is_central"]).value = None
    wb.save(filled); wb.close()
    assert pipe.validate(filled) == []


def test_a_surviving_centrality_value_is_preserved_but_not_used(filled):
    _edit(filled, "P007", "is_central", "central")
    assert pipe.validate(filled) == [], "a centrality value must never gate"
    flags = pipe.review_flags(filled)
    assert any("preserved verbatim" in f and "NOT used analytically" in f
               for f in flags), flags


def test_inconsistent_cluster_label_within_a_unit_is_rejected(filled):
    # must be two rows of the SAME cluster in the SAME unit: consistency is required
    # within the (unit_id, cluster_id) pair, not across units.
    cid, unit, ids = _cluster_with_two_rows_in_one_unit(filled)
    _edit(filled, ids[0], "cluster_label", "a different name for the same cluster")
    problems = pipe.validate(filled)
    assert any("inconsistent labels" in p for p in problems), (
        f"editing {ids[0]} in multi-member cluster {cid} was not detected")


def _cluster_with_two_rows_in_one_unit(path):
    """A cluster_id that appears twice within the SAME unit."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    cells = {}
    for r in range(2, ws.max_row + 1):
        key = (ws.cell(row=r, column=ci["cluster_id"]).value,
               ws.cell(row=r, column=ci["unit_id"]).value)
        cells.setdefault(key, []).append(
            ws.cell(row=r, column=ci["pooled_id"]).value)
    wb.close()
    for (cid, unit), ids in cells.items():
        usable = [i for i in ids if i not in pipe.CENTRALITY_MISSING]
        if len(usable) >= 2:
            return cid, unit, usable
    pytest.skip("fixture produced no cluster with two usable rows in one unit")


def _cluster_in_two_units(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    byc = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=ci["cluster_id"]).value
        unit = ws.cell(row=r, column=ci["unit_id"]).value
        pid = ws.cell(row=r, column=ci["pooled_id"]).value
        if pid not in pipe.CENTRALITY_MISSING:
            byc.setdefault(cid, {}).setdefault(unit, []).append(pid)
    wb.close()
    for cid, units in byc.items():
        if len(units) >= 2:
            (u1, i1), (u2, i2) = list(units.items())[:2]
            return cid, (u1, i1[0]), (u2, i2[0])
    pytest.skip("fixture produced no cluster spanning two units")


def test_conflicting_centrality_within_a_cluster_AND_UNIT_is_flagged_not_gated(filled):
    """
    Centrality is NOT_ASSESSED, so it cannot block. A conflict between values that
    survive in the sheet is still surfaced for a human — flagged, never repaired.
    """
    cid, unit, ids = _cluster_with_two_rows_in_one_unit(filled)
    _edit(filled, ids[0], "is_central", "central")
    _edit(filled, ids[1], "is_central", "peripheral")
    assert pipe.validate(filled) == [], "centrality must never gate"
    flags = pipe.review_flags(filled)
    assert any("conflicting centrality" in f and cid in f and unit in f
               for f in flags), flags


def test_the_same_cluster_may_be_central_in_one_unit_and_peripheral_in_another(filled):
    """
    THE POSITIVE CASE. Centrality belongs to cluster x unit: a theme can carry weight
    in one discussion and not another. Flagging that as an inconsistency would erase a
    real finding.
    """
    cid, (u1, p1), (u2, p2) = _cluster_in_two_units(filled)
    _edit(filled, p1, "is_central", "central")
    _edit(filled, p2, "is_central", "peripheral")
    problems = pipe.validate(filled)
    assert problems == [], (
        f"cluster {cid} central in {u1} and peripheral in {u2} was rejected: "
        f"{problems}")


def test_centrality_level_is_declared(filled, monkeypatch, tmp_path):
    assert pipe.CENTRALITY_LEVEL == "cluster_id x unit_id"
    cid, (u1, p1), (u2, p2) = _cluster_in_two_units(filled)
    _edit(filled, p1, "is_central", "central")
    _edit(filled, p2, "is_central", "peripheral")
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    pm = {(r["cluster_id"], r["unit_id"]): r
          for r in pipe.analyse(filled)["presence_matrix"]}
    assert pm[(cid, u1)]["is_central_in_this_unit"] == "central"
    assert pm[(cid, u2)]["is_central_in_this_unit"] == "peripheral"
    assert all(r["centrality_level"] == "cluster_id x unit_id" for r in pm.values())


def test_the_same_cluster_id_in_two_units_may_carry_different_labels(filled):
    """
    IDENTITY IS (unit_id, cluster_id). A cluster_id text reused in another unit is a
    DIFFERENT analytic cluster, so a different label there is not a defect. It is
    surfaced as a review flag so a human can confirm the reuse was intended.
    """
    cid, (u1, p1), (u2, p2) = _cluster_in_two_units(filled)
    _edit(filled, p2, "cluster_label", "renamed in the second unit")
    assert pipe.validate(filled) == [], "cross-unit label divergence must not gate"
    flags = pipe.review_flags(filled)
    assert any(cid in f and "SEPARATE clusters" in f for f in flags), flags
    assert pipe.CLUSTER_IDENTITY == "(unit_id, cluster_id)"


def _unused_cluster_label_global(filled):
    cid, (u1, p1), (u2, p2) = _cluster_in_two_units(filled)
    _edit(filled, p2, "cluster_label", "renamed in the second unit")
    assert any("inconsistent labels across units" in p
               for p in pipe.validate(filled))


def test_cell_with_only_missing_centrality_is_reported_missing(filled, monkeypatch,
                                                               tmp_path):
    """P034/P040 must never be imputed from a sibling unit's value."""
    wb = openpyxl.load_workbook(filled)
    ws = wb["Clustering"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    # give P034 its own cluster so its U04 cell contains only a missing row
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=ci["pooled_id"]).value == "P034":
            ws.cell(row=r, column=ci["cluster_id"]).value = "C99"
            ws.cell(row=r, column=ci["cluster_label"]).value = "isolated"
            break
    wb.save(filled); wb.close()
    assert pipe.validate(filled) == []
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    pm = {(r["cluster_id"], r["unit_id"]): r
          for r in pipe.analyse(filled)["presence_matrix"]}
    cell = pm[("C99", "U04")]
    assert cell["is_central_in_this_unit"] == "NOT_ASSESSED", (
        "an unassessed cell must say so — never MISSING, peripheral, false or 0")
    assert cell["n_rows_without_centrality"] == 1


def test_cluster_definitions_is_generated_not_human_input(filled, monkeypatch,
                                                          tmp_path):
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    defs = pipe.analyse(filled)["cluster_definitions"]
    assert defs, "cluster_definitions must be produced automatically"
    assert all("is_central" not in r for r in defs), (
        "centrality must not appear at cluster level — it belongs to cluster x unit")
    for r in defs:
        assert r["cluster_id"] and r["cluster_label"]
        assert r["first_seen_unit"] in pipe.SHARED_UNITS
        assert r["centrality_note"] == pipe.CENTRALITY_NOT_AVAILABLE


def test_an_odd_centrality_value_does_not_gate(filled):
    """is_central is out of the gate entirely; it cannot block a valid workbook."""
    _edit(filled, "P007", "is_central", "quite important")
    assert pipe.validate(filled) == []


def test_tampering_with_an_issued_column_is_detected(filled):
    _edit(filled, "P010", "theme_label", "rewritten by the adjudicator")
    assert any("P010: issued content was modified" in p
               for p in pipe.validate(filled))


def test_deleting_a_row_is_detected(filled):
    wb = openpyxl.load_workbook(filled)
    ws = wb["Clustering"]
    ws.delete_rows(5)
    wb.save(filled); wb.close()
    problems = pipe.validate(filled)
    assert any("row count changed" in p or "pooled_id set changed" in p
               for p in problems)


# ---------------------------------------------------------------------------
# P034 / P040 — centrality stays MISSING
# ---------------------------------------------------------------------------

def test_centrality_missing_rows_are_not_required_to_be_filled(filled):
    problems = pipe.validate(filled)
    for pid in pipe.CENTRALITY_MISSING:
        assert not any(f"{pid}: is_central is empty" in p for p in problems)


@pytest.mark.parametrize("pid", ["P034", "P040"])
def test_p034_and_p040_are_treated_exactly_like_every_other_row(filled, pid):
    """
    They once had a special exemption. Centrality is no longer assessed for ANY row,
    so no rule may single them out — in either direction.
    """
    assert pipe.validate(filled) == []
    _edit(filled, pid, "is_central", "central")
    assert pipe.validate(filled) == [], f"{pid} must not be special-cased"
    _edit(filled, pid, "is_central", None)
    assert pipe.validate(filled) == []


def test_no_gate_message_anywhere_mentions_centrality(filled):
    _edit(filled, "P005", "cluster_id", None)
    problems = pipe.validate(filled)
    assert problems, "the gate must still fire on a real defect"
    assert not any("central" in p.lower() for p in problems)


# ---------------------------------------------------------------------------
# The gate accepts, and the products are well formed
# ---------------------------------------------------------------------------

def test_a_properly_completed_workbook_validates(filled):
    problems = pipe.validate(filled)
    assert problems == [], f"a complete workbook was rejected: {problems}"


def test_products_are_produced_only_after_validation(filled, monkeypatch, tmp_path):
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    products = pipe.analyse(filled)
    assert set(products) == {"cluster_definitions", "presence_matrix",
                             "shared_and_exclusive", "cumulative_curve",
                             "codebook_comparison_scaffold"}
    assert len(products["cumulative_curve"]) == 7
    assert [r["unit_id"] for r in products["cumulative_curve"]] == pipe.SHARED_UNITS
    total = products["cumulative_curve"][-1]["cumulative_clusters"]
    assert total == len(products["shared_and_exclusive"])
    assert all(r["status"] in ("shared", "exclusive")
               for r in products["shared_and_exclusive"])
    assert all(r["closest_codebook_subtheme"] == ""
               for r in products["codebook_comparison_scaffold"]), (
        "the codebook comparison must be left for a human, done last")


def test_cumulative_curve_is_monotonic(filled, monkeypatch, tmp_path):
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    curve = pipe.analyse(filled)["cumulative_curve"]
    cum = [r["cumulative_clusters"] for r in curve]
    assert cum == sorted(cum)
    assert sum(r["n_new_clusters"] for r in curve) == cum[-1]


def test_scope_is_enforced(filled, monkeypatch, tmp_path):
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    products = pipe.analyse(filled)
    units = {r["unit_id"] for r in products["presence_matrix"]}
    assert units <= set(pipe.SHARED_UNITS)
    assert "U08" not in units
    assert not units & set(pipe.NOT_REVIEWED)


def test_no_agreement_or_saturation_figure_is_emitted(filled, monkeypatch, tmp_path):
    """Counts and a curve, yes. A reliability or saturation claim, no."""
    monkeypatch.setattr(pipe, "_DIR", tmp_path)
    products = pipe.analyse(filled)
    blob = str(products).lower()
    for forbidden in ("kappa", "alpha", "agreement_score", "saturated",
                      "saturation_reached", "codebook_coverage"):
        assert forbidden not in blob, f"{forbidden!r} must not be computed here"


# ---------------------------------------------------------------------------
# AMENDMENT 01 — P057 cluster_id C13 -> C16
# ---------------------------------------------------------------------------
#
# The researcher confirmed C13 in U06 was an identifier error. These guard the
# corrected state and the three quantities that must not be conflated.

def _returned_rows():
    return pipe._read(WB)


def test_amendment_01_is_recorded():
    import json
    rec = json.loads((pipe._DIR / "AMENDMENT_01_P057_cluster_id.json")
                     .read_text(encoding="utf-8"))
    assert rec["pooled_id"] == "P057" and rec["unit_id"] == "U06"
    assert rec["field"] == "cluster_id"
    assert (rec["previous_value"], rec["new_value"]) == ("C13", "C16")
    assert rec["cells_changed"] == 1
    assert rec["sha256_before"] != rec["sha256_after"]
    assert (pipe._DIR / rec["pre_amendment_archive"]).exists(), (
        "the pre-amendment evidence copy must be preserved")


def test_p057_now_carries_c16_with_its_original_label():
    r = next(x for x in _returned_rows() if x["pooled_id"] == "P057")
    assert r["unit_id"] == "U06"
    assert r["cluster_id"] == "C16"
    assert r["cluster_label"].strip().startswith("WOMEN ARE MORE JUDGED")


def test_u07_c13_was_not_touched():
    rows = [r for r in _returned_rows() if r["cluster_id"] == "C13"]
    assert {r["unit_id"] for r in rows} == {"U07"}
    assert {r["cluster_label"].strip() for r in rows} == {
        "OPT TO IGNORE SOCIAL EXPECTATIONS"}
    assert {r["pooled_id"] for r in rows} == {"P065", "P068", "P071"}


def test_every_cluster_id_now_carries_exactly_one_label():
    lab = {}
    for r in _returned_rows():
        lab.setdefault(r["cluster_id"], set()).add((r["cluster_label"] or "").strip())
    bad = {k: v for k, v in lab.items() if len(v) > 1}
    assert bad == {}, bad
    assert not pipe.review_flags(), "the C13 reuse flag must be withdrawn"


def test_the_three_quantities_are_distinct_and_correct():
    """
    16 thematic categories, 44 theme x unit instances, 76 original coding rows.
    Conflating them would put the wrong number in the recall denominator.
    """
    rows = _returned_rows()
    assert len(rows) == 76
    assert len({(r["unit_id"], r["cluster_id"]) for r in rows}) == 44
    assert len({r["cluster_id"] for r in rows}) == 16
