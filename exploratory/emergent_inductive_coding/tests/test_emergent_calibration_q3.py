"""
Offline guards for PRIMARY_EMERGENT_AUTOMATION_CALIBRATION_Q3.

Nothing here calls an API, and nothing opens the active clustering workbook for
writing. The workbook is read only through the module's own gate, which refuses
while it is mid-edit.

No API calls.
"""

import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import emergent_calibration_q3 as cal   # noqa: E402

OUT = ROOT / "analysis" / "production_evaluation"
CLUSTERING = OUT / "partial_emergent_clustering" / "Clustering_U01_U07.xlsx"
MATCHING = OUT / "emergent_calibration_q3" / "Emergent_Matching_Q3.xlsx"


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()

def _tokens(text: str) -> set:
    """Word tokens, so a bare-letter id like "D" is matched as a token."""
    return set(re.findall("[a-z0-9.]+", str(text).lower()))



def _flat(text: str) -> str:
    """Collapse whitespace so a line-wrapped phrase still matches."""
    return " ".join(str(text).split()).lower()


# ---------------------------------------------------------------------------
# Prompt purity
# ---------------------------------------------------------------------------

def test_prompt_contains_no_codebook_terms():
    assert cal.prompt_purity_problems() == []


def test_prompt_purity_check_actually_fires():
    """The guard must catch a planted term, not merely pass on clean text."""
    assert cal.prompt_purity_problems("mention of subtheme A.1 here")
    assert cal.prompt_purity_problems("this discussion is about gender and food")
    assert "codebook" in cal.prompt_purity_problems("consult the codebook")


def test_prompt_names_no_subtheme_ids_or_labels():
    from thematic_coding import load_codebook
    low = _flat(cal.EXTRACTION_SYSTEM_PROMPT)
    for c in load_codebook():
        sid = c["subtheme_id"]
        # "D" is a bare letter; a substring test on it matches every word with a d.
        # Tokenise instead of using a regex — no escape can be mangled in transit.
        assert sid.lower() not in _tokens(low), sid
        assert _flat(c["theme"]) not in low
    assert "4n" not in low


def test_prompt_carries_no_provenance_or_results():
    low = _flat(cal.EXTRACTION_SYSTEM_PROMPT)
    for leak in ("fg1", "fg2", "fg3", "fg4", "fg5", "enriched", "demographics",
                 "synthetic", "human transcript", "macho", "tier 1", "recall",
                 "precision", "u01", "q3"):
        assert leak not in low, f"{leak!r} leaked into the prompt"


def test_prompt_does_not_name_the_discussion_topic():
    """Naming the topic would seed the categories the model must discover."""
    low = _flat(cal.EXTRACTION_SYSTEM_PROMPT)
    for topic in ("food", "eat", "meat", "diet", "plant-based", "gender"):
        assert topic not in low


def test_prompt_states_the_theme_definition_rules():
    low = _flat(cal.EXTRACTION_SYSTEM_PROMPT)
    assert "not merely a topic" in low
    assert "merge formulations" in low
    assert "keep opposing positions separate" in low
    assert "verbatim" in low
    assert "never quote the moderator" in low
    assert "central" in low and "secondary" in low


def test_no_minimum_or_maximum_theme_count_is_imposed():
    low = _flat(cal.EXTRACTION_SYSTEM_PROMPT)
    assert "there is no target number" in low
    assert "at least" not in low.split("evidence")[0]
    for n in ("at most", "no more than", "maximum of"):
        assert n not in low, f"a cap ({n!r}) would create a ceiling artefact"


# ---------------------------------------------------------------------------
# Scope and configuration
# ---------------------------------------------------------------------------

def test_scope_is_exactly_u01_to_u07():
    assert cal.UNITS == ["U01", "U02", "U03", "U04", "U05", "U06", "U07"]
    assert cal.GUIDE_QUESTION == "Q3"
    assert "S01" not in cal.UNITS


def test_cache_key_omits_the_codebook_and_says_so():
    doc = _flat(cal.cache_key.__doc__)
    assert "there is no codebook" in doc
    eff = cal.proposed_effective_config()
    a = cal.cache_key("sha-unit", cal.prompt_sha(), eff)
    assert a != cal.cache_key("sha-other", cal.prompt_sha(), eff)
    assert a != cal.cache_key("sha-unit", "other-prompt", eff)
    assert a != cal.cache_key("sha-unit", cal.prompt_sha(),
                              dict(eff, max_output_tokens=8192))
    assert a != cal.cache_key("sha-unit", cal.prompt_sha(),
                              dict(eff, model="gemini-2.5-flash"))
    assert a != cal.cache_key("sha-unit", cal.prompt_sha(),
                              dict(eff, execution_mode="synchronous"))


def test_effective_config_records_every_transmitted_parameter():
    eff = cal.proposed_effective_config()
    for k in ("model", "execution_mode", "response_mime_type", "max_output_tokens",
              "temperature_transmitted", "thinking_config_transmitted",
              "schema_version"):
        assert k in eff
    assert eff["temperature_transmitted"] is False
    assert eff["thinking_config_transmitted"] is False


def test_model_choice_is_flagged_as_unvalidated_for_open_extraction():
    doc = _flat(cal.__doc__)
    assert "not automatically validated" in doc
    assert "different tasks" in doc
    assert "does not transfer" in doc


# ---------------------------------------------------------------------------
# Extraction output validation
# ---------------------------------------------------------------------------

UNIT_LINES = [
    "[T001] Moderator: What do you all make of that?",
    "[T002] Participant 1: I would buy it more often if it were cheaper, honestly.",
    "[T003] Participant 2: I disagree, price is not what stops me at all.",
]
CLEAN_TELEMETRY = {"finish_reasons": ["FinishReason.STOP"]}


def _theme(tid="M1", turn="T002", quote="I would buy it more often if it were cheaper",
           speaker="Participant 1", relevance="central"):
    return {"machine_theme_id": tid, "label": "cost as a barrier",
            "one_sentence_description": "Price is described as what limits the choice.",
            "relevance": relevance,
            "evidence": [{"turn_id": turn, "speaker": speaker, "quote": quote}],
            "voiced_by": [speaker]}


def test_a_clean_extraction_passes():
    v = cal.validate_extraction("U01", {"themes": [_theme()]},
                                CLEAN_TELEMETRY, UNIT_LINES)
    assert v["status"] == "COMPLETE" and v["problems"] == []
    assert v["n_themes"] == 1


def test_moderator_quote_is_rejected():
    t = _theme(turn="T001", speaker="Moderator",
               quote="What do you all make of that?")
    v = cal.validate_extraction("U01", {"themes": [t]}, CLEAN_TELEMETRY, UNIT_LINES)
    assert v["status"] == "QUARANTINE"
    assert any("quotes the moderator" in p for p in v["problems"])


def test_invented_quote_is_rejected():
    t = _theme(quote="a sentence that appears nowhere in this unit")
    v = cal.validate_extraction("U01", {"themes": [t]}, CLEAN_TELEMETRY, UNIT_LINES)
    assert any("not verbatim" in p for p in v["problems"])


def test_quote_attributed_to_the_wrong_turn_is_rejected():
    t = _theme(turn="T003")          # quote belongs to T002
    v = cal.validate_extraction("U01", {"themes": [t]}, CLEAN_TELEMETRY, UNIT_LINES)
    assert any("not verbatim" in p for p in v["problems"])


def test_unknown_turn_id_is_rejected():
    t = _theme(turn="T999")
    v = cal.validate_extraction("U01", {"themes": [t]}, CLEAN_TELEMETRY, UNIT_LINES)
    assert any("not in this unit" in p for p in v["problems"])


def test_theme_without_evidence_is_rejected():
    t = _theme()
    t["evidence"] = []
    v = cal.validate_extraction("U01", {"themes": [t]}, CLEAN_TELEMETRY, UNIT_LINES)
    assert any("no evidence" in p for p in v["problems"])


def test_duplicate_theme_ids_are_rejected():
    v = cal.validate_extraction("U01", {"themes": [_theme("M1"), _theme("M1")]},
                                CLEAN_TELEMETRY, UNIT_LINES)
    assert any("duplicate machine_theme_id" in p for p in v["problems"])


def test_incomplete_schema_is_quarantined():
    v = cal.validate_extraction("U01", {}, CLEAN_TELEMETRY, UNIT_LINES)
    assert v["status"] == "QUARANTINE"
    assert any("schema invalid" in p for p in v["problems"])


def test_truncated_output_is_quarantined():
    v = cal.validate_extraction("U01", {"themes": [_theme()]},
                                {"finish_reasons": ["FinishReason.MAX_TOKENS"]},
                                UNIT_LINES)
    assert v["status"] == "QUARANTINE"
    assert any("expected STOP" in p for p in v["problems"])


def test_bad_relevance_value_is_rejected():
    v = cal.validate_extraction("U01", {"themes": [_theme(relevance="quite")]},
                                CLEAN_TELEMETRY, UNIT_LINES)
    assert any("relevance is" in p for p in v["problems"])


def test_matching_is_refused_until_all_seven_units_are_complete():
    ok = {u: {"status": "COMPLETE"} for u in cal.UNITS[:6]}
    with pytest.raises(cal.CalibrationNotReady) as e:
        cal.assert_corpus_complete(ok)
    assert "U07" in str(e.value)

    ok["U07"] = {"status": "QUARANTINE", "problems": ["bad quote"]}
    with pytest.raises(cal.CalibrationNotReady):
        cal.assert_corpus_complete(ok)

    ok["U07"] = {"status": "COMPLETE"}
    cal.assert_corpus_complete(ok)


def test_a_unit_outside_the_scope_is_rejected():
    res = {u: {"status": "COMPLETE"} for u in cal.UNITS}
    res["S01"] = {"status": "COMPLETE"}
    with pytest.raises(cal.CalibrationNotReady) as e:
        cal.assert_corpus_complete(res)
    assert "unexpected unit: S01" in str(e.value)


# ---------------------------------------------------------------------------
# Human reference views
# ---------------------------------------------------------------------------

# NOTE: this file previously asserted that the export REFUSES because
# Clustering_U01_U07.xlsx was mid-edit. The researcher has since completed it and the
# workbook now passes its gate, so that assertion described a transient state, not a
# rule. The rule is kept by test_export_refuses_a_workbook_that_is_not_ready, which
# builds its own not-ready copy in tmp_path instead of depending on live human work.

def test_export_does_not_modify_the_active_workbook(tmp_path):
    """The export reads the human workbook; it must never write to it."""
    before = _sha(CLUSTERING)
    cal.export_human_reference(out_dir=tmp_path)
    assert _sha(CLUSTERING) == before
    assert (tmp_path / "human_reference_q3.json").exists()
    assert not list(tmp_path.glob("*.tmp")), "atomic write left a temp file"


def test_export_refuses_a_workbook_that_is_not_ready(tmp_path):
    import shutil
    import openpyxl as _o
    dst = tmp_path / "notready.xlsx"
    shutil.copy2(CLUSTERING, dst)
    wb = _o.load_workbook(dst)
    ws = wb["Clustering"]
    ci = {h: i + 1 for i, h in enumerate([c.value for c in ws[1]])}
    ws.cell(row=2, column=ci["cluster_id"]).value = None
    wb.save(dst); wb.close()
    out = tmp_path / "out"
    with pytest.raises(cal.CalibrationNotReady):
        cal.export_human_reference(clustering_workbook=dst, out_dir=out)
    assert not out.exists() or not list(out.iterdir())


def test_the_views_are_union_plus_the_two_coder_views():
    assert cal.REFERENCE_VIEWS == ("union_reference", "coder_a_view", "coder_b_view")
    assert cal.PRIMARY_VIEW == "union_reference"
    assert cal.COVERAGE_REFERENCE == "union_reference"


def test_there_is_no_central_reference():
    """
    Centrality was not assessed. A central_reference must not exist, and its absence
    must never read as "a reference containing zero central themes".
    """
    assert "central_reference" not in cal.REFERENCE_VIEWS
    assert "central_reference" in cal.FORBIDDEN_VIEWS
    assert cal.CENTRALITY_STATUS == "NOT_ASSESSED"
    ref = json.loads((OUT / "emergent_calibration_q3" /
                      "human_reference_q3.json").read_text(encoding="utf-8"))
    assert ref["central_reference"] == "NOT_AVAILABLE — CENTRALITY_NOT_ASSESSED"
    assert not isinstance(ref["central_reference"], list), (
        "an empty list would be read as zero central themes")
    assert ref["centrality_status"] == "NOT_ASSESSED"
    for rec in ref["union_reference"]:
        assert rec["centrality_status"] == "NOT_ASSESSED"


def test_the_coder_views_are_subsets_of_the_same_union():
    ref = json.loads((OUT / "emergent_calibration_q3" /
                      "human_reference_q3.json").read_text(encoding="utf-8"))
    union = {r["human_key"] for r in ref["union_reference"]}
    for view in ("coder_a_view", "coder_b_view"):
        keys = {r["human_key"] for r in ref[view]}
        assert keys <= union, f"{view} is not a subset of the union"
    # every union cluster came from at least one coder
    assert ({r["human_key"] for r in ref["coder_a_view"]} |
            {r["human_key"] for r in ref["coder_b_view"]}) == union
    for coder, d in ref["coder_recall_vs_union"].items():
        assert d["denominator"] == len(union)
        assert 0 < d["numerator"] <= d["denominator"]


def test_human_keys_are_unit_qualified():
    ref = json.loads((OUT / "emergent_calibration_q3" /
                      "human_reference_q3.json").read_text(encoding="utf-8"))
    assert ref["cluster_identity"] == "(unit_id, cluster_id)"
    for r in ref["union_reference"]:
        assert r["human_key"] == f"{r['unit_id']}::{r['cluster_id']}"
    keys = [r["human_key"] for r in ref["union_reference"]]
    assert len(keys) == len(set(keys))
    # the same cluster_id text genuinely recurs across units in this material
    ids = [r["cluster_id"] for r in ref["union_reference"]]
    assert len(set(ids)) < len(ids), (
        "if ids were unique this test would not be exercising the risk")


# ---------------------------------------------------------------------------
# Matching workbook
# ---------------------------------------------------------------------------

def test_matching_workbook_is_empty_and_scoped():
    if not MATCHING.exists():
        pytest.skip("matching workbook not built")
    wb = openpyxl.load_workbook(MATCHING, read_only=True, data_only=True)
    assert wb.sheetnames == ["Instructions", "Matching", "Coverage", "Scope"]
    rows = [r for r in wb["Matching"].iter_rows(min_row=2, values_only=True)
            if r[1]]
    blob = " ".join(str(c.value or "") for s in wb.sheetnames
                    for row in wb[s].iter_rows() for c in row).lower()
    wb.close()
    assert rows == [], "the matching sheet must ship empty"
    assert "do not consult the codebook" in _flat(blob)
    # The Scope sheet names S01-S06 deliberately, to state that the supplementary
    # sample is NEVER mixed in. That is a separation statement, not a leak — so the
    # check is that no supplementary UNIT appears as data on the Matching sheet.
    wb2 = openpyxl.load_workbook(MATCHING, read_only=True, data_only=True)
    match_blob = " ".join(str(c.value or "")
                          for row in wb2["Matching"].iter_rows() for c in row).lower()
    scope_blob = " ".join(str(c.value or "")
                          for row in wb2["Scope"].iter_rows() for c in row)
    wb2.close()
    assert not re.search(r"S0[1-6]", match_blob, re.I), (
        "a supplementary unit appears as matching data")
    assert "never mixed" in _flat(scope_blob)


def test_all_four_relations_plus_no_match_are_representable():
    for rel in ("one_to_one", "one_to_many", "many_to_one",
                "no_match_human_only", "no_match_machine_only"):
        assert rel in cal.MATCH_RELATIONS


def test_metrics_blocked_until_every_cluster_and_theme_is_decided():
    human, machine = {"U01::C1", "U01::C2"}, {"U01::M1", "U01::M2"}
    rows = [{"unit_id": "U01", "human_cluster_id": "C1", "machine_theme_id": "M1",
             "relation": "one_to_one", "decision": "match", "reasoning": ""}]
    with pytest.raises(cal.CalibrationNotReady) as e:
        cal.assert_matching_complete(rows, human, machine)
    msg = str(e.value)
    assert "U01::C2 has no decision" in msg and "U01::M2 has no decision" in msg


def test_bare_ids_are_refused_as_identities():
    """A bare id is not an identity and must not be accepted as one."""
    rows = [{"unit_id": "U01", "human_cluster_id": "C1", "machine_theme_id": "M1",
             "relation": "one_to_one", "decision": "match", "reasoning": ""}]
    with pytest.raises(cal.CalibrationNotReady) as e:
        cal.assert_matching_complete(rows, {"C1"}, {"M1"})
    assert "not (unit_id, id)" in str(e.value)


def test_many_to_one_requires_reasoning():
    rows = [{"unit_id": "U01", "human_cluster_id": "C1", "machine_theme_id": "M1",
             "relation": "many_to_one", "decision": "match", "reasoning": ""},
            {"unit_id": "U01", "human_cluster_id": "C2", "machine_theme_id": "M1",
             "relation": "many_to_one", "decision": "match", "reasoning": ""}]
    H, M = {"U01::C1", "U01::C2"}, {"U01::M1"}
    with pytest.raises(cal.CalibrationNotReady) as e:
        cal.assert_matching_complete(rows, H, M)
    assert "requires reasoning" in str(e.value)

    for r in rows:
        r["reasoning"] = "both human clusters describe the same claim"
    cal.assert_matching_complete(rows, H, M)


def test_a_complete_matching_passes():
    rows = [
        {"unit_id": "U01", "human_cluster_id": "C1", "machine_theme_id": "M1",
         "relation": "one_to_one", "decision": "match", "reasoning": ""},
        {"unit_id": "U01", "human_cluster_id": "C2", "machine_theme_id": "",
         "relation": "no_match_human_only", "decision": "no_match", "reasoning": ""},
        {"unit_id": "U01", "human_cluster_id": "", "machine_theme_id": "M2",
         "relation": "no_match_machine_only", "decision": "no_match", "reasoning": ""},
    ]
    cal.assert_matching_complete(rows, {"U01::C1", "U01::C2"},
                                 {"U01::M1", "U01::M2"})


# ---------------------------------------------------------------------------
# THE SAME ID IN TWO UNITS IS TWO DIFFERENT THEMES
# ---------------------------------------------------------------------------

def _two_unit_inventory():
    """M01 and C01 exist in BOTH U01 and U02, denoting different themes."""
    return ({"U01::C01", "U02::C01"}, {"U01::M01", "U02::M01"})


def test_adjudicating_M01_in_one_unit_does_not_complete_M01_in_the_other():
    """
    The defect this closes: keyed on a bare "M01", a decision recorded in U01 would
    satisfy the completeness check for U02's entirely different M01, and the second
    unit would never be adjudicated at all.
    """
    H, M = _two_unit_inventory()
    rows = [{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
             "relation": "one_to_one", "decision": "match", "reasoning": ""}]

    problems = cal.validate_matching(rows, H, M)
    assert "human cluster U02::C01 has no decision" in problems
    assert "machine theme U02::M01 has no decision" in problems
    assert not any("U01::M01 has no decision" in p for p in problems)
    with pytest.raises(cal.CalibrationNotReady):
        cal.assert_matching_complete(rows, H, M)

    # completing U02 leaves U01's decision untouched and closes the gate
    rows.append({"unit_id": "U02", "human_cluster_id": "C01",
                 "machine_theme_id": "M01", "relation": "no_match_human_only",
                 "decision": "no_match", "reasoning": ""})
    rows[-1]["machine_theme_id"] = ""
    rows.append({"unit_id": "U02", "human_cluster_id": "", "machine_theme_id": "M01",
                 "relation": "no_match_machine_only", "decision": "no_match",
                 "reasoning": ""})
    assert cal.validate_matching(rows, H, M) == []
    cal.assert_matching_complete(rows, H, M)

    # and the two units reached OPPOSITE decisions for the same id text
    assert rows[0]["decision"] == "match" and rows[1]["decision"] == "no_match"


def test_a_decision_in_one_unit_does_not_leak_as_a_conflict_in_another():
    """U01::M01 matched and U02::M01 not matched is legitimate, not a contradiction."""
    H, M = _two_unit_inventory()
    rows = [
        {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
         "relation": "one_to_one", "decision": "match", "reasoning": ""},
        {"unit_id": "U02", "human_cluster_id": "C01", "machine_theme_id": "",
         "relation": "no_match_human_only", "decision": "no_match", "reasoning": ""},
        {"unit_id": "U02", "human_cluster_id": "", "machine_theme_id": "M01",
         "relation": "no_match_machine_only", "decision": "no_match", "reasoning": ""},
    ]
    problems = cal.validate_matching(rows, H, M)
    assert not any("BOTH matched and not matched" in p for p in problems), problems
    assert problems == []


def test_relating_themes_across_units_is_rejected():
    H, M = {"U01::C01"}, {"U01::M01", "U02::M07"}
    rows = [{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M07",
             "relation": "one_to_one", "decision": "match", "reasoning": ""}]
    problems = cal.validate_matching(rows, H, M)
    assert any("cross-unit relation" in p and "U02" in p for p in problems), problems


@pytest.mark.parametrize("rows,needle", [
    ([{"unit_id": "U01", "human_cluster_id": "C99", "machine_theme_id": "M01",
       "relation": "one_to_one", "decision": "match"}], "unknown human cluster key"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "one_to_one", "decision": "match"},
      {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "one_to_one", "decision": "match"}], "duplicate pairing"),
    ([{"unit_id": "U01", "relation": "one_to_one", "decision": "match"}],
     "orphan row"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "decision": "match"}], "with no relation"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "one_to_one"}], "with no decision"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "one_to_one", "decision": "no_match"}], "conflicts with decision"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "one_to_many", "decision": "match", "reasoning": "x"}],
     "declared one_to_many"),
    ([{"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
       "relation": "many_to_one", "decision": "match", "reasoning": "x"}],
     "declared many_to_one"),
])
def test_every_structural_defect_is_detected(rows, needle):
    H, M = {"U01::C01"}, {"U01::M01"}
    problems = cal.validate_matching(rows, H, M)
    assert any(needle in p for p in problems), (needle, problems)


def test_an_entity_cannot_be_matched_and_unmatched_at_once():
    H, M = {"U01::C01"}, {"U01::M01"}
    rows = [
        {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
         "relation": "one_to_one", "decision": "match", "reasoning": ""},
        {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "",
         "relation": "no_match_human_only", "decision": "no_match", "reasoning": ""},
    ]
    problems = cal.validate_matching(rows, H, M)
    assert any("U01::C01: marked BOTH matched and not matched" in p
               for p in problems), problems


def test_one_to_many_and_many_to_one_are_representable():
    H = {"U01::C01", "U01::C02"}
    M = {"U01::M01", "U01::M02"}
    rows = [  # one human cluster split across two machine themes
        {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M01",
         "relation": "one_to_many", "decision": "match", "reasoning": "split a"},
        {"unit_id": "U01", "human_cluster_id": "C01", "machine_theme_id": "M02",
         "relation": "one_to_many", "decision": "match", "reasoning": "split b"},
        {"unit_id": "U01", "human_cluster_id": "C02", "machine_theme_id": "M02",
         "relation": "many_to_one", "decision": "match", "reasoning": "fused"},
    ]
    assert cal.validate_matching(rows, H, M) == []


# ---------------------------------------------------------------------------
# Nothing was executed
# ---------------------------------------------------------------------------

def test_extraction_has_run_and_every_unit_is_complete():
    """
    The extractor ran under explicit authorisation: ONE batch job, seven requests.
    All seven must be COMPLETE; a quarantined unit may not be silently dropped.
    """
    d = OUT / "emergent_calibration_q3"
    res = json.loads((d / "extraction_results_q3.json").read_text(encoding="utf-8"))
    assert res["n_units"] == 7
    assert [r["unit_id"] for r in res["results"]] == cal.UNITS
    for r in res["results"]:
        assert r["status"] == "COMPLETE", (r["unit_id"], r["problems"])
        assert r["n_themes"] > 0
        assert r["problems"] == []


def test_the_run_used_the_frozen_prompt_and_config():
    d = OUT / "emergent_calibration_q3"
    res = json.loads((d / "extraction_results_q3.json").read_text(encoding="utf-8"))
    assert res["prompt_sha256"] == cal.prompt_sha()
    assert res["response_schema_sha256"] == cal.response_schema_sha()
    cfg = res["effective_request_config"]
    assert cfg["model"] == "gemini-3.5-flash"
    assert cfg["execution_mode"] == "batch"
    assert cfg["max_output_tokens"] == 16384
    assert cfg["temperature_transmitted"] is False
    assert cfg["thinking_config_transmitted"] is False
    assert cfg["response_schema_transmitted"] is True


def test_no_output_was_truncated():
    d = OUT / "emergent_calibration_q3"
    raw = json.loads((d / "batch_raw_responses_q3.json").read_text(encoding="utf-8"))
    for r in raw["responses"]:
        assert r["error"] is None, r["unit_id"]
        assert r["telemetry"]["finish_reasons"] == ["FinishReason.STOP"], r["unit_id"]
        assert r["telemetry"]["candidates_token_count"] < 16384


def test_every_quotation_resolves_inside_its_own_cited_turn():
    """
    Independent re-check, not trusting validate_extraction: each quote must appear in
    the turn it is attributed to, and never come from the moderator.
    """
    import re as _re
    d = OUT / "emergent_calibration_q3"
    res = json.loads((d / "extraction_results_q3.json").read_text(encoding="utf-8"))
    norm = lambda t: " ".join(str(t).split())
    n_ev = 0
    for r in res["results"]:
        turns = {}
        for ln in cal.unit_lines(r["unit_id"]):
            m = _re.match(r"^\[(T\d+)\]\s+([^:]+):\s*(.*)$", ln, _re.S)
            turns[m.group(1)] = (m.group(2).strip(), m.group(3).strip())
        for th in r["themes"]:
            assert th["evidence"], th["machine_theme_id"]
            for e in th["evidence"]:
                n_ev += 1
                assert e["turn_id"] in turns, e
                speaker, body = turns[e["turn_id"]]
                assert not speaker.lower().startswith("moderator"), e
                assert norm(e["quote"]) in norm(body), (r["unit_id"], e["turn_id"])
    assert n_ev == 58, f"expected 58 quotations, found {n_ev}"


def test_machine_theme_ids_are_unique_within_each_unit():
    d = OUT / "emergent_calibration_q3"
    res = json.loads((d / "extraction_results_q3.json").read_text(encoding="utf-8"))
    for r in res["results"]:
        ids = [t["machine_theme_id"] for t in r["themes"]]
        assert len(ids) == len(set(ids)), r["unit_id"]


def test_matching_and_metrics_have_NOT_run():
    """The next prohibition: matching and the final decision are not yet permitted."""
    d = OUT / "emergent_calibration_q3"
    assert not list(d.glob("*matching_results*"))
    assert not list(d.glob("*metrics*"))
    assert not list(d.glob("*calibration_decision*"))
    # NOTE: the live Emergent_Matching_Q3_POPULATED.xlsx is NOT opened here. The
    # researcher is adjudicating in it, and a test that reads a file under active
    # human editing is both a hazard and meaningless. The builder is exercised on a
    # fresh copy instead.


def test_the_populated_workbook_carries_both_inventories_and_the_frozen_rules(tmp_path):
    """
    Built fresh in tmp_path. The live workbook is never opened: it is under active
    human editing and must not be read, copied or re-sealed.
    """
    d = OUT / "emergent_calibration_q3"
    if not (d / "extraction_results_q3.json").exists():
        pytest.skip("extraction results not present")
    wb = cal.build_populated_matching_workbook(
        out_path=tmp_path / "Emergent_Matching_Q3_POPULATED.xlsx")
    book = openpyxl.load_workbook(wb, read_only=True, data_only=True)
    assert book.sheetnames == ["Rules", "Matching", "Machine_Only_Adjudication",
                               "Fragmentation_Fusion", "Coverage", "Scope"]
    rows = [r for r in book["Matching"].iter_rows(min_row=2, values_only=True) if r[0]]
    rules = " ".join(str(c.value or "") for row in book["Rules"].iter_rows()
                     for c in row)
    book.close()
    assert sum(1 for r in rows if r[1] == "human") == 44
    assert sum(1 for r in rows if r[1] == "machine") == 30
    for k in cal.MACHINE_ONLY_VERDICTS:
        assert k in rules, f"{k} missing from the frozen rules sheet"
    assert "at least two different units" in rules
    assert "0.6364" in rules


def test_module_makes_no_api_call():
    src = (ROOT / "scripts" / "emergent_calibration_q3.py").read_text(encoding="utf-8")
    for banned in ("genai.Client", "generate_content", "batches.create",
                   "import google"):
        assert banned not in src, f"{banned!r} present — this module must not call out"


# ---------------------------------------------------------------------------
# FROZEN ADJUDICATION RULES (§6.4 = Alternative 1, qualitative gate)
# ---------------------------------------------------------------------------

def test_the_four_machine_only_verdicts_are_frozen():
    assert cal.MACHINE_ONLY_VERDICTS == (
        "VALID_NOVEL_THEME", "UNSUPPORTED_OR_SPURIOUS",
        "DUPLICATE_MACHINE_THEME", "UNCERTAIN")
    for v in cal.MACHINE_ONLY_VERDICTS:
        assert v in cal.ADJUDICATION_RULES, f"{v} has no frozen definition"


def test_each_rule_carries_its_approved_definition():
    r = cal.ADJUDICATION_RULES
    assert "not sufficient textual evidence" in r["UNSUPPORTED_OR_SPURIOUS"]
    assert "contradicts the text" in r["SEVERE_UNSUPPORTED_THEME"]
    assert "incidental mention" in r["SEVERE_UNSUPPORTED_THEME"]
    assert "at least two different units" in r["RECURRENT_UNSUPPORTED_PATTERN"]
    assert "NOT as a statistical test" in r["RECURRENT_UNSUPPORTED_PATTERN"]
    assert "within one unit" in r["DUPLICATE_MACHINE_THEME"]
    assert "more than one reasonable reading" in r["UNCERTAIN"]
    assert "clearly supported by the text" in r["VALID_NOVEL_THEME"]


def test_uncertain_is_neither_correct_nor_error():
    assert "neither correct nor incorrect" in cal.ADJUDICATION_RULES["UNCERTAIN"]


def test_every_decision_must_retain_quote_unit_category_and_justification():
    for f in ("unit_id", "machine_theme_id", "verdict", "quote", "human_justification"):
        assert f in cal.ADJUDICATION_REQUIRED_FIELDS


def test_recall_alone_cannot_produce_a_pass():
    """The benchmark is necessary, not sufficient. Three further conditions apply."""
    assert len(cal.PASS_CONDITIONS) == 4
    joined = " ".join(cal.PASS_CONDITIONS).lower()
    assert "0.6364" in joined
    assert "no recurrent severe unsupported errors" in joined
    assert "complete adjudication of every machine-only theme" in joined
    assert "fragmentation and fusion" in joined
    assert abs(cal.COVERAGE_BENCHMARK - 28 / 44) < 1e-12


def test_the_four_final_states_are_frozen():
    assert cal.FINAL_STATES == (
        "PASS_WITH_SAMPLED_HUMAN_VERIFICATION",
        "BORDERLINE — FALL_BACK_TO_ASSISTIVE_REVIEW",
        "FAIL — FALL_BACK_TO_ASSISTIVE_REVIEW",
        "UNRESOLVED_AT_THIS_SAMPLE_SIZE")


# ---------------------------------------------------------------------------
# The transmitted unit text
# ---------------------------------------------------------------------------

def test_all_seven_units_load_and_carry_no_provenance():
    for u in cal.UNITS:
        assert cal.unit_text_problems(u) == []
        lines = cal.unit_lines(u)
        assert lines, f"{u} produced no turns"
        for ln in lines:
            assert ln.startswith("[T"), ln[:40]


def test_the_unit_banner_is_stripped():
    """"UNIT U01" is provenance and must not reach the model."""
    for u in cal.UNITS:
        t = cal.unit_text(u)
        assert "UNIT U" not in t
        assert "=====" not in t
        assert u not in t


def test_unit_text_matches_the_excerpt_the_coders_saw():
    """The model must read exactly what the humans read, or the comparison is void."""
    import openpyxl as _o
    pkg = OUT / "gold_standard_package" / "Coder_A_Part1_Emergent.xlsx"
    if not pkg.exists():
        pytest.skip("package not present")
    wb = _o.load_workbook(pkg, read_only=True, data_only=True)
    seen = {}
    for r in wb["Units"].iter_rows(min_row=2, values_only=True):
        if r[0] in cal.UNITS:
            seen.setdefault(r[0], []).append(f"[{r[1]}] {r[2]}: {r[4]}")
    wb.close()
    for u in cal.UNITS:
        assert cal.unit_lines(u) == seen[u], f"{u} differs from the issued excerpt"
