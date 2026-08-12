"""
Raw-row import, sealing, and the hardened consolidation freeze.

TWO ACTIVE HUMAN WORKBOOKS ARE OFF LIMITS while people are editing them:
`Clustering_U01_U07.xlsx` and `Transportability_Emergent_SingleCoder.xlsx`. Nothing
here opens either for writing. The non-mutation test below hashes them at runtime,
runs the constructors with every output redirected to tmp_path, and re-hashes —
rather than pinning a hash that would fail the moment a researcher saves legitimate
work.

Every fixture builds its own synthetic coder workbook via pkg.build(). No real human
return is copied, read or used as a template. No real human return is
imported.

No API calls.
"""

import hashlib
import json
import pathlib
import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import build_transportability_consolidation as con   # noqa: E402
import build_transportability_package as pkg         # noqa: E402
import build_transportability_sample as smp          # noqa: E402

OUT = ROOT / "analysis" / "production_evaluation"
ISSUED = OUT / "transportability_sample" / "Transportability_Emergent_SingleCoder.xlsx"

ACTIVE_HUMAN_WORKBOOKS = [
    OUT / "partial_emergent_clustering" / "Clustering_U01_U07.xlsx",
    OUT / "transportability_sample" / "Transportability_Emergent_SingleCoder.xlsx",
    OUT / "open_coding_adjudication" / "OCA-001_adjudication.xlsx",
]


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


# ---------------------------------------------------------------------------
# Non-mutation, measured at runtime — never a pinned hash
# ---------------------------------------------------------------------------

def test_constructors_do_not_mutate_the_active_human_workbooks(tmp_path, monkeypatch):
    """
    A pinned hash would fail as soon as a researcher saved authorised work, which
    makes it a trap rather than a guard. Hash immediately before, run the builders
    with outputs redirected, hash immediately after.
    """
    present = [p for p in ACTIVE_HUMAN_WORKBOOKS if p.exists()]
    assert present, "no active workbook found to protect"
    before = {p: _sha(p) for p in present}

    sample_dir = tmp_path / "sample"
    sealed_dir = tmp_path / "sealed"
    sample_dir.mkdir()
    sealed_dir.mkdir()
    monkeypatch.setattr(smp, "_DIR", sample_dir)
    monkeypatch.setattr(smp, "_SEALED_DIR", sealed_dir)
    monkeypatch.setattr(pkg, "_DIR", sample_dir)
    monkeypatch.setattr(pkg, "_WB", sample_dir / "coder.xlsx")
    monkeypatch.setattr(pkg, "_SEAL", sealed_dir / "pkg_seal.json")
    monkeypatch.setattr(con, "_DIR", sample_dir)
    monkeypatch.setattr(con, "_WB", sample_dir / "consolidation.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", sealed_dir / "raw_seal.json")

    assert smp.main() == 0
    pkg.build()
    con.build()

    after = {p: _sha(p) for p in present}
    changed = [p.name for p in present if before[p] != after[p]]
    assert not changed, f"constructors mutated active human workbooks: {changed}"


def test_the_builders_wrote_only_into_the_redirected_workspace(tmp_path, monkeypatch):
    sample_dir, sealed_dir = tmp_path / "s", tmp_path / "z"
    sample_dir.mkdir(); sealed_dir.mkdir()
    monkeypatch.setattr(pkg, "_DIR", sample_dir)
    monkeypatch.setattr(pkg, "_WB", sample_dir / "coder.xlsx")
    monkeypatch.setattr(pkg, "_SEAL", sealed_dir / "pkg_seal.json")
    monkeypatch.setattr(con, "_DIR", sample_dir)
    monkeypatch.setattr(con, "_WB", sample_dir / "consolidation.xlsx")
    shutil.copy2(OUT / "transportability_sample" / "_units_for_packaging.json",
                 sample_dir / "_units_for_packaging.json")
    pkg.build()
    con.build()
    assert (sample_dir / "coder.xlsx").exists()
    assert (sample_dir / "consolidation.xlsx").exists()
    assert (sealed_dir / "pkg_seal.json").exists()


# ---------------------------------------------------------------------------
# A synthetic returned coder workbook
# ---------------------------------------------------------------------------

@pytest.fixture
def workspace(tmp_path, monkeypatch):
    """Isolated consolidation workspace with an imported, sealed raw-row set."""
    sample_dir, sealed_dir = tmp_path / "sample", tmp_path / "sealed"
    sample_dir.mkdir(); sealed_dir.mkdir()
    monkeypatch.setattr(con, "_DIR", sample_dir)
    monkeypatch.setattr(con, "_WB", sample_dir / "consolidation.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", sealed_dir / "raw_seal.json")

    # A returned coder workbook: two themes in S01 (to be merged), one elsewhere.
    #
    # This is BUILT, not copied from the live coder file. It used to do
    # shutil.copy2(ISSUED, ...), which took the real returned workbook as a template
    # and overwrote only the first slot or two per unit. Once the coder actually
    # completed the sample, every row she had filled survived into the fixture, and
    # those rows failed the return gate — 37 tests errored at setup. A file that
    # legitimately changes cannot be a fixture template.
    monkeypatch.setattr(pkg, "_DIR", sample_dir)
    monkeypatch.setattr(pkg, "_WB", sample_dir / "coder.xlsx")
    monkeypatch.setattr(pkg, "_SEAL", sealed_dir / "pkg_seal.json")
    shutil.copy2(OUT / "transportability_sample" / "_units_for_packaging.json",
                 sample_dir / "_units_for_packaging.json")
    pkg.build()
    returned = tmp_path / "returned.xlsx"
    shutil.copy2(sample_dir / "coder.xlsx", returned)
    wb = openpyxl.load_workbook(returned)
    units = {}
    for r in wb["Units"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            units.setdefault(r[0], []).append(str(r[3]))
    ws = wb["Emergent_Coding"]
    filled = {}
    for row in range(2, ws.max_row + 1):
        uid = ws.cell(row=row, column=1).value
        slot = ws.cell(row=row, column=2).value
        if not uid:
            continue
        n = filled.get(uid, 0)
        want = 2 if uid == "S01" else 1          # S01 gets two rows to consolidate
        if n >= want:
            continue
        filled[uid] = n + 1
        ws.cell(row=row, column=3).value = f"{uid} theme {slot}"
        ws.cell(row=row, column=4).value = "One sentence."
        ws.cell(row=row, column=5).value = " ".join(units[uid][0].split()[:8])
        ws.cell(row=row, column=6).value = "central"
    wb.save(returned)
    wb.close()

    con.build()
    seal = con.import_raw_rows(returned)
    return {"dir": sample_dir, "wb": con._WB, "seal": seal,
            "seal_path": con._RAW_SEAL, "returned": returned}


def _consolidate(workspace, mapping, themes):
    """Write a consolidation state: mapping = {source_row_id: theme_id}."""
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    ws.protection.sheet = False
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=2).value
        if sid in mapping:
            ws.cell(row=r, column=7).value = mapping[sid]
    ws2 = wb["Consolidated_Themes"]
    ws2.delete_rows(2, ws2.max_row)
    for i, t in enumerate(themes, start=2):
        for j, v in enumerate(t, start=1):
            ws2.cell(row=i, column=j, value=v)
    wb.save(workspace["wb"])
    wb.close()


def _valid_state(workspace):
    """The canonical valid consolidation for the fixture's raw rows."""
    rows = workspace["seal"]["rows"]
    by_unit = {}
    for sid, meta in rows.items():
        by_unit.setdefault(meta["blind_unit_id"], []).append(sid)
    mapping, themes = {}, []
    for uid, sids in sorted(by_unit.items()):
        tid = f"{uid}_T1"                 # ids are unit-prefixed by rule
        for sid in sids:
            mapping[sid] = tid
        themes.append((uid, tid, f"{uid} consolidated", "One sentence.",
                       "central", "|".join(sorted(sids)), ""))
    return mapping, themes


# ---------------------------------------------------------------------------
# Import and seal
# ---------------------------------------------------------------------------

def test_import_requires_a_workbook_that_passed_the_return_gate(tmp_path, monkeypatch):
    """
    A freshly issued (empty) workbook has no complete theme, so import must refuse.

    It builds its own empty workbook rather than pointing at the live coder file: that
    file's gate status legitimately changes as the coder works, and a test whose
    outcome depends on someone else's editing state proves nothing.
    """
    sample_dir, sealed_dir = tmp_path / "s", tmp_path / "z"
    sample_dir.mkdir(); sealed_dir.mkdir()
    monkeypatch.setattr(pkg, "_DIR", sample_dir)
    monkeypatch.setattr(pkg, "_WB", sample_dir / "coder.xlsx")
    monkeypatch.setattr(pkg, "_SEAL", sealed_dir / "pkg_seal.json")
    monkeypatch.setattr(con, "_WB", tmp_path / "c.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", tmp_path / "seal.json")
    monkeypatch.setattr(con, "_DIR", tmp_path)
    shutil.copy2(OUT / "transportability_sample" / "_units_for_packaging.json",
                 sample_dir / "_units_for_packaging.json")
    pkg.build()
    con.build()
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.import_raw_rows(sample_dir / "coder.xlsx")   # empty: no themes yet
    assert "return gate" in str(e.value)


def test_no_test_uses_the_live_coder_workbook_as_a_template():
    """
    Guard against the defect that broke 37 tests: a live human return being copied and
    treated as a blank template. The path constant may exist; it must never be READ.

    Matches call sites, not mentions — an earlier version of this guard searched for
    the bare token and so failed on its own source.
    """
    import re as _re
    src = pathlib.Path(__file__).read_text(encoding="utf-8")
    readers = _re.compile(
        r"(copy2|copyfile|copy|load_workbook|import_raw_rows|read_bytes|read_text|"
        r"open)\s*\(\s*ISSUED\b")
    hits = [ln.strip() for ln in src.splitlines()
            if readers.search(ln) and not ln.strip().startswith("#")]
    assert hits == [], f"the live coder workbook is being read: {hits}"

def test_import_assigns_stable_unique_ids(workspace):
    ids = list(workspace["seal"]["rows"])
    assert len(ids) == len(set(ids))
    assert all(i.startswith("S") and ("_slot_" in i or "_ovf_" in i) for i in ids)
    for sid, meta in workspace["seal"]["rows"].items():
        assert sid.startswith(meta["blind_unit_id"]), (
            "the id must encode its unit so a row cannot be reassigned")


def test_import_is_deterministic(workspace, tmp_path, monkeypatch):
    again = con.import_raw_rows(workspace["returned"])
    assert set(again["rows"]) == set(workspace["seal"]["rows"])
    for sid in again["rows"]:
        assert again["rows"][sid]["content_sha256"] == \
            workspace["seal"]["rows"][sid]["content_sha256"]


def test_seal_records_content_hashes_and_column_protection(workspace):
    s = workspace["seal"]
    assert s["protected_columns"] == list(con.RAW_COLS)
    assert s["editable_columns"] == list(con.EDITABLE_COLS)
    assert all(len(v["content_sha256"]) == 64 for v in s["rows"].values())
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    assert ws.protection.sheet is True
    assert ws.cell(row=2, column=3).protection.locked is True     # raw label
    assert ws.cell(row=2, column=7).protection.locked is False    # theme id
    wb.close()


def test_no_raw_rows_are_invented_before_the_return(tmp_path, monkeypatch):
    monkeypatch.setattr(con, "_DIR", tmp_path)
    monkeypatch.setattr(con, "_WB", tmp_path / "c.xlsx")
    con.build()
    wb = openpyxl.load_workbook(tmp_path / "c.xlsx", read_only=True, data_only=True)
    rows = [r for r in wb["Raw_To_Consolidated"].iter_rows(min_row=2, values_only=True)
            if r[1]]
    wb.close()
    assert rows == []


# ---------------------------------------------------------------------------
# Freeze — positive
# ---------------------------------------------------------------------------

def test_a_valid_consolidation_freezes(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"])
    assert out["n_consolidated_themes"] == 6
    assert out["calibration_grade"] is False
    assert out["is_inter_coder_agreement"] is False
    assert {t["blind_unit_id"] for t in out["human_reference_themes"]} == set(con.UNITS)


def test_two_raw_rows_merged_count_once(workspace):
    """S01 has two raw rows; consolidated they are ONE theme in the denominator."""
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"])
    s01 = [t for t in out["human_reference_themes"] if t["blind_unit_id"] == "S01"]
    assert len(s01) == 1, "two raw rows must consolidate to one theme"
    assert len(s01[0]["source_row_ids"]) == 2, "both raw rows stay recorded"
    assert out["n_raw_rows"] > out["n_consolidated_themes"]
    assert "NEVER a denominator" in out["denominator_rule"]


# ---------------------------------------------------------------------------
# Freeze — negative
# ---------------------------------------------------------------------------

def _expect(workspace, fragment):
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    assert fragment in str(e.value), f"expected {fragment!r} in:\n{e.value}"


def test_rejects_a_deleted_raw_row(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    wb = openpyxl.load_workbook(workspace["wb"])
    wb["Raw_To_Consolidated"].delete_rows(2)
    wb.save(workspace["wb"]); wb.close()
    _expect(workspace, "raw rows missing from the mapping")


def test_rejects_an_added_raw_row(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="S01")
    ws.cell(row=r, column=2, value="S01_slot_99")
    ws.cell(row=r, column=7, value="T1")
    wb.save(workspace["wb"]); wb.close()
    _expect(workspace, "raw rows not present in the seal")


def test_rejects_a_duplicated_raw_row(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    first_id = ws.cell(row=2, column=2).value
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=ws.cell(row=2, column=1).value)
    ws.cell(row=r, column=2, value=first_id)
    ws.cell(row=r, column=7, value="T1")
    wb.save(workspace["wb"]); wb.close()
    _expect(workspace, "duplicated raw rows in the mapping")


def test_rejects_an_unassigned_raw_row(workspace):
    mapping, themes = _valid_state(workspace)
    first = sorted(mapping)[0]
    del mapping[first]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "no consolidated_theme_id")


def test_rejects_a_nonexistent_source_row_id(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [(u, t, l, d, c, src + "|S01_slot_99", n)
              if u == "S01" else (u, t, l, d, c, src, n)
              for (u, t, l, d, c, src, n) in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "does not exist")


def test_rejects_a_source_row_claimed_twice(workspace):
    mapping, themes = _valid_state(workspace)
    s01 = next(t for t in themes if t[0] == "S01")
    sids = s01[5].split("|")
    themes = [t for t in themes if t[0] != "S01"] + [
        ("S01", "S01_T1", "a", "One sentence.", "central", "|".join(sids), ""),
        ("S01", "S01_T2", "b", "One sentence.", "central", sids[0], "")]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "is claimed by")


def test_rejects_mapping_and_declared_ids_that_disagree(workspace):
    mapping, themes = _valid_state(workspace)
    s01 = next(t for t in themes if t[0] == "S01")
    sids = s01[5].split("|")
    themes = [t if t[0] != "S01" else
              ("S01", "S01_T1", "a", "One sentence.", "central", sids[0], "")
              for t in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "disagree with the mapping")


def test_rejects_a_duplicate_theme_id_within_a_unit(workspace):
    mapping, themes = _valid_state(workspace)
    s01 = next(t for t in themes if t[0] == "S01")
    sids = s01[5].split("|")
    for sid in sids:
        mapping[sid] = "S01_T1"
    themes = [t for t in themes if t[0] != "S01"] + [
        ("S01", "S01_T1", "a", "One sentence.", "central", sids[0], ""),
        ("S01", "S01_T1", "b", "One sentence.", "central", sids[1], "")]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "appears more than once")


def test_rejects_a_theme_id_that_is_not_unit_prefixed(workspace):
    """Unit-prefixed ids are what make 'never crosses units' checkable."""
    mapping, themes = _valid_state(workspace)
    s01 = [sid for sid, m in workspace["seal"]["rows"].items()
           if m["blind_unit_id"] == "S01"]
    for sid in s01:
        mapping[sid] = "T1"
    themes = [t if t[0] != "S01" else
              ("S01", "T1", "a", "One sentence.", "central", "|".join(sorted(s01)), "")
              for t in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "not unit-prefixed")


def test_rejects_a_theme_claiming_a_row_from_another_unit(workspace):
    mapping, themes = _valid_state(workspace)
    other = next(sid for sid, m in workspace["seal"]["rows"].items()
                 if m["blind_unit_id"] == "S02")
    themes = [(u, t, l, d, c, (src + "|" + other) if u == "S01" else src, n)
              for (u, t, l, d, c, src, n) in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "belongs to")


def test_the_same_local_number_in_different_units_is_fine(workspace):
    """S01_T1 and S02_T1 are different themes and must both be allowed."""
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"])
    ids = [t["consolidated_theme_id"] for t in out["human_reference_themes"]]
    assert "S01_T1" in ids and "S02_T1" in ids
    assert len(set(ids)) == len(ids), "unit-prefixed ids must be globally unique"


def test_rejects_invalid_centrality(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [(u, t, l, d, "quite central" if u == "S01" else c, src, n)
              for (u, t, l, d, c, src, n) in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "expected one of")


def test_rejects_an_unknown_unit_rather_than_ignoring_it(workspace):
    mapping, themes = _valid_state(workspace)
    themes = themes + [("S99", "S99_T1", "a", "One sentence.", "central",
                        "S99_slot_01", "")]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "is not a known unit")


def test_rejects_a_theme_with_no_raw_rows(workspace):
    mapping, themes = _valid_state(workspace)
    themes = themes + [("S01", "S01_T9", "empty", "One sentence.", "central", " ", "")]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "partially completed")


def test_rejects_a_partially_completed_theme(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [(u, t, l, "", c, src, n) if u == "S01" else (u, t, l, d, c, src, n)
              for (u, t, l, d, c, src, n) in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "partially completed")


def test_rejects_a_unit_with_no_consolidated_theme(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [t for t in themes if t[0] != "S06"]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "S06: no consolidated theme")


def test_refuses_without_a_seal(workspace, tmp_path):
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], tmp_path / "absent.json")
    assert "no raw-row seal" in str(e.value)


# ---------------------------------------------------------------------------
# Nothing real was created
# ---------------------------------------------------------------------------

def test_no_real_human_reference_was_written():
    assert not (OUT / "transportability_sample" /
                "human_reference_themes.json").exists()


def test_no_real_raw_seal_was_written():
    assert not (OUT / "gold_standard_sealed" /
                "transportability_raw_rows_seal.json").exists()


# ---------------------------------------------------------------------------
# Content-hash verification — sheet protection is not a guarantee
# ---------------------------------------------------------------------------

def _edit_raw(workspace, column_name, value, row=2):
    """Edit an imported column directly, as any xlsx writer could."""
    col = con.RAW_COLS.index(column_name) + 1
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    ws.protection.sheet = False          # exactly what an attacker or a slip does
    sid = ws.cell(row=row, column=2).value
    ws.cell(row=row, column=col).value = value
    wb.save(workspace["wb"])
    wb.close()
    return sid


@pytest.mark.parametrize("column,value", [
    ("raw_theme_label", "a label the coder never wrote"),
    ("raw_theme_description", "a description the coder never wrote"),
    ("raw_supporting_quote", "a quote the coder never wrote"),
    ("raw_relevance", "secondary"),
])
def test_rejects_any_edit_to_an_imported_column(workspace, column, value):
    """
    Each imported column is covered by the content hash. Excel sheet protection can
    be switched off; the hash is what actually holds.
    """
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    sid = _edit_raw(workspace, column, value)
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    msg = str(e.value)
    assert "raw content was modified after import" in msg
    assert sid in msg, "the offending source_row_id must be named"
    assert "expected" in msg and "observed" in msg


def test_unedited_rows_pass_the_hash_check(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"])
    assert out["n_consolidated_themes"] == 6


def test_the_hash_covers_exactly_the_imported_columns(workspace):
    """Editing an EDITABLE column must not trip the content hash."""
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    wb = openpyxl.load_workbook(workspace["wb"])
    ws = wb["Raw_To_Consolidated"]
    ws.protection.sheet = False
    ws.cell(row=2, column=8).value = "a consolidator note"      # editable column
    wb.save(workspace["wb"])
    wb.close()
    con.freeze_reference(workspace["wb"], workspace["seal_path"])


# ---------------------------------------------------------------------------
# Orphan assignments and three-way set equality
# ---------------------------------------------------------------------------

def test_rejects_a_mapping_row_pointing_at_a_nonexistent_theme(workspace):
    """
    The hole this closes: a mapping key with no Consolidated_Themes row never entered
    the comparison, so the raw row silently vanished from the denominator.
    """
    mapping, themes = _valid_state(workspace)
    s01 = sorted(sid for sid, m in workspace["seal"]["rows"].items()
                 if m["blind_unit_id"] == "S01")
    mapping[s01[0]] = "S01_T7"                     # no such theme row
    themes = [t if t[0] != "S01" else
              ("S01", "S01_T1", "a", "One sentence.", "central",
               "|".join(s01[1:]), "")
              for t in themes]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "no row in Consolidated_Themes")


def test_rejects_a_consolidated_theme_with_no_mapping_rows(workspace):
    mapping, themes = _valid_state(workspace)
    s01 = sorted(sid for sid, m in workspace["seal"]["rows"].items()
                 if m["blind_unit_id"] == "S01")
    themes = themes + [("S01", "S01_T5", "unmapped", "One sentence.", "central",
                        s01[0], "")]
    _consolidate(workspace, mapping, themes)
    _expect(workspace, "no raw row assigned in the mapping")


def test_three_way_set_equality_is_enforced(workspace):
    mapping, themes = _valid_state(workspace)
    s01 = sorted(sid for sid, m in workspace["seal"]["rows"].items()
                 if m["blind_unit_id"] == "S01")
    # drop one row from the declared ids only: sealed == assigned != declared
    themes = [t if t[0] != "S01" else
              ("S01", "S01_T1", "a", "One sentence.", "central", s01[0], "")
              for t in themes]
    _consolidate(workspace, mapping, themes)
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    msg = str(e.value)
    assert ("three raw-row sets are not identical" in msg
            or "disagree with the mapping" in msg)


def test_every_sealed_row_appears_exactly_once_in_the_final_reference(workspace):
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"])
    declared = [sid for t in out["human_reference_themes"]
                for sid in t["source_row_ids"]]
    assert sorted(declared) == sorted(workspace["seal"]["rows"])
    assert len(declared) == len(set(declared)), "a raw row appears twice"


# ---------------------------------------------------------------------------
# Atomicity and the overwrite policy
# ---------------------------------------------------------------------------

def test_a_failed_freeze_does_not_touch_an_existing_reference(workspace):
    """
    POLICY: an existing frozen reference is never overwritten by a failing run, and
    is refused even by a passing one unless force=True. A frozen reference fixes the
    recall denominator; replacing it silently would let the denominator move after
    results existed.
    """
    sentinel = workspace["dir"] / "human_reference_themes.json"
    sentinel.write_text('{"sentinel": "do not clobber"}', encoding="utf-8")
    before = sentinel.read_text(encoding="utf-8")

    mapping, themes = _valid_state(workspace)
    themes = [t for t in themes if t[0] != "S06"]          # guaranteed failure
    _consolidate(workspace, mapping, themes)
    with pytest.raises(con.ConsolidationNotReady):
        con.freeze_reference(workspace["wb"], workspace["seal_path"], force=True)

    assert sentinel.read_text(encoding="utf-8") == before
    assert not list(workspace["dir"].glob("*.tmp")), "a partial temp file was left"


def test_an_existing_reference_is_refused_without_force(workspace):
    sentinel = workspace["dir"] / "human_reference_themes.json"
    sentinel.write_text('{"sentinel": "already frozen"}', encoding="utf-8")
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    assert "already exists" in str(e.value)
    assert sentinel.read_text(encoding="utf-8") == '{"sentinel": "already frozen"}'


def test_force_replaces_it_and_leaves_no_temp_file(workspace):
    sentinel = workspace["dir"] / "human_reference_themes.json"
    sentinel.write_text('{"sentinel": "old"}', encoding="utf-8")
    mapping, themes = _valid_state(workspace)
    _consolidate(workspace, mapping, themes)
    out = con.freeze_reference(workspace["wb"], workspace["seal_path"], force=True)
    assert json.loads(sentinel.read_text(encoding="utf-8"))["n_consolidated_themes"] == 6
    assert out["n_consolidated_themes"] == 6
    assert not list(workspace["dir"].glob("*.tmp"))


def test_a_failed_freeze_writes_nothing_at_all(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [t for t in themes if t[0] != "S03"]
    _consolidate(workspace, mapping, themes)
    with pytest.raises(con.ConsolidationNotReady):
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    assert not (workspace["dir"] / "human_reference_themes.json").exists()
    assert not list(workspace["dir"].glob("*.tmp"))


def test_the_error_enumerates_every_problem(workspace):
    mapping, themes = _valid_state(workspace)
    themes = [t for t in themes if t[0] not in ("S05", "S06")]
    _consolidate(workspace, mapping, themes)
    with pytest.raises(con.ConsolidationNotReady) as e:
        con.freeze_reference(workspace["wb"], workspace["seal_path"])
    msg = str(e.value)
    assert "S05" in msg and "S06" in msg
    assert msg.count("  - ") >= 2, "problems must be enumerated, not summarised"


# ---------------------------------------------------------------------------
# Overflow_Themes obeys the same completeness rule as Emergent_Coding
# ---------------------------------------------------------------------------
#
# The defect: this branch still demanded ("theme_label", "theme_description",
# "supporting_quote", "relevance") after relevance became NOT_ASSESSED by decision, so
# every overflow theme a coder wrote would have been silently skipped — not rejected
# with a message, just dropped from the raw-row set and therefore from the denominator.

def _overflow_workbook(tmp_path, monkeypatch, rows):
    """An issued workbook with one complete slot per unit plus given overflow rows."""
    sample, sealed = tmp_path / "s", tmp_path / "z"
    sample.mkdir(exist_ok=True); sealed.mkdir(exist_ok=True)
    monkeypatch.setattr(pkg, "_DIR", sample)
    monkeypatch.setattr(pkg, "_WB", sample / "coder.xlsx")
    monkeypatch.setattr(pkg, "_SEAL", sealed / "pkg_seal.json")
    shutil.copy2(OUT / "transportability_sample" / "_units_for_packaging.json",
                 sample / "_units_for_packaging.json")
    pkg.build()

    returned = tmp_path / "returned_ovf.xlsx"
    shutil.copy2(sample / "coder.xlsx", returned)
    wb = openpyxl.load_workbook(returned)
    units = {}
    for r in wb["Units"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            units.setdefault(r[0], []).append(str(r[3]))

    ws = wb["Emergent_Coding"]
    done = set()
    for row in range(2, ws.max_row + 1):
        uid = ws.cell(row=row, column=1).value
        if not uid or uid in done:
            continue
        done.add(uid)
        ws.cell(row=row, column=3).value = f"{uid} theme"
        ws.cell(row=row, column=4).value = "One sentence."
        ws.cell(row=row, column=5).value = " ".join(units[uid][0].split()[:8])
        # relevance deliberately left blank: NOT_ASSESSED

    ovf = wb["Overflow_Themes"]
    for i, (uid, label, desc, quote, rel) in enumerate(rows, start=2):
        ovf.cell(row=i, column=1).value = uid
        ovf.cell(row=i, column=2).value = label
        ovf.cell(row=i, column=3).value = desc
        ovf.cell(row=i, column=4).value = (
            " ".join(units[uid][0].split()[:10]) if quote == "REAL" else quote)
        ovf.cell(row=i, column=5).value = rel
    wb.save(returned)
    wb.close()
    return returned


def _raw_rows_written(consolidation_wb):
    """Raw_To_Consolidated as written. import_raw_rows returns a SEAL, whose `rows`
    is a dict of source_row_id -> {blind_unit_id, content_sha256}; the imported values
    live in the workbook."""
    wb = openpyxl.load_workbook(consolidation_wb, read_only=True, data_only=True)
    ws = wb["Raw_To_Consolidated"]
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)
            if r[1]]
    wb.close()
    return rows


def test_a_complete_overflow_row_with_empty_relevance_is_imported(tmp_path, monkeypatch):
    """THE FIX. Substantively complete overflow must be imported, not skipped."""
    returned = _overflow_workbook(
        tmp_path, monkeypatch,
        [("S01", "extra overflow theme", "One sentence.", "REAL", None)])
    assert pkg.validate(returned) == []

    monkeypatch.setattr(con, "_DIR", tmp_path)
    monkeypatch.setattr(con, "_WB", tmp_path / "consolidation.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", tmp_path / "raw_seal.json")
    con.build()
    out = con.import_raw_rows(returned)

    assert "S01_ovf_01" in out["rows"], (
        f"the overflow row was dropped from the seal: {sorted(out['rows'])}")

    written = _raw_rows_written(con._WB)
    ovf = [r for r in written if "_ovf_" in str(r["source_row_id"])]
    assert len(ovf) == 1, f"the overflow row was dropped: {written}"
    assert ovf[0]["blind_unit_id"] == "S01"
    assert ovf[0]["source_row_id"] == "S01_ovf_01"
    assert ovf[0]["raw_theme_label"] == "extra overflow theme"
    assert ovf[0]["raw_relevance"] == pkg.RELEVANCE_STATUS == "NOT_ASSESSED"

    # and it is treated exactly like a slot row
    slot = [r for r in written if r["source_row_id"] == "S01_slot_01"][0]
    assert slot["raw_relevance"] == ovf[0]["raw_relevance"]


@pytest.mark.parametrize("label,desc,quote,missing", [
    (None, "One sentence.", "REAL", "theme_label"),
    ("a label", None, "REAL", "theme_description"),
    ("a label", "One sentence.", None, "supporting_quote"),
])
def test_an_incomplete_overflow_row_does_not_enter_the_raw_rows(
        tmp_path, monkeypatch, label, desc, quote, missing):
    """
    Current policy: an overflow row missing a REQUIRED field is not imported. The
    return gate is what reports it — the importer must not invent a partial theme.
    """
    returned = _overflow_workbook(tmp_path, monkeypatch,
                                  [("S01", label, desc, quote, None)])
    problems = pkg.validate(returned)
    assert any("partially completed" in p and missing in p for p in problems), problems

    monkeypatch.setattr(con, "_DIR", tmp_path)
    monkeypatch.setattr(con, "_WB", tmp_path / "consolidation.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", tmp_path / "raw_seal.json")
    con.build()
    # the gate blocks import outright, so the partial row can never be imported
    with pytest.raises(con.ConsolidationNotReady):
        con.import_raw_rows(returned)


def test_a_strange_overflow_relevance_is_preserved_and_flagged_not_used(
        tmp_path, monkeypatch):
    """A surviving value is kept verbatim, surfaced as a flag, never used."""
    returned = _overflow_workbook(
        tmp_path, monkeypatch,
        [("S01", "overflow theme", "One sentence.", "REAL", "quite important")])
    assert pkg.validate(returned) == [], "an odd relevance must not gate"

    flags = pkg.review_flags(returned)
    assert any("quite important" in f for f in flags), flags
    assert any("NOT used analytically" in f for f in flags), flags

    monkeypatch.setattr(con, "_DIR", tmp_path)
    monkeypatch.setattr(con, "_WB", tmp_path / "consolidation.xlsx")
    monkeypatch.setattr(con, "_RAW_SEAL", tmp_path / "raw_seal.json")
    con.build()
    con.import_raw_rows(returned)
    written = _raw_rows_written(con._WB)
    ovf = [r for r in written if "_ovf_" in str(r["source_row_id"])]
    assert len(ovf) == 1
    assert ovf[0]["raw_relevance"] == "quite important", (
        "a surviving value must be preserved verbatim, not overwritten with "
        "NOT_ASSESSED")


def test_both_import_branches_use_the_same_required_fields():
    """Guard against the two branches drifting apart again."""
    import inspect
    import re as _re
    src = inspect.getsource(con.import_raw_rows)
    filters = _re.findall(r"if not all\(r\.get\(f\) for f in ([^)]*)\)", src)
    assert len(filters) == 2, filters
    assert all("REQUIRED_THEME_FIELDS" in f for f in filters), filters
    assert "relevance" not in " ".join(filters)
