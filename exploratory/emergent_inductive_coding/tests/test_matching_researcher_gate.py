"""
The researcher-facing matching workbook (V2) and its return gate.

Three defects found by direct validation of V1 are held closed here:

  1. evidence was truncated at 180/200 characters, mid-sentence — every text is now
     compared CHARACTER BY CHARACTER against human_reference_q3.json and
     extraction_results_q3.json;
  2. the immutable columns were unsealed — each one has its own tamper test;
  3. keys named on an UNCERTAIN row were treated as matches — they are now proved not to
     enter confirmed_links, not to move recall, and not to clear the adjudication queue.

No API calls. Every fixture builds its own workbook in tmp_path; no live workbook is
opened.
"""

import json
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import emergent_matching_researcher as mr   # noqa: E402
import emergent_calibration_q3 as cal       # noqa: E402

OUT = ROOT / "analysis" / "production_evaluation" / "emergent_calibration_q3"

pytestmark = pytest.mark.skipif(
    not (OUT / "extraction_results_q3.json").exists(),
    reason="extraction results not present",
)


@pytest.fixture()
def book(tmp_path):
    return mr.build(tmp_path / "Emergent_Matching_Q3_RESEARCHER_V2.xlsx")


@pytest.fixture(scope="module")
def sources():
    ref = json.loads((OUT / "human_reference_q3.json").read_text(encoding="utf-8"))
    res = json.loads((OUT / "extraction_results_q3.json").read_text(encoding="utf-8"))
    return ref, res


def _cols(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


def _fill(path, fn):
    wb = openpyxl.load_workbook(path)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=col["unit_id"]).value:
            continue
        rec = {h: ws.cell(row=r, column=i).value for h, i in col.items()}
        dec, keys, why = fn(rec)
        ws.cell(row=r, column=col["human_decision"]).value = dec
        ws.cell(row=r, column=col["matched_machine_keys"]).value = keys
        ws.cell(row=r, column=col["researcher_reasoning"]).value = why
    wb.save(path)
    wb.close()
    return path


def _edit(path, human_key, column, value):
    wb = openpyxl.load_workbook(path)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=col["human_key"]).value == human_key:
            ws.cell(row=r, column=col[column]).value = value
            break
    wb.save(path)
    wb.close()
    return path


def _all_no_match(rec):
    return "NO_MATCH_HUMAN_ONLY", None, None


def _keys_of(rec):
    return [k.strip() for k in str(rec["available_machine_keys"]).split(";") if k.strip()]


# ---------------------------------------------------------------------------
# 1. Evidence is complete — character by character against the frozen sources
# ---------------------------------------------------------------------------

def test_every_human_quote_appears_in_full(book, sources):
    ref, _ = sources
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    cells = {}
    for r in ws.iter_rows(min_row=2):
        rec = {h: r[i - 1].value for h, i in col.items()}
        if rec["human_key"]:
            cells[rec["human_key"]] = rec
    wb.close()
    checked = 0
    for h in ref["union_reference"]:
        blob = cells[h["human_key"]]["human_supporting_quotes"] or ""
        for q in h.get("supporting_quotes", []):
            assert q["quote"] in blob, (
                f"{h['human_key']}: quote is missing or abbreviated")
            checked += 1
    assert checked > 0


def test_every_machine_quote_appears_in_full_on_its_units_rows(book, sources):
    _, res = sources
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    by_unit = {}
    for r in ws.iter_rows(min_row=2):
        unit = r[col["unit_id"] - 1].value
        if unit:
            by_unit.setdefault(unit, []).append(r[col["available_machine_quotes"] - 1].value)
    wb.close()
    checked = 0
    for unit_res in res["results"]:
        u = unit_res["unit_id"]
        for th in unit_res["themes"]:
            for e in th["evidence"]:
                for blob in by_unit[u]:
                    assert e["quote"] in (blob or ""), (
                        f"{u} {th['machine_theme_id']}: quote abbreviated on a row")
                checked += 1
    assert checked == 58


def test_machine_labels_and_descriptions_appear_in_full(book, sources):
    _, res = sources
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    by_unit = {}
    for r in ws.iter_rows(min_row=2):
        unit = r[col["unit_id"] - 1].value
        if unit:
            by_unit.setdefault(unit, []).append(r[col["available_machine_labels"] - 1].value)
    wb.close()
    for unit_res in res["results"]:
        for th in unit_res["themes"]:
            for blob in by_unit[unit_res["unit_id"]]:
                assert th["label"] in (blob or "")
                assert th["one_sentence_description"] in (blob or "")


def test_machine_themes_sheet_keeps_every_quote_in_full(book, sources):
    _, res = sources
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    rows = {r[1]: r for r in wb["Machine_Themes"].iter_rows(min_row=2, values_only=True)
            if r[0]}
    wb.close()
    n = 0
    for unit_res in res["results"]:
        for th in unit_res["themes"]:
            key = cal.machine_key(unit_res["unit_id"], th["machine_theme_id"])
            row = rows[key]
            assert row[2] == th["label"]
            assert row[3] == th["one_sentence_description"]
            for e in th["evidence"]:
                assert e["quote"] in row[4]
                n += 1
    assert n == 58 and len(rows) == 30


def test_nothing_anywhere_is_truncated(book, sources):
    """
    Every model-authored text in the workbook must equal its source EXACTLY.

    An earlier version of this test also failed any cell whose length was 180 or 200,
    on the theory that those were leftover slice widths. Two genuine
    one_sentence_description values are exactly that long in extraction_results_q3.json
    and end in a full stop, so the heuristic flagged faithful text. Character-by-character
    equality against the source is both stricter and correct.
    """
    _, res = sources
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    rows = {r[1]: r for r in wb["Machine_Themes"].iter_rows(min_row=2, values_only=True)
            if r[0]}
    for sheet in ("Human_Matching", "Machine_Themes"):
        for r in wb[sheet].iter_rows(min_row=2):
            for c in r:
                if isinstance(c.value, str):
                    assert not c.value.rstrip().endswith(("...", "\u2026"))
    wb.close()

    n = 0
    for unit_res in res["results"]:
        for th in unit_res["themes"]:
            row = rows[cal.machine_key(unit_res["unit_id"], th["machine_theme_id"])]
            assert row[2] == th["label"], "label is not character-identical"
            assert row[3] == th["one_sentence_description"], (
                "description is not character-identical")
            n += 1
    assert n == 30

def test_no_cell_exceeds_the_excel_limit(book):
    assert mr.oversized_cells() == []
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    for sheet in wb.sheetnames:
        for r in wb[sheet].iter_rows():
            for c in r:
                if isinstance(c.value, str):
                    assert len(c.value) <= mr.EXCEL_CELL_LIMIT
    wb.close()


# ---------------------------------------------------------------------------
# 2. Every immutable column is sealed
# ---------------------------------------------------------------------------

def test_a_clean_build_matches_the_canonical_representation(book):
    _fill(book, _all_no_match)
    assert mr.validate(book) == []


@pytest.mark.parametrize("column", list(mr.IMMUTABLE_COLS))
def test_tampering_with_each_immutable_column_is_rejected(book, column):
    """Sheet protection can be switched off; the canonical comparison is what holds."""
    _fill(book, _all_no_match)
    assert mr.validate(book) == []
    _edit(book, "U01::C01", column, "tampered value")
    problems = mr.validate(book)
    named = any(column in p and "modified" in p for p in problems)
    reordered = any("reordered or replaced" in p for p in problems)
    assert named or reordered, (column, problems)


@pytest.mark.parametrize("column", list(mr.IMMUTABLE_COLS))
def test_deleting_the_content_of_each_immutable_column_is_rejected(book, column):
    _fill(book, _all_no_match)
    _edit(book, "U03::C02", column, None)
    assert mr.validate(book) != []


def test_a_deleted_row_is_rejected(book):
    _fill(book, _all_no_match)
    wb = openpyxl.load_workbook(book)
    wb["Human_Matching"].delete_rows(5)
    wb.save(book)
    wb.close()
    problems = mr.validate(book)
    assert any("added or deleted" in p for p in problems), problems


def test_an_added_row_is_rejected(book):
    _fill(book, _all_no_match)
    wb = openpyxl.load_workbook(book)
    ws = wb["Human_Matching"]
    n = ws.max_row + 1
    ws.cell(row=n, column=1, value="U01")
    ws.cell(row=n, column=2, value="U01::C99")
    wb.save(book)
    wb.close()
    problems = mr.validate(book)
    assert any("added" in p for p in problems), problems


def test_reordering_rows_is_rejected(book):
    _fill(book, _all_no_match)
    wb = openpyxl.load_workbook(book)
    ws = wb["Human_Matching"]
    a = [c.value for c in ws[2]]
    b = [c.value for c in ws[3]]
    for i, v in enumerate(b, start=1):
        ws.cell(row=2, column=i).value = v
    for i, v in enumerate(a, start=1):
        ws.cell(row=3, column=i).value = v
    wb.save(book)
    wb.close()
    problems = mr.validate(book)
    assert any("reordered or replaced" in p or "modified" in p for p in problems)


def test_the_relation_formula_is_sealed_too(book):
    _fill(book, _all_no_match)
    _edit(book, "U02::C03", "relation_derived", "=1+1")
    assert any("relation_derived" in p for p in mr.validate(book))


def test_only_the_three_intended_columns_are_unlocked(book):
    wb = openpyxl.load_workbook(book)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    assert ws.protection.sheet is True
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=col["unit_id"]).value:
            continue
        for h, i in col.items():
            locked = ws.cell(row=r, column=i).protection.locked
            assert locked == (h not in mr.EDITABLE), f"{h} lock state is wrong"
    wb.close()


def test_the_workbook_ships_empty_with_44_rows_and_30_themes(book):
    wb = openpyxl.load_workbook(book, read_only=True, data_only=False)
    ws = wb["Human_Matching"]
    col = _cols(ws)
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    n_machine = len([r for r in wb["Machine_Themes"].iter_rows(min_row=2,
                                                              values_only=True) if r[0]])
    wb.close()
    assert len(rows) == 44 and n_machine == 30
    for h in mr.EDITABLE:
        assert all(r[col[h] - 1] is None for r in rows), f"{h} was pre-filled"


# ---------------------------------------------------------------------------
# 3. UNCERTAIN is never a match
# ---------------------------------------------------------------------------

def _one_matched_rest_no_match(rec):
    if rec["human_key"] == "U01::C01":
        return "MATCHED", _keys_of(rec)[0], None
    return "NO_MATCH_HUMAN_ONLY", None, None


def test_an_uncertain_key_never_enters_confirmed_links(book):
    def fn(rec):
        if rec["human_key"] == "U01::C02":
            return "UNCERTAIN", _keys_of(rec)[0], "could be either theme"
        return _one_matched_rest_no_match(rec)
    _fill(book, fn)
    assert mr.validate(book) == []
    d = mr.derive(book)
    cand = _keys_of({"available_machine_keys":
                     "; ".join(d["candidate_uncertain_links"])}) or \
        list(d["candidate_uncertain_links"])
    assert d["candidate_uncertain_links"], "the candidate was not recorded"
    for k, humans in d["candidate_uncertain_links"].items():
        assert "U01::C02" in humans
        assert "U01::C02" not in d["confirmed_links"].get(k, []), (
            "an UNCERTAIN key leaked into confirmed_links")


def test_an_uncertain_key_does_not_change_recall(book):
    base_fn = _one_matched_rest_no_match
    _fill(book, base_fn)
    before = mr.derive(book)

    def fn(rec):
        if rec["human_key"] == "U01::C02":
            return "UNCERTAIN", _keys_of(rec)[0], "could be either theme"
        return base_fn(rec)
    _fill(book, fn)
    after = mr.derive(book)
    assert (after["recall_numerator_matched_human_instances"] ==
            before["recall_numerator_matched_human_instances"] == 1)
    assert after["recall_denominator_union_reference"] == 44


def test_an_uncertain_key_does_not_clear_the_adjudication_queue(book):
    def fn(rec):
        if rec["unit_id"] == "U01":
            return "UNCERTAIN", _keys_of(rec)[0], "weighing it"
        return "NO_MATCH_HUMAN_ONLY", None, None
    _fill(book, fn)
    d = mr.derive(book)
    assert len(d["machine_only_requiring_adjudication"]) == 30, (
        "a candidate key removed a machine theme from the queue")
    assert d["confirmed_links"] == {}


def test_uncertain_produces_no_confirmed_fusion_or_fragmentation(book):
    def fn(rec):
        if rec["unit_id"] == "U01":
            ks = _keys_of(rec)
            return "UNCERTAIN", "; ".join(ks[:2]), "two candidates"
        return "NO_MATCH_HUMAN_ONLY", None, None
    _fill(book, fn)
    d = mr.derive(book)
    assert d["possible_fusion_one_machine_many_human"] == {}
    assert d["possible_fragmentation_one_human_many_machine"] == {}


def test_any_uncertain_row_keeps_the_state_unresolved(book):
    _fill(book, _one_matched_rest_no_match)
    assert mr.derive(book)["resolution_state"] == "RESOLVED"

    def fn(rec):
        if rec["human_key"] == "U05::C01":
            return "UNCERTAIN", None, "genuinely ambiguous"
        return _one_matched_rest_no_match(rec)
    _fill(book, fn)
    d = mr.derive(book)
    assert d["resolution_state"] == mr.UNRESOLVED
    assert d["n_uncertain"] == 1
    assert "do not raise recall" in d["resolution_note"]


def test_uncertain_without_keys_is_allowed_with_a_reason(book):
    _fill(book, lambda rec: ("UNCERTAIN", None, "two readings are defensible"))
    assert mr.validate(book) == []
    d = mr.derive(book)
    assert d["candidate_uncertain_links"] == {}
    assert d["n_uncertain"] == 44
    assert d["recall_numerator_matched_human_instances"] == 0


def test_uncertain_without_a_reason_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "UNCERTAIN")
    assert any("UNCERTAIN requires a short reason" in p for p in mr.validate(book))


# ---------------------------------------------------------------------------
# Decision rules that must not have loosened
# ---------------------------------------------------------------------------

def test_matched_without_a_machine_key_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "MATCHED")
    assert any("MATCHED but no machine key" in p for p in mr.validate(book))


def test_no_match_with_a_machine_key_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "matched_machine_keys", "U01::M1")
    assert any("must leave matched_machine_keys empty" in p for p in mr.validate(book))


def test_an_unknown_machine_key_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "MATCHED")
    _edit(book, "U01::C01", "matched_machine_keys", "U01::M99")
    assert any("unknown machine key" in p for p in mr.validate(book))


def test_a_machine_key_from_another_unit_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "MATCHED")
    _edit(book, "U01::C01", "matched_machine_keys", "U02::M1")
    assert any("belongs to U02, not U01" in p for p in mr.validate(book))


def test_duplicate_machine_keys_are_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "MATCHED")
    _edit(book, "U01::C01", "matched_machine_keys", "U01::M1; U01::M1")
    assert any("duplicate machine keys" in p for p in mr.validate(book))


def test_an_incomplete_or_invalid_decision_is_rejected(book):
    _fill(book, _all_no_match)
    _edit(book, "U03::C02", "human_decision", None)
    assert any("no decision" in p for p in mr.validate(book))
    _edit(book, "U03::C02", "human_decision", "probably")
    assert any("is not one of" in p for p in mr.validate(book))


def test_derivation_is_refused_while_invalid(book):
    _fill(book, _all_no_match)
    _edit(book, "U01::C01", "human_decision", "MATCHED")
    with pytest.raises(mr.MatchingNotReady):
        mr.derive(book)


def test_confirmed_fusion_and_fragmentation_still_derive(book):
    def fn(rec):
        ks = _keys_of(rec)
        if rec["human_key"] == "U01::C01":
            return "MATCHED", "; ".join(ks[:2]), "one idea split in two"
        if rec["unit_id"] == "U01":
            return "MATCHED", ks[0], None
        return "NO_MATCH_HUMAN_ONLY", None, None
    _fill(book, fn)
    assert mr.validate(book) == []
    d = mr.derive(book)
    assert "U01::C01" in d["possible_fragmentation_one_human_many_machine"]
    assert any(len(v) > 1 for v in d["possible_fusion_one_machine_many_human"].values())
    assert d["resolution_state"] == "RESOLVED"


def test_v1_workbooks_are_not_overwritten():
    for name in ("Emergent_Matching_Q3_POPULATED.xlsx",
                 "Emergent_Matching_Q3_RESEARCHER.xlsx"):
        p = OUT / name
        if p.exists():
            assert p.name != mr._WB.name, "V2 must be a distinct file"
