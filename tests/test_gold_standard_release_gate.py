"""
Negative tests for the gold-standard Part-2 release gate.

The gate exists to stop a damaged or edited Part-1 workbook from unlocking the
deductive stage. A gate that has never been shown to fail is not evidence of
anything, so every rejection path is exercised here against a real workbook.

Covered:
  1. Units text altered
  2. Units turn deleted
  3. Units turns reordered
  4. Overflow_Themes row populated with no unit_id
  5. Emergent_Coding row deleted
  6. Emergent_Coding partly filled theme
  7. Quote not a literal substring of the excerpt
  8. Quote that only matches text the coder edited INTO their Units sheet
  9. Unit with no theme recorded
 10. A fully valid submission is accepted

No API calls. Each test works on a temp copy; the shipped package is untouched.
"""

import re
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import build_gold_standard_package as gs  # noqa: E402

PKG = ROOT / "analysis" / "production_evaluation" / "gold_standard_package"
RETURNED_LIVE = PKG / "Coder_A_Part1_Emergent.xlsx"

# ---------------------------------------------------------------------------
# DETERMINISTIC_CANONICAL_TEST_FIXTURE
# ---------------------------------------------------------------------------
#
# These tests once used PKG/Coder_A_Part1_Emergent.xlsx as if it were a blank issued
# template. It is not: the coders returned their completed work by saving in place, so
# that file now holds 43 real themes. Planting a "partly filled" row into slot 2 was
# therefore overwriting a slot that already had a full theme in it — nothing was partly
# filled, the gate correctly found no defect, and the test failed for the right reason.
#
# A live file carrying completed human work can never be the reference for an
# immutability check: it is exactly the thing that is allowed to change. The fixture
# below is rebuilt deterministically from the canonical sources instead —
# gold_standard_package/U*.txt (untouched) plus the sealed (unit_id, theme_slot) grid.
# It is genuinely empty, so an empty cell is genuinely empty.
CANONICAL_UNITS = [f"U{i:02d}" for i in range(1, 16)]
CANONICAL_SLOTS = 12
_TURN = re.compile(r"^\[(T\d+)\]\s+([^:]+):\s*(.*)$")


def _canonical_turns() -> list[tuple]:
    """Parse the canonical excerpt files. This is the immutable source of truth."""
    out = []
    for uid in CANONICAL_UNITS:
        src = PKG / f"{uid}.txt"
        cur = None
        for line in src.read_text(encoding="utf-8").splitlines():
            m = _TURN.match(line)
            if m:
                if cur:
                    out.append(cur)
                tid, speaker, text = m.group(1), m.group(2).strip(), m.group(3)
                cur = [uid, tid, speaker, text]
            elif cur is not None:
                # A blank line is a PARAGRAPH BREAK inside a turn. Dropping it
                # collapses the text and the fixture stops matching the source.
                cur[3] += "\n" + line
        if cur:
            out.append(cur)
    out = [[u, t, sp, tx.rstrip()] for u, t, sp, tx in out]
    return [(u, t, s, len(x.split()), x) for u, t, s, x in out]


def build_canonical_fixture(path: Path) -> Path:
    """An EMPTY, deterministic Part-1 workbook. No human content whatsoever."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws.cell(row=1, column=1, value="DETERMINISTIC_CANONICAL_TEST_FIXTURE")

    ws = wb.create_sheet("Units")
    for j, h in enumerate(["unit_id", "turn_id", "speaker", "words", "turn_text"], 1):
        ws.cell(row=1, column=j, value=h)
    for i, row in enumerate(_canonical_turns(), start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    ws = wb.create_sheet("Emergent_Coding")
    for j, h in enumerate(["unit_id", "theme_slot", "theme_label", "theme_description",
                           "supporting_quote", "relevance", "coder_notes"], 1):
        ws.cell(row=1, column=j, value=h)
    i = 2
    for uid in CANONICAL_UNITS:
        for slot in range(1, CANONICAL_SLOTS + 1):
            ws.cell(row=i, column=1, value=uid)
            ws.cell(row=i, column=2, value=slot)
            i += 1

    ws = wb.create_sheet("Overflow_Themes")
    for j, h in enumerate(["unit_id", "theme_label", "theme_description",
                           "supporting_quote", "relevance", "coder_notes"], 1):
        ws.cell(row=1, column=j, value=h)

    wb.save(path)
    wb.close()
    return path


@pytest.fixture(scope="module")
def canonical(tmp_path_factory) -> Path:
    return build_canonical_fixture(
        tmp_path_factory.mktemp("canonical") / "Coder_A_Part1_Emergent.xlsx")


pytestmark = pytest.mark.skipif(
    not (PKG / "U01.txt").exists(),
    reason="gold-standard package not built in this checkout",
)


def _unit_first_words(wb, unit_id: str, n: int = 6) -> str:
    """A genuine quote: the opening words of a real turn in that unit."""
    for r in wb["Units"].iter_rows(min_row=2):
        if r[0].value == unit_id and (r[4].value or "").strip():
            return " ".join((r[4].value or "").split()[:n])
    return "x"


def _complete(tmp_path: Path, canonical: Path = None) -> Path:
    """A valid Part-1 submission: one complete theme for every unit."""
    dst = tmp_path / "Coder_A_Part1_Emergent.xlsx"
    shutil.copy(canonical, dst)
    wb = load_workbook(dst)
    ws = wb["Emergent_Coding"]
    for r in ws.iter_rows(min_row=2):
        if r[1].value == 1:
            r[2].value = "Gender framed as irrelevant"
            r[3].value = "Participants deny that gender shapes what they eat."
            r[4].value = _unit_first_words(wb, r[0].value)
            r[5].value = "central"
    wb.save(dst)
    return dst


@pytest.fixture()
def gate(tmp_path, monkeypatch, canonical):
    """Point the gate's returned/withheld/output dirs at a temp workspace."""
    returned = tmp_path / "returned"
    withheld = tmp_path / "withheld"
    out = tmp_path / "released"
    for d in (returned, withheld, out):
        d.mkdir()
    shutil.copy(ROOT / "analysis" / "production_evaluation"
                / "gold_standard_part2_withheld" / "Coder_A_Part2_Deductive.xlsx",
                withheld / "Coder_A_Part2_Deductive.xlsx")
    shutil.copy(canonical, out / "Coder_A_Part1_Emergent.xlsx")
    monkeypatch.setattr(gs, "_RETURNED_DIR", returned)
    monkeypatch.setattr(gs, "_WITHHELD_DIR", withheld)
    monkeypatch.setattr(gs, "_PKG_DIR", out)
    return returned


def _submit(gate_dir: Path, path: Path) -> int:
    shutil.copy(path, gate_dir / "Coder_A_Part1_Emergent.xlsx")
    return gs.release_part2("A")


# --- 10: the happy path must actually pass -------------------------------

def test_valid_submission_is_released(gate, tmp_path, canonical):
    assert _submit(gate, _complete(tmp_path, canonical)) == 0
    assert (gs._PKG_DIR / "Coder_A_Part2_Deductive.xlsx").exists()


# --- 1: Units text altered ------------------------------------------------

def test_altered_units_text_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    ws = wb["Units"]
    ws.cell(row=2, column=5).value = "TAMPERED " + (ws.cell(row=2, column=5).value or "")
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "Units row 2 altered" in capsys.readouterr().out
    assert not (gs._PKG_DIR / "Coder_A_Part2_Deductive.xlsx").exists()


# --- 2: Units turn deleted -------------------------------------------------

def test_deleted_units_turn_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    wb["Units"].delete_rows(3)
    wb.save(p)
    assert _submit(gate, p) == 3
    out = capsys.readouterr().out
    assert "turn deleted or added" in out


# --- 3: Units turns reordered ---------------------------------------------

def test_reordered_units_turns_are_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    ws = wb["Units"]
    a = [c.value for c in ws[2]]
    b = [c.value for c in ws[3]]
    for i, v in enumerate(b, start=1):
        ws.cell(row=2, column=i).value = v
    for i, v in enumerate(a, start=1):
        ws.cell(row=3, column=i).value = v
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "Units row" in capsys.readouterr().out


# --- 4: overflow populated without unit_id --------------------------------

def test_overflow_without_unit_id_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    wb["Overflow_Themes"].append(
        [None, "An extra theme", "Its description", "some quote", "central", ""])
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "unit_id is empty" in capsys.readouterr().out


def test_overflow_fully_blank_row_is_allowed(gate, tmp_path, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    wb["Overflow_Themes"].append([None, None, None, None, None, None])
    wb.save(p)
    assert _submit(gate, p) == 0


# --- 5-7: coding-grid and quote integrity ---------------------------------

def test_deleted_coding_row_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    wb["Emergent_Coding"].delete_rows(5)
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "rows added or deleted" in capsys.readouterr().out


def test_partly_filled_theme_is_rejected(gate, tmp_path, capsys, canonical):
    """
    The planted row must be genuinely empty first. Against the live returned workbook
    slot 2 already held a complete human theme, so overwriting its label left a fully
    populated row and there was nothing partly filled to detect.
    """
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    planted = False
    for r in wb["Emergent_Coding"].iter_rows(min_row=2):
        if r[1].value == 2:
            assert all((c.value or "") == "" for c in r[2:6]), (
                "slot 2 is not empty in the fixture — the test would be vacuous")
            r[2].value = "Only a label"
            planted = True
            break
    assert planted, "no theme_slot 2 row exists in the fixture"
    wb.save(p)
    assert _submit(gate, p) == 3
    out = capsys.readouterr().out
    assert "partly filled theme" in out
    assert "theme_description" in out and "supporting_quote" in out


def test_fabricated_quote_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    for r in wb["Emergent_Coding"].iter_rows(min_row=2):
        if r[1].value == 1:
            r[4].value = "this sentence appears in no excerpt anywhere"
            break
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "not a literal substring" in capsys.readouterr().out


# --- 8: quotes are checked against the ISSUED copy, not the returned one ---

def test_quote_matching_only_self_edited_units_text_is_rejected(gate, tmp_path, capsys, canonical):
    """
    The decisive case. A coder edits their own Units sheet to contain a sentence,
    then quotes it. Validating against the returned sheet would accept this;
    validating against the issued copy must not.
    """
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    forged = "a sentence that was never in the original excerpt"
    ws = wb["Units"]
    target_unit = ws.cell(row=2, column=1).value
    ws.cell(row=2, column=5).value = (ws.cell(row=2, column=5).value or "") + " " + forged
    for r in wb["Emergent_Coding"].iter_rows(min_row=2):
        if r[0].value == target_unit and r[1].value == 1:
            r[4].value = forged
            break
    wb.save(p)
    assert _submit(gate, p) == 3
    out = capsys.readouterr().out
    assert "Units row 2 altered" in out
    assert "not a literal substring" in out


# --- 9: every unit needs at least one theme -------------------------------

def test_unit_with_no_theme_is_rejected(gate, tmp_path, capsys, canonical):
    p = _complete(tmp_path, canonical)
    wb = load_workbook(p)
    ws = wb["Emergent_Coding"]
    first_unit = ws.cell(row=2, column=1).value
    for r in ws.iter_rows(min_row=2):
        if r[0].value == first_unit:
            for c in r[2:6]:
                c.value = None
    wb.save(p)
    assert _submit(gate, p) == 3
    assert "no theme recorded" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# The fixture is faithful, and it is not the coders' work
# ---------------------------------------------------------------------------

def test_canonical_fixture_reproduces_the_issued_units_grid(canonical):
    """
    The fixture is only trustworthy if it reconstructs the real Units grid exactly.
    Compared against the live returned workbook, whose Units sheet was independently
    verified to reconcile with the untouched U*.txt sources.
    """
    if not RETURNED_LIVE.exists():
        pytest.skip("returned workbook not present")
    fx = load_workbook(canonical, read_only=True, data_only=True)
    lv = load_workbook(RETURNED_LIVE, read_only=True, data_only=True)
    a = [tuple(r) for r in fx["Units"].iter_rows(min_row=1, values_only=True)]
    b = [tuple(r) for r in lv["Units"].iter_rows(min_row=1, values_only=True)]
    fx.close(); lv.close()
    assert len(a) == len(b), f"fixture has {len(a)} Units rows, issued has {len(b)}"
    for i, (x, y) in enumerate(zip(a, b), start=1):
        assert x == y, f"Units row {i} differs:\n  fixture={x}\n  issued ={y}"


def test_canonical_fixture_grid_matches_the_sealed_shape(canonical):
    fx = load_workbook(canonical, read_only=True, data_only=True)
    grid = [(r[0], r[1]) for r in
            fx["Emergent_Coding"].iter_rows(min_row=2, values_only=True)]
    fx.close()
    assert grid == [(u, s) for u in CANONICAL_UNITS
                    for s in range(1, CANONICAL_SLOTS + 1)]
    assert len(grid) == 180


def test_canonical_fixture_carries_no_human_coding(canonical):
    """It must be EMPTY. A fixture with content cannot test 'partly filled'."""
    fx = load_workbook(canonical, read_only=True, data_only=True)
    filled = [r for r in fx["Emergent_Coding"].iter_rows(min_row=2, values_only=True)
              if any((c or "") != "" for c in r[2:])]
    fx.close()
    assert filled == [], f"{len(filled)} fixture rows carry content"


def test_the_gate_never_reads_the_coders_returned_workbook(gate, tmp_path, canonical,
                                                           capsys):
    """
    Immutable content is compared against the canonical fixture, never against a live
    file that legitimately contains completed human work.
    """
    before = RETURNED_LIVE.read_bytes() if RETURNED_LIVE.exists() else None
    p = _complete(tmp_path, canonical)
    assert _submit(gate, p) == 0
    if before is not None:
        assert RETURNED_LIVE.read_bytes() == before, (
            "the coders' returned workbook was touched")


def test_acceptance_and_rejection_are_both_exercised(gate, tmp_path, canonical,
                                                     capsys):
    """A gate that only ever accepts, or only ever rejects, proves nothing."""
    ok = _complete(tmp_path, canonical)
    assert _submit(gate, ok) == 0
    capsys.readouterr()

    bad = _complete(tmp_path / "bad", canonical) if (tmp_path / "bad").exists() else None
    (tmp_path / "bad").mkdir(exist_ok=True)
    bad = _complete(tmp_path / "bad", canonical)
    wb = load_workbook(bad)
    ws = wb["Units"]
    ws.cell(row=2, column=5).value = "TAMPERED " + (ws.cell(row=2, column=5).value or "")
    wb.save(bad); wb.close()
    assert _submit(gate, bad) == 3
    assert "altered" in capsys.readouterr().out
