"""
OCA-001 must be blind and must carry enough context to distinguish A.1 from A.3.

Three isolated quote fragments could not separate the two subthemes: A.1 and A.3
differ in STANCE, and the sentence that establishes stance sat outside the fragment.
The form therefore shows full turns plus the minimum preceding context.

No API calls.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

OUT = ROOT / "analysis" / "production_evaluation"
FORM = OUT / "open_coding_adjudication" / "OCA-001_adjudication.xlsx"
SEALED = OUT / "gold_standard_sealed" / "open_coding_item_mapping.json"


@pytest.fixture(scope="module")
def text():
    if not FORM.exists():
        pytest.skip("form not generated")
    wb = openpyxl.load_workbook(FORM)
    ws = wb["Adjudication"]
    blob = " ".join(str(ws.cell(row=r, column=c).value or "")
                    for r in range(1, ws.max_row + 1) for c in (1, 2))
    wb.close()
    return blob


@pytest.fixture(scope="module")
def sealed():
    return json.loads(SEALED.read_text(encoding="utf-8"))


def test_no_roster_name_appears(text, sealed):
    """Word-boundary, not substring: 'marked' must not be read as the name 'Mark'."""
    offenders = {nm: len(re.findall(rf"\b{re.escape(nm)}\b", text, re.I))
                 for nm in sealed["roster_names_redacted"]}
    assert all(n == 0 for n in offenders.values()), f"roster names present: {offenders}"
    assert sealed["roster_names_redacted"], "the scan must run against a real roster"


def test_no_provenance_leaks(text):
    for leak in ("fg4", "demograph", "enriched", "synthetic", "macho", "run01",
                 "recall", "precision", "condition"):
        assert leak not in text.lower(), f"{leak!r} leaked onto the form"


def test_it_does_not_claim_the_passages_are_verbatim(text):
    assert "NOT verbatim" in text
    assert "minimal redaction" in text


def test_full_turns_and_context_are_present(text, sealed):
    item = sealed["items"][0]
    assert item["turns_shown"] == ["T020", "T021", "T022", "T023", "T026", "T027"]
    assert item["turns_cited"] == ["T021", "T023", "T027"]
    for t in item["turns_shown"]:
        assert t in text, f"{t} missing from the form"
    context_only = set(item["turns_shown"]) - set(item["turns_cited"])
    assert context_only, "no context turns were included"


def test_the_stance_sentence_survives(text):
    """
    T021 opens by acknowledging an influence before describing the household. That
    opening is what separates A.1 from A.3, and the original three-quote fragment
    omitted it.
    """
    assert "probably does, doesn't it" in text
    assert "hadn't really thought about it before you asked" in text


def test_both_boundary_definitions_are_shown(text):
    assert "acknowledged a meaningful influence of gender" in text     # A.1
    assert "rejected or were unsure of the idea of a gender influence" in text  # A.3
    assert "Does influence" in text and "No influence, but" in text


def test_the_question_remains_about_a1(text):
    assert "supports subtheme A.1" in text


def test_the_three_options_are_both_visible_and_enforced():
    """
    Visible as text AND enforced by the dropdown. A dropdown alone is invisible to a
    reviewer who never clicks the cell, and to anyone reading the file as text.
    """
    wb = openpyxl.load_workbook(FORM)
    ws = wb["Adjudication"]
    blob = " ".join(str(ws.cell(row=r, column=c).value or "")
                    for r in range(1, ws.max_row + 1) for c in (1, 2))
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    wb.close()
    for opt in ("SUPPORTS_A1", "DOES_NOT_SUPPORT_A1", "UNCERTAIN"):
        assert opt in blob, f"{opt} is not visible as text"
        assert any(opt in f for f in formulas), f"{opt} is not in the dropdown"


def test_speaker_labels_are_generic(text):
    assert re.search(r"\bParticipant \d\b", text)
    assert "Moderator" in text


def test_sealed_side_keeps_original_and_presented_text(sealed):
    item = sealed["items"][0]
    assert sealed["redaction_rule"]
    for t in item["turn_provenance"]:
        assert t["original_text"] and t["presented_text"]
        assert t["original_sha256"] and t["presented_sha256"]
        assert len(t["original_sha256"]) == 64
        if t["redactions"]:
            assert t["original_text"] != t["presented_text"]
            assert t["original_sha256"] != t["presented_sha256"]
    assert item["internal_id"] == "FG4-DEMO-R01-A1"
    assert item["form_item_id"] == "OCA-001"


def test_redaction_preserves_reference_between_speakers(sealed):
    """Every name maps to that speaker's own label, not to one generic phrase."""
    subs = [s for t in sealed["items"][0]["turn_provenance"] for s in t["redactions"]]
    assert subs
    targets = {s.split(" -> ")[1].rstrip("'s") for s in subs}
    assert len(targets) > 1, (
        "all names collapsed to a single phrase — who-refers-to-whom is lost")
    assert all(re.fullmatch(r"Participant \d", t) for t in targets), targets
